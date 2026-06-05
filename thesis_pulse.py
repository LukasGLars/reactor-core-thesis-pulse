# -*- coding: utf-8 -*-
"""
Reactor Core Thesis Pulse v2.0
Daily thesis monitoring for 8-position portfolio.
Runs via GitHub Actions — sends email with raw data + recession tracker.
"""
import requests, json, os, sys, smtplib, time, csv
from datetime import date
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

sys.stdout.reconfigure(encoding="utf-8")

# ── CONFIG ─────────────────────────────────────────────────
EDGAR_HEADERS     = {"User-Agent": "ThesisPulse research@example.com"}
FRED_API_KEY      = os.environ.get("FRED_API_KEY", "")
ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY", "")
EMAIL_ADDRESS     = os.environ.get("EMAIL_ADDRESS", "")
EMAIL_PASSWORD    = os.environ.get("EMAIL_PASSWORD", "")
RECIPIENT_EMAIL   = os.environ.get("RECIPIENT_EMAIL", EMAIL_ADDRESS)
SMTP_SERVER       = os.environ.get("SMTP_SERVER", "smtp.gmail.com")
SMTP_PORT         = int(os.environ.get("SMTP_PORT", "587"))
WGC_EMAIL         = os.environ.get("WGC_EMAIL", "")
WGC_PASSWORD      = os.environ.get("WGC_PASSWORD", "")
# Legacy cookie fallback (kept for local dev — login is preferred in Actions)
WGC_API_AUTH      = os.environ.get("WGC_API_AUTH", "")
WGC_AUTH_COOKIE   = os.environ.get("WGC_AUTH_COOKIE", "")
WGC_AUTH_SESSION  = os.environ.get("WGC_AUTH_SESSION", "")
WGC_XSRF          = os.environ.get("WGC_XSRF", "")

_dir = os.path.dirname(os.path.abspath(__file__))

# ── PORTFOLIO WEIGHTS (v3) ──────────────────────────────────
V3_WEIGHTS = {
    "gold": 0.25, "silver": 0.10, "lly": 0.15, "wmt": 0.15,
    "vrt":  0.10, "ccj":   0.10, "avgo": 0.09, "jnj": 0.06,
}
with open(os.path.join(_dir, "thesis_v3.md"),       encoding="utf-8") as f: THESIS_DOC       = f.read()
with open(os.path.join(_dir, "invalidation_v3.md"), encoding="utf-8") as f: INVALIDATION_DOC = f.read()

# ── FRED ───────────────────────────────────────────────────
def fred_url(series_id):
    return (f"https://api.stlouisfed.org/fred/series/observations"
            f"?series_id={series_id}&api_key={FRED_API_KEY}&file_type=json&sort_order=asc")

def _make_session(max_retries=5, backoff=0.5):
    session = requests.Session()
    retry = Retry(
        total=max_retries,
        backoff_factor=backoff,
        status_forcelist=[429, 500, 502, 503, 504],
        allowed_methods=frozenset(["GET", "HEAD"]),
        raise_on_status=False,
    )
    adapter = HTTPAdapter(max_retries=retry)
    session.mount("https://", adapter)
    session.mount("http://", adapter)
    return session

def fred_latest(series_id):
    session = _make_session()
    for attempt in range(1, 6):
        try:
            r = session.get(
                fred_url(series_id),
                headers={"User-Agent": "thesis-pulse/1.0"},
                timeout=30,
            )
            r.raise_for_status()
            obs = r.json().get("observations", [])
            rows = [(o["date"], float(o["value"])) for o in obs
                    if o.get("value") not in (".", "")]
            if len(rows) >= 2:
                return rows[-1][1], rows[-2][1], rows[-1][0]
            if rows:
                return rows[-1][1], None, rows[-1][0]
            return None, None, None
        except Exception as e:
            wait = backoff * (2 ** (attempt - 1)) if (backoff := 0.5) else 0.5
            print(f"  FRED {series_id} retry {attempt}/5: {e}")
            time.sleep(wait)
    return None, None, None

def fred_recent(series_id, lookback=20):
    """Returns (latest, prev, lookback_val, latest_date) — lookback in business days."""
    session = _make_session()
    for attempt in range(1, 6):
        try:
            r = session.get(
                fred_url(series_id),
                headers={"User-Agent": "thesis-pulse/1.0"},
                timeout=30,
            )
            r.raise_for_status()
            obs = r.json().get("observations", [])
            rows = [(o["date"], float(o["value"])) for o in obs
                    if o.get("value") not in (".", "")]
            if not rows:
                return None, None, None, None
            latest_val, latest_date = rows[-1][1], rows[-1][0]
            prev_val   = rows[-2][1]  if len(rows) >= 2        else None
            lb_val     = rows[-lookback][1] if len(rows) >= lookback else rows[0][1]
            return latest_val, prev_val, lb_val, latest_date
        except Exception as e:
            wait = 0.5 * (2 ** (attempt - 1))
            print(f"  FRED {series_id} retry {attempt}/5: {e}")
            time.sleep(wait)
    return None, None, None, None

# ── YAHOO FINANCE ──────────────────────────────────────────
def yahoo_history(symbol):
    url = f"https://query1.finance.yahoo.com/v8/finance/chart/{symbol}?interval=1d&range=1y"
    try:
        r = requests.get(url, headers={"User-Agent": "Mozilla/5.0"}, timeout=10)
        if r.status_code != 200:
            return None
        closes = r.json()["chart"]["result"][0]["indicators"]["quote"][0]["close"]
        closes = [c for c in closes if c is not None]
        if len(closes) < 2:
            return None
        curr     = closes[-1]
        prev     = closes[-2]
        high_52w = max(closes)
        ma_200 = sum(closes[-200:]) / min(200, len(closes))
        import statistics as _stats
        med_3w      = _stats.median(closes[-15:])    if len(closes) >= 15 else None
        prev_med_3w = _stats.median(closes[-30:-15]) if len(closes) >= 30 else None
        volumes = r.json()["chart"]["result"][0]["indicators"]["quote"][0].get("volume", [])
        volumes = [v for v in volumes if v is not None]
        avg_vol = sum(volumes[-20:]) / min(20, len(volumes)) if len(volumes) >= 2 else None
        rel_vol = volumes[-1] / avg_vol if avg_vol and avg_vol > 0 else None

        return {
            "price":        curr,
            "prev_close":   prev,
            "closes_30":    closes[-30:],
            "chg_1d":       (curr - prev) / prev * 100,
            "pts_1d":       curr - prev,
            "chg_1m":       (curr - closes[-22]) / closes[-22] * 100 if len(closes) >= 22 else None,
            "pts_4w":       curr - closes[-22] if len(closes) >= 22 else None,
            "pts_8w":       closes[-22] - closes[-44] if len(closes) >= 44 else None,
            "chg_3m":       (curr - closes[-63]) / closes[-63] * 100 if len(closes) >= 63 else None,
            "high_52w":     high_52w,
            "dd_52w":       (curr - high_52w) / high_52w * 100,
            "ma_200":       ma_200,
            "vs_ma_200":    (curr - ma_200) / ma_200 * 100,
            "med_3w":       med_3w,
            "prev_med_3w":  prev_med_3w,
            "rel_vol":      rel_vol,
        }
    except Exception:
        return None

# ── EDGAR ──────────────────────────────────────────────────
CIKS = {
    "MSFT":  "CIK0000789019",
    "GOOGL": "CIK0001652044",
    "AMZN":  "CIK0001018724",
    "META":  "CIK0001326801",
    "VRT":   "CIK0001674101",
    "AVGO":  "CIK0001730168",
    "WMT":   "CIK0000104169",
    "LLY":   "CIK0000059478",
    "JNJ":   "CIK0000200406",
    "NVDA":  "CIK0001045810",
}

def edgar_concept(ticker, concept, namespace="us-gaap"):
    cik = CIKS.get(ticker)
    if not cik:
        return None, None
    url = f"https://data.sec.gov/api/xbrl/companyconcept/{cik}/{namespace}/{concept}.json"
    try:
        r = requests.get(url, headers=EDGAR_HEADERS, timeout=15)
        if r.status_code != 200:
            return None, None
        units = r.json().get("units", {})
        entries = units.get("USD", units.get("USD/shares", units.get("shares", units.get("pure", []))))
        filings = [e for e in entries if e.get("form") in ("10-K","10-Q","20-F","40-F")]
        filings.sort(key=lambda x: x.get("end",""), reverse=True)
        seen, unique = set(), []
        for f in filings:
            key = f.get("end")
            if key not in seen:
                seen.add(key)
                unique.append(f)
        if not unique:
            return None, None
        curr = unique[0]
        curr_fp, curr_form, curr_end = curr.get("fp",""), curr.get("form",""), curr.get("end","")
        prior = None
        for f in unique[1:]:
            if (f.get("fp") == curr_fp and f.get("form") == curr_form
                    and f.get("end","")[:4] < curr_end[:4]):
                prior = f
                break
        if prior is None:
            for f in unique[1:]:
                if f.get("form") == curr_form:
                    prior = f
                    break
        return curr, prior
    except Exception:
        return None, None

def edgar_revenue(ticker):
    from datetime import timedelta
    cutoff = (date.today() - timedelta(days=548)).isoformat()
    curr, prev = edgar_concept(ticker, "RevenueFromContractWithCustomerExcludingAssessedTax")
    if curr and curr.get("end","") >= cutoff:
        return curr, prev
    curr2, prev2 = edgar_concept(ticker, "Revenues")
    if curr2 and curr2.get("end","") >= cutoff:
        return curr2, prev2
    if curr and curr2:
        return (curr, prev) if curr.get("end","") > curr2.get("end","") else (curr2, prev2)
    return (curr, prev) if curr else (curr2, prev2)

# ── CENTRAL BANK GOLD (IMF IFS primary, WGC cookie fallback) ───────────────
def imf_central_bank_gold():
    """
    Query IMF IFS series RAFAGOLDV (monthly gold holdings, fine troy oz) for all countries.
    Computes world net purchase TTM vs prior 12m — same underlying data as WGC.
    No auth required. Returns (ttm_tonnes, prev_ttm_tonnes, latest_date_str, lag_days).
    """
    TROY_OZ_PER_TONNE = 32150.75
    # Try HTTPS SDMX first, then data.imf.org JSON API
    candidates = [
        "https://dataservices.imf.org/REST/SDMX_JSON.svc/CompactData/IFS/M..RAFAGOLDV?startPeriod=2022-01",
        "https://data.imf.org/api/SDMX/1.0/rest/data/IFS/M..RAFAGOLDV?startPeriod=2022-01&format=jsondata",
    ]
    r = None
    for url in candidates:
        try:
            resp = requests.get(url, headers={"User-Agent": "thesis-pulse/1.0"}, timeout=60)
            if resp.status_code == 200:
                r = resp
                break
            print(f"  IMF gold: {url[30:70]} -> {resp.status_code}")
        except Exception as ex:
            print(f"  IMF gold: {url[30:70]} -> {type(ex).__name__}")
    if r is None:
        return None, None, None, None
    try:
        raw = r.json()
        levels = {}  # {country: {period: float}}

        # Format A: classic SDMX_JSON (dataservices.imf.org)
        if "CompactData" in raw:
            series_raw = raw["CompactData"]["DataSet"].get("Series", [])
            if isinstance(series_raw, dict):
                series_raw = [series_raw]
            for s in series_raw:
                country = s.get("@REF_AREA", "")
                obs = s.get("Obs", [])
                if isinstance(obs, dict):
                    obs = [obs]
                for o in obs:
                    period = o.get("@TIME_PERIOD", "")
                    val_s  = o.get("@OBS_VALUE")
                    if val_s is None:
                        continue
                    try:
                        levels.setdefault(country, {})[period] = float(val_s)
                    except ValueError:
                        pass

        # Format B: SDMX 2.1 jsondata (data.imf.org)
        elif "data" in raw:
            ds = raw["data"]
            # dimensions order: FREQ, REF_AREA, INDICATOR, ...
            dim_ids = [d["id"] for d in ds.get("structure", {}).get("dimensions", {}).get("series", [])]
            ref_area_idx = dim_ids.index("REF_AREA") if "REF_AREA" in dim_ids else 1
            time_periods  = [p["id"] for p in ds.get("structure", {}).get("dimensions", {}).get("observation", [{}])[0].get("values", [])]
            area_values   = ds.get("structure", {}).get("dimensions", {}).get("series", [{}])[ref_area_idx].get("values", [])
            for key_str, obs_dict in ds.get("dataSets", [{}])[0].get("series", {}).items():
                parts = key_str.split(":")
                area_idx = int(parts[ref_area_idx]) if len(parts) > ref_area_idx else 0
                country = area_values[area_idx]["id"] if area_idx < len(area_values) else key_str
                for t_idx_s, val_list in obs_dict.get("observations", {}).items():
                    t_idx = int(t_idx_s)
                    val   = val_list[0] if val_list else None
                    if val is None or t_idx >= len(time_periods):
                        continue
                    try:
                        levels.setdefault(country, {})[time_periods[t_idx]] = float(val)
                    except (TypeError, ValueError):
                        pass

        if not levels:
            return None, None, None, None

        # All periods in ascending order
        all_periods = sorted(set(p for c in levels.values() for p in c))
        if len(all_periods) < 14:
            return None, None, None, None

        # Monthly world net change: sum(level[t] - level[t-1]) across all countries
        monthly_change = {}
        for i in range(1, len(all_periods)):
            t0, t1 = all_periods[i-1], all_periods[i]
            net = 0.0
            for cdata in levels.values():
                v0 = cdata.get(t0)
                v1 = cdata.get(t1)
                if v0 is not None and v1 is not None:
                    net += v1 - v0
            monthly_change[t1] = net

        periods = sorted(monthly_change)
        if len(periods) < 13:
            return None, None, None, None

        ttm_periods  = periods[-12:]
        prev_periods = periods[-24:-12]
        latest_date  = ttm_periods[-1]

        ttm_oz  = sum(monthly_change[p] for p in ttm_periods)
        prev_oz = sum(monthly_change[p] for p in prev_periods) if len(prev_periods) == 12 else None

        ttm_t  = round(ttm_oz  / TROY_OZ_PER_TONNE, 0)
        prev_t = round(prev_oz / TROY_OZ_PER_TONNE, 0) if prev_oz is not None else None

        from datetime import date as _date
        lag = (_date.today() - _date.fromisoformat(latest_date + "-01")).days

        return ttm_t, prev_t, latest_date, lag

    except Exception as e:
        print(f"  IMF gold error: {e}")
        return None, None, None, None


def _wgc_login():
    """
    Log in to user.gold.org and return a requests.Session with auth cookies set.
    Uses WGC_EMAIL + WGC_PASSWORD env vars (primary) or legacy cookie env vars (fallback).
    Returns session on success, None on failure.
    """
    import re as _re
    hdrs = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/124",
        "Accept":     "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Referer":    "https://www.gold.org/",
    }

    # ── primary: automated login ──────────────────────────────
    if WGC_EMAIL and WGC_PASSWORD:
        s = requests.Session()
        try:
            # GET login page — captures XSRF-TOKEN + wgcAuth_session
            r = s.get("https://user.gold.org/login", headers=hdrs, timeout=15)
            if r.status_code != 200:
                print(f"  WGC login GET: {r.status_code}")
                return None
            token = _re.search(r'name="_token"\s+value="([^"]+)"', r.text)
            if not token:
                print("  WGC login: _token not found")
                return None
            # POST credentials
            payload = {
                "_token":   token.group(1),
                "ema":      WGC_EMAIL,
                "password": WGC_PASSWORD,
                "remember": "1",
                "log":      "1",
            }
            post_hdrs = dict(hdrs)
            post_hdrs.update({
                "Content-Type": "application/x-www-form-urlencoded",
                "Referer":      "https://user.gold.org/login",
                "Origin":       "https://user.gold.org",
            })
            r2 = s.post("https://user.gold.org/log", data=payload,
                        headers=post_hdrs, timeout=15, allow_redirects=True)
            # Success: session now has wgcApiAuth_cookie + wgcAuth_cookie
            if "wgcApiAuth_cookie" in s.cookies or r2.status_code in (200, 302):
                print("  WGC login: OK")
                return s
            print(f"  WGC login POST: {r2.status_code} — auth cookie not set")
            return None
        except Exception as e:
            print(f"  WGC login error: {e}")
            return None

    # ── fallback: legacy static cookies ──────────────────────
    if WGC_AUTH_SESSION:
        s = requests.Session()
        s.cookies.set("wgcApiAuth_cookie", WGC_API_AUTH,  domain=".gold.org")
        s.cookies.set("wgcAuth_cookie",    WGC_AUTH_COOKIE, domain=".gold.org")
        s.cookies.set("wgcAuth_session",   WGC_AUTH_SESSION, domain=".gold.org")
        s.cookies.set("XSRF-TOKEN",        WGC_XSRF,      domain=".gold.org")
        return s

    return None


def wgc_central_banks():
    """
    Download WGC monthly central bank gold changes (IFS source, ~2-month lag).
    Logs in via WGC_EMAIL/WGC_PASSWORD (auto-refresh) or falls back to static cookies.
    Returns (ttm_tonnes, prev_ttm_tonnes, latest_date_str, lag_days).
    """
    import io
    try:
        import openpyxl
    except ImportError:
        print("  openpyxl not installed — skipping WGC")
        return None, None, None, None
    from datetime import date as _date

    session = _wgc_login()
    if session is None:
        return None, None, None, None

    _MONTHS = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
    hdrs = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        "Referer":    "https://www.gold.org/goldhub/data/gold-reserves-by-country",
        "Accept":     "application/octet-stream,*/*",
    }

    content = None
    today = _date.today()
    for delta in range(4):
        m = today.month - 1 - delta
        y = today.year
        while m < 0:
            m += 12
            y -= 1
        url = (f"https://www.gold.org/download/file/7741/"
               f"Changes_latest_as_of_{_MONTHS[m]}{y}_IFS.xlsx")
        try:
            r = session.get(url, headers=hdrs, timeout=30)
            if r.status_code == 200 and len(r.content) > 50000:
                content = r.content
                break
            print(f"  WGC {_MONTHS[m]}{y}: status {r.status_code}")
        except Exception as e:
            print(f"  WGC {_MONTHS[m]}{y}: {e}")

    if content is None:
        return None, None, None, None

    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb["Monthly"]
        header_row = list(ws.iter_rows(min_row=8, max_row=8, values_only=True))[0]
        date_cols = [(i, v) for i, v in enumerate(header_row) if hasattr(v, 'year')]
        if not date_cols:
            return None, None, None, None

        all_rows = list(ws.iter_rows(min_row=9, values_only=True))

        last_ci = date_cols[0][0]
        for ci, _ in date_cols:
            has_data = any(
                isinstance(row[ci], (int, float)) and row[ci] != 0
                for row in all_rows if len(row) > ci
            )
            if has_data:
                last_ci = ci

        populated = [(ci, dt) for ci, dt in date_cols if ci <= last_ci]
        if len(populated) < 13:
            return None, None, None, None

        ttm_cols  = populated[-12:]
        prev_cols = populated[-24:-12]
        latest_dt = ttm_cols[-1][1]

        def _sum(cols):
            total = 0.0
            for ci, _ in cols:
                for row in all_rows:
                    v = row[ci] if len(row) > ci else None
                    if isinstance(v, (int, float)):
                        total += float(v)
            return total

        ttm  = _sum(ttm_cols)
        prev = _sum(prev_cols) if len(prev_cols) == 12 else None
        latest_date_str = latest_dt.strftime("%Y-%m")
        lag = (_date.today() - latest_dt.date()).days

        return round(ttm, 0), round(prev, 0) if prev is not None else None, latest_date_str, lag

    except Exception as e:
        print(f"  WGC parse error: {e}")
        return None, None, None, None


# ── URANIUM ────────────────────────────────────────────────
def get_uranium():
    """IMF uranium price via FRED (PURANUSDM). Monthly, ~6wk lag. Zero scraping risk."""
    val, prev, as_of = fred_latest("PURANUSDM")
    return val, prev, as_of

# ── OIL TERM SPREAD ────────────────────────────────────────
def get_oil_term_spread():
    """
    WTI spot (FRED DCOILWTICO) vs 12-month forward (Yahoo dynamic contract).
    Positive spread = backwardation = physical supply stress.
    Thesis signal: spread >$20 STRESS | $10-20 ELEVATED | $0-10 NORMAL | <$0 CONTANGO.
    Returns spread_chg_4w: 4-week change in the spread (spot_4w - fwd_4w).
    """
    spot, _, spot_4w, spot_date = fred_recent("DCOILWTICO", lookback=20)
    _, _, spot_8w, _            = fred_recent("DCOILWTICO", lookback=40)

    MONTH_CODES = "FGHJKMNQUVXZ"
    today = date.today()
    target_month = today.month
    target_year  = today.year + 1
    code = MONTH_CODES[target_month - 1]
    fwd_ticker = f"CL{code}{str(target_year)[2:]}.NYM"

    fwd = fwd_prev = fwd_closes = None
    url = f"https://query1.finance.yahoo.com/v8/finance/chart/{fwd_ticker}?interval=1d&range=3mo"
    try:
        r = requests.get(url, headers={"User-Agent": "Mozilla/5.0"}, timeout=10)
        if r.status_code == 200:
            closes = r.json()["chart"]["result"][0]["indicators"]["quote"][0]["close"]
            closes = [c for c in closes if c is not None]
            fwd       = closes[-1]  if closes           else None
            fwd_prev  = closes[-2]  if len(closes) >= 2 else None
            fwd_closes = closes[-31:] if len(closes) >= 2 else None
    except Exception:
        pass

    spot_rows  = fred_last_n("DCOILWTICO", n=31)
    spot_vals  = [v for _, v in spot_rows]

    spread      = round(spot          - fwd,      2) if spot      and fwd      else None
    spread_prev = round(spot_vals[-2] - fwd_prev, 2) if len(spot_vals) >= 2 and fwd_prev else None

    import statistics as _stats
    spread_med_3w = spread_prev_med_3w = None
    if fwd_closes and len(spot_vals) >= 30 and len(fwd_closes) >= 30:
        n = min(len(spot_vals), len(fwd_closes))
        paired = [s - f for s, f in zip(spot_vals[-n:], fwd_closes[-n:])]
        spread_med_3w      = round(_stats.median(paired[-15:]),    2) if len(paired) >= 15 else None
        spread_prev_med_3w = round(_stats.median(paired[-30:-15]), 2) if len(paired) >= 30 else None

    return spot, fwd, spread, spread_prev, spread_med_3w, spread_prev_med_3w, spot_date, fwd_ticker

# ── HELPERS ────────────────────────────────────────────────
def pct(a, b):
    if a and b and b != 0:
        return (a - b) / abs(b) * 100
    return None


def fmt(val, decimals=2, prefix="", suffix=""):
    if val is None:
        return "n/a"
    return f"{prefix}{val:.{decimals}f}{suffix}"

def fmt_bn(val):
    if val is None:
        return "n/a"
    if isinstance(val, dict):
        val = val.get("val")
    if val is None:
        return "n/a"
    return f"${val/1e9:.1f}B"

def fmt_px(d):
    if not d:
        return "n/a"
    m = f"{d['chg_1m']:+.1f}%" if d["chg_1m"] is not None else "n/a"
    q = f"{d['chg_3m']:+.1f}%" if d["chg_3m"] is not None else "n/a"
    return (f"${d['price']:.2f}  1d {d['chg_1d']:+.1f}%  1m {m}  3m {q}  "
            f"52wH ${d['high_52w']:.2f} ({d['dd_52w']:+.1f}%)")

def _f(d, key, decimals=1, prefix="", suffix=""):
    if not d:
        return "n/a"
    v = d.get(key)
    return fmt(v, decimals, prefix, suffix) if v is not None else "n/a"

# ── INTERPRETATION ─────────────────────────────────────────
def get_interpretation(facts):
    prompt = f"""You are a portfolio analyst reviewing daily thesis pulse data for an 8-position portfolio.
Your job: assess whether each thesis is intact, weakening, or at a trigger threshold.

THESIS DOCUMENT:
{THESIS_DOC}

TODAY'S PRE-COMPUTED FACTS ({facts['today']}):

MACRO:
- Real yield: {facts['ry']} — {facts['ry_dist']}bps to 3.0% invalidation — {facts['ry_signal']}
  Momentum: {facts['ry_chg_1d']} today, {facts['ry_chg_4w']} over 4 weeks — {facts['ry_weeks_to_inv']} to invalidation at current pace
- DXY: {facts['dxy']} — {facts['dxy_dist']}pts to 115 invalidation — {facts['dxy_signal']}
  Momentum: {facts['dxy_chg_1d']} today, {facts['dxy_chg_4w']} over 4 weeks — {facts['dxy_weeks_to_inv']}

HEDGES:
- Gold: {facts['gold_px']} | 1m {facts['gold_1m']} | 3m {facts['gold_3m']} | {facts['gold_dd']} from 52wH
- Silver: {facts['silver_px']} | 1m {facts['silver_1m']} | 3m {facts['silver_3m']} | {facts['silver_dd']} from 52wH
- G/S ratio: {facts['gs']} (deploy trigger <55 = {facts['gs_dist_deploy']} pts away | invalidation >90 = {facts['gs_dist_inv']} pts away)
  Momentum: {facts['gs_chg_1d']} today, {facts['gs_chg_4w']} over 4 weeks — {facts['gs_velocity_label']}

CARRY:
- LLY: price {facts['lly_px']} ({facts['lly_dd']} from 52wH) | revenue {facts['lly_rev']} {facts['lly_rev_yoy']} YoY ({facts['lly_rev_date']})
- WMT: price {facts['wmt_px']} ({facts['wmt_dd']} from 52wH) | revenue {facts['wmt_rev']} {facts['wmt_rev_yoy']} YoY ({facts['wmt_rev_date']})
- JNJ: price {facts['jnj_px']} ({facts['jnj_dd']} from 52wH) | revenue {facts['jnj_rev']} {facts['jnj_rev_yoy']} YoY | div/share {facts['jnj_div']} {facts['jnj_div_yoy']} YoY

CYCLICAL:
- CCJ: price {facts['ccj_px']} ({facts['ccj_dd']} from 52wH)
- Uranium: {facts['uranium']} (monthly IMF series, {facts['uranium_lag']}d lag — treat as directional, not spot). Invalidation threshold: <$50/lb — currently ${facts['uranium_dist']:.2f} above.
- Oil term spread (WTI spot vs 12M forward): {facts['oil_spread']} backwardation | spot {facts['oil_spot']} (as of {facts['oil_spot_date']}) | 12M fwd {facts['oil_fwd']} ({facts['oil_fwd_ticker']}) [{facts['oil_signal']}]
  Interpretation: spread narrows on physical fall = thesis weakening | spread narrows on futures rise = thesis strengthening | spread stays wide (>$20) = supply stress intact

CONVEXITY:
- VRT: price {facts['vrt_px']} ({facts['vrt_dd']} from 52wH) | revenue {facts['vrt_rev']} {facts['vrt_rev_yoy']} YoY (as of {facts['vrt_rev_date']})
- AVGO: price {facts['avgo_px']} ({facts['avgo_dd']} from 52wH) | revenue {facts['avgo_rev']} {facts['avgo_rev_yoy']} YoY (as of {facts['avgo_rev_date']})
- NVDA revenue: {facts['nvda_rev']} {facts['nvda_rev_yoy']} YoY (as of {facts['nvda_rev_date']}, AI spend proxy)
- Hyperscaler capex: {facts['capex']} {facts['capex_yoy']} YoY (as of {facts['capex_date']}, invalidation threshold: <-30% YoY)

INSTRUCTIONS:
- Use ONLY the pre-computed facts above. Do not recalculate or restate raw numbers beyond what is given.
- Be direct and clinical. No filler. Each section max 2-3 sentences.
- Oil term spread belongs in CYCLICAL only. Do not reference it in CONVEXITY or any other section. You MUST mention it in the CYCLICAL output section.
- Real yield direction: falling real yield = moving AWAY from 3.0% invalidation = favorable. Rising real yield = moving TOWARD invalidation = unfavorable. Never describe a falling real yield as "approaching invalidation."
- Price moves alone are never invalidation — always tie to thesis conditions.
- Write only what the numbered facts above directly show. Every clause must trace back to a specific fact value.
- Do not comment on conditions, risks, or events that have no corresponding fact — not even to say they are absent, unconfirmed, or not contradicted. If there is no fact for it, it does not appear in your output.
- Do not restate or paraphrase thesis rationale, structural arguments, or position logic from the thesis document. Use the thesis document only to know what each number means, not as content to quote or summarize.

OUTPUT FORMAT (use exactly these headers, no deviations):

OVERALL: [INTACT | ONE FLAG | REVIEW NEEDED] — [one-line reason]

MACRO: [2 sentences. Are macro conditions favorable, neutral, or headwind for the thesis?]

HEDGES: [2-3 sentences covering gold and silver. Thesis intact? G/S ratio context?]

CARRY: [2-3 sentences covering LLY, WMT, JNJ. Revenue trends vs thesis requirements?]

CYCLICAL: [2-3 sentences. Uranium thesis intact? Include oil term spread signal and its current reading.]

CONVEXITY: [2-3 sentences covering VRT and AVGO. AI capex confirming? Price/fundamental divergence notable?]
"""
    try:
        r = requests.post(
            "https://api.anthropic.com/v1/messages",
            headers={
                "x-api-key": ANTHROPIC_API_KEY,
                "anthropic-version": "2023-06-01",
                "content-type": "application/json",
            },
            json={
                "model": "claude-haiku-4-5-20251001",
                "max_tokens": 900,
                "temperature": 0.2,
                "messages": [{"role": "user", "content": prompt}],
            },
            timeout=30,
        )
        if r.status_code == 200:
            return r.json()["content"][0]["text"]
        return f"[Claude error {r.status_code}]"
    except Exception as e:
        return f"[Claude error: {e}]"

# ── EMAIL ──────────────────────────────────────────────────
def send_email(subject, body):
    if not all([EMAIL_ADDRESS, EMAIL_PASSWORD, RECIPIENT_EMAIL]):
        print("Email credentials not set — skipping.")
        return
    try:
        msg = MIMEMultipart()
        msg["From"]    = EMAIL_ADDRESS
        msg["To"]      = RECIPIENT_EMAIL
        msg["Subject"] = subject
        msg.attach(MIMEText(body, "plain", "utf-8"))
        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as s:
            s.starttls()
            s.login(EMAIL_ADDRESS, EMAIL_PASSWORD)
            s.send_message(msg)
        print(f"Email sent to {RECIPIENT_EMAIL}")
    except Exception as e:
        print(f"Email error: {e}")

# ── STALE FLAG ─────────────────────────────────────────────
def stale_flag(date_str):
    if not date_str:
        return ""
    try:
        d = date.fromisoformat(str(date_str)[:10])
        delta = (date.today() - d).days
        return f"  S-{delta}D" if delta > 0 else ""
    except Exception:
        return ""


# ── FRED LAST N ────────────────────────────────────────────
def fred_last_n(series_id, n=15):
    session = _make_session()
    for attempt in range(1, 4):
        try:
            r = session.get(fred_url(series_id),
                            headers={"User-Agent": "thesis-pulse/1.0"}, timeout=30)
            r.raise_for_status()
            obs = r.json().get("observations", [])
            rows = [(o["date"], float(o["value"])) for o in obs
                    if o.get("value") not in (".", "")]
            return rows[-n:] if rows else []
        except Exception:
            time.sleep(0.5 * (2 ** (attempt - 1)))
    return []


def fred_3w_medians(series_id):
    """Returns (med_3w, prev_med_3w) using last 30 daily observations."""
    import statistics as _stats
    rows = fred_last_n(series_id, n=30)
    vals = [v for _, v in rows]
    med_3w      = _stats.median(vals[-15:])  if len(vals) >= 15 else None
    prev_med_3w = _stats.median(vals[-30:-15]) if len(vals) >= 30 else None
    return med_3w, prev_med_3w


# ── TECHNICAL LEVELS ──────────────────────────────────────
def get_tech_levels(ticker):
    """Returns (sup, res, tgt, brk) from weekly closes. sup/res bracket current price."""
    try:
        url = f"https://query1.finance.yahoo.com/v8/finance/chart/{ticker}?interval=1wk&range=5y"
        r = requests.get(url, headers={"User-Agent": "Mozilla/5.0"}, timeout=10)
        if r.status_code != 200:
            return None, None, None, None
        closes = [c for c in r.json()["chart"]["result"][0]["indicators"]["quote"][0]["close"] if c is not None]
        if len(closes) < 20:
            return None, None, None, None

        price  = closes[-1]
        high5y = max(closes)
        low2y  = min(closes[-104:]) if len(closes) >= 104 else min(closes)

        def _cluster(levels, pct=0.04):
            if not levels:
                return []
            s = sorted(levels)
            groups = [[s[0]]]
            for v in s[1:]:
                if (v - groups[-1][-1]) / groups[-1][-1] < pct:
                    groups[-1].append(v)
                else:
                    groups.append([v])
            return [sum(g) / len(g) for g in groups]

        peaks   = [closes[i] for i in range(2, len(closes)-2)
                   if closes[i] > closes[i-1] and closes[i] > closes[i-2]
                   and closes[i] > closes[i+1] and closes[i] > closes[i+2]]
        troughs = [closes[i] for i in range(2, len(closes)-2)
                   if closes[i] < closes[i-1] and closes[i] < closes[i-2]
                   and closes[i] < closes[i+1] and closes[i] < closes[i+2]]

        c_peaks   = _cluster(peaks)
        c_troughs = _cluster(troughs)

        above = [c for c in c_peaks if c > price * 1.01]
        res   = above[0] if above else None
        tgt   = next((c for c in above if c > res * 1.05), None) if res else None

        below_t = sorted([c for c in c_troughs if c < price * 0.99], reverse=True)
        sup = next((c for c in below_t if c > price * 0.75), None)
        if sup is None:
            sup = high5y - (high5y - low2y) * 0.236

        brk = high5y - (high5y - low2y) * 0.382

        return sup, res, tgt, brk
    except Exception:
        return None, None, None, None


# ── CAPE SCRAPER ───────────────────────────────────────────
def get_cape():
    try:
        import re
        r = requests.get("https://www.multpl.com/shiller-pe",
                         headers={"User-Agent": "Mozilla/5.0"}, timeout=10)
        if r.status_code == 200:
            m = re.search(r'<div[^>]+id=["\']current-value["\'][^>]*>\s*([0-9]+(?:\.[0-9]+)?)', r.text)
            if m:
                return float(m.group(1))
    except Exception:
        pass
    return None


# ── RECESSION TRACKER ──────────────────────────────────────
def compute_recession_signals(ry_val, ry_date, ry_prev=None, ry_4w=None):
    import json as _json
    config_path = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                               "calibration", "recession_config.json")
    try:
        with open(config_path, encoding="utf-8") as f:
            cfg = _json.load(f)
    except Exception:
        return None
    ind_cfg = cfg["indicators"]
    results = {}

    def _bps(curr, prev):
        if curr is None or prev is None: return "n/a"
        return f"{(curr - prev) * 100:+.0f}bps"

    # T10Y3M
    val, prev_val, lb_val, dt = fred_recent("T10Y3M", lookback=20)
    results["T10Y3M"] = {
        "val":        f"{val:.2f}" if val is not None else "n/a",
        "thr":        "0.0",
        "sig":        1 if val is not None and val < 0 else 0,
        "prev_sig":   1 if prev_val is not None and prev_val < 0 else 0,
        "level_str":  f"{val:.2f}" if val is not None else "n/a",
        "chg_4w_str": _bps(val, lb_val),
        "dist_str":   f"{val - 0.0:+.2f}" if val is not None else "n/a",
    }

    # T10Y2Y
    val, prev_val, lb_val, dt = fred_recent("T10Y2Y", lookback=20)
    results["T10Y2Y"] = {
        "val":        f"{val:.2f}" if val is not None else "n/a",
        "thr":        "0.0",
        "sig":        1 if val is not None and val < 0 else 0,
        "prev_sig":   1 if prev_val is not None and prev_val < 0 else 0,
        "level_str":  f"{val:.2f}" if val is not None else "n/a",
        "chg_4w_str": _bps(val, lb_val),
        "dist_str":   f"{val - 0.0:+.2f}" if val is not None else "n/a",
    }

    # DFII10
    results["DFII10"] = {
        "val":        f"{ry_val:.2f}" if ry_val is not None else "n/a",
        "thr":        "1.0%",
        "sig":        1 if ry_val is not None and ry_val > 1.0 else 0,
        "prev_sig":   1 if ry_prev is not None and ry_prev > 1.0 else 0,
        "level_str":  f"{ry_val:.2f}%" if ry_val is not None else "n/a",
        "chg_4w_str": _bps(ry_val, ry_4w),
        "dist_str":   f"{ry_val - 1.0:+.2f}" if ry_val is not None else "n/a",
    }

    # ICSA
    val, prev_val, lb_val, dt = fred_recent("ICSA", lookback=4)
    thr = ind_cfg["ICSA"]["threshold_val"]
    results["ICSA"] = {
        "val":        f"{val/1000:.1f}K" if val is not None else "n/a",
        "thr":        f"{thr/1000:.0f}K",
        "sig":        1 if val is not None and val > thr else 0,
        "prev_sig":   1 if prev_val is not None and prev_val > thr else 0,
        "level_str":  f"{val/1000:.0f}K" if val is not None else "n/a",
        "chg_4w_str": f"{(val - lb_val)/1000:+.0f}K" if val is not None and lb_val is not None else "n/a",
        "dist_str":   f"{(val - thr)/1000:+.0f}K" if val is not None else "n/a",
    }

    # UMCSENT: FP rate >50% at all thresholds, permanent exclusion confirmed empirically
    # INDPRO
    rows = fred_last_n("INDPRO", n=7)
    sig, prev_sig = 0, 0
    if len(rows) >= 4:
        recent = [v for _, v in rows[-4:]]
        sig = 1 if all(recent[i] < recent[i-1] for i in range(1, len(recent))) else 0
    if len(rows) >= 5:
        prev_r = [v for _, v in rows[-5:-1]]
        prev_sig = 1 if all(prev_r[i] < prev_r[i-1] for i in range(1, len(prev_r))) else 0
    curr_v = rows[-1][1] if rows else None
    prev_v = rows[-2][1] if len(rows) >= 2 else None
    results["INDPRO"] = {
        "val":        f"{curr_v:.1f}" if curr_v else "n/a",
        "thr":        "3mo↓",
        "sig":        sig,
        "prev_sig":   prev_sig,
        "level_str":  f"{curr_v:.1f}" if curr_v else "n/a",
        "chg_4w_str": f"{curr_v - prev_v:+.1f}" if curr_v and prev_v else "n/a",
        "dist_str":   "n/a",
    }

    # MANEMP
    rows = fred_last_n("MANEMP", n=7)
    sig, prev_sig = 0, 0
    if len(rows) >= 4:
        recent = [v for _, v in rows[-4:]]
        sig = 1 if all(recent[i] < recent[i-1] for i in range(1, len(recent))) else 0
    if len(rows) >= 5:
        prev_r = [v for _, v in rows[-5:-1]]
        prev_sig = 1 if all(prev_r[i] < prev_r[i-1] for i in range(1, len(prev_r))) else 0
    curr_v = rows[-1][1] if rows else None
    prev_v = rows[-2][1] if len(rows) >= 2 else None
    results["MANEMP"] = {
        "val":        f"{curr_v/1000:.2f}M" if curr_v else "n/a",
        "thr":        "3mo↓",
        "sig":        sig,
        "prev_sig":   prev_sig,
        "level_str":  f"{curr_v/1000:.2f}M" if curr_v else "n/a",
        "chg_4w_str": f"{(curr_v - prev_v)/1000:+.2f}M" if curr_v and prev_v else "n/a",
        "dist_str":   "n/a",
    }

    # PCEPILFE
    rows = fred_last_n("PCEPILFE", n=15)
    sig, prev_sig = 0, 0
    yoy_str = "n/a"
    curr_yoy = prev_yoy = None
    if len(rows) >= 13:
        curr_v, prev12 = rows[-1][1], rows[-13][1]
        curr_yoy = (curr_v - prev12) / prev12 * 100 if prev12 else 0
        yoy_str = f"{curr_yoy:.1f}%"
        sig = 1 if curr_yoy > ind_cfg["PCEPILFE"]["threshold_val"] else 0
    if len(rows) >= 14:
        p_curr, p_prev = rows[-2][1], rows[-14][1]
        prev_yoy = (p_curr - p_prev) / p_prev * 100 if p_prev else 0
        prev_sig = 1 if prev_yoy > ind_cfg["PCEPILFE"]["threshold_val"] else 0
    results["PCEPILFE"] = {
        "val":        yoy_str,
        "thr":        "2.0%",
        "sig":        sig,
        "prev_sig":   prev_sig,
        "level_str":  yoy_str,
        "chg_4w_str": f"{curr_yoy - prev_yoy:+.1f}%" if curr_yoy is not None and prev_yoy is not None else "n/a",
        "dist_str":   f"{curr_yoy - 2.0:+.1f}%" if curr_yoy is not None else "n/a",
    }

    # DFF
    val_dff, _, dt_dff = fred_latest("DFF")
    rows = fred_last_n("FEDFUNDS", n=7)
    sig, prev_sig = 0, 0
    if len(rows) >= 4:
        recent = [v for _, v in rows[-4:]]
        sig = 1 if all(recent[i] < recent[i-1] for i in range(1, len(recent))) else 0
    if len(rows) >= 5:
        prev_r = [v for _, v in rows[-5:-1]]
        prev_sig = 1 if all(prev_r[i] < prev_r[i-1] for i in range(1, len(prev_r))) else 0
    ff_curr = rows[-1][1] if rows else None
    ff_prev = rows[-2][1] if len(rows) >= 2 else None
    if ff_curr is not None and ff_prev is not None:
        dff_d = ff_curr - ff_prev
        dff_chg = "flat" if abs(dff_d) < 0.01 else f"{dff_d:+.2f}%"
    else:
        dff_chg = "n/a"
    results["DFF"] = {
        "val":        f"{val_dff:.2f}%" if val_dff is not None else "n/a",
        "thr":        "cut 3mo",
        "sig":        sig,
        "prev_sig":   prev_sig,
        "level_str":  f"{val_dff:.2f}%" if val_dff is not None else "n/a",
        "chg_4w_str": dff_chg,
        "dist_str":   "n/a",
    }

    # VIXCLS
    val, prev_val, lb_val, dt = fred_recent("VIXCLS", lookback=20)
    thr = ind_cfg["VIXCLS"]["threshold_val"]
    results["VIXCLS"] = {
        "val":        f"{val:.1f}" if val is not None else "n/a",
        "thr":        f"{thr:.0f}.0",
        "sig":        1 if val is not None and val > thr else 0,
        "prev_sig":   1 if prev_val is not None and prev_val > thr else 0,
        "level_str":  f"{val:.1f}" if val is not None else "n/a",
        "chg_4w_str": f"{val - lb_val:+.1f}" if val is not None and lb_val is not None else "n/a",
        "dist_str":   f"{val - thr:+.1f}" if val is not None else "n/a",
    }

    # SP500
    rows = fred_last_n("SP500", n=13)
    sig, prev_sig = 0, 0
    thr_str = "<10mo MA"
    val_str = "n/a"
    sp_val = ma10 = prev_sp = None
    if rows:
        sp_val = rows[-1][1]
        val_str = f"{sp_val:.0f}"
        if len(rows) >= 10:
            ma10 = sum(v for _, v in rows[-10:]) / 10
            thr_str = f"<{ma10:.0f}"
            sig = 1 if sp_val < ma10 else 0
        if len(rows) >= 11:
            prev_sp = rows[-2][1]
            prev_ma = sum(v for _, v in rows[-11:-1]) / 10
            prev_sig = 1 if prev_sp < prev_ma else 0
    results["SP500"] = {
        "val":        val_str,
        "thr":        thr_str,
        "sig":        sig,
        "prev_sig":   prev_sig,
        "level_str":  val_str,
        "chg_4w_str": f"{sp_val - prev_sp:+.0f}" if sp_val and prev_sp else "n/a",
        "dist_str":   f"{sp_val - ma10:+.0f}" if sp_val and ma10 else "n/a",
    }

    # BAA_AAA spread (daily: DBAA/DAAA)
    baa, prev_baa, _ = fred_latest("DBAA")
    aaa, prev_aaa, _ = fred_latest("DAAA")
    spread      = round(baa - aaa, 2)             if baa is not None and aaa is not None else None
    prev_spread = round(prev_baa - prev_aaa, 2)   if prev_baa is not None and prev_aaa is not None else None
    thr = ind_cfg["CREDIT_SPREAD"]["threshold_val"]
    baa_rows = fred_last_n("DBAA", n=2)
    aaa_rows = fred_last_n("DAAA", n=2)
    if len(baa_rows) >= 2 and len(aaa_rows) >= 2:
        prev_cs = round(baa_rows[-2][1] - aaa_rows[-2][1], 2)
        cs_chg  = f"{spread - prev_cs:+.2f}%" if spread is not None else "n/a"
    else:
        cs_chg = "n/a"
    results["BAA_AAA_spread"] = {
        "val":        f"{spread:.2f}%" if spread is not None else "n/a",
        "thr":        f"{thr}%",
        "sig":        1 if spread is not None and spread > thr else 0,
        "prev_sig":   1 if prev_spread is not None and prev_spread > thr else 0,
        "level_str":  f"{spread:.2f}%" if spread is not None else "n/a",
        "chg_4w_str": cs_chg,
        "dist_str":   f"{spread - thr:+.2f}%" if spread is not None else "n/a",
    }

    # CAPE: multpl.com fetch works but valuation elevated persistently since 1990s -- not a timing signal. Permanent exclusion confirmed empirically.

    composite      = sum(r["sig"]      for r in results.values())
    prev_composite = sum(r["prev_sig"] for r in results.values())
    denominator    = len(results)
    return {
        "indicators":     results,
        "composite":      composite,
        "prev_composite": prev_composite,
        "denominator":    denominator,
    }

# ── LOG HELPERS ────────────────────────────────────────────
def _safe(val, decimals=4):
    """Format value for CSV; empty string for None."""
    if val is None:
        return ""
    if isinstance(val, int):
        return str(val)
    if isinstance(val, float):
        return f"{val:.{decimals}f}"
    return str(val)

def _num(s):
    """Parse numeric string with optional %, K (×1e3), M (×1e6) suffix."""
    if not s or s == "n/a":
        return None
    s = str(s).strip().lstrip("+")
    mult = 1.0
    if s.endswith("K"):
        mult, s = 1e3, s[:-1]
    elif s.endswith("M"):
        mult, s = 1e6, s[:-1]
    try:
        return float(s.replace("%", "")) * mult
    except Exception:
        return None

def _migrate_macro_log():
    path = os.path.join(_dir, "macro_log.csv")
    if not os.path.exists(path):
        return
    with open(path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        if not reader.fieldnames or "ry_cvstc" in reader.fieldnames:
            return
        rows = list(reader)
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=_MACRO_HEADER, extrasaction="ignore")
        w.writeheader()
        for row in rows:
            w.writerow(row)
    print("  macro_log.csv: migrated — added CvsTC columns")


def _read_prev_cvstc():
    path = os.path.join(_dir, "macro_log.csv")
    if not os.path.exists(path):
        return None, None, None, None
    try:
        with open(path, newline="", encoding="utf-8") as f:
            rows = list(csv.DictReader(f))
        today = date.today().isoformat()
        prev_rows = [r for r in rows if r.get("date") != today]
        if not prev_rows:
            return None, None, None, None
        last = prev_rows[-1]
        def _g(key):
            v = last.get(key, "")
            return float(v) if v else None
        return _g("ry_cvstc"), _g("dxy_cvstc"), _g("gsr_cvstc"), _g("oil_cvstc")
    except Exception:
        return None, None, None, None


def _append_csv(path, header, row):
    """Append one row to a CSV file, writing the header if the file is new. Skips if date already present."""
    write_header = not os.path.exists(path)
    if not write_header:
        with open(path, newline="", encoding="utf-8") as f:
            if any(r[0] == row[0] for r in csv.reader(f) if r):
                print(f"  {os.path.basename(path)}: {row[0]} already logged — skipping")
                return
    with open(path, "a", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        if write_header:
            w.writerow(header)
        w.writerow(row)

# Columns and units for macro_log.csv:
#   t10y2y / t10y3m       — percentage points (e.g. 0.49)
#   real_yield_10y        — percentage points (DFII10)
#   core_pce_yoy_pct      — percent YoY (e.g. 3.2)
#   icsa_claims           — raw initial claims count (e.g. 200000)
#   fed_funds_pct         — percent (e.g. 3.64)
#   credit_spread_pct     — percent (e.g. 0.61)
#   manemp_k              — thousands of persons (FRED native unit, e.g. 12340)
#   oil_spread            — USD/bbl backwardation (positive = backwardation)
_MACRO_HEADER = [
    "date", "dxy", "real_yield_10y", "t10y2y", "t10y3m",
    "core_pce_yoy_pct", "icsa_claims", "fed_funds_pct", "vix",
    "sp500", "credit_spread_pct", "indpro", "manemp_k",
    "oil_spread", "gs_ratio", "recession_count",
    "ry_cvstc", "dxy_cvstc", "gsr_cvstc", "oil_cvstc",
]

# Columns for asset_log.csv:
#   *_usd    — closing price in USD
#   *_sek    — closing price in SEK (USD price × USDSEK rate)
#   *_ret    — 1-day % return (chg_1d from Yahoo, e.g. 1.23 means +1.23%)
#   portfolio_ret — weighted 1-day % return at v3 target weights
_ASSET_HEADER = [
    "date",
    "gold_usd", "silver_usd", "lly_usd", "wmt_usd",
    "vrt_usd",  "ccj_usd",   "avgo_usd", "jnj_usd",
    "usdsek", "gold_sek", "silver_sek",
    "gold_ret", "silver_ret", "lly_ret", "wmt_ret",
    "vrt_ret",  "ccj_ret",   "avgo_ret", "jnj_ret",
    "portfolio_ret",
]

def append_macro_log(today, dxy_price, ry_val, rec, gs_ratio, oil_spread,
                     ry_cvstc=None, dxy_cvstc=None, gsr_cvstc=None, oil_cvstc=None):
    ind = (rec or {}).get("indicators", {})

    def _iv(key):
        return _num((ind.get(key) or {}).get("level_str", ""))

    # MANEMP level_str is "12.34M" where the unit is already millions of persons,
    # but FRED's MANEMP is in thousands — so _num("12.34M")=12340000, /1000 → 12340k.
    manemp_raw = _iv("MANEMP")
    manemp_k   = manemp_raw / 1000 if manemp_raw is not None else None

    row = [
        today,
        _safe(dxy_price, 4),
        _safe(ry_val, 4),
        _safe(_iv("T10Y2Y"), 4),
        _safe(_iv("T10Y3M"), 4),
        _safe(_iv("PCEPILFE"), 2),
        _safe(_iv("ICSA"), 0),          # raw claims e.g. 200000
        _safe(_iv("DFF"), 4),
        _safe(_iv("VIXCLS"), 2),
        _safe(_iv("SP500"), 0),
        _safe(_iv("BAA_AAA_spread"), 4),
        _safe(_iv("INDPRO"), 2),
        _safe(manemp_k, 0),
        _safe(oil_spread, 2),
        _safe(gs_ratio, 4),
        (rec or {}).get("composite", ""),
        _safe(ry_cvstc, 4),
        _safe(dxy_cvstc, 4),
        _safe(gsr_cvstc, 4),
        _safe(oil_cvstc, 4),
    ]
    _append_csv(os.path.join(_dir, "macro_log.csv"), _MACRO_HEADER, row)

def append_asset_log(today, gold, silver, lly_px, wmt_px, vrt_px, ccj_px,
                     avgo_px, jnj_px, usdsek_price):
    def _p(d):
        return d["price"] if d else None

    def _r(d):
        return d["chg_1d"] if d else None  # already in % terms (e.g. 1.23 = +1.23%)

    gold_p  = _p(gold);    silver_p = _p(silver)
    lly_p   = _p(lly_px);  wmt_p   = _p(wmt_px)
    vrt_p   = _p(vrt_px);  ccj_p   = _p(ccj_px)
    avgo_p  = _p(avgo_px); jnj_p   = _p(jnj_px)

    gold_sek   = gold_p   * usdsek_price if gold_p   and usdsek_price else None
    silver_sek = silver_p * usdsek_price if silver_p and usdsek_price else None

    rets = {
        "gold":   _r(gold),    "silver": _r(silver),
        "lly":    _r(lly_px),  "wmt":    _r(wmt_px),
        "vrt":    _r(vrt_px),  "ccj":    _r(ccj_px),
        "avgo":   _r(avgo_px), "jnj":    _r(jnj_px),
    }
    # Weighted portfolio return; excludes any position with missing data
    valid = {k: v for k, v in rets.items() if v is not None}
    if valid:
        total_w = sum(V3_WEIGHTS[k] for k in valid)
        port_ret = sum(V3_WEIGHTS[k] * v for k, v in valid.items()) / total_w if total_w else None
    else:
        port_ret = None

    row = [
        today,
        _safe(gold_p, 4),   _safe(silver_p, 4),
        _safe(lly_p, 4),    _safe(wmt_p, 4),
        _safe(vrt_p, 4),    _safe(ccj_p, 4),
        _safe(avgo_p, 4),   _safe(jnj_p, 4),
        _safe(usdsek_price, 4),
        _safe(gold_sek, 2), _safe(silver_sek, 2),
        _safe(rets["gold"],   4), _safe(rets["silver"], 4),
        _safe(rets["lly"],    4), _safe(rets["wmt"],    4),
        _safe(rets["vrt"],    4), _safe(rets["ccj"],    4),
        _safe(rets["avgo"],   4), _safe(rets["jnj"],    4),
        _safe(port_ret, 4),
    ]
    _append_csv(os.path.join(_dir, "asset_log.csv"), _ASSET_HEADER, row)

# ── MAIN ───────────────────────────────────────────────────
def main():
    today = date.today().isoformat()
    print(f"Thesis Pulse | {today}")

    print("Fetching Yahoo Finance...")
    gold    = yahoo_history("GC%3DF");  time.sleep(1)
    silver  = yahoo_history("SI%3DF");  time.sleep(1)
    dxy     = yahoo_history("DX-Y.NYB"); time.sleep(1)
    lly_px  = yahoo_history("LLY");     time.sleep(1)
    wmt_px  = yahoo_history("WMT");     time.sleep(1)
    jnj_px  = yahoo_history("JNJ");     time.sleep(1)
    ccj_px  = yahoo_history("CCJ");     time.sleep(1)
    vrt_px  = yahoo_history("VRT");     time.sleep(1)
    avgo_px = yahoo_history("AVGO");  time.sleep(1)
    usdsek  = yahoo_history("USDSEK=X")

    print("Fetching FRED...")
    ry_val, ry_prev, ry_4w, ry_date = fred_recent("DFII10", lookback=20)

    print("Fetching EDGAR...")
    lly_c,     lly_p     = edgar_revenue("LLY")
    wmt_c,     wmt_p     = edgar_revenue("WMT")
    jnj_c,     jnj_p     = edgar_revenue("JNJ")
    vrt_c,     vrt_p     = edgar_revenue("VRT")
    avgo_c,    avgo_p    = edgar_revenue("AVGO")
    nvda_c,    nvda_p    = edgar_revenue("NVDA")
    jnj_div_c, jnj_div_p = edgar_concept("JNJ", "CommonStockDividendsPerShareCashPaid")
    msft_c,    msft_p    = edgar_concept("MSFT",  "PaymentsToAcquirePropertyPlantAndEquipment")
    googl_c,   googl_p   = edgar_concept("GOOGL", "PaymentsToAcquirePropertyPlantAndEquipment")
    amzn_c,    amzn_p    = edgar_concept("AMZN",  "PaymentsToAcquireProductiveAssets")
    meta_c,    meta_p    = edgar_concept("META",  "PaymentsToAcquirePropertyPlantAndEquipment")

    print("Fetching uranium...")
    uranium, uranium_prev, uranium_date = get_uranium()

    print("Fetching oil term spread...")
    oil_spot, oil_fwd, oil_spread, oil_spread_prev, oil_spread_med_3w, oil_spread_prev_med_3w, oil_spot_date, oil_fwd_ticker = get_oil_term_spread()

    print("Fetching recession indicators...")
    rec = compute_recession_signals(ry_val, ry_date, ry_prev=ry_prev, ry_4w=ry_4w)

    # Compute derived values
    gs_ratio      = gold["price"] / silver["price"]             if gold and silver else None
    gs_ratio_prev = gold["prev_close"] / silver["prev_close"]   if gold and silver and gold.get("prev_close") and silver.get("prev_close") else None

    import statistics as _stats
    gold_c30   = gold["closes_30"]   if gold   and gold.get("closes_30")   else []
    silver_c30 = silver["closes_30"] if silver and silver.get("closes_30") else []
    if len(gold_c30) >= 30 and len(silver_c30) >= 30:
        gsr_series      = [g / s for g, s in zip(gold_c30, silver_c30)]
        gsr_med_3w      = _stats.median(gsr_series[-15:])
        gsr_prev_med_3w = _stats.median(gsr_series[-30:-15])
    else:
        gsr_med_3w = gsr_prev_med_3w = None

    capex_vals       = [x["val"] for x in [msft_c, googl_c, amzn_c, meta_c] if x]
    capex_prevs      = [x["val"] for x in [msft_p, googl_p, amzn_p, meta_p] if x]
    capex_total      = sum(capex_vals)  if capex_vals  else None
    capex_total_prev = sum(capex_prevs) if capex_prevs else None

    ry_med_3w, ry_prev_med_3w = fred_3w_medians("DFII10")

    ry_dist   = int(round(300 - ry_val * 100)) if ry_val is not None else "n/a"
    dxy_price = dxy["price"] if dxy else None

    uranium_mom = fmt(pct(uranium, uranium_prev), 1) if uranium and uranium_prev else "n/a"
    capex_yoy   = fmt(pct(capex_total, capex_total_prev), 1) if capex_total and capex_total_prev else "n/a"

    # Build output
    lines = []
    lines.append("=" * 68)
    lines.append(f"  THESIS PULSE  |  {today}")
    lines.append("=" * 68)
    lines.append("")
    try:
        from vol_events import get_vol_events
        lines.append(get_vol_events())
    except Exception as e:
        lines.append(f"  PULSE  n/a ({e})")
    lines.append("")

    # THESIS
    lines.append("  THESIS")
    lines.append(f"  {'-'*38}")

    prev_ry_cvstc, prev_dxy_cvstc, prev_gsr_cvstc, prev_oil_cvstc = _read_prev_cvstc()

    def _cvstc(curr, med, pmed):
        if curr is None or med is None or pmed is None:
            return None
        mid = (med + pmed) / 2
        return (curr - mid) / mid * 100 if mid != 0 else None

    ry_cvstc  = _cvstc(ry_val,    ry_med_3w,          ry_prev_med_3w)
    dxy_cvstc = _cvstc(dxy_price, dxy["med_3w"] if dxy else None, dxy["prev_med_3w"] if dxy else None)
    gsr_cvstc = _cvstc(gs_ratio,  gsr_med_3w,         gsr_prev_med_3w)
    oil_cvstc = _cvstc(oil_spread, oil_spread_med_3w, oil_spread_prev_med_3w)

    def _thesis_block(label, curr_s, prev_s, med_s, pmed_s, cvstc, prev_cvstc=None, extra=None):
        if cvstc is not None:
            c = f"CvsTC {cvstc:+.2f}% ({prev_cvstc:+.2f}%)" if prev_cvstc is not None else f"CvsTC {cvstc:+.2f}%"
        else:
            c = ""
        lines.append(f"  {label:<8}{curr_s} {prev_s:<10} {c}".rstrip())
        lines.append(f"          3wM {med_s}  prv {pmed_s}".rstrip())
        if extra:
            lines.append(f"          {extra}")
        lines.append("")

    _thesis_block(
        "RY",
        f"{ry_val:.2f}%"          if ry_val          is not None else "n/a",
        f"({ry_prev:.2f})"         if ry_prev         is not None else "",
        f"{ry_med_3w:.2f}%"        if ry_med_3w       is not None else "n/a",
        f"{ry_prev_med_3w:.2f}%"   if ry_prev_med_3w  is not None else "n/a",
        ry_cvstc, prev_ry_cvstc,
    )
    _thesis_block(
        "DXY",
        f"{dxy_price:.2f}"                if dxy_price                      is not None else "n/a",
        f"({dxy['prev_close']:.2f})"      if dxy and dxy.get("prev_close")  is not None else "",
        f"{dxy['med_3w']:.2f}"            if dxy and dxy.get("med_3w")      is not None else "n/a",
        f"{dxy['prev_med_3w']:.2f}"       if dxy and dxy.get("prev_med_3w") is not None else "n/a",
        dxy_cvstc, prev_dxy_cvstc,
    )
    _thesis_block(
        "GSR",
        f"{gs_ratio:.1f}"             if gs_ratio        is not None else "n/a",
        f"({gs_ratio_prev:.1f})"      if gs_ratio_prev   is not None else "",
        f"{gsr_med_3w:.1f}"           if gsr_med_3w      is not None else "n/a",
        f"{gsr_prev_med_3w:.1f}"      if gsr_prev_med_3w is not None else "n/a",
        gsr_cvstc, prev_gsr_cvstc,
        extra="T1 83.4  T2 86.5",
    )
    _thesis_block(
        "Oil",
        f"${oil_spread:.1f}"               if oil_spread              is not None else "n/a",
        f"(${oil_spread_prev:.1f})"        if oil_spread_prev         is not None else "",
        f"${oil_spread_med_3w:.1f}"        if oil_spread_med_3w       is not None else "n/a",
        f"${oil_spread_prev_med_3w:.1f}"   if oil_spread_prev_med_3w  is not None else "n/a",
        oil_cvstc, prev_oil_cvstc,
    )

    if capex_total:
        prev_s = f" ({fmt_bn(capex_total_prev)})" if capex_total_prev else ""
        lines.append(f"  Capex   {fmt_bn(capex_total)}{prev_s}  {capex_yoy}% YoY")
    else:
        lines.append(f"  Capex   n/a")
    lines.append("")

    # POSITIONS
    lines.append(f"  {'POSITIONS':<18} {'price':<10} {'vs 200MA':<9} {'vs 52wH':<9} {'sup':<8} {'res':<8} vol")
    lines.append(f"  {'-'*68}")

    def _pos_line(name, px_dict, levels=None):
        if px_dict is None:
            return f"  {name:<18} {'n/a':<10} {'n/a':<9} {'n/a':<9} {'—':<8} {'—':<8} —"
        price_s = f"${px_dict['price']:.2f}"
        ma_s    = f"{px_dict['vs_ma_200']:+.1f}%" if px_dict.get("vs_ma_200") is not None else "n/a"
        hi_s    = f"{px_dict['dd_52w']:+.1f}%"    if px_dict.get("dd_52w")    is not None else "n/a"
        rv      = px_dict.get("rel_vol")
        vol_s   = f"{rv:.1f}x" if rv is not None else "—"
        if levels and any(v is not None for v in levels):
            sup, res, tgt, brk = levels
            sup_s = f"${sup:.0f}" if sup is not None else "—"
            res_s = f"${res:.0f}" if res is not None else "—"
        else:
            sup_s = res_s = "—"
        return f"  {name:<18} {price_s:<10} {ma_s:<9} {hi_s:<9} {sup_s:<8} {res_s:<8} {vol_s}"

    print("Fetching technical levels...")
    gold_levels   = get_tech_levels("GC%3DF");  time.sleep(1)
    silver_levels = get_tech_levels("SI%3DF");  time.sleep(1)
    lly_levels    = get_tech_levels("LLY");     time.sleep(1)
    wmt_levels    = get_tech_levels("WMT");     time.sleep(1)
    jnj_levels    = get_tech_levels("JNJ");     time.sleep(1)
    ccj_levels    = get_tech_levels("CCJ");     time.sleep(1)
    vrt_levels    = get_tech_levels("VRT");     time.sleep(1)
    avgo_levels   = get_tech_levels("AVGO")

    lines.append(_pos_line("Gold",   gold,   gold_levels))
    lines.append(_pos_line("Silver", silver, silver_levels))
    lines.append(_pos_line("LLY",    lly_px, lly_levels))
    lines.append(_pos_line("WMT",    wmt_px, wmt_levels))
    lines.append(_pos_line("JNJ",    jnj_px, jnj_levels))
    lines.append(_pos_line("CCJ",    ccj_px, ccj_levels))
    lines.append(_pos_line("VRT",    vrt_px, vrt_levels))
    lines.append(_pos_line("AVGO",   avgo_px, avgo_levels))
    lines.append("")

    # RECESSION TRACKER
    lines.append("  RECESSION TRACKER")
    lines.append(f"  {'-'*64}")
    if rec:
        prev_str = f"prev {rec['prev_composite']}/{rec['denominator']}"
        lines.append(f"  Composite    {rec['composite']}/{rec['denominator']}    {prev_str}")
        lines.append("")
        lines.append(f"  {'Signal':<18} {'Level':<11} {'Threshold':<13} {'4wk':<11} Distance")
        lines.append(f"  {'-'*64}")
        order = ["T10Y3M", "T10Y2Y", "DFII10", "ICSA",
                 "INDPRO", "MANEMP", "PCEPILFE", "DFF", "VIXCLS",
                 "SP500", "BAA_AAA_spread"]
        for ind in order:
            r = rec["indicators"].get(ind)
            if r is None:
                continue
            level_s  = r.get("level_str", r.get("val", "n/a"))
            thr_s    = r.get("thr", "n/a")
            chg_s    = r.get("chg_4w_str", "n/a")
            dist_s   = r.get("dist_str", "n/a")
            bullet   = "  ●" if r.get("sig", 0) else ""
            lines.append(f"  {ind:<18} {level_s:<11} {thr_s:<13} {chg_s:<11} {dist_s}{bullet}")
    else:
        lines.append("  n/a  (recession_config.json not found)")
    lines.append("")
    lines.append("=" * 68)

    body = "\n".join(lines)
    print(body)

    print("Writing daily logs...")
    _migrate_macro_log()
    usdsek_price = usdsek["price"] if usdsek else None
    append_macro_log(today, dxy_price, ry_val, rec, gs_ratio, oil_spread,
                     ry_cvstc=ry_cvstc, dxy_cvstc=dxy_cvstc,
                     gsr_cvstc=gsr_cvstc, oil_cvstc=oil_cvstc)
    append_asset_log(today, gold, silver, lly_px, wmt_px, vrt_px, ccj_px,
                     avgo_px, jnj_px, usdsek_price)

    subject = f"Thesis Pulse {today}"
    send_email(subject, body)


if __name__ == "__main__":
    main()
