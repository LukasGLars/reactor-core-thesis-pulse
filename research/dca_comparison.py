import pandas as pd
import numpy as np
import os, sys, io, warnings
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference
warnings.filterwarnings("ignore")
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

DATA_DIR = r"C:\Users\lukas.larsson\Desktop\Privat\Project Reactor Core\data"
REPORT   = r"C:\Users\lukas.larsson\Desktop\Privat\Project Reactor Core\DCA_Comparison.xlsx"

INITIAL  = 1_000_000   # kr
MONTHLY  =     6_000   # kr
START    = "2016-04-01"

TICKERS_8 = ["xauusd","lly.us","wmt.us","avgo.us","cost.us","ccj.us","vrt.us","jnj.us"]
TICKERS_9 = TICKERS_8 + ["xagusd"]

W8 = {"xauusd":0.250,"wmt.us":0.227,"lly.us":0.197,"vrt.us":0.090,
      "avgo.us":0.079,"ccj.us":0.057,"jnj.us":0.050,"cost.us":0.050}

W9 = {"xauusd":0.250,"wmt.us":0.204,"lly.us":0.188,"vrt.us":0.085,
      "avgo.us":0.073,"ccj.us":0.050,"jnj.us":0.050,"cost.us":0.050,"xagusd":0.050}

NAMES = {"xauusd":"Gold","lly.us":"Eli Lilly","wmt.us":"Walmart","avgo.us":"Broadcom",
         "cost.us":"Costco","ccj.us":"Cameco","vrt.us":"Vertiv","jnj.us":"J&J","xagusd":"Silver"}

def load(ticker):
    df = pd.read_csv(os.path.join(DATA_DIR, ticker.replace("-","_") + ".csv"))
    df["Date"] = pd.to_datetime(df["Date"])
    return df.set_index("Date").sort_index()["Close"].replace(0, np.nan).dropna()

# Load prices
all_tickers = list(set(TICKERS_9))
prices = pd.DataFrame({t: load(t) for t in all_tickers})
prices = prices[prices.index >= START].dropna()

# Monthly resample (end of month)
monthly = prices.resample("ME").last()
monthly_rets = monthly.pct_change().dropna()

def simulate(weights, rets):
    w = np.array([weights[t] for t in rets.columns])
    records = []
    value = INITIAL
    contributed = INITIAL
    for date, row in rets.iterrows():
        port_ret = np.dot(w, row.values)
        value   *= (1 + port_ret)
        value   += MONTHLY
        contributed += MONTHLY
        profit   = value - contributed
        records.append({
            "date":        date,
            "value":       round(value),
            "contributed": round(contributed),
            "profit":      round(profit),
            "return_pct":  round((value / contributed - 1) * 100, 1),
        })
    return pd.DataFrame(records).set_index("date")

# Run simulations — align columns to intersection
cols_8 = [t for t in TICKERS_8 if t in monthly_rets.columns]
cols_9 = [t for t in TICKERS_9 if t in monthly_rets.columns]
rets_8 = monthly_rets[cols_8]
rets_9 = monthly_rets[cols_9]

sim8 = simulate({t: W8[t] for t in cols_8}, rets_8)
sim9 = simulate({t: W9[t] for t in cols_9}, rets_9)

# Yearly snapshots
def yearly(sim):
    return sim.resample("YE").last()

y8 = yearly(sim8)
y9 = yearly(sim9)

print("=" * 72)
print("  DCA Simulation — 1,000,000 kr initial + 6,000 kr/month")
print(f"  Period: {monthly_rets.index[0].date()} → {monthly_rets.index[-1].date()}")
print("=" * 72)
print(f"\n  {'Year':<6}  {'Contributed':>13}  {'8-pos Value':>13}  "
      f"{'9-pos Value':>13}  {'Diff':>10}  {'8-pos Ret':>10}  {'9-pos Ret':>10}")

years = y8.index.union(y9.index)
for dt in years:
    if dt in y8.index and dt in y9.index:
        c   = y8.loc[dt, "contributed"]
        v8  = y8.loc[dt, "value"]
        v9  = y9.loc[dt, "value"]
        r8  = y8.loc[dt, "return_pct"]
        r9  = y9.loc[dt, "return_pct"]
        diff = v9 - v8
        print(f"  {dt.year:<6}  {c:>13,.0f}  {v8:>13,.0f}  "
              f"{v9:>13,.0f}  {diff:>+10,.0f}  {r8:>9.1f}%  {r9:>9.1f}%")

# Final
v8f = sim8.iloc[-1]["value"]; v9f = sim9.iloc[-1]["value"]
cf  = sim8.iloc[-1]["contributed"]
print(f"\n  Final contributed:  {cf:>13,.0f} kr")
print(f"  8-pos final value:  {v8f:>13,.0f} kr  (+{v8f-cf:,.0f} kr profit)")
print(f"  9-pos final value:  {v9f:>13,.0f} kr  (+{v9f-cf:,.0f} kr profit)")
print(f"  Silver cost/gain:   {v9f-v8f:>+13,.0f} kr")

# ── Excel ─────────────────────────────────────────────────────────────────────
DARK="1F4E79"; MID="2E75B6"; LIGHT="D6E4F0"; ALT="EBF3FB"
GRN="C6EFCE"; RED="FFC7CE"; YEL="FFEB9C"; WHT="FFFFFF"; GREY="F2F2F2"
thin   = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr(ws, row, col, val, bg=DARK, fg=WHT):
    c = ws.cell(row=row, column=col, value=val)
    c.fill = PatternFill("solid", fgColor=bg)
    c.font = Font(bold=True, color=fg, size=10)
    c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    c.border = BORDER; return c

def cel(ws, row, col, val, bg=None, bold=False, fmt=None, ha="center"):
    c = ws.cell(row=row, column=col, value=val)
    if bg: c.fill = PatternFill("solid", fgColor=bg)
    c.font = Font(bold=bold)
    c.alignment = Alignment(vertical="center", horizontal=ha)
    c.border = BORDER
    if fmt: c.number_format = fmt
    return c

wb = openpyxl.Workbook()

# ━━ SHEET 1: Yearly Comparison ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ws1 = wb.active; ws1.title = "Yearly Comparison"
ws1.merge_cells("A1:I1")
c1 = ws1["A1"]
c1.value = "DCA Simulation — 1,000,000 kr Initial + 6,000 kr/Month  |  8-pos vs 9-pos (+Silver)"
c1.font = Font(bold=True, color=WHT, size=13)
c1.fill = PatternFill("solid", fgColor=DARK)
c1.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 34

ws1.merge_cells("A2:I2")
c2 = ws1["A2"]
c2.value = (f"8-pos = Reactor Core v2 original  |  9-pos = v2.1 with Silver at 5%  |  "
            f"Period: {monthly_rets.index[0].date()} → {monthly_rets.index[-1].date()}  |  "
            f"Simple DCA, no rebalancing, SEK")
c2.font = Font(italic=True, size=9, color="444444")
c2.fill = PatternFill("solid", fgColor=LIGHT)
c2.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws1.row_dimensions[2].height = 22

hdrs = ["Year","Contributed","8-pos Value","8-pos Profit","8-pos Ret%",
        "9-pos Value","9-pos Profit","9-pos Ret%","Silver Diff"]
for ci, h in enumerate(hdrs, 1): hdr(ws1, 3, ci, h)
ws1.row_dimensions[3].height = 28

fmt_kr = '#,##0'
fmt_pct = '0.0"%"'

chart_rows = []
for ri, dt in enumerate(years, 4):
    if dt not in y8.index or dt not in y9.index: continue
    bg   = ALT if ri % 2 == 0 else WHT
    c_   = int(y8.loc[dt,"contributed"])
    v8_  = int(y8.loc[dt,"value"])
    v9_  = int(y9.loc[dt,"value"])
    p8_  = v8_ - c_
    p9_  = v9_ - c_
    r8_  = y8.loc[dt,"return_pct"]
    r9_  = y9.loc[dt,"return_pct"]
    diff = v9_ - v8_
    diff_bg = GRN if diff >= 0 else RED

    cel(ws1, ri, 1, dt.year,   bg=GREY, bold=True)
    cel(ws1, ri, 2, c_,        bg=bg,   fmt=fmt_kr)
    cel(ws1, ri, 3, v8_,       bg=bg,   fmt=fmt_kr, bold=True)
    cel(ws1, ri, 4, p8_,       bg=GRN if p8_>=0 else RED, fmt=fmt_kr)
    cel(ws1, ri, 5, r8_,       bg=bg,   fmt=fmt_pct)
    cel(ws1, ri, 6, v9_,       bg=bg,   fmt=fmt_kr, bold=True)
    cel(ws1, ri, 7, p9_,       bg=GRN if p9_>=0 else RED, fmt=fmt_kr)
    cel(ws1, ri, 8, r9_,       bg=bg,   fmt=fmt_pct)
    cel(ws1, ri, 9, diff,      bg=diff_bg, fmt=fmt_kr, bold=True)
    ws1.row_dimensions[ri].height = 22
    chart_rows.append(ri)

# Summary row
ri_s = len(years) + 5
ws1.merge_cells(f"A{ri_s}:I{ri_s}")
c_s = ws1.cell(row=ri_s, column=1,
    value=(f"Total contributed: {int(cf):,} kr  |  "
           f"8-pos final: {int(v8f):,} kr (+{int(v8f-cf):,} kr)  |  "
           f"9-pos final: {int(v9f):,} kr (+{int(v9f-cf):,} kr)  |  "
           f"Silver net effect: {int(v9f-v8f):+,} kr"))
c_s.font = Font(bold=True, size=10)
c_s.fill = PatternFill("solid", fgColor=YEL if v9f >= v8f else "FFC7CE")
c_s.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws1.row_dimensions[ri_s].height = 28

# Column widths
for i, w_ in enumerate([8,14,14,14,11,14,14,11,13], 1):
    ws1.column_dimensions[get_column_letter(i)].width = w_

# ━━ SHEET 2: Monthly Detail ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ws2 = wb.create_sheet("Monthly Detail")
ws2.merge_cells("A1:G1")
c_m = ws2["A1"]
c_m.value = "Monthly DCA — Full Detail"
c_m.font = Font(bold=True, color=WHT, size=12)
c_m.fill = PatternFill("solid", fgColor=DARK)
c_m.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 30

for ci, h in enumerate(["Month","Contributed","8-pos Value","8-pos Profit",
                          "9-pos Value","9-pos Profit","Silver Diff"], 1):
    hdr(ws2, 2, ci, h)
ws2.row_dimensions[2].height = 26

for ri, (dt, row8) in enumerate(sim8.iterrows(), 3):
    if dt not in sim9.index: continue
    row9 = sim9.loc[dt]
    bg   = ALT if ri % 2 == 0 else WHT
    diff = int(row9["value"]) - int(row8["value"])
    cel(ws2, ri, 1, str(dt.date()),        bg=GREY)
    cel(ws2, ri, 2, int(row8["contributed"]), bg=bg, fmt=fmt_kr)
    cel(ws2, ri, 3, int(row8["value"]),    bg=bg, fmt=fmt_kr, bold=True)
    cel(ws2, ri, 4, int(row8["profit"]),   bg=GRN if row8["profit"]>=0 else RED, fmt=fmt_kr)
    cel(ws2, ri, 5, int(row9["value"]),    bg=bg, fmt=fmt_kr, bold=True)
    cel(ws2, ri, 6, int(row9["profit"]),   bg=GRN if row9["profit"]>=0 else RED, fmt=fmt_kr)
    cel(ws2, ri, 7, diff,                   bg=GRN if diff>=0 else RED, fmt=fmt_kr)
    ws2.row_dimensions[ri].height = 16

for i, w_ in enumerate([12,14,14,14,14,14,13], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w_

wb.save(REPORT)
print(f"\n  Report saved: {REPORT}")
print("=" * 72)
