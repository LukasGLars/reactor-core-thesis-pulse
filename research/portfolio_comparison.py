import requests, pandas as pd, numpy as np
import os, sys, io, warnings, time
from io import StringIO
warnings.filterwarnings("ignore")
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DATA_DIR   = r"C:\Users\lukas.larsson\Desktop\Privat\Project Reactor Core\data"
REPORT     = r"C:\Users\lukas.larsson\Desktop\Privat\Project Reactor Core\Portfolio_Comparison.xlsx"

# ── Download missing tickers ──────────────────────────────────────────────────
SESSION = requests.Session()
SESSION.headers.update({"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120.0.0.0 Safari/537.36", "Referer": "https://stooq.com"})
SESSION.get("https://stooq.com", timeout=15)

def download_and_save(ticker, fname):
    path = os.path.join(DATA_DIR, fname)
    if os.path.exists(path):
        return
    r = SESSION.get(f"https://stooq.com/q/d/l/?s={ticker}&i=d", timeout=15)
    df = pd.read_csv(StringIO(r.text))
    df["Date"] = pd.to_datetime(df["Date"])
    df = df.sort_values("Date")
    df.to_csv(path, index=False)
    print(f"  Downloaded {ticker} -> {fname}")
    time.sleep(0.4)

download_and_save("xmld.uk", "xmld.uk.csv")
download_and_save("0rq9.uk", "0rq9.uk.csv")

# ── Portfolio definitions ─────────────────────────────────────────────────────
# Combined weights: GE=82%, SR=18%
# Cash earns 0%
CURRENT_PORTFOLIO = {
    # GE (82%)
    "acwi.us":  0.70 * 0.82,   # IUSQ.DE proxy
    "xmld.uk":  0.15 * 0.82,   # XMLD.DE direct
    "gld.us":   0.07 * 0.82,   # PPFB.DE proxy
    # GE cash 8% * 82% = 6.56% -> 0
    # SR (18%)
    "ura.us":   0.30 * 0.18,   # URNU.DE proxy
    "smh.us":   0.20 * 0.18,   # VVSM.DE proxy
    "0rq9.uk":  0.18 * 0.18,   # Lundin Mining direct
    "paas.us":  0.07 * 0.18,   # Pan American Silver
    # SR cash 25% * 18% = 4.5% -> 0
}
# Cash total: (0.08*0.82 + 0.25*0.18) = 0.1106 -> not invested, so normalize
CURRENT_TOTAL = sum(CURRENT_PORTFOLIO.values())
CURRENT_PORTFOLIO = {t: w/CURRENT_TOTAL for t, w in CURRENT_PORTFOLIO.items()}

CURRENT_NAMES = {
    "acwi.us":  "MSCI ACWI (IUSQ proxy)",
    "xmld.uk":  "L&G AI ETF (XMLD)",
    "gld.us":   "Gold ETC (PPFB proxy)",
    "ura.us":   "Uranium ETF (URNU proxy)",
    "smh.us":   "Semiconductors (VVSM proxy)",
    "0rq9.uk":  "Lundin Mining",
    "paas.us":  "Pan American Silver",
}

REACTOR_PORTFOLIO = {
    "xauusd":  0.379, "lly.us":  0.156, "wmt.us":  0.114,
    "avgo.us": 0.050, "lite.us": 0.050, "cost.us": 0.050,
    "tsla.us": 0.050, "ccj.us":  0.050, "vrt.us":  0.050,
    "jnj.us":  0.050,
}

REACTOR_NAMES = {
    "xauusd":  "Gold",        "lly.us":  "Eli Lilly",
    "wmt.us":  "Walmart",     "avgo.us": "Broadcom",
    "lite.us": "Lumentum",    "cost.us": "Costco",
    "tsla.us": "Tesla",       "ccj.us":  "Cameco",
    "vrt.us":  "Vertiv",      "jnj.us":  "J&J",
}

REGIMES = {
    "Pre-COVID Bull":      ("2016-04-01", "2020-01-31"),
    "COVID Crash":         ("2020-02-01", "2020-03-31"),
    "COVID Recovery":      ("2020-04-01", "2021-12-31"),
    "Rate Hike/Inflation": ("2022-01-01", "2023-07-31"),
    "Post-Hike/AI Bull":   ("2023-08-01", "2024-08-31"),
    "Rate Cut":            ("2024-09-01", "2026-04-01"),
}

# ── Load ──────────────────────────────────────────────────────────────────────
def load(ticker):
    fname = ticker.replace("^","").replace("-","_").replace(".","_") + ".csv"
    # try exact match first
    for f in [ticker.replace("^","").replace("-","_") + ".csv",
              ticker.replace("^","").replace("-","_").replace(".","_") + ".csv"]:
        path = os.path.join(DATA_DIR, f)
        if os.path.exists(path):
            df = pd.read_csv(path)
            df["Date"] = pd.to_datetime(df["Date"])
            df = df.set_index("Date").sort_index()
            return df["Close"].replace(0, np.nan).dropna()
    return None

def load_spx():
    path = os.path.join(DATA_DIR, "spx.csv")
    if os.path.exists(path):
        df = pd.read_csv(path)
        df["Date"] = pd.to_datetime(df["Date"])
        df = df.set_index("Date").sort_index()
        return df["Close"].replace(0, np.nan).dropna()
    return None

# ── Portfolio metrics ─────────────────────────────────────────────────────────
def build_portfolio(weights, start="2021-01-01"):
    price_dict = {}
    for t in weights:
        s = load(t)
        if s is not None:
            s = s[s.index >= start]
            if len(s) > 10:
                price_dict[t] = s
    if not price_dict:
        return None
    df = pd.DataFrame(price_dict).ffill().dropna()
    used = list(df.columns)
    total_w = sum(weights[t] for t in used)
    w_arr = np.array([weights[t]/total_w for t in used])
    normed = df / df.iloc[0]
    port = normed.dot(w_arr)
    return port

def metrics(port):
    if port is None or len(port) < 10:
        return {}
    rets = port.pct_change().dropna()
    n_years = len(port) / 252
    ann_ret = (port.iloc[-1] ** (1/n_years)) - 1
    ann_vol = rets.std() * np.sqrt(252)
    sharpe  = ann_ret / ann_vol if ann_vol > 0 else 0
    roll_max = port.cummax()
    max_dd  = ((port - roll_max) / roll_max).min()
    calmar  = ann_ret / abs(max_dd) if max_dd < 0 else 0
    total   = port.iloc[-1] - 1
    return {
        "ann_return":  round(ann_ret*100, 2),
        "ann_vol":     round(ann_vol*100, 2),
        "sharpe":      round(sharpe, 3),
        "max_dd":      round(max_dd*100, 2),
        "calmar":      round(calmar, 3),
        "total_return":round(total*100, 2),
        "start":       str(port.index[0].date()),
        "end":         str(port.index[-1].date()),
        "n_days":      len(port),
    }

def regime_metrics(port, start, end):
    if port is None: return None
    s = port.loc[start:end]
    if len(s) < 5: return None
    s = s / s.iloc[0]
    ret = s.iloc[-1] - 1
    max_dd = ((s - s.cummax()) / s.cummax()).min()
    vol = s.pct_change().std() * np.sqrt(252)
    return {"ret": round(ret*100,2), "max_dd": round(max_dd*100,2), "vol": round(vol*100,2)}

# ── Build portfolios ──────────────────────────────────────────────────────────
print("Building portfolios...")

# Common start: 2021 (earliest both portfolios can share)
START = "2021-04-01"

port_current = build_portfolio(CURRENT_PORTFOLIO, start=START)
port_reactor  = build_portfolio(REACTOR_PORTFOLIO, start=START)
spx = load_spx()
if spx is not None:
    spx_s = spx[spx.index >= START]
    spx_norm = spx_s / spx_s.iloc[0]
else:
    spx_norm = None

m_current = metrics(port_current)
m_reactor  = metrics(port_reactor)
m_spx      = metrics(spx_norm) if spx_norm is not None else {}

print(f"  Current portfolio: {m_current}")
print(f"  Reactor Core:      {m_reactor}")

# ── Regime comparison ─────────────────────────────────────────────────────────
regime_rows = []
for rname, (rstart, rend) in REGIMES.items():
    rc = regime_metrics(port_current, rstart, rend)
    rr = regime_metrics(port_reactor,  rstart, rend)
    rs = regime_metrics(spx_norm, rstart, rend) if spx_norm is not None else None
    regime_rows.append({"regime": rname, "current": rc, "reactor": rr, "spx": rs})

# ── Excel output ──────────────────────────────────────────────────────────────
DARK_BLUE="1F4E79"; MID_BLUE="2E75B6"; LIGHT_BLUE="D6E4F0"
ALT="EBF3FB"; GREEN="C6EFCE"; RED="FFC7CE"; YELLOW="FFEB9C"; ORANGE="FFCC99"
WHITE="FFFFFF"; GREY="F2F2F2"
CURRENT_COLOR = "FFF2CC"
REACTOR_COLOR = "E2EFDA"
SPX_COLOR     = "DDEBF7"

thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr(ws, row, col, val, bg=DARK_BLUE, fg=WHITE, bold=True, size=11, halign="center"):
    c = ws.cell(row=row, column=col, value=val)
    c.fill = PatternFill("solid", fgColor=bg); c.font = Font(bold=bold, color=fg, size=size)
    c.alignment = Alignment(wrap_text=True, vertical="center", horizontal=halign)
    c.border = BORDER; return c

def cel(ws, row, col, val, bg=None, bold=False, halign="center", color="000000"):
    c = ws.cell(row=row, column=col, value=val)
    if bg: c.fill = PatternFill("solid", fgColor=bg)
    c.font = Font(bold=bold, color=color)
    c.alignment = Alignment(wrap_text=True, vertical="center", horizontal=halign)
    c.border = BORDER; return c

def set_cols(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

wb = openpyxl.Workbook()

# ━━ SHEET 1: Head-to-Head ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ws1 = wb.active; ws1.title = "Head-to-Head"

ws1.merge_cells("A1:D1")
c = ws1["A1"]; c.value = "Portfolio Comparison — Current vs Reactor Core vs S&P 500"
c.font = Font(bold=True, color=WHITE, size=14)
c.fill = PatternFill("solid", fgColor=DARK_BLUE)
c.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 40

ws1.merge_cells("A2:D2")
c = ws1["A2"]; c.value = f"Common window: {START} -> present  |  Cash excluded from portfolio returns (earns 0%)"
c.font = Font(italic=True, color="444444", size=10)
c.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
c.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[2].height = 18

for c2, (h, bg) in enumerate([("Metric","F2F2F2"),("Current Portfolio",CURRENT_COLOR),("Reactor Core (10pos)",REACTOR_COLOR),("S&P 500",SPX_COLOR)], 1):
    hdr(ws1, 3, c2, h, bg=bg, fg="000000" if bg!="1F4E79" else WHITE, bold=True)
ws1.row_dimensions[3].height = 28

def winner(a, b, higher_better=True):
    if a is None or b is None: return None, None
    if higher_better: return (GREEN if a > b else RED), (GREEN if b >= a else RED)
    else: return (GREEN if a < b else RED), (GREEN if b <= a else RED)

metrics_display = [
    ("Ann. Return",   "ann_return",   True,  lambda v: f"{v:+.2f}%"),
    ("Ann. Volatility","ann_vol",     False, lambda v: f"{v:.2f}%"),
    ("Sharpe Ratio",  "sharpe",       True,  lambda v: f"{v:.3f}"),
    ("Max Drawdown",  "max_dd",       False, lambda v: f"{v:.2f}%"),
    ("Calmar Ratio",  "calmar",       True,  lambda v: f"{v:.3f}"),
    ("Total Return",  "total_return", True,  lambda v: f"{v:+.1f}%"),
    ("Period Start",  "start",        None,  lambda v: str(v)),
    ("N Trading Days","n_days",       None,  lambda v: str(v)),
]

for i, (label, key, higher_better, fmt) in enumerate(metrics_display, 4):
    bg = ALT if i % 2 == 0 else WHITE
    cel(ws1, i, 1, label, bg=GREY, bold=True, halign="left")
    vc = m_current.get(key); vr = m_reactor.get(key); vs = m_spx.get(key)
    if higher_better is not None:
        bgc, bgr = winner(vc, vr, higher_better)
    else:
        bgc = bgr = bg
    cel(ws1, i, 2, fmt(vc) if vc is not None else "N/A", bg=bgc or bg, bold=(bgc==GREEN))
    cel(ws1, i, 3, fmt(vr) if vr is not None else "N/A", bg=bgr or bg, bold=(bgr==GREEN))
    cel(ws1, i, 4, fmt(vs) if vs is not None else "N/A", bg=bg)
    ws1.row_dimensions[i].height = 22

set_cols(ws1, [22, 22, 22, 18])

# ━━ SHEET 2: Regime Comparison ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ws2 = wb.create_sheet("Regime Comparison")

ws2.merge_cells("A1:J1")
c = ws2["A1"]; c.value = "Regime-by-Regime Comparison"
c.font = Font(bold=True, color=WHITE, size=13)
c.fill = PatternFill("solid", fgColor=DARK_BLUE)
c.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 30

hdrs2 = ["Regime", "Current Ret", "Reactor Ret", "S&P 500 Ret",
         "Current MaxDD", "Reactor MaxDD",
         "vs S&P (Current)", "vs S&P (Reactor)", "Winner"]
for c2, h in enumerate(hdrs2, 1): hdr(ws2, 2, c2, h)
ws2.row_dimensions[2].height = 30

for i, row in enumerate(regime_rows, 3):
    bg = ALT if i % 2 == 0 else WHITE
    rc = row["current"]; rr = row["reactor"]; rs = row["spx"]
    cel(ws2, i, 1, row["regime"], bg=bg, halign="left", bold=True)

    # Returns
    rc_ret = rc["ret"] if rc else None
    rr_ret = rr["ret"] if rr else None
    rs_ret = rs["ret"] if rs else None

    bgc, bgr = winner(rc_ret, rr_ret, True)
    cel(ws2, i, 2, f"{rc_ret:+.1f}%" if rc_ret is not None else "N/A", bg=bgc or bg)
    cel(ws2, i, 3, f"{rr_ret:+.1f}%" if rr_ret is not None else "N/A", bg=bgr or bg)
    cel(ws2, i, 4, f"{rs_ret:+.1f}%" if rs_ret is not None else "N/A", bg=bg)

    # MaxDD
    rc_dd = rc["max_dd"] if rc else None
    rr_dd = rr["max_dd"] if rr else None
    bgdc, bgdr = winner(rc_dd, rr_dd, False)
    cel(ws2, i, 5, f"{rc_dd:.1f}%" if rc_dd is not None else "N/A", bg=bgdc or bg)
    cel(ws2, i, 6, f"{rr_dd:.1f}%" if rr_dd is not None else "N/A", bg=bgdr or bg)

    # vs SPX
    vc_spx = round(rc_ret - rs_ret, 1) if rc_ret and rs_ret else None
    vr_spx = round(rr_ret - rs_ret, 1) if rr_ret and rs_ret else None
    cel(ws2, i, 7, f"{vc_spx:+.1f}%" if vc_spx is not None else "N/A",
        bg=GREEN if vc_spx and vc_spx > 0 else (RED if vc_spx and vc_spx < 0 else bg))
    cel(ws2, i, 8, f"{vr_spx:+.1f}%" if vr_spx is not None else "N/A",
        bg=GREEN if vr_spx and vr_spx > 0 else (RED if vr_spx and vr_spx < 0 else bg))

    # Winner
    if rc_ret is not None and rr_ret is not None:
        w_label = "Reactor Core" if rr_ret > rc_ret else ("Current" if rc_ret > rr_ret else "Tie")
        w_bg = REACTOR_COLOR if w_label == "Reactor Core" else (CURRENT_COLOR if w_label == "Current" else YELLOW)
    else:
        w_label, w_bg = "N/A", bg
    cel(ws2, i, 9, w_label, bg=w_bg, bold=True)
    ws2.row_dimensions[i].height = 22

set_cols(ws2, [22, 14, 14, 14, 14, 14, 18, 18, 14])

# ━━ SHEET 3: Portfolio Holdings ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ws3 = wb.create_sheet("Holdings")
ws3.merge_cells("A1:E1")
c = ws3["A1"]; c.value = "Current Portfolio — Holdings (excl. cash)"
c.font = Font(bold=True, color=WHITE, size=12)
c.fill = PatternFill("solid", fgColor=DARK_BLUE)
c.alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 28

for c2, h in enumerate(["Ticker","Name","Sub-portfolio","Raw Weight","Normalized Weight"], 1):
    hdr(ws3, 2, c2, h, bg=CURRENT_COLOR, fg="000000")
ws3.row_dimensions[2].height = 26

ge_tickers = {"acwi.us": ("GE",0.70*0.82), "xmld.uk": ("GE",0.15*0.82), "gld.us": ("GE",0.07*0.82)}
sr_tickers = {"ura.us": ("SR",0.30*0.18), "smh.us": ("SR",0.20*0.18), "0rq9.uk": ("SR",0.18*0.18), "paas.us": ("SR",0.07*0.18)}
all_holdings = {**ge_tickers, **sr_tickers}

for i, (ticker, (sub, raw_w)) in enumerate(all_holdings.items(), 3):
    bg = ALT if i % 2 == 0 else WHITE
    sub_bg = CURRENT_COLOR if sub == "GE" else REACTOR_COLOR
    cel(ws3, i, 1, ticker, bg=bg)
    cel(ws3, i, 2, CURRENT_NAMES.get(ticker, ticker), bg=bg, halign="left")
    cel(ws3, i, 3, sub, bg=sub_bg, bold=True)
    cel(ws3, i, 4, f"{raw_w*100:.1f}%", bg=bg)
    cel(ws3, i, 5, f"{CURRENT_PORTFOLIO[ticker]*100:.1f}%", bg=bg, bold=True)
    ws3.row_dimensions[i].height = 22

# Reactor holdings
r = len(all_holdings) + 5
ws3.merge_cells(f"A{r}:E{r}")
c = ws3.cell(row=r, column=1, value="Reactor Core — Holdings")
c.font = Font(bold=True, color=WHITE, size=12)
c.fill = PatternFill("solid", fgColor=DARK_BLUE)
c.alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[r].height = 28
r += 1

for c2, h in enumerate(["Ticker","Name","Sector","—","Weight"], 1):
    hdr(ws3, r, c2, h, bg=REACTOR_COLOR, fg="000000")
ws3.row_dimensions[r].height = 26
r += 1

REACTOR_SECTORS = {
    "xauusd":"Commodity","lly.us":"Defensive","wmt.us":"Defensive",
    "avgo.us":"Semiconductor","lite.us":"Semiconductor","cost.us":"Defensive",
    "tsla.us":"Wildcard","ccj.us":"Commodity","vrt.us":"Infrastructure","jnj.us":"Defensive"
}
for i, (ticker, w) in enumerate(sorted(REACTOR_PORTFOLIO.items(), key=lambda x: x[1], reverse=True), 1):
    bg = ALT if i % 2 == 0 else WHITE
    cel(ws3, r, 1, ticker, bg=bg)
    cel(ws3, r, 2, REACTOR_NAMES.get(ticker, ticker), bg=bg, halign="left")
    cel(ws3, r, 3, REACTOR_SECTORS.get(ticker, ""), bg=bg)
    cel(ws3, r, 4, "", bg=bg)
    cel(ws3, r, 5, f"{w*100:.1f}%", bg=bg, bold=True)
    ws3.row_dimensions[r].height = 22
    r += 1

set_cols(ws3, [14, 28, 18, 8, 14])

wb.save(REPORT)
print(f"\nReport saved: {REPORT}")
