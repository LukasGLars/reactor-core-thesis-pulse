import pandas as pd
import numpy as np
import os, sys, io, warnings
from scipy.optimize import minimize
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
warnings.filterwarnings("ignore")
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

DATA_DIR = r"C:\Users\lukas.larsson\Desktop\Privat\Project Reactor Core\data"
REPORT   = r"C:\Users\lukas.larsson\Desktop\Privat\Project Reactor Core\Gold_Stress_Test.xlsx"

TICKERS = ["xauusd","lly.us","wmt.us","avgo.us","cost.us","ccj.us","vrt.us","jnj.us"]
NAMES   = {
    "xauusd":"Gold","lly.us":"Eli Lilly","wmt.us":"Walmart",
    "avgo.us":"Broadcom","cost.us":"Costco","ccj.us":"Cameco",
    "vrt.us":"Vertiv","jnj.us":"J&J",
}
MIN_W, MAX_W = 0.05, 0.40

# Gold correction scenarios — based on historical precedents
# 2011-2015: gold fell -44% peak to trough over 4 years
# 2020 COVID: gold fell -12% in 6 weeks then recovered
# 1980-1982: gold fell -65% from peak
SHOCKS = [-0.10, -0.20, -0.30, -0.40, -0.50]

# Gold cap allocations to test
GOLD_CAPS_TEST = [0.10, 0.15, 0.20, 0.25, 0.30, 0.35, 0.379]  # 37.9% = current

def load(ticker):
    fname = ticker.replace("^","").replace("-","_") + ".csv"
    path  = os.path.join(DATA_DIR, fname)
    if not os.path.exists(path): return None
    df = pd.read_csv(path)
    df["Date"] = pd.to_datetime(df["Date"])
    df = df.set_index("Date").sort_index()
    return df["Close"].replace(0, np.nan).dropna()

def optimize(df, gold_cap=None, n_starts=80):
    rets = df.pct_change().dropna()
    mu   = rets.mean(); cov = rets.cov()
    n    = len(df.columns)
    bounds = [(MIN_W, min(MAX_W, gold_cap) if (t == "xauusd" and gold_cap) else MAX_W)
              for t in df.columns]
    def neg_sharpe(w):
        r = np.dot(w, mu)*252; v = np.sqrt(w @ cov @ w)*np.sqrt(252)
        return -r/v if v > 0 else 0
    constraints = [{"type":"eq","fun":lambda w: np.sum(w)-1}]
    best = None
    for _ in range(n_starts):
        w0 = np.random.dirichlet(np.ones(n))
        w0 = np.clip(w0, [b[0] for b in bounds], [b[1] for b in bounds])
        w0 /= w0.sum()
        res = minimize(neg_sharpe, w0, method="SLSQP", bounds=bounds,
                       constraints=constraints, options={"maxiter":1000,"ftol":1e-12})
        if res.success and (best is None or res.fun < best.fun):
            best = res
    return best.x if best else None

# ── Build base price data (5Y window for scenarios) ───────────────────────────
START = "2021-04-01"
price_dict = {t: load(t) for t in TICKERS if load(t) is not None}
price_df   = pd.DataFrame({t: s[s.index >= START] for t, s in price_dict.items()}).dropna()

print("=" * 68)
print("  Gold Stress Test — portfolio impact at different allocations")
print(f"  Base window: {START} to present  |  Scenarios: {[f'{s*100:.0f}%' for s in SHOCKS]}")
print("=" * 68)

# ── For each gold cap level, get optimized weights then apply shocks ───────────
all_results = []

for gold_cap in GOLD_CAPS_TEST:
    w_arr = optimize(price_df, gold_cap=gold_cap)
    if w_arr is None: continue
    w_dict   = dict(zip(price_df.columns, w_arr))
    gold_w   = w_dict.get("xauusd", 0)
    cap_label = f"{gold_cap*100:.0f}%" if gold_cap != 0.379 else "37.9% (current)"

    # Baseline portfolio (no shock)
    normed_base = price_df / price_df.iloc[0]
    port_base   = normed_base.dot(w_arr)
    base_total  = round((port_base.iloc[-1]-1)*100, 1)
    base_ret_ann= round(((port_base.iloc[-1])**(252/len(port_base))-1)*100, 2)
    base_dd     = round(((port_base - port_base.cummax())/port_base.cummax()).min()*100, 2)
    base_vol    = round(port_base.pct_change().dropna().std()*np.sqrt(252)*100, 2)
    base_sharpe = round(base_ret_ann/base_vol*100/100, 3) if base_vol > 0 else 0

    row_base = {
        "gold_cap": cap_label, "gold_weight": round(gold_w*100,1),
        "shock": "None (baseline)", "shock_pct": 0,
        "port_total": base_total, "port_ann": base_ret_ann,
        "port_dd": base_dd, "port_vol": base_vol, "port_sharpe": base_sharpe,
        "gold_contribution": round(gold_w * 0 * 100, 1),   # 0 shock
        "port_impact_from_gold": 0,
    }
    all_results.append(row_base)

    for shock in SHOCKS:
        # Shock = gold drops X% from TODAY's price (prospective, not historical)
        # Method: append one extra day where gold price = last_price * (1 + shock)
        shocked_df = price_df.copy()
        extra_row  = shocked_df.iloc[[-1]].copy()
        extra_row.index = [extra_row.index[0] + pd.Timedelta(days=1)]
        extra_row["xauusd"] = extra_row["xauusd"].values[0] * (1 + shock)
        shocked_df = pd.concat([shocked_df, extra_row])

        normed_s = shocked_df / shocked_df.iloc[0]
        port_s   = normed_s.dot(w_arr)

        total_s  = round((port_s.iloc[-1]-1)*100, 1)
        ann_s    = round(((port_s.iloc[-1])**(252/len(port_s))-1)*100, 2)
        dd_s     = round(((port_s - port_s.cummax())/port_s.cummax()).min()*100, 2)
        vol_s    = round(port_s.pct_change().dropna().std()*np.sqrt(252)*100, 2)
        sh_s     = round(ann_s/vol_s*100/100, 3) if vol_s > 0 else 0

        # Direct impact: gold weight * shock (applied to terminal value)
        direct_impact = round(gold_w * shock * 100, 1)

        row = {
            "gold_cap": cap_label, "gold_weight": round(gold_w*100,1),
            "shock": f"Gold {shock*100:.0f}%", "shock_pct": shock*100,
            "port_total": total_s, "port_ann": ann_s,
            "port_dd": dd_s, "port_vol": vol_s, "port_sharpe": sh_s,
            "port_impact_from_gold": direct_impact,
            "delta_total": round(total_s - base_total, 1),
            "delta_sharpe": round(sh_s - base_sharpe, 3),
        }
        all_results.append(row)

    print(f"\n  Gold cap {cap_label:12}  (actual weight {gold_w*100:.1f}%)")
    print(f"  {'Shock':15}  {'Port Total':>11}  {'Delta':>8}  {'Ann Ret':>8}  "
          f"{'MaxDD':>8}  {'Sharpe':>8}  {'Gold drag':>10}")
    for r in [x for x in all_results if x["gold_cap"] == cap_label]:
        delta = r.get("delta_total","")
        delta_str = f"{delta:+.1f}pp" if delta != "" else "—"
        drag  = r.get("port_impact_from_gold",0)
        drag_str = f"{drag:+.1f}pp" if drag != 0 else "—"
        print(f"  {r['shock']:15}  {r['port_total']:>+10.1f}%  {delta_str:>8}  "
              f"{r['port_ann']:>7.1f}%  {r['port_dd']:>7.1f}%  "
              f"{r['port_sharpe']:>8.3f}  {drag_str:>10}")

# ── Crossover analysis: at what gold drop does each cap level underperform? ───
print(f"\n{'=' * 68}")
print("  Portfolio total return by gold cap level vs shock intensity")
print(f"  {'Gold cap':>12}  " + "  ".join(f"{'Gold'+str(int(s*100))+'%':>10}" for s in SHOCKS))
print(f"  {'-'*68}")

for gold_cap in GOLD_CAPS_TEST:
    cap_label = f"{gold_cap*100:.0f}%" if gold_cap != 0.379 else "37.9%"
    rows_cap  = [r for r in all_results if r["gold_cap"].startswith(cap_label.split()[0])]
    vals = []
    for shock in SHOCKS:
        match = [r for r in rows_cap if abs(r["shock_pct"] - shock*100) < 0.1]
        vals.append(f"{match[0]['port_total']:>+.1f}%" if match else "—")
    print(f"  {cap_label:>12}  " + "  ".join(f"{v:>10}" for v in vals))

# ── Excel ─────────────────────────────────────────────────────────────────────
DARK="1F4E79"; MID="2E75B6"; LIGHT="D6E4F0"; ALT="EBF3FB"
GRN="C6EFCE"; RED="FFC7CE"; YEL="FFEB9C"; WHT="FFFFFF"; GREY="F2F2F2"
ORG="FFCC99"

thin   = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr(ws, row, col, val, bg=DARK, fg=WHT, bold=True, sz=10, ha="center"):
    c = ws.cell(row=row, column=col, value=val)
    c.fill = PatternFill("solid", fgColor=bg); c.font = Font(bold=bold, color=fg, size=sz)
    c.alignment = Alignment(wrap_text=True, vertical="center", horizontal=ha)
    c.border = BORDER; return c

def cel(ws, row, col, val, bg=None, bold=False, ha="center", color="000000"):
    c = ws.cell(row=row, column=col, value=val)
    if bg: c.fill = PatternFill("solid", fgColor=bg)
    c.font = Font(bold=bold, color=color)
    c.alignment = Alignment(wrap_text=True, vertical="center", horizontal=ha)
    c.border = BORDER; return c

def title(ws, row, nc, text, bg=DARK, sz=13):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=nc)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(bold=True, color=WHT, size=sz)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 34

def cw(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

wb = openpyxl.Workbook()

# ━━ SHEET 1: Full Scenario Table ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ws1 = wb.active; ws1.title = "Scenario Detail"
title(ws1, 1, 9, "Gold Stress Test — Portfolio Impact Across Allocation Levels & Shock Scenarios")
ws1.merge_cells("A2:I2")
c = ws1["A2"]
c.value = ("Each gold allocation is re-optimized. Shock applied to gold price series — other assets unchanged.  "
           "Historical context: 2011-2015 gold fell -44% over 4 years | 1980-1982 fell -65%.")
c.font = Font(italic=True, size=9, color="444444")
c.fill = PatternFill("solid", fgColor=LIGHT)
c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws1.row_dimensions[2].height = 26

for ci, h in enumerate(["Gold Cap","Gold Weight","Shock","Port Total","Delta vs Base",
                          "Ann Return","Max DD","Sharpe","Gold Drag"], 1):
    hdr(ws1, 3, ci, h)
ws1.row_dimensions[3].height = 28

ri = 4
for r in all_results:
    is_base   = r["shock"] == "None (baseline)"
    shock_v   = r.get("shock_pct", 0)
    bg = "FFF2CC" if is_base else (ALT if ri%2==0 else WHT)

    # colour shock column by severity
    if shock_v == 0:       sh_bg = bg
    elif shock_v >= -20:   sh_bg = YEL
    elif shock_v >= -30:   sh_bg = ORG
    else:                  sh_bg = RED

    delta = r.get("delta_total","")
    delta_str = f"{delta:+.1f}pp" if delta != "" else "—"
    delta_bg  = (GRN if isinstance(delta,float) and delta >= 0 else
                 RED if isinstance(delta,float) and delta < -10 else
                 ORG if isinstance(delta,float) and delta < 0 else bg)

    drag  = r.get("port_impact_from_gold", 0)
    drag_str = f"{drag:+.1f}pp" if drag != 0 else "—"

    cel(ws1, ri, 1, r["gold_cap"],        bg=bg, bold=is_base, ha="left")
    cel(ws1, ri, 2, f"{r['gold_weight']:.1f}%", bg=bg, bold=is_base)
    cel(ws1, ri, 3, r["shock"],           bg=sh_bg)
    cel(ws1, ri, 4, f"{r['port_total']:+.1f}%", bg=bg, bold=is_base)
    cel(ws1, ri, 5, delta_str,            bg=delta_bg, bold=not is_base)
    cel(ws1, ri, 6, f"{r['port_ann']:.1f}%",    bg=bg)
    cel(ws1, ri, 7, f"{r['port_dd']:.1f}%",     bg=bg)
    cel(ws1, ri, 8, r["port_sharpe"],     bg=bg)
    cel(ws1, ri, 9, drag_str,             bg=sh_bg if drag != 0 else bg, bold=(drag != 0))
    ws1.row_dimensions[ri].height = 18
    ri += 1

cw(ws1, [16,12,14,13,14,12,10,10,12])

# ━━ SHEET 2: Crossover Matrix ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ws2 = wb.create_sheet("Crossover Matrix")
title(ws2, 1, len(SHOCKS)+2, "Portfolio Total Return Matrix — Gold Cap vs Shock Severity")
ws2.merge_cells(f"A2:{get_column_letter(len(SHOCKS)+2)}2")
c = ws2["A2"]
c.value = ("Read: for a given gold allocation (row) and gold shock (column), what is the portfolio total return?  "
           "Green = above 100% total | Yellow = 50-100% | Orange = 0-50% | Red = negative")
c.font = Font(italic=True, size=9, color="444444")
c.fill = PatternFill("solid", fgColor=LIGHT)
c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws2.row_dimensions[2].height = 26

# Headers
hdr(ws2, 3, 1, "Gold Cap")
hdr(ws2, 3, 2, "Actual Weight")
for ci, s in enumerate(SHOCKS, 3):
    hdr(ws2, 3, ci, f"Gold {s*100:.0f}%")
ws2.row_dimensions[3].height = 28

# Add baseline column header
# Also add delta columns
hdr(ws2, 3, len(SHOCKS)+3, "Baseline\n(no shock)")

for ri_off, gold_cap in enumerate(GOLD_CAPS_TEST, 4):
    cap_label = f"{gold_cap*100:.0f}%" if gold_cap != 0.379 else "37.9%"
    rows_cap  = [r for r in all_results if r["gold_cap"].startswith(cap_label.split()[0])]
    base_row  = next((r for r in rows_cap if r["shock"] == "None (baseline)"), None)
    bg_row    = "FFF2CC" if gold_cap == 0.379 else (ALT if ri_off%2==0 else WHT)
    is_cur    = gold_cap == 0.379

    cel(ws2, ri_off, 1, cap_label + (" ← current" if is_cur else ""),
        bg=bg_row, bold=is_cur, ha="left")
    cel(ws2, ri_off, 2, f"{rows_cap[0]['gold_weight']:.1f}%" if rows_cap else "—",
        bg=bg_row, bold=is_cur)

    for ci, shock in enumerate(SHOCKS, 3):
        match = next((r for r in rows_cap if abs(r["shock_pct"] - shock*100) < 0.1), None)
        if match:
            v = match["port_total"]
            v_bg = GRN if v >= 100 else (YEL if v >= 50 else (ORG if v >= 0 else RED))
            cel(ws2, ri_off, ci, f"{v:+.1f}%", bg=v_bg, bold=is_cur)
        else:
            cel(ws2, ri_off, ci, "—", bg=bg_row)

    if base_row:
        v = base_row["port_total"]
        cel(ws2, ri_off, len(SHOCKS)+3, f"{v:+.1f}%", bg=GRN, bold=True)

    ws2.row_dimensions[ri_off].height = 22

# ━━ SHEET 3: Sharpe Matrix ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ws3 = wb.create_sheet("Sharpe Matrix")
title(ws3, 1, len(SHOCKS)+3, "Sharpe Ratio Matrix — Gold Cap vs Shock Severity")
ws3.merge_cells(f"A2:{get_column_letter(len(SHOCKS)+3)}2")
c = ws3["A2"]
c.value = "Same structure as Crossover Matrix but showing Sharpe ratio. Green ≥1.5 | Yellow 1.0-1.5 | Orange 0.5-1.0 | Red <0.5"
c.font = Font(italic=True, size=9, color="444444")
c.fill = PatternFill("solid", fgColor=LIGHT)
c.alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[2].height = 22

hdr(ws3, 3, 1, "Gold Cap"); hdr(ws3, 3, 2, "Gold Weight")
for ci, s in enumerate(SHOCKS, 3): hdr(ws3, 3, ci, f"Gold {s*100:.0f}%")
hdr(ws3, 3, len(SHOCKS)+3, "Baseline")
ws3.row_dimensions[3].height = 28

for ri_off, gold_cap in enumerate(GOLD_CAPS_TEST, 4):
    cap_label = f"{gold_cap*100:.0f}%" if gold_cap != 0.379 else "37.9%"
    rows_cap  = [r for r in all_results if r["gold_cap"].startswith(cap_label.split()[0])]
    base_row  = next((r for r in rows_cap if r["shock"] == "None (baseline)"), None)
    bg_row    = "FFF2CC" if gold_cap == 0.379 else (ALT if ri_off%2==0 else WHT)
    is_cur    = gold_cap == 0.379

    cel(ws3, ri_off, 1, cap_label + (" ← current" if is_cur else ""),
        bg=bg_row, bold=is_cur, ha="left")
    cel(ws3, ri_off, 2, f"{rows_cap[0]['gold_weight']:.1f}%" if rows_cap else "—",
        bg=bg_row, bold=is_cur)

    for ci, shock in enumerate(SHOCKS, 3):
        match = next((r for r in rows_cap if abs(r["shock_pct"] - shock*100) < 0.1), None)
        if match:
            v = match["port_sharpe"]
            v_bg = GRN if v >= 1.5 else (YEL if v >= 1.0 else (ORG if v >= 0.5 else RED))
            cel(ws3, ri_off, ci, str(v), bg=v_bg, bold=is_cur)
        else:
            cel(ws3, ri_off, ci, "—", bg=bg_row)

    if base_row:
        v = base_row["port_sharpe"]
        v_bg = GRN if v >= 1.5 else YEL
        cel(ws3, ri_off, len(SHOCKS)+3, str(v), bg=v_bg, bold=True)
    ws3.row_dimensions[ri_off].height = 22

for ws in [ws2, ws3]:
    cw(ws, [18,13] + [12]*len(SHOCKS) + [13])

wb.save(REPORT)
print(f"\n  Report saved: {REPORT}")
print("=" * 68)
