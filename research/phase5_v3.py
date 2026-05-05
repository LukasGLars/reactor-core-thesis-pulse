import pandas as pd
import numpy as np
import os, sys, io, warnings
warnings.filterwarnings("ignore")
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DATA_DIR   = r"C:\Users\lukas.larsson\Desktop\Privat\Project Reactor Core\data"
REPORT     = r"C:\Users\lukas.larsson\Desktop\Privat\Project Reactor Core\Reactor_Core_v3.xlsx"

# ── v3 Portfolio Definition ───────────────────────────────────────────────────
# Changes vs v2: Costco removed (abundance business), Silver added at 10%,
# weights rebalanced toward scarcity thesis. No structural cash — DCA reserve held externally.
TICKERS = {
    "xauusd":  "Gold",
    "xagusd":  "Silver",
    "lly.us":  "Eli Lilly",
    "wmt.us":  "Walmart",
    "ccj.us":  "Cameco",
    "vrt.us":  "Vertiv",
    "avgo.us": "Broadcom",
    "jnj.us":  "J&J",
}

WEIGHTS_10Y = {
    "xauusd": 0.250, "xagusd": 0.100, "lly.us": 0.150,
    "wmt.us": 0.150, "ccj.us": 0.100, "vrt.us": 0.100,
    "avgo.us": 0.090, "jnj.us": 0.060,
}

# Live prices as of 2026-03-31 (USD); USD/SEK as of 2026-04-01
# Gold: Guld AVA certificate (Avanza 1054955) — SEK-native, price in SEK directly
LIVE_PRICES_USD = {
    "xauusd": None,   "xagusd": 74.985, "lly.us": 919.77,
    "wmt.us": 124.28, "ccj.us": 108.61, "vrt.us": 250.58,
    "avgo.us": 309.51, "jnj.us": 244.44,
}
LIVE_PRICES_SEK_OVERRIDE = {
    "xauusd": 1125.09,  # Guld AVA (Avanza 1054955), SEK-denominated certificate
}
USDSEK = 9.3943
PRICE_DATE = "2026-03-31"

ROLES = {
    "xauusd": "Hedge",     "xagusd":  "Hedge",
    "lly.us": "Carry",     "wmt.us":  "Carry",
    "ccj.us": "Cyclical",  "vrt.us":  "Convexity",
    "avgo.us":"Convexity", "jnj.us":  "Carry",
}

ROLE_COLORS = {
    "Hedge":     "FFF2CC",
    "Carry":     "E2EFDA",
    "Convexity": "DDEBF7",
    "Cyclical":  "FCE4D6",
}

REGIMES = {
    "Pre-COVID Bull":      ("2016-04-01","2020-01-31"),
    "COVID Crash":         ("2020-02-01","2020-03-31"),
    "COVID Recovery":      ("2020-04-01","2021-12-31"),
    "Rate Hike/Inflation": ("2022-01-01","2023-07-31"),
    "Post-Hike/AI Bull":   ("2023-08-01","2024-08-31"),
    "Rate Cut":            ("2024-09-01","2026-04-01"),
}

WINDOWS = {"3Y": "2023-04-01", "5Y": "2021-04-01", "10Y": "2016-04-01"}

DARK="1F4E79"; MID="2E75B6"; LIGHT="D6E4F0"
ALT="EBF3FB"; GRN="C6EFCE"; RED="FFC7CE"; YEL="FFEB9C"; ORG="FFCC99"
GREY="F2F2F2"; WHT="FFFFFF"

thin   = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr(ws, row, col, val, bg=DARK, fg=WHT, bold=True, sz=11, ha="center"):
    c = ws.cell(row=row, column=col, value=val)
    c.fill = PatternFill("solid", fgColor=bg); c.font = Font(bold=bold, color=fg, size=sz)
    c.alignment = Alignment(wrap_text=True, vertical="center", horizontal=ha)
    c.border = BORDER; return c

def cel(ws, row, col, val, bg=None, bold=False, ha="center", color="000000"):
    c = ws.cell(row=row, column=col, value=val)
    if bg: c.fill = PatternFill("solid", fgColor=bg)
    c.font = Font(bold=bold, color=color)
    c.alignment = Alignment(wrap_text=True, vertical="center", horizontal=ha)
    c.border = BORDER; return c

def cw(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ── Data loading ──────────────────────────────────────────────────────────────
def load(ticker):
    fname = ticker.replace("^","").replace("-","_") + ".csv"
    path  = os.path.join(DATA_DIR, fname)
    if not os.path.exists(path): return None
    df = pd.read_csv(path)
    df["Date"] = pd.to_datetime(df["Date"])
    df = df.set_index("Date").sort_index()
    return df["Close"].replace(0, np.nan).dropna()

price_dict = {t: load(t) for t in TICKERS if load(t) is not None}

def build_df(start):
    df = pd.DataFrame({t: s[s.index >= start] for t, s in price_dict.items()}).dropna()
    return df

def port_metrics(df, weights):
    tickers = list(df.columns)
    w = np.array([weights.get(t, 0) for t in tickers])
    w /= w.sum()
    rets    = df.pct_change().dropna()
    mu      = rets.mean(); cov = rets.cov()
    ann_ret = np.dot(w, mu) * 252
    ann_vol = np.sqrt(w @ cov @ w) * np.sqrt(252)
    sharpe  = ann_ret / ann_vol if ann_vol > 0 else 0
    port    = (df / df.iloc[0]).dot(w)
    max_dd  = ((port - port.cummax()) / port.cummax()).min()
    calmar  = ann_ret / abs(max_dd) if max_dd < 0 else 0
    total   = port.iloc[-1] - 1
    return {"sharpe": round(sharpe,3), "ann_ret": round(ann_ret*100,2),
            "ann_vol": round(ann_vol*100,2), "max_dd": round(max_dd*100,2),
            "calmar": round(calmar,3), "total": round(total*100,1)}

def asset_metrics(series, start):
    s = series[series.index >= start].dropna()
    if len(s) < 20: return {}
    rets    = s.pct_change().dropna()
    n       = len(s) / 252
    ann_ret = (s.iloc[-1] ** (1/n) - 1) if n > 0 else 0
    ann_vol = rets.std() * np.sqrt(252)
    sharpe  = ann_ret / ann_vol if ann_vol > 0 else 0
    max_dd  = ((s - s.cummax()) / s.cummax()).min()
    return {"ann_ret": round(ann_ret*100,2), "sharpe": round(sharpe,3),
            "max_dd": round(max_dd*100,2)}

def regime_port_ret(df, weights, rstart, rend):
    s = df.loc[rstart:rend].dropna(how="all")
    if len(s) < 5: return None
    tickers = list(s.columns)
    w = np.array([weights.get(t,0) for t in tickers]); w /= w.sum()
    port = (s / s.iloc[0]).dot(w)
    ret  = port.iloc[-1] - 1
    dd   = ((port - port.cummax()) / port.cummax()).min()
    return {"ret": round(ret*100,1), "max_dd": round(dd*100,1)}

def regime_asset_ret(series, rstart, rend):
    s = series.loc[rstart:rend].dropna()
    if len(s) < 5: return None
    return round((s.iloc[-1]/s.iloc[0]-1)*100, 1)

# ── Compute all metrics ───────────────────────────────────────────────────────
print("Computing metrics...")
window_metrics = {}
for wname, wstart in WINDOWS.items():
    df = build_df(wstart)
    window_metrics[wname] = port_metrics(df, WEIGHTS_10Y)
    print(f"  {wname}: Sharpe {window_metrics[wname]['sharpe']}  "
          f"AnnRet {window_metrics[wname]['ann_ret']}%  MaxDD {window_metrics[wname]['max_dd']}%")

# SPX benchmark
spx = load("^spx")
if spx is None: spx = load("spx")

# Regime portfolio performance
# Note: silver data starts ~2018; pre-2018 regimes use available tickers only
regime_results = []
df_full = build_df("2016-04-01")
actual_start = df_full.index[0].strftime("%Y-%m-%d") if len(df_full) > 0 else "2016-04-01"
for rname, (rs, re) in REGIMES.items():
    pm = regime_port_ret(df_full, WEIGHTS_10Y, rs, re)
    spx_s = spx.loc[rs:re].dropna() if spx is not None else None
    spx_r = round((spx_s.iloc[-1]/spx_s.iloc[0]-1)*100,1) if spx_s is not None and len(spx_s)>5 else None
    vs    = round(pm["ret"] - spx_r, 1) if pm and spx_r is not None else None
    regime_results.append({"regime": rname, "start": rs, "end": re,
                            "port_ret": pm["ret"] if pm else None,
                            "port_dd":  pm["max_dd"] if pm else None,
                            "spx_ret":  spx_r, "vs_spx": vs})

# Per-asset regime returns
asset_regime = {}
for t in TICKERS:
    s = price_dict.get(t)
    if s is None: continue
    rets = []
    asset_regime[t] = {}
    for rname, (rs, re) in REGIMES.items():
        r = regime_asset_ret(s, rs, re)
        asset_regime[t][rname] = r
        if r is not None: rets.append(r)
    asset_regime[t]["wins"] = sum(1 for v in rets if v > 0)
    asset_regime[t]["avg"]  = round(np.mean(rets), 1) if rets else None

# ── Excel ─────────────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

# ━━ SHEET 1: Summary ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ws1 = wb.active; ws1.title = "Summary"

ws1.merge_cells("A1:F1")
c = ws1["A1"]; c.value = "PROJECT REACTOR CORE — v3"
c.font = Font(bold=True, color=WHT, size=16)
c.fill = PatternFill("solid", fgColor=DARK)
c.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 44

ws1.merge_cells("A2:F2")
c = ws1["A2"]
c.value = ("8 positions | Short Abundance, Long Scarcity | "
           "Gold + Silver = 35% scarcity hedge | Cameco 10% uranium | "
           "Costco removed | DCA reserve held externally")
c.font = Font(italic=True, color="444444", size=10)
c.fill = PatternFill("solid", fgColor=LIGHT)
c.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[2].height = 20

def block(ws, r, title, rows_data, ncols=6):
    ws.merge_cells(f"A{r}:{get_column_letter(ncols)}{r}")
    c = ws.cell(row=r, column=1, value=title)
    c.font = Font(bold=True, color=WHT, size=11)
    c.fill = PatternFill("solid", fgColor=MID)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[r].height = 22; r += 1
    for label, val, bg in rows_data:
        cel(ws, r, 1, label, bg=GREY, bold=True, ha="left")
        ws.merge_cells(f"B{r}:{get_column_letter(ncols)}{r}")
        c2 = ws.cell(row=r, column=2, value=val)
        c2.fill = PatternFill("solid", fgColor=bg or WHT)
        c2.alignment = Alignment(vertical="center", indent=1); c2.border = BORDER
        ws.row_dimensions[r].height = 20; r += 1
    return r + 1

m10 = window_metrics["10Y"]
r = 4
r = block(ws1, r, "v3 Changes vs v2", [
    ("Removed",      "Costco — abundance business (membership retailer, consumer discretionary proxy). "
                     "Thesis filter: Short Abundance means avoiding names that benefit from cheap/plentiful goods.", RED),
    ("Added",        "Silver (xagusd) at 10% — structural supply deficit, companion to gold. "
                     "Gold/silver ratio compression thesis (current ~65, target <55 for re-entry signal). "
                     "Gold + Silver = 35% total scarcity hedge.", GRN),
    ("Rebalanced",   "Walmart 22.7%→15%, Lilly 19.7%→15% (freed weight to scarcity assets). "
                     "Cameco 5.7%→10% (uranium — central thesis). Vertiv 9.0%→10%. Broadcom 7.9%→8%.", YEL),
    ("Cash",         "2% permanent cash floor — structural feature, not deployed, not part of optimization. "
                     "Investable weights sum to 98%.", GREY),
    ("Net result",   "Thesis Sharpe 1.783 vs v2 1.851 (-0.068). Trade-off accepted: thesis wins scarcity/commodity years "
                     "(2018, 2020, 2025). 35% hedge vs 25%. Silver adds asymmetric upside if ratio compresses. "
                     "Note: silver data from ~2018 — 10Y window limited to available overlap.", YEL),
])

r = block(ws1, r, f"Performance (data from {actual_start})", [
    ("Ann. Return",     f"{m10['ann_ret']}%",   GRN),
    ("Ann. Volatility", f"{m10['ann_vol']}%",   YEL),
    ("Sharpe Ratio",    f"{m10['sharpe']}",     GRN),
    ("Max Drawdown",    f"{m10['max_dd']}%",    YEL),
    ("Calmar Ratio",    f"{m10['calmar']}",     GRN),
    ("Total Return",    f"+{m10['total']}%",    GRN),
])

regime_lines = []
for row in regime_results:
    vs  = row["vs_spx"]
    ret = row["port_ret"]
    bg  = GRN if vs and vs > 0 else (RED if vs and vs < -5 else YEL)
    vs_str = f"{vs:+.1f}% vs S&P" if vs is not None else ""
    ret_str = f"Portfolio {ret:+.1f}%  |  {vs_str}" if ret is not None else "Insufficient data (silver gap)"
    regime_lines.append((
        f"{row['regime']} ({row['start'][:4]}-{row['end'][:4]})",
        ret_str,
        bg if ret is not None else GREY,
    ))
r = block(ws1, r, "Portfolio vs S&P 500 by Regime", regime_lines)

holding_lines = [
    ("Gold (25.0%)",    "Scarcity anchor. Monetary hedge. Capped at 25%. SEK-native via Guld AVA (Avanza 1054955).", YEL),
    ("Silver (10.0%)",  "Scarcity anchor. Structural supply deficit — above-ground stocks declining, industrial demand rising. "
                        "Gold/silver ratio compression thesis: current ~65, target <55. "
                        "Companion to gold — adds asymmetric upside in commodity supercycles.", YEL),
    ("Eli Lilly (15.0%)","Carry compounder. Patent-protected GLP-1 monopoly. Reduced from 19.7% to fund scarcity rebalance. "
                        "Constrained manufacturing capacity.", GRN),
    ("Walmart (15.0%)", "Carry compounder. Pricing power moat. Reduced from 22.7% to fund scarcity rebalance. "
                        "Positive 5/6 regimes.", GRN),
    ("Cameco (10.0%)",  "Cyclical scarcity. Uranium supply deficit, decade-long mine lead times. "
                        "Increased from 5.7% — central thesis holding. Long Scarcity.", GRN),
    ("Vertiv (10.0%)",  "Convexity. Power infrastructure scarcity. Cannot replicate data center power at this scale. "
                        "Increased from 9.0%.", GRN),
    ("Broadcom (9.0%)", "Convexity. Custom silicon + networking. Positive 5/6 regimes. Highest single-stock conviction.", GRN),
    ("J&J (6.0%)",      "Carry defensive. Portfolio stabiliser. Near-zero drawdown in most regimes.", GRN),
]
r = block(ws1, r, "Holdings", holding_lines)

cw(ws1, [24, 14, 14, 14, 14, 14])

# ━━ SHEET 2: Final Portfolio ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ws2 = wb.create_sheet("Final Portfolio")
ws2.merge_cells("A1:K1")
c = ws2["A1"]; c.value = "Final Portfolio — Reactor Core v3 (Short Abundance, Long Scarcity)"
c.font = Font(bold=True, color=WHT, size=13)
c.fill = PatternFill("solid", fgColor=DARK)
c.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 30

# Subtitle row with FX note
ws2.merge_cells("A2:K2")
c2 = ws2["A2"]
c2.value = (f"Live prices as of {PRICE_DATE}  |  USD/SEK {USDSEK}  |  "
            f"Min capital = price of most expensive position / its weight  |  "
            f"Silver data from ~2018 (limits historical window)")
c2.font = Font(italic=True, size=9, color="444444")
c2.fill = PatternFill("solid", fgColor=LIGHT)
c2.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws2.row_dimensions[2].height = 22

for ci, h in enumerate(["#","Name","Ticker","Role","Weight","Regime Wins","Avg Regime Ret",
                         "10Y Sharpe","Price (USD)","Price (SEK)","Min capital (kr)"], 1):
    hdr(ws2, 3, ci, h)
ws2.row_dimensions[3].height = 28

sorted_w = sorted(WEIGHTS_10Y.items(), key=lambda x: x[1], reverse=True)
for i, (t, w) in enumerate(sorted_w, 1):
    bg   = ALT if i%2==0 else WHT
    role = ROLES.get(t,"")
    wins = asset_regime.get(t,{}).get("wins",0)
    avg  = asset_regime.get(t,{}).get("avg",None)
    am   = asset_metrics(price_dict[t], "2016-04-01") if t in price_dict else {}
    win_bg = GRN if wins >= 5 else (YEL if wins >= 4 else ORG)
    p_usd  = LIVE_PRICES_USD.get(t)
    if t in LIVE_PRICES_SEK_OVERRIDE:
        p_sek   = LIVE_PRICES_SEK_OVERRIDE[t]
        p_usd_display = "SEK-native"
    else:
        p_sek   = round(p_usd * USDSEK, 0)
        p_usd_display = f"${p_usd:,.2f}"
    min_cap = round(p_sek / w, 0) if w > 0 else 0
    def _min(tt, ww):
        if tt in LIVE_PRICES_SEK_OVERRIDE: return round(LIVE_PRICES_SEK_OVERRIDE[tt]/ww, 0)
        pu = LIVE_PRICES_USD.get(tt, 0); return round(pu*USDSEK/ww, 0) if ww > 0 else 0
    max_min_cap = max(_min(tt, ww) for tt, ww in WEIGHTS_10Y.items())
    cel(ws2, i+3, 1,  i,                                    bg=bg)
    cel(ws2, i+3, 2,  TICKERS[t],                           bg=bg, bold=True, ha="left")
    cel(ws2, i+3, 3,  t,                                    bg=bg)
    cel(ws2, i+3, 4,  role,                                 bg=ROLE_COLORS.get(role,WHT))
    cel(ws2, i+3, 5,  f"{w*100:.1f}%",                     bg=bg, bold=True)
    cel(ws2, i+3, 6,  f"{wins}/6",                          bg=win_bg, bold=True)
    cel(ws2, i+3, 7,  f"{avg:+.1f}%" if avg else "N/A",    bg=bg)
    cel(ws2, i+3, 8,  am.get("sharpe","N/A"),               bg=bg)
    cel(ws2, i+3, 9,  p_usd_display,                         bg=bg)
    cel(ws2, i+3, 10, int(round(p_sek)),                     bg=bg)
    cel(ws2, i+3, 11, int(min_cap),                           bg=YEL if min_cap == max_min_cap else bg)
    ws2.row_dimensions[i+3].height = 22

# Min capital summary row
def _sek_price(tt):
    if tt in LIVE_PRICES_SEK_OVERRIDE: return LIVE_PRICES_SEK_OVERRIDE[tt]
    return (LIVE_PRICES_USD[tt] or 0) * USDSEK
def _min_cap(tt, ww): return round(_sek_price(tt)/ww, 0) if ww > 0 else 0
binding_asset = max(WEIGHTS_10Y.items(), key=lambda x: _min_cap(x[0], x[1]))
min_total_kr  = _min_cap(binding_asset[0], binding_asset[1])
b_price_str   = (f"{LIVE_PRICES_SEK_OVERRIDE[binding_asset[0]]:,.2f} kr/share (SEK-native)"
                 if binding_asset[0] in LIVE_PRICES_SEK_OVERRIDE
                 else f"{LIVE_PRICES_USD[binding_asset[0]]:,.2f} USD/share")
ri_min = len(WEIGHTS_10Y) + 5
ws2.merge_cells(f"A{ri_min}:K{ri_min}")
c_min = ws2.cell(row=ri_min, column=1,
    value=(f"Minimum initial capital to hold >=1 share of every position: "
           f"{int(min_total_kr)} kr  "
           f"(binding: {TICKERS[binding_asset[0]]} at {binding_asset[1]*100:.1f}% weight, "
           f"{b_price_str})"))
c_min.font  = Font(bold=True, size=10, color="000000")
c_min.fill  = PatternFill("solid", fgColor=YEL)
c_min.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws2.row_dimensions[ri_min].height = 28

# Role summary
ri = len(TICKERS) + 7
ws2.merge_cells(f"A{ri}:K{ri}")
c = ws2.cell(row=ri, column=1, value="Role Distribution")
c.font = Font(bold=True, color=WHT); c.fill = PatternFill("solid", fgColor=MID)
c.alignment = Alignment(horizontal="center"); ws2.row_dimensions[ri].height = 22; ri += 1

from collections import Counter
role_counts = Counter(ROLES.values())
role_weights = {}
for t, w in WEIGHTS_10Y.items():
    role_weights[ROLES[t]] = role_weights.get(ROLES[t],0) + w

for ci, h in enumerate(["Role","Count","Total Weight",""], 1): hdr(ws2, ri, ci, h, bg=GREY, fg="000000")
ws2.row_dimensions[ri].height = 22; ri += 1

for role, cnt in sorted(role_counts.items()):
    wt = role_weights.get(role,0)
    bg = ROLE_COLORS.get(role, WHT)
    cel(ws2, ri, 1, role, bg=bg, bold=True)
    cel(ws2, ri, 2, cnt,  bg=bg)
    cel(ws2, ri, 3, f"{wt*100:.1f}%", bg=bg, bold=True)
    cel(ws2, ri, 4, "", bg=bg)
    ws2.row_dimensions[ri].height = 20; ri += 1

cw(ws2, [4, 18, 12, 13, 11, 13, 16, 12, 13, 14, 18])

# ━━ SHEET 3: Regime Analysis ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ws3 = wb.create_sheet("Regime Analysis")
rnames = list(REGIMES.keys())
ncols3 = 3 + len(rnames)

ws3.merge_cells(f"A1:{get_column_letter(ncols3)}1")
c = ws3["A1"]; c.value = "Asset Returns by Regime — Reactor Core v3"
c.font = Font(bold=True, color=WHT, size=13)
c.fill = PatternFill("solid", fgColor=DARK)
c.alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 30

for ci, h in enumerate(["Name","Ticker","Wins"] + rnames, 1): hdr(ws3, 2, ci, h)
ws3.row_dimensions[2].height = 36

for i, (t, name) in enumerate(TICKERS.items(), 1):
    bg  = ALT if i%2==0 else WHT
    ar  = asset_regime.get(t,{})
    wins = ar.get("wins",0)
    win_bg = GRN if wins>=5 else (YEL if wins>=4 else (ORG if wins>=3 else RED))
    cel(ws3, i+2, 1, name,     bg=bg, ha="left")
    cel(ws3, i+2, 2, t,        bg=bg)
    cel(ws3, i+2, 3, f"{wins}/6", bg=win_bg, bold=True)
    for j, rname in enumerate(rnames, 4):
        v = ar.get(rname)
        if v is None: cel(ws3, i+2, j, "N/A", bg=bg)
        else:
            cbg = GRN if v>20 else (YEL if v>0 else (ORG if v>-20 else RED))
            cel(ws3, i+2, j, f"{v:+.1f}%", bg=cbg)
    ws3.row_dimensions[i+2].height = 22

# Portfolio vs SPX summary
ri = len(TICKERS) + 5
ws3.merge_cells(f"A{ri}:{get_column_letter(ncols3)}{ri}")
c = ws3.cell(row=ri, column=1, value="Portfolio (v3) vs S&P 500 by Regime")
c.font = Font(bold=True, color=WHT); c.fill = PatternFill("solid", fgColor=MID)
c.alignment = Alignment(horizontal="center"); ws3.row_dimensions[ri].height = 22; ri += 1

for label, key in [("Portfolio", "port_ret"), ("S&P 500", "spx_ret"), ("vs S&P", "vs_spx")]:
    cel(ws3, ri, 1, label, bg=GREY, bold=True, ha="left")
    cel(ws3, ri, 2, "",    bg=WHT)
    cel(ws3, ri, 3, "",    bg=WHT)
    for ci, row in enumerate(regime_results, 4):
        v = row.get(key)
        if v is None: cel(ws3, ri, ci, "N/A")
        else:
            cbg = GRN if v>0 else RED
            cel(ws3, ri, ci, f"{v:+.1f}%", bg=cbg, bold=(label=="vs S&P"))
    ws3.row_dimensions[ri].height = 20; ri += 1

cw(ws3, [18, 12, 8] + [14]*len(rnames))

# ━━ SHEET 4: Weight Sensitivity ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ws4 = wb.create_sheet("Weight Sensitivity")
ws4.merge_cells("A1:H1")
c = ws4["A1"]; c.value = "Weight Sensitivity — v3 Weights Across Windows + Stress Context"
c.font = Font(bold=True, color=WHT, size=12)
c.fill = PatternFill("solid", fgColor=DARK)
c.alignment = Alignment(horizontal="center", vertical="center")
ws4.row_dimensions[1].height = 30

for ci, h in enumerate(["Name","Ticker","Role","3Y","5Y","10Y (final)","Gold -40% impact",""], 1):
    hdr(ws4, 2, ci, h)
ws4.row_dimensions[2].height = 26

from scipy.optimize import minimize as sp_min

def quick_opt(df, gold_cap=0.25, n=60):
    rets = df.pct_change().dropna(); mu=rets.mean(); cov=rets.cov(); n_a=len(df.columns)
    bounds=[(0.05, min(0.40, gold_cap) if t=="xauusd" else 0.40) for t in df.columns]
    def f(w): r=np.dot(w,mu)*252; v=np.sqrt(w@cov@w)*np.sqrt(252); return -r/v if v>0 else 0
    con=[{"type":"eq","fun":lambda w:np.sum(w)-1}]; best=None
    for _ in range(n):
        w0=np.random.dirichlet(np.ones(n_a)); w0=np.clip(w0,[b[0] for b in bounds],[b[1] for b in bounds]); w0/=w0.sum()
        res=sp_min(f,w0,method="SLSQP",bounds=bounds,constraints=con,options={"maxiter":1000,"ftol":1e-12})
        if res.success and (best is None or res.fun<best.fun): best=res
    return dict(zip(df.columns, best.x)) if best else {}

print("Computing window weights for sensitivity sheet...")
w3y = quick_opt(build_df("2023-04-01"))
w5y = quick_opt(build_df("2021-04-01"))

for i, (t, name) in enumerate(TICKERS.items(), 1):
    bg   = ALT if i%2==0 else WHT
    role = ROLES.get(t,"")
    v3   = w3y.get(t,0); v5 = w5y.get(t,0); v10 = WEIGHTS_10Y.get(t,0)
    gold_drag = round(v10 * (-0.40) * 100, 1) if t == "xauusd" else "—"
    gold_drag_str = f"{gold_drag:.1f}pp" if isinstance(gold_drag, float) else gold_drag

    def wbg(v): return GRN if v>0.20 else (YEL if v>0.10 else (GREY if v>0.05 else WHT))
    cel(ws4, i+2, 1, name,              bg=bg, ha="left")
    cel(ws4, i+2, 2, t,                 bg=bg)
    cel(ws4, i+2, 3, role,              bg=ROLE_COLORS.get(role,WHT))
    cel(ws4, i+2, 4, f"{v3*100:.1f}%",  bg=wbg(v3))
    cel(ws4, i+2, 5, f"{v5*100:.1f}%",  bg=wbg(v5))
    cel(ws4, i+2, 6, f"{v10*100:.1f}%", bg=wbg(v10), bold=True)
    cel(ws4, i+2, 7, gold_drag_str,     bg=RED if isinstance(gold_drag,float) else bg, bold=isinstance(gold_drag,float))
    cel(ws4, i+2, 8, "",                bg=bg)
    ws4.row_dimensions[i+2].height = 22

# Window performance summary
ri = len(TICKERS) + 5
for ci, h in enumerate(["Window","Sharpe","Ann Return","Max DD","Calmar","Total Return","",""], 1):
    hdr(ws4, ri, ci, h, bg=GREY, fg="000000")
ws4.row_dimensions[ri].height = 24; ri += 1

for wname in ["3Y","5Y","10Y"]:
    m  = window_metrics[wname]
    bg = ALT if ri%2==0 else WHT
    is_10 = wname == "10Y"
    cel(ws4, ri, 1, wname,                  bg=GREY, bold=True)
    cel(ws4, ri, 2, m["sharpe"],            bg=GRN if m["sharpe"]>=1.5 else YEL, bold=is_10)
    cel(ws4, ri, 3, f"{m['ann_ret']}%",    bg=bg, bold=is_10)
    cel(ws4, ri, 4, f"{m['max_dd']}%",     bg=bg, bold=is_10)
    cel(ws4, ri, 5, m["calmar"],            bg=bg, bold=is_10)
    cel(ws4, ri, 6, f"+{m['total']}%",     bg=GRN, bold=is_10)
    cel(ws4, ri, 7, "", bg=bg); cel(ws4, ri, 8, "", bg=bg)
    ws4.row_dimensions[ri].height = 22; ri += 1

cw(ws4, [18, 10, 12, 10, 12, 13, 14, 8])

# ━━ SHEET 5: Asset Metrics ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ws5 = wb.create_sheet("Asset Metrics")
ws5.merge_cells("A1:L1")
c = ws5["A1"]; c.value = "Individual Asset Metrics — 3Y / 5Y / 10Y"
c.font = Font(bold=True, color=WHT, size=13)
c.fill = PatternFill("solid", fgColor=DARK)
c.alignment = Alignment(horizontal="center", vertical="center")
ws5.row_dimensions[1].height = 30

for ci, h in enumerate(["Name","Ticker","Role",
                         "Ann Ret 10Y","Sharpe 10Y","MaxDD 10Y",
                         "Ann Ret 5Y","Sharpe 5Y","MaxDD 5Y",
                         "Ann Ret 3Y","Sharpe 3Y","MaxDD 3Y"], 1):
    hdr(ws5, 2, ci, h)
ws5.row_dimensions[2].height = 28

for i, (t, name) in enumerate(TICKERS.items(), 1):
    bg = ALT if i%2==0 else WHT
    cel(ws5, i+2, 1, name,              bg=bg, ha="left")
    cel(ws5, i+2, 2, t,                 bg=bg)
    cel(ws5, i+2, 3, ROLES.get(t,""),  bg=ROLE_COLORS.get(ROLES.get(t,""),WHT))
    ci_off = 4
    for wstart in ["2016-04-01","2021-04-01","2023-04-01"]:
        am = asset_metrics(price_dict.get(t, pd.Series(dtype=float)), wstart)
        for col in ["ann_ret","sharpe","max_dd"]:
            v = am.get(col)
            if v is None: cel(ws5, i+2, ci_off, "N/A", bg=bg)
            else:
                fv = f"{v:+.1f}%" if col in ("ann_ret","max_dd") else f"{v:.3f}"
                if col=="max_dd":   cbg = RED if v<-50 else (ORG if v<-30 else (YEL if v<-15 else GRN))
                elif col=="ann_ret":cbg = GRN if v>20 else (YEL if v>10 else (ORG if v>0 else RED))
                else:               cbg = GRN if v>0.9 else (YEL if v>0.5 else (ORG if v>0 else RED))
                cel(ws5, i+2, ci_off, fv, bg=cbg)
            ci_off += 1
    ws5.row_dimensions[i+2].height = 22

cw(ws5, [18, 12, 13, 11,10,10, 11,10,10, 11,10,10])

wb.save(REPORT)
print(f"\nReport saved: {REPORT}")
