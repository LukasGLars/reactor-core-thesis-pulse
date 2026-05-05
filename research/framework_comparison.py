import pandas as pd
import numpy as np
import os, sys, io, warnings
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
warnings.filterwarnings("ignore")
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

DATA_DIR  = r"C:\Users\lukas.larsson\Desktop\Privat\Project Reactor Core\data"
MACRO_DIR = r"C:\Users\lukas.larsson\Desktop\Privat\macro_data\portfolio_data"
REPORT    = r"C:\Users\lukas.larsson\Desktop\Privat\Project Reactor Core\Framework_Comparison.xlsx"

# ── Parameters ────────────────────────────────────────────────────────────────
START          = "2021-04-01"
INITIAL        = 1_000_000       # SEK (normalized)
MONTHLY        = 6_000           # SEK/month DCA
SR_CASH_FLOOR  = 0.25            # permanent SR floor, never deployed
SR_WAR_TARGET  = 0.40            # war chest target during CAUTION (% of SR total)

SOP_MAX = 249; T1_MIN = 320; T2_MIN = 360

# ── Portfolio definitions ─────────────────────────────────────────────────────
# Current portfolio — GE/SR split, proxies
CURRENT_GE = {"acwi.us": 0.70, "xmld.uk": 0.15, "gld.us": 0.07}   # cash 8% excluded
CURRENT_SR = {"ura.us": 0.28,  "smh.us":  0.22, "0rq9.uk": 0.15,
              "paas.us": 0.10}                                        # cash_floor 25% excluded

# Reactor Core v2 — natural GE/SR mapping (77/23)
REACTOR_GE = {"xauusd": 0.250, "wmt.us": 0.227, "lly.us": 0.197,
              "jnj.us": 0.050, "cost.us": 0.050}   # total = 0.774
REACTOR_SR = {"vrt.us": 0.090, "avgo.us": 0.079, "ccj.us": 0.057}  # total = 0.226

# Normalize within each pool
def norm(d): s = sum(d.values()); return {k: v/s for k, v in d.items()}
C_GE = norm(CURRENT_GE); C_SR = norm(CURRENT_SR)
R_GE = norm(REACTOR_GE); R_SR = norm(REACTOR_SR)

# GE/SR capital split
C_GE_FRAC = 0.82; C_SR_FRAC = 0.18
R_GE_FRAC  = sum(REACTOR_GE.values()); R_SR_FRAC = sum(REACTOR_SR.values())

# ── Load helpers ──────────────────────────────────────────────────────────────
def load(ticker):
    for fname in [ticker.replace("^","").replace("-","_")+".csv",
                  ticker.replace("^","").replace("-","_").replace(".","_")+".csv"]:
        path = os.path.join(DATA_DIR, fname)
        if os.path.exists(path):
            df = pd.read_csv(path); df["Date"] = pd.to_datetime(df["Date"])
            return df.set_index("Date").sort_index()["Close"].replace(0,np.nan).dropna()
    return None

def pool_returns(weights, start=START):
    d = {t: load(t) for t in weights if load(t) is not None}
    df = pd.DataFrame({t: s[s.index >= start] for t,s in d.items()}).ffill().dropna()
    used = [t for t in weights if t in df.columns]
    w = np.array([weights[t] for t in used]); w /= w.sum()
    return (df[used] / df[used].iloc[0]).dot(w).pct_change().dropna()

# ── Load HY-IG spread ─────────────────────────────────────────────────────────
hy = pd.read_csv(os.path.join(MACRO_DIR,"BAMLH0A0HYM2 (3).csv"))
ig = pd.read_csv(os.path.join(MACRO_DIR,"BAMLC0A0CM (4).csv"))
hy["observation_date"] = pd.to_datetime(hy["observation_date"])
ig["observation_date"] = pd.to_datetime(ig["observation_date"])
spread_bps = ((hy.set_index("observation_date")["BAMLH0A0HYM2"] -
               ig.set_index("observation_date")["BAMLC0A0CM"]) * 100).dropna()

END = pd.Timestamp("2026-04-01")
if spread_bps.index[-1] < END:
    extra = pd.date_range(spread_bps.index[-1]+pd.Timedelta(days=1), END, freq="B")
    spread_bps = pd.concat([spread_bps, pd.Series(spread_bps.iloc[-1], index=extra)])

def regime(bps):
    if bps <= SOP_MAX:  return "SOP"
    elif bps < T1_MIN:  return "CAUTION"
    elif bps < T2_MIN:  return "T1"
    else:               return "T2"

# ── Simulation engine ─────────────────────────────────────────────────────────
def simulate(ge_weights, sr_weights, ge_frac, sr_frac, label):
    """
    Two-pool framework simulation.
    GE pool: receives monthly DCA in SOP, holds cash in CAUTION/T1/T2
    SR pool:  permanent 25% cash floor + war chest builds to 40% in CAUTION
    T1: deploy 50% ge_cash + 50% sr_war_chest
    T2: deploy remaining; DD-triggered tranches at -15% and -20%
    RE-ENTRY: weekly tranches over 4 weeks
    """
    ge_rets = pool_returns(ge_weights)
    sr_rets = pool_returns(sr_weights)

    common = ge_rets.index.intersection(sr_rets.index).intersection(spread_bps.index)
    ge_r = ge_rets[common]; sr_r = sr_rets[common]
    sp   = spread_bps.reindex(common, method="ffill")

    # Initial allocation
    ge_inv   = INITIAL * ge_frac * (1 - 0.08)     # 8% GE cash initially
    ge_cash  = INITIAL * ge_frac * 0.08
    sr_total = INITIAL * sr_frac
    sr_floor = sr_total * SR_CASH_FLOOR            # permanent, never moves
    sr_inv   = sr_total * (1 - SR_CASH_FLOOR)      # 75% invested initially
    sr_war   = 0.0                                  # war chest starts empty

    prev_regime = regime(sp.iloc[0])
    prev_month  = common[0].to_period("M")

    # RE-ENTRY state
    re_ge_snap = 0.0; re_sr_snap = 0.0; re_days = 0

    # T2 DD tracking
    t2_active = False; t2_peak = None
    t2_ge_remaining = 0.0; t2_sr_remaining = 0.0

    total_vals = []; ge_vals = []; sr_vals = []
    log = []

    for date, (gr, srr) in pd.DataFrame({"ge": ge_r, "sr": sr_r}).iterrows():
        sp_val = sp.get(date, sp.iloc[-1])
        reg    = regime(sp_val)

        # ── Monthly DCA ───────────────────────────────────────────────────
        if date.to_period("M") != prev_month:
            prev_month = date.to_period("M")
            if reg == "SOP" and re_ge_snap <= 0:
                ge_inv  += MONTHLY           # full DCA into GE
            elif reg == "CAUTION":
                ge_cash += MONTHLY           # hold as cash; gold exception is minor
                log.append({"Date":date.date(),"Event":f"Monthly held ({reg})",
                            "GE_cash":round(ge_cash),"SR_war":round(sr_war),
                            "Spread":round(sp_val,1)})
            else:
                ge_cash += MONTHLY           # T1/T2: accumulate

        # ── CAUTION: build SR war chest to 40% of SR total ────────────────
        if reg == "CAUTION":
            sr_total_now = sr_inv + sr_floor + sr_war
            war_target   = sr_total_now * SR_WAR_TARGET
            if sr_war < war_target and sr_inv > 0:
                # Skim from SR invested to build war chest (capped at 2% per day)
                skim = min(sr_inv * 0.002, war_target - sr_war)
                sr_inv -= skim; sr_war += skim

        # ── Regime transitions ────────────────────────────────────────────
        if prev_regime in ("CAUTION","T1","T2") and reg == "SOP":
            # RE-ENTRY: deploy over 4 weeks
            if ge_cash > 0 or sr_war > 0:
                re_ge_snap = ge_cash; re_sr_snap = sr_war; re_days = 0
                log.append({"Date":date.date(),"Event":"RE-ENTRY triggered",
                            "GE_cash":round(ge_cash),"SR_war":round(sr_war),
                            "Spread":round(sp_val,1)})

        elif prev_regime == "CAUTION" and reg == "T1":
            deploy_ge = ge_cash * 0.50; deploy_sr = sr_war * 0.50
            ge_inv += deploy_ge; ge_cash -= deploy_ge
            sr_inv += deploy_sr; sr_war  -= deploy_sr
            log.append({"Date":date.date(),"Event":"T1 — deploy 50% GE cash + 50% SR war",
                        "GE_cash":round(ge_cash),"SR_war":round(sr_war),
                        "Spread":round(sp_val,1)})

        elif prev_regime in ("CAUTION","T1") and reg == "T2":
            # Deploy remaining; -15%/-20% DD tranches handled below
            t2_active = True
            t2_ge_remaining = ge_cash; t2_sr_remaining = sr_war
            t2_peak = ge_inv + ge_cash + sr_inv + sr_war + sr_floor
            log.append({"Date":date.date(),"Event":"T2 triggered — DD tranches active",
                        "GE_cash":round(ge_cash),"SR_war":round(sr_war),
                        "Spread":round(sp_val,1)})

        # ── RE-ENTRY gradual deployment ───────────────────────────────────
        if re_ge_snap > 0 and (ge_cash > 0 or sr_war > 0):
            re_days += 1
            deploy_ge = min(ge_cash, re_ge_snap / 20)
            deploy_sr = min(sr_war,  re_sr_snap / 20)
            ge_inv += deploy_ge; ge_cash -= deploy_ge
            sr_inv += deploy_sr; sr_war  -= deploy_sr
            if ge_cash <= 1 and sr_war <= 1 or re_days >= 20:
                re_ge_snap = re_sr_snap = 0; re_days = 0

        # ── T2 DD-triggered deployment ────────────────────────────────────
        if t2_active and t2_peak is not None:
            total_now = ge_inv + ge_cash + sr_inv + sr_war + sr_floor
            dd = (total_now - t2_peak) / t2_peak if t2_peak > 0 else 0
            if dd <= -0.15 and t2_ge_remaining > 0:
                deploy_ge = t2_ge_remaining * 0.50
                deploy_sr = t2_sr_remaining * 0.50
                ge_inv += deploy_ge; ge_cash -= deploy_ge
                sr_inv += deploy_sr; sr_war  -= deploy_sr
                t2_ge_remaining -= deploy_ge; t2_sr_remaining -= deploy_sr
                log.append({"Date":date.date(),"Event":f"T2 -15% DD tranche (DD={dd*100:.1f}%)",
                            "GE_cash":round(ge_cash),"SR_war":round(sr_war),
                            "Spread":round(sp_val,1)})
            if dd <= -0.20 and t2_ge_remaining > 0:
                ge_inv += t2_ge_remaining; ge_cash -= t2_ge_remaining
                sr_inv += t2_sr_remaining; sr_war  -= t2_sr_remaining
                t2_ge_remaining = t2_sr_remaining = 0; t2_active = False
                log.append({"Date":date.date(),"Event":f"T2 -20% DD tranche (DD={dd*100:.1f}%)",
                            "GE_cash":round(ge_cash),"SR_war":round(sr_war),
                            "Spread":round(sp_val,1)})
            if reg != "T2": t2_active = False

        # ── Apply daily returns ───────────────────────────────────────────
        ge_inv *= (1 + gr)
        sr_inv *= (1 + srr)
        sr_floor *= (1 + srr)   # floor grows with SR assets

        total = ge_inv + ge_cash + sr_inv + sr_war + sr_floor
        total_vals.append(total)
        ge_vals.append(ge_inv + ge_cash)
        sr_vals.append(sr_inv + sr_war + sr_floor)
        prev_regime = reg

    idx = common
    return (pd.Series(total_vals, index=idx),
            pd.Series(ge_vals,    index=idx),
            pd.Series(sr_vals,    index=idx),
            pd.DataFrame(log))

# ── Simple DCA simulation (Reactor only — already know result but run for chart)
def simple_dca(ge_weights, sr_weights, ge_frac, sr_frac):
    ge_rets = pool_returns(ge_weights); sr_rets = pool_returns(sr_weights)
    common  = ge_rets.index.intersection(sr_rets.index)
    ge_r = ge_rets[common]; sr_r = sr_rets[common]
    ge_inv = INITIAL * ge_frac; sr_inv = INITIAL * sr_frac
    prev_month = common[0].to_period("M")
    vals = []
    for date, (gr, srr) in pd.DataFrame({"ge": ge_r, "sr": sr_r}).iterrows():
        if date.to_period("M") != prev_month:
            prev_month = date.to_period("M")
            ge_inv += MONTHLY
        ge_inv *= (1 + gr); sr_inv *= (1 + srr)
        vals.append(ge_inv + sr_inv)
    return pd.Series(vals, index=common)

# ── Run ───────────────────────────────────────────────────────────────────────
print("Running simulations...")
cur_total,  cur_ge,  cur_sr,  cur_log  = simulate(C_GE, C_SR, C_GE_FRAC, C_SR_FRAC, "Current")
rxr_total,  rxr_ge,  rxr_sr,  rxr_log  = simulate(R_GE, R_SR, R_GE_FRAC, R_SR_FRAC, "Reactor")
rxr_simple                              = simple_dca(R_GE, R_SR, R_GE_FRAC, R_SR_FRAC)

def val_metrics(s):
    d = s.pct_change().dropna()
    d_clean = d[d.abs() < 0.10]
    n = len(s)/252
    ann_vol = d_clean.std()*np.sqrt(252)
    final   = s.iloc[-1]
    contrib = INITIAL + MONTHLY * n * 12
    cagr    = (final/INITIAL)**(1/n)-1
    sh      = cagr/ann_vol if ann_vol>0 else 0
    roll    = s.cummax(); dd = ((s-roll)/roll).min()
    calmar  = cagr/abs(dd) if dd<0 else 0
    return {"final": round(final), "contrib": round(contrib),
            "profit": round(final-contrib), "cagr": round(cagr*100,2),
            "vol": round(ann_vol*100,2), "sharpe": round(sh,3),
            "max_dd": round(dd*100,2), "calmar": round(calmar,3),
            "total_ret": round((final/INITIAL-1)*100,1)}

mc = val_metrics(cur_total)
mr = val_metrics(rxr_total)
ms = val_metrics(rxr_simple)

print(f"\n  {'':32} {'Current+FW':>14} {'Reactor+FW':>14} {'Reactor DCA':>14}")
print(f"  {'Final value (kr)':32} {mc['final']:>14,} {mr['final']:>14,} {ms['final']:>14,}")
print(f"  {'CAGR':32} {mc['cagr']:>13.2f}% {mr['cagr']:>13.2f}% {ms['cagr']:>13.2f}%")
print(f"  {'Sharpe':32} {mc['sharpe']:>14.3f} {mr['sharpe']:>14.3f} {ms['sharpe']:>14.3f}")
print(f"  {'Max DD':32} {mc['max_dd']:>13.2f}% {mr['max_dd']:>13.2f}% {ms['max_dd']:>13.2f}%")
print(f"  {'Total return vs initial':32} {mc['total_ret']:>13.1f}% {mr['total_ret']:>13.1f}% {ms['total_ret']:>13.1f}%")

# ── Excel ─────────────────────────────────────────────────────────────────────
DARK="1F4E79"; MID="2E75B6"; LIGHT="D6E4F0"; ALT="EBF3FB"
GRN="C6EFCE"; RED="FFC7CE"; YEL="FFEB9C"; ORG="FFCC99"; WHT="FFFFFF"; GREY="F2F2F2"
C_CUR="FFF2CC"; C_RXR="E2EFDA"; C_DCA="DDEBF7"

thin=Side(style="thin",color="BFBFBF"); BORDER=Border(left=thin,right=thin,top=thin,bottom=thin)

def hdr(ws,row,col,val,bg=DARK,fg=WHT,bold=True,sz=10,ha="center"):
    c=ws.cell(row=row,column=col,value=val)
    c.fill=PatternFill("solid",fgColor=bg); c.font=Font(bold=bold,color=fg,size=sz)
    c.alignment=Alignment(wrap_text=True,vertical="center",horizontal=ha)
    c.border=BORDER; return c

def cel(ws,row,col,val,bg=None,bold=False,ha="center",color="000000"):
    c=ws.cell(row=row,column=col,value=val)
    if bg: c.fill=PatternFill("solid",fgColor=bg)
    c.font=Font(bold=bold,color=color)
    c.alignment=Alignment(wrap_text=True,vertical="center",horizontal=ha)
    c.border=BORDER; return c

def title(ws,row,nc,text,bg=DARK,sz=13):
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=nc)
    c=ws.cell(row=row,column=1,value=text)
    c.font=Font(bold=True,color=WHT,size=sz); c.fill=PatternFill("solid",fgColor=bg)
    c.alignment=Alignment(horizontal="center",vertical="center")
    ws.row_dimensions[row].height=34

def cw(ws,widths):
    for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w

wb = openpyxl.Workbook()

# ━━ SHEET 1: Head-to-Head ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ws1=wb.active; ws1.title="Head-to-Head"
title(ws1,1,4,"Framework Comparison — Current Portfolio vs Reactor Core v2")

ws1.merge_cells("A2:D2")
c=ws1["A2"]
c.value=(f"Start: {START}  |  Initial: {INITIAL:,} kr  |  Monthly DCA: {MONTHLY:,} kr  |  "
         f"Framework v4.12: SOP≤{SOP_MAX} / CAUTION 250-{T1_MIN-1} / T1 {T1_MIN}-{T2_MIN-1} / T2≥{T2_MIN} bps  |  "
         f"SR cash floor: {SR_CASH_FLOOR*100:.0f}%  |  SR war chest target: {SR_WAR_TARGET*100:.0f}%")
c.font=Font(italic=True,size=9,color="444444"); c.fill=PatternFill("solid",fgColor=LIGHT)
c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
ws1.row_dimensions[2].height=30

for ci,(h,bg) in enumerate([("Metric",GREY),("Current + Framework",C_CUR),
                              ("Reactor v2 + Framework",C_RXR),("Reactor v2 Simple DCA",C_DCA)],1):
    hdr(ws1,3,ci,h,bg=bg,fg="000000")
ws1.row_dimensions[3].height=28

rows=[
    ("Final Portfolio Value",  "final",    True, lambda v:f"{v:,.0f} kr"),
    ("Total Contributed",      "contrib",  None, lambda v:f"{v:,.0f} kr"),
    ("Total Profit",           "profit",   True, lambda v:f"{v:,.0f} kr"),
    ("Total Return vs Initial","total_ret",True, lambda v:f"{v:+.1f}%"),
    ("CAGR",                   "cagr",     True, lambda v:f"{v:.2f}%"),
    ("Ann. Volatility",        "vol",      False,lambda v:f"{v:.2f}%"),
    ("Sharpe Ratio",           "sharpe",   True, lambda v:f"{v:.3f}"),
    ("Max Drawdown",           "max_dd",   False,lambda v:f"{v:.2f}%"),
    ("Calmar Ratio",           "calmar",   True, lambda v:f"{v:.3f}"),
]

for ri,(label,key,hb,fmt) in enumerate(rows,4):
    bg=ALT if ri%2==0 else WHT
    vc=mc[key]; vr=mr[key]; vd=ms[key]
    cel(ws1,ri,1,label,bg=GREY,bold=True,ha="left")
    # winner highlight between Current+FW and Reactor+FW only
    if hb is True:
        bgc=GRN if vc>=vr else RED; bgr=GRN if vr>=vc else RED
    elif hb is False:
        bgc=GRN if vc<=vr else RED; bgr=GRN if vr<=vc else RED
    else:
        bgc=bgr=bg
    cel(ws1,ri,2,fmt(vc),bg=bgc or bg,bold=(bgc==GRN))
    cel(ws1,ri,3,fmt(vr),bg=bgr or bg,bold=(bgr==GRN))
    cel(ws1,ri,4,fmt(vd),bg=C_DCA)
    ws1.row_dimensions[ri].height=22

# Framework advantage block
ri=len(rows)+6
ws1.merge_cells(f"A{ri}:D{ri}")
c=ws1.cell(row=ri,column=1,value="Framework Effect on Reactor (Framework vs Simple DCA)")
c.font=Font(bold=True,color=WHT,size=10); c.fill=PatternFill("solid",fgColor=MID)
c.alignment=Alignment(horizontal="center"); ws1.row_dimensions[ri].height=24; ri+=1

for ci,h in enumerate(["Metric","—","Framework value","DCA value"],1):
    hdr(ws1,ri,ci,h,bg=GREY,fg="000000")
ws1.row_dimensions[ri].height=22; ri+=1

for label,key,_,fmt in rows:
    if key in ("contrib",): continue
    fw=mr[key]; dca=ms[key]
    diff=fw-dca
    if key in ("vol","max_dd"): better=diff<0
    else: better=diff>0
    diff_str=(f"+{diff:.1f}%" if diff>=0 else f"{diff:.1f}%") if isinstance(diff,float) else f"{diff:+,.0f}"
    if key in ("final","profit"): diff_str=f"{diff:+,.0f} kr"
    elif key in ("sharpe","calmar"): diff_str=f"{diff:+.3f}"
    elif key in ("cagr","vol","max_dd","total_ret"): diff_str=f"{diff:+.2f}pp"
    bg=ALT if ri%2==0 else WHT
    cel(ws1,ri,1,label,bg=GREY,bold=True,ha="left")
    cel(ws1,ri,2,"—",bg=bg)
    cel(ws1,ri,3,fmt(fw),bg=bg)
    cel(ws1,ri,4,f"{diff_str}  {'↑ FW better' if better else '↓ DCA better'}",
        bg=GRN if better else RED,bold=True)
    ws1.row_dimensions[ri].height=20; ri+=1

cw(ws1,[28,22,22,22])

# ━━ SHEET 2: Framework Event Log ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ws2=wb.create_sheet("Event Log")
title(ws2,1,6,"Framework Deployment Events — Current & Reactor")

for ci,h in enumerate(["Date","Portfolio","Event","GE Cash (kr)","SR War Chest (kr)","Spread (bps)"],1):
    hdr(ws2,2,ci,h)
ws2.row_dimensions[2].height=26

CAUTION_C="FFE0CC"; T1_C="FFD0D0"; T2_C="FF9999"; SOP_C="CCFFCC"

all_log=[]
for row in cur_log.to_dict("records"):
    row["portfolio"]="Current"; all_log.append(row)
for row in rxr_log.to_dict("records"):
    row["portfolio"]="Reactor"; all_log.append(row)
all_log.sort(key=lambda x:str(x["Date"]))

for ri,row in enumerate(all_log,3):
    ev=row.get("Event","")
    bg=WHT
    if "T2"      in ev: bg=T2_C
    elif "T1"    in ev: bg=T1_C
    elif "RE-ENTRY" in ev: bg=SOP_C
    elif "held"  in ev: bg=CAUTION_C
    elif "CAUTION" in ev: bg=CAUTION_C
    alt=ALT if ri%2==0 else bg
    cel(ws2,ri,1,str(row.get("Date","")),bg=alt)
    cel(ws2,ri,2,row.get("portfolio",""),bg=alt,bold=True)
    cel(ws2,ri,3,ev,bg=bg,ha="left")
    cel(ws2,ri,4,f"{row.get('GE_cash',0):,.0f}",bg=alt)
    cel(ws2,ri,5,f"{row.get('SR_war',0):,.0f}",bg=alt)
    cel(ws2,ri,6,row.get("Spread",""),bg=alt)
    ws2.row_dimensions[ri].height=16

cw(ws2,[12,10,44,16,18,13])

# ━━ SHEET 3: Assumptions ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ws3=wb.create_sheet("Assumptions")
title(ws3,1,2,"Simulation Assumptions & Pool Mapping")

notes=[
    ("Starting capital",        f"{INITIAL:,} kr (normalized — not actual current TPV)"),
    ("Monthly contribution",    f"{MONTHLY:,} kr/month (midpoint of 5-7k guideline)"),
    ("Start date",              START),
    ("Cash return",             "0% — conservative; ignores HYSA/MMF yield"),
    ("SR cash floor",           f"{SR_CASH_FLOOR*100:.0f}% of SR total permanently held"),
    ("SR war chest target",     f"{SR_WAR_TARGET*100:.0f}% of SR total during CAUTION (built via daily 0.2% skim from SR invested)"),
    ("T1 deployment",           "50% GE cash + 50% SR war chest on first day spread ≥320"),
    ("T2 deployment",           "Remaining deployed in two DD tranches: 50% at -15% portfolio DD, remainder at -20%"),
    ("RE-ENTRY",                "Deploy GE cash + SR war chest over 20 trading days (~4 weeks)"),
    ("Gold exception (CAUTION)","Not modelled separately — minor effect on monthly DCA amount"),
    ("Skim mechanics",          "Approximated as 0.2%/day SR invested → war chest during CAUTION"),
    ("Current GE proxies",      "ACWI.US (IUSQ), GLD.US (PPFB), XMLD.UK (XMLD.DE)"),
    ("Current SR proxies",      "URA.US (URNU), SMH.US (VVSM), 0RQ9.UK, PAAS.US"),
    ("Reactor GE pool",         "Gold 32.3% | Walmart 29.3% | Eli Lilly 25.5% | J&J 6.5% | Costco 6.5% (normalized from total weights)"),
    ("Reactor SR pool",         "Vertiv 39.8% | Broadcom 35.0% | Cameco 25.2% (normalized)"),
    ("Reactor GE/SR split",     f"{R_GE_FRAC*100:.1f}% / {R_SR_FRAC*100:.1f}% (natural from 10Y optimized weights)"),
    ("Current GE/SR split",     "82% / 18% (as defined)"),
    ("FX",                      "All prices USD. Returns in USD. Monthly contribution in kr = USD equivalent at constant rate. FX effect equal for both portfolios."),
]

hdr(ws3,2,1,"Parameter",bg=GREY,fg="000000"); hdr(ws3,2,2,"Value",bg=GREY,fg="000000")
ws3.row_dimensions[2].height=22

for ri,(k,v) in enumerate(notes,3):
    bg=ALT if ri%2==0 else WHT
    cel(ws3,ri,1,k,bg=GREY,bold=True,ha="left")
    cel(ws3,ri,2,v,bg=bg,ha="left")
    ws3.row_dimensions[ri].height=20

cw(ws3,[30,65])

wb.save(REPORT)
print(f"\nReport saved: {REPORT}")
