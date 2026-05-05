import pandas as pd
import numpy as np
import os, sys, io, warnings
warnings.filterwarnings("ignore")
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_DIR = r"C:\Users\lukas.larsson\Desktop\Privat\Project Reactor Core\output"
REPORT     = r"C:\Users\lukas.larsson\Desktop\Privat\Project Reactor Core\Reactor_Core_10pos.xlsx"

TICKERS = {
    "xauusd":  "Gold",
    "ccj.us":  "Cameco",
    "avgo.us": "Broadcom",
    "lite.us": "Lumentum",
    "vrt.us":  "Vertiv",
    "lly.us":  "Eli Lilly",
    "jnj.us":  "J&J",
    "wmt.us":  "Walmart",
    "cost.us": "Costco",
    "tsla.us": "Tesla",
}

SECTORS = {
    "xauusd":  "Commodity",     "ccj.us":  "Commodity",
    "avgo.us": "Semiconductor", "lite.us": "Semiconductor",
    "vrt.us":  "Infrastructure","lly.us":  "Defensive",
    "jnj.us":  "Defensive",     "wmt.us":  "Defensive",
    "cost.us": "Defensive",     "tsla.us": "Wildcard",
}

SECTOR_COLORS = {
    "Commodity": "FFF2CC", "Semiconductor": "E2EFDA",
    "Infrastructure": "DDEBF7", "Defensive": "FCE4D6", "Wildcard": "EAD1DC",
}

DARK_BLUE="1F4E79"; MID_BLUE="2E75B6"; LIGHT_BLUE="D6E4F0"
ALT="EBF3FB"; GREEN="C6EFCE"; RED="FFC7CE"; YELLOW="FFEB9C"; ORANGE="FFCC99"; GREY="F2F2F2"; WHITE="FFFFFF"
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr(ws, row, col, val, bg=DARK_BLUE, fg=WHITE, bold=True, size=11, halign="center"):
    c = ws.cell(row=row, column=col, value=val)
    c.fill = PatternFill("solid", fgColor=bg)
    c.font = Font(bold=bold, color=fg, size=size)
    c.alignment = Alignment(wrap_text=True, vertical="center", horizontal=halign)
    c.border = BORDER
    return c

def cel(ws, row, col, val, bg=None, bold=False, halign="center", color="000000"):
    c = ws.cell(row=row, column=col, value=val)
    if bg: c.fill = PatternFill("solid", fgColor=bg)
    c.font = Font(bold=bold, color=color)
    c.alignment = Alignment(wrap_text=True, vertical="center", horizontal=halign)
    c.border = BORDER
    return c

def set_cols(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# Load data
w3  = pd.read_csv(os.path.join(OUTPUT_DIR,"weights10_3Y.csv"))
w5  = pd.read_csv(os.path.join(OUTPUT_DIR,"weights10_5Y.csv"))
w10 = pd.read_csv(os.path.join(OUTPUT_DIR,"weights10_10Y.csv"))
regime_sum = pd.read_csv(os.path.join(OUTPUT_DIR,"regime_summary10.csv"))
regime_mat = pd.read_csv(os.path.join(OUTPUT_DIR,"asset_regime_matrix10.csv")).set_index("ticker")
metrics_10 = pd.read_csv(os.path.join(OUTPUT_DIR,"metrics_10Y.csv"), index_col="ticker")
metrics_5  = pd.read_csv(os.path.join(OUTPUT_DIR,"metrics_5Y.csv"),  index_col="ticker")
metrics_3  = pd.read_csv(os.path.join(OUTPUT_DIR,"metrics_3Y.csv"),  index_col="ticker")

def get_w(df, obj):
    return df[df["objective"]==obj].set_index("ticker")["weight"]

# Build final weights: avg of Sharpe weights across 3Y/5Y/10Y
ws3s  = get_w(w3,  "sharpe")
ws5s  = get_w(w5,  "sharpe")
ws10s = get_w(w10, "sharpe")

weight_df = pd.DataFrame({"3Y": ws3s, "5Y": ws5s, "10Y": ws10s}).fillna(0)
weight_df["avg"] = weight_df.mean(axis=1)
weight_df["name"] = [TICKERS.get(t, t) for t in weight_df.index]
weight_df["sector"] = [SECTORS.get(t, "") for t in weight_df.index]
weight_df = weight_df.sort_values("avg", ascending=False)

wb = openpyxl.Workbook()

# ━━ SHEET 1: Summary ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ws_sum = wb.active
ws_sum.title = "Summary"

ws_sum.merge_cells("A1:F1")
c = ws_sum["A1"]; c.value = "PROJECT REACTOR CORE — 10-Position Portfolio"
c.font = Font(bold=True, color=WHITE, size=16)
c.fill = PatternFill("solid", fgColor=DARK_BLUE)
c.alignment = Alignment(horizontal="center", vertical="center")
ws_sum.row_dimensions[1].height = 44

ws_sum.merge_cells("A2:F2")
c = ws_sum["A2"]; c.value = "Minimum positions for maximum Sharpe | 10Y / 5Y / 3Y windows | 5% min / 40% max per position"
c.font = Font(italic=True, color="444444", size=10)
c.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
c.alignment = Alignment(horizontal="center", vertical="center")
ws_sum.row_dimensions[2].height = 20

def block(ws, r, title, rows_data):
    ws.merge_cells(f"A{r}:F{r}")
    c = ws.cell(row=r, column=1, value=title)
    c.font = Font(bold=True, color=WHITE, size=11)
    c.fill = PatternFill("solid", fgColor=MID_BLUE)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[r].height = 22
    r += 1
    for label, val, bg in rows_data:
        cel(ws, r, 1, label, bg=GREY, bold=True, halign="left")
        ws.merge_cells(f"B{r}:F{r}")
        c2 = ws.cell(row=r, column=2, value=val)
        c2.fill = PatternFill("solid", fgColor=bg or WHITE)
        c2.alignment = Alignment(vertical="center", indent=1); c2.border = BORDER
        ws.row_dimensions[r].height = 20
        r += 1
    return r + 1

r = 4
r = block(ws_sum, r, "Why 10 positions?", [
    ("Method",   "Iterative Sharpe optimization: drop lowest-weight ticker, re-optimize, repeat", LIGHT_BLUE),
    ("Finding",  "Sharpe increases as positions are removed — 32-position portfolio was over-diversified", LIGHT_BLUE),
    ("Peak",     "Sharpe peaks at 10 positions (1.882) then declines — this is the efficient minimum", GREEN),
    ("Knee",     "Removing below 10 costs >0.03 Sharpe per step — diminishing returns accelerate", YELLOW),
])

r = block(ws_sum, r, "10Y Sharpe-Optimal Performance", [
    ("Ann. Return",     "30.1%",  GREEN),
    ("Ann. Volatility", "18.0%",  YELLOW),
    ("Sharpe Ratio",    "1.908",  GREEN),
    ("Max Drawdown",    "-18.0%", YELLOW),
    ("Calmar Ratio",    "1.674",  GREEN),
    ("Total Return",    "+313%",  GREEN),
])

r = block(ws_sum, r, "Portfolio vs S&P 500 by Regime", [
    ("Pre-COVID Bull (2016-2020)",      "-20.0% vs S&P — underperforms in calm bull (low vol focus costs upside)", ORANGE),
    ("COVID Crash (Feb-Mar 2020)",      "+14.5% vs S&P — portfolio -5.9% vs S&P -20.4%", GREEN),
    ("COVID Recovery (2020-2021)",      "+23.5% vs S&P — portfolio +116% vs S&P +93%", GREEN),
    ("Rate Hike/Inflation (2022-2023)", "+21.1% vs S&P — portfolio +17% vs S&P -4%", GREEN),
    ("Post-Hike/AI Bull (2023-2024)",   "+28.5% vs S&P — portfolio +52% vs S&P +23%", GREEN),
    ("Rate Cut (2024-present)",         "+107.5% vs S&P — portfolio +126% vs S&P +18%", GREEN),
])

r = block(ws_sum, r, "Holdings", [
    ("Gold (37.9%)",      "Anchor. Positive 5/6 regimes. Hedge against everything.", GREEN),
    ("Eli Lilly (15.6%)", "Best risk-adj in defensive bucket. Sharpe 1.03 over 10Y.", GREEN),
    ("Walmart (11.4%)",   "Defensive compounder. Positive 5/6 regimes. Low vol.", GREEN),
    ("Broadcom (5%)",     "Semi compounder. Sharpe 0.98. Positive 5/6 regimes.", GREEN),
    ("Vertiv (5%)",       "Infrastructure/AI play. Explosive in AI bull and rate cut.", YELLOW),
    ("Cameco (5%)",       "Uranium. Strong in rate hike and rate cut. Uncorrelated.", YELLOW),
    ("Lumentum (5%)",     "High variance. -51% in rate hike, +1120% in rate cut.", ORANGE),
    ("Tesla (5%)",        "High variance. Keep small. Adds tail upside.", ORANGE),
    ("J&J (5%)",          "Defensive anchor. Near-zero drawdown in most regimes.", GREEN),
    ("Costco (5%)",       "Steady compounder. Positive 4/6 regimes.", GREEN),
])

set_cols(ws_sum, [22, 14, 14, 14, 14, 14])

# ━━ SHEET 2: Final Portfolio ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ws_port = wb.create_sheet("Final Portfolio")
ws_port.merge_cells("A1:J1")
c = ws_port["A1"]; c.value = "Final Portfolio — Sharpe-Optimal Weights (avg 3Y/5Y/10Y)"
c.font = Font(bold=True, color=WHITE, size=13)
c.fill = PatternFill("solid", fgColor=DARK_BLUE)
c.alignment = Alignment(horizontal="center", vertical="center")
ws_port.row_dimensions[1].height = 30

hdrs = ["#","Name","Ticker","Sector","Avg Weight","3Y","5Y","10Y","Pos Regimes","Avg Regime Ret"]
for c2, h in enumerate(hdrs, 1): hdr(ws_port, 2, c2, h)
ws_port.row_dimensions[2].height = 28

for i, (ticker, row) in enumerate(weight_df.iterrows(), 1):
    r = i + 2
    bg = ALT if i % 2 == 0 else WHITE
    cel(ws_port, r, 1, i, bg=bg)
    cel(ws_port, r, 2, row["name"], bg=bg, bold=True, halign="left")
    cel(ws_port, r, 3, ticker, bg=bg)
    cel(ws_port, r, 4, row["sector"], bg=SECTOR_COLORS.get(row["sector"], WHITE))
    cel(ws_port, r, 5, f"{row['avg']*100:.1f}%", bg=bg, bold=True)
    cel(ws_port, r, 6, f"{row['3Y']*100:.1f}%", bg=bg)
    cel(ws_port, r, 7, f"{row['5Y']*100:.1f}%", bg=bg)
    cel(ws_port, r, 8, f"{row['10Y']*100:.1f}%", bg=bg)
    pos = regime_mat.loc[ticker, "positive_regimes"] if ticker in regime_mat.index else "N/A"
    pos_bg = GREEN if pos >= 5 else (YELLOW if pos >= 4 else ORANGE)
    cel(ws_port, r, 9, f"{int(pos)}/6", bg=pos_bg, bold=True)
    avg_r = regime_mat.loc[ticker, "avg_return"] if ticker in regime_mat.index else None
    cel(ws_port, r, 10, f"{avg_r:+.1f}%" if avg_r else "N/A", bg=bg)
    ws_port.row_dimensions[r].height = 22

set_cols(ws_port, [4, 20, 12, 16, 11, 11, 11, 11, 12, 16])

# ━━ SHEET 3: Regime Analysis ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ws_reg = wb.create_sheet("Regime Analysis")
REGIMES_ORDER = ["Pre-COVID Bull","COVID Crash","COVID Recovery",
                 "Rate Hike/Inflation","Post-Hike/AI Bull","Rate Cut"]

ws_reg.merge_cells(f"A1:{get_column_letter(3+len(REGIMES_ORDER))}1")
c = ws_reg["A1"]; c.value = "Asset Returns by Regime"
c.font = Font(bold=True, color=WHITE, size=13)
c.fill = PatternFill("solid", fgColor=DARK_BLUE)
c.alignment = Alignment(horizontal="center", vertical="center")
ws_reg.row_dimensions[1].height = 30

for c2, h in enumerate(["Name","Ticker","Pos/6"] + REGIMES_ORDER, 1):
    hdr(ws_reg, 2, c2, h)
ws_reg.row_dimensions[2].height = 36

for i, (ticker, name) in enumerate(TICKERS.items(), 1):
    r = i + 2
    bg = ALT if i % 2 == 0 else WHITE
    cel(ws_reg, r, 1, name, bg=bg, halign="left")
    cel(ws_reg, r, 2, ticker, bg=bg)
    pos = int(regime_mat.loc[ticker, "positive_regimes"]) if ticker in regime_mat.index else 0
    pos_bg = GREEN if pos >= 5 else (YELLOW if pos >= 4 else (ORANGE if pos >= 3 else RED))
    cel(ws_reg, r, 3, f"{pos}/6", bg=pos_bg, bold=True)
    for j, rname in enumerate(REGIMES_ORDER, 4):
        v = regime_mat.loc[ticker, rname] if ticker in regime_mat.index else None
        if v is None or (isinstance(v, float) and np.isnan(v)):
            cel(ws_reg, r, j, "N/A", bg=bg)
        else:
            cbg = GREEN if v > 20 else (YELLOW if v > 0 else (ORANGE if v > -20 else RED))
            cel(ws_reg, r, j, f"{v:+.1f}%", bg=cbg)
    ws_reg.row_dimensions[r].height = 22

# Portfolio vs SPX rows
r2 = len(TICKERS) + 4
ws_reg.merge_cells(f"A{r2}:{get_column_letter(3+len(REGIMES_ORDER))}{r2}")
c = ws_reg.cell(row=r2, column=1, value="Portfolio vs S&P 500")
c.font = Font(bold=True, color=WHITE); c.fill = PatternFill("solid", fgColor=MID_BLUE)
c.alignment = Alignment(horizontal="center", vertical="center")
ws_reg.row_dimensions[r2].height = 22
r2 += 1

for col_idx in range(1, 4+len(REGIMES_ORDER)+1):
    hdr(ws_reg, r2, col_idx, "", bg=MID_BLUE)
ws_reg.row_dimensions[r2].height = 6

for label, key in [("Portfolio", "portfolio_ret"), ("S&P 500", "spx_ret"), ("vs S&P 500", "vs_spx")]:
    r2 += 1
    cel(ws_reg, r2, 1, label, bold=True, halign="left")
    cel(ws_reg, r2, 2, "", bg=WHITE); cel(ws_reg, r2, 3, "", bg=WHITE)
    for c2, (_, row_s) in enumerate(regime_sum.iterrows(), 4):
        v = row_s[key]
        if pd.isna(v): cel(ws_reg, r2, c2, "N/A")
        else:
            cbg = GREEN if v > 0 else RED
            cel(ws_reg, r2, c2, f"{v:+.1f}%", bg=cbg, bold=(label=="vs S&P 500"))
    ws_reg.row_dimensions[r2].height = 20

set_cols(ws_reg, [20, 12, 8] + [14]*len(REGIMES_ORDER))

# ━━ SHEET 4: Weight Sensitivity ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ws_sens = wb.create_sheet("Weight Sensitivity")
ws_sens.merge_cells("A1:G1")
c = ws_sens["A1"]; c.value = "Weight Sensitivity — Sharpe-Optimal Weights Across Windows"
c.font = Font(bold=True, color=WHITE, size=12)
c.fill = PatternFill("solid", fgColor=DARK_BLUE)
c.alignment = Alignment(horizontal="center", vertical="center")
ws_sens.row_dimensions[1].height = 30

for c2, h in enumerate(["Name","Ticker","Sector","3Y","5Y","10Y","Avg"], 1):
    hdr(ws_sens, 2, c2, h)
ws_sens.row_dimensions[2].height = 26

for i, (ticker, row) in enumerate(weight_df.iterrows(), 1):
    r = i + 2
    bg = ALT if i % 2 == 0 else WHITE
    cel(ws_sens, r, 1, row["name"], bg=bg, halign="left")
    cel(ws_sens, r, 2, ticker, bg=bg)
    cel(ws_sens, r, 3, row["sector"], bg=SECTOR_COLORS.get(row["sector"], WHITE))
    for c2, wv in [(4, row["3Y"]), (5, row["5Y"]), (6, row["10Y"])]:
        wbg = GREEN if wv > 0.20 else (YELLOW if wv > 0.10 else (GREY if wv > 0.05 else WHITE))
        cel(ws_sens, r, c2, f"{wv*100:.1f}%", bg=wbg)
    cel(ws_sens, r, 7, f"{row['avg']*100:.1f}%", bg=bg, bold=True)
    ws_sens.row_dimensions[r].height = 22

set_cols(ws_sens, [22, 12, 16, 12, 12, 12, 12])

# ━━ SHEET 5: Asset Metrics ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ws_met = wb.create_sheet("Asset Metrics")
ws_met.merge_cells("A1:L1")
c = ws_met["A1"]; c.value = "Individual Asset Metrics — 3Y / 5Y / 10Y"
c.font = Font(bold=True, color=WHITE, size=13)
c.fill = PatternFill("solid", fgColor=DARK_BLUE)
c.alignment = Alignment(horizontal="center", vertical="center")
ws_met.row_dimensions[1].height = 30

mhdrs = ["Name","Ticker","Sector","Ann Ret 10Y","Sharpe 10Y","MaxDD 10Y",
         "Ann Ret 5Y","Sharpe 5Y","MaxDD 5Y","Ann Ret 3Y","Sharpe 3Y","MaxDD 3Y"]
for c2, h in enumerate(mhdrs, 1): hdr(ws_met, 2, c2, h)
ws_met.row_dimensions[2].height = 28

for i, (ticker, name) in enumerate(TICKERS.items(), 1):
    r = i + 2
    bg = ALT if i % 2 == 0 else WHITE
    cel(ws_met, r, 1, name, bg=bg, halign="left")
    cel(ws_met, r, 2, ticker, bg=bg)
    cel(ws_met, r, 3, SECTORS.get(ticker,""), bg=SECTOR_COLORS.get(SECTORS.get(ticker,""), WHITE))
    for ci, (mdf, col) in enumerate([(metrics_10,"ann_return"),(metrics_10,"sharpe"),(metrics_10,"max_dd"),
                                      (metrics_5,"ann_return"),(metrics_5,"sharpe"),(metrics_5,"max_dd"),
                                      (metrics_3,"ann_return"),(metrics_3,"sharpe"),(metrics_3,"max_dd")], 4):
        v = mdf.loc[ticker, col] if ticker in mdf.index else None
        if v is None or (isinstance(v, float) and np.isnan(v)):
            cel(ws_met, r, ci, "N/A", bg=bg)
        else:
            fv = f"{v:+.1f}%" if col in ("ann_return","max_dd") else f"{v:.3f}"
            if col=="max_dd": cbg = RED if v<-50 else (ORANGE if v<-30 else (YELLOW if v<-15 else GREEN))
            elif col=="ann_return": cbg = GREEN if v>20 else (YELLOW if v>10 else (ORANGE if v>0 else RED))
            else: cbg = GREEN if v>0.9 else (YELLOW if v>0.5 else (ORANGE if v>0 else RED))
            cel(ws_met, r, ci, fv, bg=cbg)
    ws_met.row_dimensions[r].height = 22

set_cols(ws_met, [20, 12, 16, 12, 10, 10, 12, 10, 10, 12, 10, 10])

wb.save(REPORT)
print(f"Report saved: {REPORT}")
