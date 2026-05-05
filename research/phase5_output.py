import pandas as pd
import numpy as np
import os
import sys
import io
import warnings
warnings.filterwarnings("ignore")

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

OUTPUT_DIR = r"C:\Users\lukas.larsson\Desktop\Privat\Project Reactor Core\output"
REPORT     = r"C:\Users\lukas.larsson\Desktop\Privat\Project Reactor Core\Reactor_Core_Report.xlsx"

TICKERS = {
    "xauusd":   "Gold",
    "xagusd":   "Silver",
    "copx.us":  "Copper Miners ETF",
    "paas.us":  "Pan American Silver",
    "xom.us":   "Exxon Mobil",
    "cvx.us":   "Chevron",
    "ccj.us":   "Cameco",
    "nvda.us":  "NVIDIA",
    "tsm.us":   "TSMC",
    "asml.us":  "ASML",
    "avgo.us":  "Broadcom",
    "amd.us":   "AMD",
    "lrcx.us":  "Lam Research",
    "qcom.us":  "Qualcomm",
    "adi.us":   "Analog Devices",
    "ter.us":   "Teradyne",
    "mu.us":    "Micron",
    "lite.us":  "Lumentum",
    "glw.us":   "Corning",
    "vrt.us":   "Vertiv",
    "etn.us":   "Eaton",
    "brk-b.us": "Berkshire Hathaway B",
    "lly.us":   "Eli Lilly",
    "jnj.us":   "J&J",
    "v.us":     "Visa",
    "ma.us":    "Mastercard",
    "wmt.us":   "Walmart",
    "cost.us":  "Costco",
    "jpm.us":   "JP Morgan",
    "pstg.us":  "Pure Storage",
    "tsla.us":  "Tesla",
    "aapl.us":  "Apple",
}

SECTORS = {
    "xauusd":   "Commodity", "xagusd": "Commodity",
    "copx.us":  "Commodity", "paas.us": "Commodity", "xom.us": "Commodity",
    "cvx.us":   "Commodity", "ccj.us": "Commodity",
    "nvda.us":  "Semiconductor", "tsm.us": "Semiconductor", "asml.us": "Semiconductor",
    "avgo.us":  "Semiconductor", "amd.us": "Semiconductor", "lrcx.us": "Semiconductor",
    "qcom.us":  "Semiconductor", "adi.us": "Semiconductor", "ter.us": "Semiconductor",
    "mu.us":    "Semiconductor", "lite.us": "Semiconductor", "glw.us": "Semiconductor",
    "vrt.us":   "Infrastructure", "etn.us": "Infrastructure",
    "brk-b.us": "Defensive", "lly.us": "Defensive", "jnj.us": "Defensive",
    "v.us":     "Defensive", "ma.us": "Defensive", "wmt.us": "Defensive",
    "cost.us":  "Defensive", "jpm.us": "Defensive",
    "pstg.us":  "Wildcard", "tsla.us": "Wildcard", "aapl.us": "Wildcard",
}

# ── Styling helpers ───────────────────────────────────────────────────────────
DARK_BLUE  = "1F4E79"
MID_BLUE   = "2E75B6"
LIGHT_BLUE = "D6E4F0"
ALT_ROW    = "EBF3FB"
GREEN      = "C6EFCE"
RED        = "FFC7CE"
YELLOW     = "FFEB9C"
ORANGE     = "FFCC99"
GREY       = "F2F2F2"
WHITE      = "FFFFFF"

thin = Side(style="thin", color="BFBFBF")
med  = Side(style="medium", color="2E75B6")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
BORDER_TOP = Border(left=thin, right=thin, top=med, bottom=thin)

def hdr(ws, row, col, value, bg=DARK_BLUE, fg=WHITE, bold=True, size=11, wrap=True, halign="center"):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = PatternFill("solid", fgColor=bg)
    c.font = Font(bold=bold, color=fg, size=size)
    c.alignment = Alignment(wrap_text=wrap, vertical="center", horizontal=halign)
    c.border = BORDER
    return c

def cell(ws, row, col, value, bg=None, bold=False, halign="center", fmt=None, color=None):
    c = ws.cell(row=row, column=col, value=value)
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    c.font = Font(bold=bold, color=color or "000000")
    c.alignment = Alignment(wrap_text=True, vertical="center", horizontal=halign)
    c.border = BORDER
    if fmt:
        c.number_format = fmt
    return c

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def set_row_heights(ws, start, end, h):
    for r in range(start, end + 1):
        ws.row_dimensions[r].height = h

# ── Load data ─────────────────────────────────────────────────────────────────
metrics_10y = pd.read_csv(os.path.join(OUTPUT_DIR, "metrics_10Y.csv"), index_col="ticker")
metrics_5y  = pd.read_csv(os.path.join(OUTPUT_DIR, "metrics_5Y.csv"),  index_col="ticker")
metrics_3y  = pd.read_csv(os.path.join(OUTPUT_DIR, "metrics_3Y.csv"),  index_col="ticker")
regime_df   = pd.read_csv(os.path.join(OUTPUT_DIR, "asset_regime_matrix.csv"))
regime_sum  = pd.read_csv(os.path.join(OUTPUT_DIR, "regime_summary.csv"))
weights_10y = pd.read_csv(os.path.join(OUTPUT_DIR, "weights_10Y.csv"))
weights_5y  = pd.read_csv(os.path.join(OUTPUT_DIR, "weights_5Y.csv"))
weights_3y  = pd.read_csv(os.path.join(OUTPUT_DIR, "weights_3Y.csv"))

# Build final portfolio: avg of Sharpe weights across 3Y/5Y/10Y, then score
def get_sharpe_weights(df):
    s = df[df["objective"] == "sharpe"].set_index("ticker")["weight"]
    return s

w3  = get_sharpe_weights(weights_3y)
w5  = get_sharpe_weights(weights_5y)
w10 = get_sharpe_weights(weights_10y)

all_t = sorted(set(w3.index) | set(w5.index) | set(w10.index))
weight_matrix = pd.DataFrame({
    "3Y":  w3.reindex(all_t, fill_value=0),
    "5Y":  w5.reindex(all_t, fill_value=0),
    "10Y": w10.reindex(all_t, fill_value=0),
})
weight_matrix["avg"] = weight_matrix.mean(axis=1)
weight_matrix["consistency"] = (weight_matrix[["3Y","5Y","10Y"]] > 0.02).sum(axis=1)
weight_matrix = weight_matrix[weight_matrix["avg"] > 0.005].sort_values("avg", ascending=False)
weight_matrix["name"] = [TICKERS.get(t, t) for t in weight_matrix.index]
weight_matrix["sector"] = [SECTORS.get(t, "Other") for t in weight_matrix.index]

# Confidence: 3/3 windows = High, 2/3 = Medium, 1/3 = Low
def confidence(c):
    return {3: "High", 2: "Medium", 1: "Low"}.get(c, "Low")

weight_matrix["confidence"] = weight_matrix["consistency"].apply(confidence)

# Regime scores
regime_df_idx = regime_df.set_index("ticker")
weight_matrix["pos_regimes"] = regime_df_idx.reindex(weight_matrix.index)["positive_regimes"].values
weight_matrix["avg_regime_ret"] = regime_df_idx.reindex(weight_matrix.index)["avg_return"].values

wb = openpyxl.Workbook()

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# SHEET 1: Final Portfolio
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ws1 = wb.active
ws1.title = "Final Portfolio"

ws1.merge_cells("A1:J1")
t = ws1["A1"]
t.value = "PROJECT REACTOR CORE — Final Portfolio"
t.font = Font(bold=True, color=WHITE, size=14)
t.fill = PatternFill("solid", fgColor=DARK_BLUE)
t.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 36

ws1.merge_cells("A2:J2")
t2 = ws1["A2"]
t2.value = "Sharpe-optimized weights averaged across 3Y / 5Y / 10Y windows | Min 1% / Max 25% per position"
t2.font = Font(italic=True, color="444444", size=10)
t2.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
t2.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[2].height = 20

headers_p = ["#", "Name", "Ticker", "Sector", "Avg Weight", "3Y Weight", "5Y Weight",
             "10Y Weight", "Confidence", "Positive Regimes"]
for c, h in enumerate(headers_p, 1):
    hdr(ws1, 3, c, h)
ws1.row_dimensions[3].height = 30

conf_colors = {"High": GREEN, "Medium": YELLOW, "Low": ORANGE}
sector_colors = {
    "Commodity": "FFF2CC", "Semiconductor": "E2EFDA", "Infrastructure": "DDEBF7",
    "Defensive": "FCE4D6", "China/EM": "F4CCCC", "Wildcard": "EAD1DC",
}

row = 4
total_weight = 0
for i, (ticker, row_data) in enumerate(weight_matrix.iterrows(), 1):
    bg = ALT_ROW if i % 2 == 0 else WHITE
    sec_bg = sector_colors.get(row_data["sector"], WHITE)
    cell(ws1, row, 1, i, bg=bg, halign="center")
    cell(ws1, row, 2, row_data["name"], bg=bg, bold=True, halign="left")
    cell(ws1, row, 3, ticker, bg=bg, halign="center")
    cell(ws1, row, 4, row_data["sector"], bg=sec_bg, halign="center")
    cell(ws1, row, 5, f"{row_data['avg']*100:.1f}%", bg=bg, bold=True, halign="center")
    cell(ws1, row, 6, f"{row_data['3Y']*100:.1f}%", bg=bg, halign="center")
    cell(ws1, row, 7, f"{row_data['5Y']*100:.1f}%", bg=bg, halign="center")
    cell(ws1, row, 8, f"{row_data['10Y']*100:.1f}%", bg=bg, halign="center")
    conf = row_data["confidence"]
    cell(ws1, row, 9, conf, bg=conf_colors.get(conf, WHITE), bold=True, halign="center")
    pos = row_data["pos_regimes"]
    pos_bg = GREEN if pos >= 5 else (YELLOW if pos >= 4 else ORANGE)
    cell(ws1, row, 10, f"{int(pos)}/6" if not np.isnan(pos) else "N/A", bg=pos_bg, halign="center")
    total_weight += row_data["avg"]
    row += 1
    ws1.row_dimensions[row-1].height = 22

# Total row
cell(ws1, row, 1, "", bg=LIGHT_BLUE)
cell(ws1, row, 2, "TOTAL", bg=LIGHT_BLUE, bold=True, halign="left")
for c in [3,4]: cell(ws1, row, c, "", bg=LIGHT_BLUE)
cell(ws1, row, 5, f"{total_weight*100:.1f}%", bg=LIGHT_BLUE, bold=True)
for c in [6,7,8,9,10]: cell(ws1, row, c, "", bg=LIGHT_BLUE)
ws1.row_dimensions[row].height = 22

set_col_widths(ws1, [4, 26, 12, 16, 11, 11, 11, 11, 12, 16])

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# SHEET 2: Asset Metrics
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ws2 = wb.create_sheet("Asset Metrics")

ws2.merge_cells("A1:L1")
t = ws2["A1"]
t.value = "Individual Asset Metrics — 3Y / 5Y / 10Y"
t.font = Font(bold=True, color=WHITE, size=13)
t.fill = PatternFill("solid", fgColor=DARK_BLUE)
t.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 30

hdrs2 = ["Name", "Ticker", "Sector",
         "Ann Ret 10Y", "Sharpe 10Y", "MaxDD 10Y",
         "Ann Ret 5Y",  "Sharpe 5Y",  "MaxDD 5Y",
         "Ann Ret 3Y",  "Sharpe 3Y",  "MaxDD 3Y"]
for c, h in enumerate(hdrs2, 1):
    hdr(ws2, 2, c, h)
ws2.row_dimensions[2].height = 28

def get_metric(df, ticker, col):
    if ticker in df.index and col in df.columns:
        v = df.loc[ticker, col]
        return v if not pd.isna(v) else None
    return None

for i, (ticker, name) in enumerate(TICKERS.items(), 1):
    r = i + 2
    bg = ALT_ROW if i % 2 == 0 else WHITE
    sec_bg = sector_colors.get(SECTORS.get(ticker, ""), WHITE)
    cell(ws2, r, 1, name, bg=bg, halign="left")
    cell(ws2, r, 2, ticker, bg=bg, halign="center")
    cell(ws2, r, 3, SECTORS.get(ticker, ""), bg=sec_bg, halign="center")

    for col_idx, (mdf, col) in enumerate([
        (metrics_10y, "ann_return"), (metrics_10y, "sharpe"), (metrics_10y, "max_dd"),
        (metrics_5y,  "ann_return"), (metrics_5y,  "sharpe"), (metrics_5y,  "max_dd"),
        (metrics_3y,  "ann_return"), (metrics_3y,  "sharpe"), (metrics_3y,  "max_dd"),
    ], 4):
        v = get_metric(mdf, ticker, col)
        if v is None:
            cell(ws2, r, col_idx, "N/A", bg=bg)
        else:
            fmt_v = f"{v:+.1f}%" if col in ("ann_return","max_dd") else f"{v:.3f}"
            is_ret = col == "ann_return"
            is_dd  = col == "max_dd"
            is_sh  = col == "sharpe"
            if is_dd:
                cbg = RED if v < -50 else (ORANGE if v < -30 else (YELLOW if v < -15 else GREEN))
            elif is_ret:
                cbg = GREEN if v > 20 else (YELLOW if v > 10 else (ORANGE if v > 0 else RED))
            elif is_sh:
                cbg = GREEN if v > 0.9 else (YELLOW if v > 0.5 else (ORANGE if v > 0 else RED))
            else:
                cbg = bg
            cell(ws2, r, col_idx, fmt_v, bg=cbg, halign="center")
    ws2.row_dimensions[r].height = 20

set_col_widths(ws2, [24, 12, 14, 12, 10, 10, 12, 10, 10, 12, 10, 10])

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# SHEET 3: Regime Analysis
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ws3 = wb.create_sheet("Regime Analysis")

REGIMES_ORDER = ["Pre-COVID Bull","COVID Crash","COVID Recovery",
                 "Rate Hike / Inflation","Post-Hike / AI Bull","Rate Cut"]

ws3.merge_cells(f"A1:{get_column_letter(3+len(REGIMES_ORDER))}1")
t = ws3["A1"]
t.value = "Asset Returns by Regime"
t.font = Font(bold=True, color=WHITE, size=13)
t.fill = PatternFill("solid", fgColor=DARK_BLUE)
t.alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 30

regime_hdrs = ["Name", "Ticker", "Pos/6"] + REGIMES_ORDER
for c, h in enumerate(regime_hdrs, 1):
    hdr(ws3, 2, c, h)
ws3.row_dimensions[2].height = 36

regime_df_idx2 = regime_df.set_index("ticker")
for i, (ticker, name) in enumerate(TICKERS.items(), 1):
    if ticker == "^spx":
        continue
    r = i + 2
    bg = ALT_ROW if i % 2 == 0 else WHITE
    cell(ws3, r, 1, name, bg=bg, halign="left")
    cell(ws3, r, 2, ticker, bg=bg, halign="center")
    pos = int(regime_df_idx2.loc[ticker, "positive_regimes"]) if ticker in regime_df_idx2.index else 0
    pos_bg = GREEN if pos >= 5 else (YELLOW if pos >= 4 else (ORANGE if pos >= 3 else RED))
    cell(ws3, r, 3, f"{pos}/6", bg=pos_bg, bold=True, halign="center")
    for j, rname in enumerate(REGIMES_ORDER, 4):
        v = regime_df_idx2.loc[ticker, rname] if ticker in regime_df_idx2.index else None
        if v is None or pd.isna(v):
            cell(ws3, r, j, "N/A", bg=bg)
        else:
            cbg = GREEN if v > 20 else (YELLOW if v > 0 else (ORANGE if v > -20 else RED))
            cell(ws3, r, j, f"{v:+.1f}%", bg=cbg, halign="center")
    ws3.row_dimensions[r].height = 20

# Portfolio vs SPX summary block
r_start = len(TICKERS) + 4
ws3.merge_cells(f"A{r_start}:{get_column_letter(3+len(REGIMES_ORDER))}{r_start}")
t = ws3.cell(row=r_start, column=1, value="Portfolio vs S&P 500 by Regime")
t.font = Font(bold=True, color=WHITE, size=11)
t.fill = PatternFill("solid", fgColor=MID_BLUE)
t.alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[r_start].height = 24

sum_hdrs = ["", "", ""] + [r["regime"] for _, r in regime_sum.iterrows()]
for c, h in enumerate(sum_hdrs, 1):
    hdr(ws3, r_start+1, c, h, bg=MID_BLUE)

for label, key, good_positive in [
    ("Portfolio", "portfolio_ret", True),
    ("S&P 500",   "spx_ret",       True),
    ("vs S&P 500","vs_spx",        True),
]:
    r_start += 2
    cell(ws3, r_start, 1, label, bold=True, halign="left")
    cell(ws3, r_start, 2, "", bg=WHITE)
    cell(ws3, r_start, 3, "", bg=WHITE)
    for c, (_, row_s) in enumerate(regime_sum.iterrows(), 4):
        v = row_s[key]
        if pd.isna(v):
            cell(ws3, r_start, c, "N/A")
        else:
            cbg = GREEN if v > 0 else RED
            cell(ws3, r_start, c, f"{v:+.1f}%", bg=cbg, bold=(label=="vs S&P 500"), halign="center")
    ws3.row_dimensions[r_start].height = 20

set_col_widths(ws3, [24, 12, 8] + [14]*len(REGIMES_ORDER))

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# SHEET 4: Weight Sensitivity
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ws4 = wb.create_sheet("Weight Sensitivity")

ws4.merge_cells("A1:H1")
t = ws4["A1"]
t.value = "Weight Sensitivity — How Allocations Shift Across Lookback Windows (Sharpe Objective)"
t.font = Font(bold=True, color=WHITE, size=12)
t.fill = PatternFill("solid", fgColor=DARK_BLUE)
t.alignment = Alignment(horizontal="center", vertical="center")
ws4.row_dimensions[1].height = 30

hdrs4 = ["Name", "Ticker", "Sector", "3Y Weight", "5Y Weight", "10Y Weight", "Avg Weight", "Stability"]
for c, h in enumerate(hdrs4, 1):
    hdr(ws4, 2, c, h)
ws4.row_dimensions[2].height = 28

all_tickers_sens = sorted(set(w3.index) | set(w5.index) | set(w10.index))
sens_rows = []
for ticker in all_tickers_sens:
    w3v = w3.get(ticker, 0)
    w5v = w5.get(ticker, 0)
    w10v = w10.get(ticker, 0)
    avg = (w3v + w5v + w10v) / 3
    std = np.std([w3v, w5v, w10v])
    sens_rows.append((ticker, w3v, w5v, w10v, avg, std))

sens_rows.sort(key=lambda x: x[4], reverse=True)

for i, (ticker, w3v, w5v, w10v, avg, std) in enumerate(sens_rows, 1):
    r = i + 2
    bg = ALT_ROW if i % 2 == 0 else WHITE
    name = TICKERS.get(ticker, ticker)
    sector = SECTORS.get(ticker, "Other")
    cell(ws4, r, 1, name, bg=bg, halign="left")
    cell(ws4, r, 2, ticker, bg=bg, halign="center")
    cell(ws4, r, 3, sector, bg=sector_colors.get(sector, WHITE), halign="center")
    for col, wv in [(4, w3v), (5, w5v), (6, w10v)]:
        wbg = GREEN if wv > 0.10 else (YELLOW if wv > 0.03 else (GREY if wv > 0.01 else WHITE))
        cell(ws4, r, col, f"{wv*100:.1f}%", bg=wbg, halign="center")
    cell(ws4, r, 7, f"{avg*100:.1f}%", bg=bg, bold=True, halign="center")
    stability = "Stable" if std < 0.03 else ("Variable" if std < 0.08 else "Unstable")
    stab_bg = GREEN if stability == "Stable" else (YELLOW if stability == "Variable" else RED)
    cell(ws4, r, 8, stability, bg=stab_bg, halign="center")
    ws4.row_dimensions[r].height = 20

set_col_widths(ws4, [26, 12, 14, 11, 11, 11, 11, 12])

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# SHEET 5: Summary
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ws5 = wb.create_sheet("Summary", 0)

ws5.merge_cells("A1:F1")
t = ws5["A1"]
t.value = "PROJECT REACTOR CORE"
t.font = Font(bold=True, color=WHITE, size=18)
t.fill = PatternFill("solid", fgColor=DARK_BLUE)
t.alignment = Alignment(horizontal="center", vertical="center")
ws5.row_dimensions[1].height = 48

ws5.merge_cells("A2:F2")
t2 = ws5["A2"]
t2.value = "Data-driven portfolio optimization | No thesis | Let Sharpe, Calmar, and total return decide"
t2.font = Font(italic=True, color="444444", size=11)
t2.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
t2.alignment = Alignment(horizontal="center", vertical="center")
ws5.row_dimensions[2].height = 24

# Key stats block
def stat_block(ws, start_row, title, stats):
    ws.merge_cells(f"A{start_row}:F{start_row}")
    t = ws.cell(row=start_row, column=1, value=title)
    t.font = Font(bold=True, color=WHITE, size=11)
    t.fill = PatternFill("solid", fgColor=MID_BLUE)
    t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[start_row].height = 24
    r = start_row + 1
    for label, value, bg in stats:
        cell(ws, r, 1, label, bg=GREY, bold=True, halign="left")
        ws.merge_cells(f"B{r}:F{r}")
        c = ws.cell(row=r, column=2, value=value)
        c.fill = PatternFill("solid", fgColor=bg or WHITE)
        c.font = Font(size=11)
        c.alignment = Alignment(vertical="center", indent=1)
        c.border = BORDER
        ws.row_dimensions[r].height = 20
        r += 1
    return r + 1

r = 4
r = stat_block(ws5, r, "Optimization Setup", [
    ("Universe",        "37 liquid USD-denominated tickers across 6 sectors", LIGHT_BLUE),
    ("Windows tested",  "3Y (2023-), 5Y (2021-), 10Y (2016-)", LIGHT_BLUE),
    ("Objectives",      "Max Sharpe | Max Calmar | Blended", LIGHT_BLUE),
    ("Constraints",     "Min 1% / Max 25% per position", LIGHT_BLUE),
    ("Final weights",   "Average of Sharpe-optimal weights across all 3 windows", LIGHT_BLUE),
])

r = stat_block(ws5, r, "Optimal Portfolio Performance (10Y Sharpe)", [
    ("Ann. Return",     "30.4%",     GREEN),
    ("Ann. Volatility", "17.6%",     YELLOW),
    ("Sharpe Ratio",    "1.72",      GREEN),
    ("Max Drawdown",    "-17.3%",    YELLOW),
    ("Calmar Ratio",    "1.76",      GREEN),
    ("Total Return",    "+290%",     GREEN),
])

r = stat_block(ws5, r, "Portfolio vs S&P 500 by Regime", [
    ("Pre-COVID Bull (2016-2020)",       "Portfolio -19% vs S&P — gives up upside in calm equity bull", ORANGE),
    ("COVID Crash (Feb-Mar 2020)",       "Portfolio -7% vs S&P -20% — +13% outperformance", GREEN),
    ("COVID Recovery (2020-2021)",       "Portfolio +126% vs S&P +93% — +33% outperformance", GREEN),
    ("Rate Hike/Inflation (2022-2023)",  "Portfolio +20% vs S&P -4% — +25% outperformance", GREEN),
    ("Post-Hike/AI Bull (2023-2024)",    "Portfolio +62% vs S&P +23% — +39% outperformance", GREEN),
    ("Rate Cut (2024-present)",          "Portfolio +64% vs S&P +18% — +46% outperformance", GREEN),
])

r = stat_block(ws5, r, "All-Weather Core (positive in 5-6/6 regimes)", [
    ("NVIDIA",          "6/6 regimes positive | Sharpe 1.40 | Ann ret 70%", GREEN),
    ("Gold",            "5/6 regimes positive | Sharpe 0.91 | Low correlation hedge", GREEN),
    ("Walmart",         "5/6 regimes positive | Sharpe 0.93 | Defensive compounder", GREEN),
    ("Eli Lilly",       "4/6 regimes positive | Sharpe 1.03 | Best risk-adj in defensive bucket", YELLOW),
    ("Lam Research",    "5/6 regimes positive | Strong semi cyclical with low drawdown vs peers", GREEN),
    ("Broadcom",        "5/6 regimes positive | Sharpe 0.98 | Consistent compounder", GREEN),
])

r = stat_block(ws5, r, "Avoid or Size Carefully", [
    ("US Oil Fund (USO)",  "-87% drawdown in COVID crash. Regime-dependent, low Calmar 0.06", RED),
    ("Alibaba / JD.com",   "2-4/6 regimes positive. China political risk, persistent underperformance", RED),
    ("Dow Inc",            "2/6 positive. Negative ann return 10Y. No regime where it clearly wins", RED),
    ("LyondellBasell",     "Ann ret 2.8% over 10Y. -68% max DD. Capital destruction vs alternatives", RED),
])

set_col_widths(ws5, [24, 16, 16, 16, 16, 16])

wb.save(REPORT)
print(f"Report saved: {REPORT}")
