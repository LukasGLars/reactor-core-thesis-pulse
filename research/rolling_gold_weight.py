import pandas as pd
import numpy as np
import os, sys, io, warnings
from scipy.optimize import minimize
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
warnings.filterwarnings("ignore")
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

DATA_DIR = r"C:\Users\lukas.larsson\Desktop\Privat\Project Reactor Core\data"
REPORT   = r"C:\Users\lukas.larsson\Desktop\Privat\Project Reactor Core\Rolling_Gold_Weight.xlsx"

TICKERS = ["xauusd","lly.us","wmt.us","avgo.us","cost.us","ccj.us","vrt.us","jnj.us"]
NAMES   = {
    "xauusd":"Gold","lly.us":"Eli Lilly","wmt.us":"Walmart",
    "avgo.us":"Broadcom","cost.us":"Costco","ccj.us":"Cameco",
    "vrt.us":"Vertiv","jnj.us":"J&J",
}
MIN_W, MAX_W = 0.05, 0.40

def load(ticker):
    fname = ticker.replace("^","").replace("-","_") + ".csv"
    path  = os.path.join(DATA_DIR, fname)
    if not os.path.exists(path): return None
    df = pd.read_csv(path)
    df["Date"] = pd.to_datetime(df["Date"])
    df = df.set_index("Date").sort_index()
    return df["Close"].replace(0, np.nan).dropna()

def optimize(df, n_starts=60):
    rets = df.pct_change().dropna()
    mu   = rets.mean(); cov = rets.cov()
    n    = len(df.columns)
    bounds = [(MIN_W, MAX_W)] * n
    def neg_sharpe(w):
        r = np.dot(w, mu)*252
        v = np.sqrt(w @ cov @ w)*np.sqrt(252)
        return -r/v if v > 0 else 0
    constraints = [{"type":"eq","fun":lambda w: np.sum(w)-1}]
    best = None
    for _ in range(n_starts):
        w0 = np.random.dirichlet(np.ones(n))
        w0 = np.clip(w0, MIN_W, MAX_W); w0 /= w0.sum()
        res = minimize(neg_sharpe, w0, method="SLSQP", bounds=bounds,
                       constraints=constraints, options={"maxiter":1000,"ftol":1e-12})
        if res.success and (best is None or res.fun < best.fun):
            best = res
    return best.x if best else None

def port_metrics(df, w):
    wa   = np.array(w)
    rets = df.pct_change().dropna()
    mu   = rets.mean(); cov = rets.cov()
    ann_ret = np.dot(wa, mu)*252
    ann_vol = np.sqrt(wa @ cov @ wa)*np.sqrt(252)
    sharpe  = ann_ret/ann_vol if ann_vol > 0 else 0
    port    = (df/df.iloc[0]).dot(wa)
    max_dd  = ((port - port.cummax())/port.cummax()).min()
    total   = port.iloc[-1]-1
    return {"sharpe":round(sharpe,3),"ann_ret":round(ann_ret*100,2),
            "max_dd":round(max_dd*100,2),"total":round(total*100,1)}

# ── Build full price matrix ───────────────────────────────────────────────────
price_dict = {t: load(t) for t in TICKERS if load(t) is not None}
price_df   = pd.DataFrame(price_dict).dropna()

# ── Rolling 3Y windows — step every quarter ──────────────────────────────────
WINDOW_DAYS = 252 * 3   # 3 years
STEP_DAYS   = 63        # ~1 quarter

print("=" * 65)
print("  Rolling 3Y optimization — gold weight over time")
print("  Step: quarterly   Min/Max per position: 5% / 40%")
print("=" * 65)
print(f"\n  {'Window start':12}  {'Window end':12}  {'Gold%':>7}  "
      f"{'Sharpe':>7}  {'AnnRet':>8}  {'MaxDD':>8}  {'Gold price':>10}")

results = []
idx = price_df.index

for start_i in range(0, len(idx) - WINDOW_DAYS, STEP_DAYS):
    end_i   = start_i + WINDOW_DAYS
    if end_i >= len(idx): break
    w_start = idx[start_i]
    w_end   = idx[end_i]
    df_w    = price_df.iloc[start_i:end_i+1]

    # Need enough data for all tickers in this window
    df_w = df_w.dropna()
    if len(df_w) < 200 or "xauusd" not in df_w.columns: continue

    w = optimize(df_w)
    if w is None: continue

    w_dict    = dict(zip(df_w.columns, w))
    m         = port_metrics(df_w, w)
    gold_w    = round(w_dict.get("xauusd",0)*100, 1)
    gold_price= round(df_w["xauusd"].iloc[-1], 0)

    row = {
        "window_start": w_start, "window_end": w_end,
        "gold_weight":  gold_w,
        "sharpe":       m["sharpe"],
        "ann_ret":      m["ann_ret"],
        "max_dd":       m["max_dd"],
        "gold_price_end": gold_price,
        "gold_ret_window": round((df_w["xauusd"].iloc[-1]/df_w["xauusd"].iloc[0]-1)*100,1),
    }
    # All weights
    for t in TICKERS:
        row[f"w_{t}"] = round(w_dict.get(t,0)*100,1)

    results.append(row)
    print(f"  {str(w_start.date()):12}  {str(w_end.date()):12}  "
          f"{gold_w:>6.1f}%  {m['sharpe']:>7.3f}  {m['ann_ret']:>7.1f}%  "
          f"{m['max_dd']:>7.1f}%  ${gold_price:>9,.0f}")

df_res = pd.DataFrame(results)

# ── Summary stats ─────────────────────────────────────────────────────────────
gold_weights = df_res["gold_weight"]
print(f"\n  Gold weight statistics across {len(df_res)} rolling windows:")
print(f"    Min:    {gold_weights.min():.1f}%")
print(f"    Max:    {gold_weights.max():.1f}%")
print(f"    Mean:   {gold_weights.mean():.1f}%")
print(f"    Median: {gold_weights.median():.1f}%")
print(f"    Stdev:  {gold_weights.std():.1f}%")

# Split: pre-2020 vs post-2020
pre  = df_res[df_res["window_end"] <  "2020-01-01"]["gold_weight"]
post = df_res[df_res["window_end"] >= "2020-01-01"]["gold_weight"]
print(f"\n  Pre-2020 windows ({len(pre)}):   mean {pre.mean():.1f}%  median {pre.median():.1f}%")
print(f"  Post-2020 windows ({len(post)}):  mean {post.mean():.1f}%  median {post.median():.1f}%")

# Correlation: gold weight vs gold price return in that window
corr = round(df_res["gold_weight"].corr(df_res["gold_ret_window"]), 3)
print(f"\n  Correlation: gold weight vs gold return in window: {corr}")
print(f"  (positive = optimizer chases gold after it has already run)")

# ── Excel output ──────────────────────────────────────────────────────────────
DARK="1F4E79"; MID="2E75B6"; LIGHT="D6E4F0"; ALT="EBF3FB"
GRN="C6EFCE"; RED="FFC7CE"; YEL="FFEB9C"; WHT="FFFFFF"; GREY="F2F2F2"
ORG="FFCC99"

thin   = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr(ws, row, col, val, bg=DARK, fg=WHT, bold=True, sz=10, ha="center"):
    c = ws.cell(row=row, column=col, value=val)
    c.fill = PatternFill("solid", fgColor=bg); c.font = Font(bold=bold, color=fg, size=sz)
    c.alignment = Alignment(wrap_text=True, vertical="center", horizontal=ha)
    c.border = BORDER; return c

def cel(ws, row, col, val, bg=None, bold=False, ha="center", color="000000"):
    c = ws.cell(row=row, column=col, value=val)
    if bg: c.fill = PatternFill("solid", fgColor=bg)
    c.font = Font(bold=bold, color=color)
    c.alignment = Alignment(wrap_text=True, vertical="center", horizontal=ha)
    c.border = BORDER; return c

def title(ws, row, nc, text, bg=DARK, sz=13):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=nc)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(bold=True, color=WHT, size=sz)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 34

def cols_w(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

wb = openpyxl.Workbook()

# ━━ SHEET 1: Rolling Results ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ws1 = wb.active; ws1.title = "Rolling Gold Weight"
title(ws1, 1, 14, "Rolling 3Y Optimization — Gold Weight Over Time (Quarterly Steps)")

ws1.merge_cells("A2:N2")
c = ws1["A2"]
c.value = (f"Each row = 3Y optimization window ending on that date.  "
           f"Gold weight highlighted: green <25% | yellow 25-35% | red >35%.  "
           f"Pre-2020 mean: {pre.mean():.1f}%  |  Post-2020 mean: {post.mean():.1f}%  |  "
           f"Corr(gold weight, gold return): {corr}")
c.font = Font(italic=True, size=9, color="444444")
c.fill = PatternFill("solid", fgColor=LIGHT)
c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws1.row_dimensions[2].height = 28

hdrs1 = (["Window End","Gold Ret%","Gold Price","Gold Weight","Sharpe","AnnRet%","MaxDD%"] +
         [NAMES[t] for t in TICKERS if t != "xauusd"])
for ci, h in enumerate(hdrs1, 1): hdr(ws1, 3, ci, h)
ws1.row_dimensions[3].height = 30

for ri, row in enumerate(results, 4):
    bg = ALT if ri%2==0 else WHT
    gw = row["gold_weight"]
    gw_bg = GRN if gw <= 25 else (YEL if gw <= 35 else RED)
    era_bg = "E8F5E9" if row["window_end"] < pd.Timestamp("2020-01-01") else WHT

    cel(ws1, ri, 1, str(row["window_end"].date()),         bg=era_bg)
    cel(ws1, ri, 2, f"{row['gold_ret_window']:+.1f}%",
        bg=GRN if row["gold_ret_window"]>0 else RED)
    cel(ws1, ri, 3, f"${row['gold_price_end']:,.0f}",      bg=era_bg)
    cel(ws1, ri, 4, f"{gw:.1f}%",                          bg=gw_bg, bold=True)
    cel(ws1, ri, 5, row["sharpe"],                          bg=era_bg)
    cel(ws1, ri, 6, f"{row['ann_ret']:.1f}%",              bg=era_bg)
    cel(ws1, ri, 7, f"{row['max_dd']:.1f}%",               bg=era_bg)
    for ci_off, t in enumerate([t for t in TICKERS if t != "xauusd"], 8):
        cel(ws1, ri, ci_off, f"{row.get(f'w_{t}',0):.1f}%", bg=bg)
    ws1.row_dimensions[ri].height = 18

# Summary block
ri_s = len(results) + 6
title(ws1, ri_s, 14, "Gold Weight Distribution Summary", bg=MID, sz=11)
ws1.row_dimensions[ri_s].height = 26; ri_s += 1

summ_rows = [
    ("All windows",       f"{gold_weights.min():.1f}%", f"{gold_weights.max():.1f}%",
                          f"{gold_weights.mean():.1f}%", f"{gold_weights.median():.1f}%",
                          f"{gold_weights.std():.1f}%"),
    ("Pre-2020",          f"{pre.min():.1f}%", f"{pre.max():.1f}%",
                          f"{pre.mean():.1f}%", f"{pre.median():.1f}%", f"{pre.std():.1f}%"),
    ("Post-2020",         f"{post.min():.1f}%", f"{post.max():.1f}%",
                          f"{post.mean():.1f}%", f"{post.median():.1f}%", f"{post.std():.1f}%"),
]
for ci, h in enumerate(["Period","Min","Max","Mean","Median","Stdev"], 1):
    hdr(ws1, ri_s, ci, h, bg=GREY, fg="000000")
ws1.row_dimensions[ri_s].height = 24; ri_s += 1

for i, (period, mn, mx, mean, med, std) in enumerate(summ_rows):
    bg = ALT if i%2==0 else WHT
    is_pre = "Pre" in period
    row_bg = "E8F5E9" if is_pre else bg
    cel(ws1, ri_s, 1, period, bg=GREY, bold=True, ha="left")
    cel(ws1, ri_s, 2, mn,   bg=row_bg)
    cel(ws1, ri_s, 3, mx,   bg=row_bg)
    cel(ws1, ri_s, 4, mean, bg=row_bg, bold=True)
    cel(ws1, ri_s, 5, med,  bg=row_bg)
    cel(ws1, ri_s, 6, std,  bg=row_bg)
    ws1.row_dimensions[ri_s].height = 22; ri_s += 1

# Optimizer-chasing note
ri_s += 1
ws1.merge_cells(f"A{ri_s}:N{ri_s}")
c = ws1.cell(row=ri_s, column=1,
             value=(f"Optimizer momentum signal: correlation between gold weight assigned and gold return "
                    f"in that window = {corr:.3f}. "
                    f"{'High positive correlation confirms the optimizer is chasing gold performance.' if corr > 0.5 else 'Moderate correlation — gold weight partially but not entirely driven by recent performance.'}"))
c.font = Font(italic=True, size=10, bold=True,
              color="FF0000" if corr > 0.5 else "FF6600")
c.fill = PatternFill("solid", fgColor="FFF2CC")
c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws1.row_dimensions[ri_s].height = 32

cols_w(ws1, [13,11,13,13,9,10,10, 9,9,10,8,9,9,8])

wb.save(REPORT)
print(f"\n  Report saved: {REPORT}")
print("=" * 65)
