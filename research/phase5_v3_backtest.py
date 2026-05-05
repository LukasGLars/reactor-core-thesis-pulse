import pandas as pd
import numpy as np
import os, sys, io, warnings
from scipy.optimize import minimize
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
warnings.filterwarnings("ignore")
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

DATA_DIR = r"C:\Users\lukas.larsson\Desktop\Privat\Project Reactor Core\data"
REPORT   = r"C:\Users\lukas.larsson\Desktop\Privat\Project Reactor Core\Reactor_Core_v3_Backtest.xlsx"

# ── Portfolio Definitions ─────────────────────────────────────────────────────
V3_TICKERS = ["xauusd","xagusd","lly.us","wmt.us","ccj.us","vrt.us","avgo.us","jnj.us"]
V3_NAMES   = {"xauusd":"Gold","xagusd":"Silver","lly.us":"Eli Lilly","wmt.us":"Walmart",
               "ccj.us":"Cameco","vrt.us":"Vertiv","avgo.us":"Broadcom","jnj.us":"J&J"}
V3_ROLES   = {"xauusd":"Hedge","xagusd":"Hedge","lly.us":"Carry","wmt.us":"Carry",
               "ccj.us":"Cyclical","vrt.us":"Convexity","avgo.us":"Convexity","jnj.us":"Carry"}
# Raw weights (sum=0.98, 2% permanent cash)
V3_RAW  = {"xauusd":0.250,"xagusd":0.100,"lly.us":0.150,"wmt.us":0.150,
            "ccj.us":0.100,"vrt.us":0.100,"avgo.us":0.090,"jnj.us":0.060}
V3_NORM = V3_RAW  # weights sum to 1.0, no cash

V2_TICKERS = ["xauusd","wmt.us","lly.us","vrt.us","avgo.us","ccj.us","jnj.us","cost.us"]
V2_WEIGHTS = {"xauusd":0.250,"wmt.us":0.227,"lly.us":0.197,"vrt.us":0.090,
              "avgo.us":0.079,"ccj.us":0.057,"jnj.us":0.050,"cost.us":0.050}
V2_NAMES   = {"xauusd":"Gold","wmt.us":"Walmart","lly.us":"Eli Lilly","vrt.us":"Vertiv",
              "avgo.us":"Broadcom","ccj.us":"Cameco","jnj.us":"J&J","cost.us":"Costco"}

WINDOWS      = {"3Y":"2023-04-01","5Y":"2021-04-01","10Y":"2016-04-01"}
GOLD_CAP     = 0.25
N_RESTARTS   = 80
IS_START     = "2016-04-01"
IS_END       = "2026-03-31"
DCA_START    = "2018-08-31"
INITIAL_KR   = 1_000_000
MONTHLY_KR   = 6_000
GOLD_CAPS_T  = [0.10, 0.15, 0.20, 0.25, 0.30, 0.35, 0.40]
SHOCK_LEVELS = [-0.10, -0.20, -0.30, -0.40, -0.50]

REGIMES = {
    "Pre-COVID Bull":      ("2016-04-01","2020-02-19"),
    "COVID Crash":         ("2020-02-20","2020-03-23"),
    "COVID Recovery":      ("2020-03-24","2021-12-31"),
    "Rate Hike/Inflation": ("2022-01-01","2023-06-30"),
    "Post-Hike/AI Bull":   ("2023-07-01","2024-08-31"),
    "Rate Cut":            ("2024-09-01","2026-03-31"),
}

# ── Styling ───────────────────────────────────────────────────────────────────
DARK="1F4E79"; MID="2E75B6"; LIGHT="D6E4F0"
ALT="EBF3FB"; GRN="C6EFCE"; RED="FFC7CE"; YEL="FFEB9C"; ORG="FFCC99"
GREY="F2F2F2"; WHT="FFFFFF"
thin   = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr(ws, row, col, val, bg=DARK, fg=WHT, bold=True, sz=10, ha="center"):
    c = ws.cell(row=row, column=col, value=val)
    c.fill = PatternFill("solid", fgColor=bg)
    c.font = Font(bold=bold, color=fg, size=sz)
    c.alignment = Alignment(wrap_text=True, vertical="center", horizontal=ha)
    c.border = BORDER; return c

def cel(ws, row, col, val, bg=None, bold=False, ha="center", color="000000"):
    c = ws.cell(row=row, column=col, value=val)
    if bg: c.fill = PatternFill("solid", fgColor=bg)
    c.font = Font(bold=bold, color=color)
    c.alignment = Alignment(wrap_text=True, vertical="center", horizontal=ha)
    c.border = BORDER; return c

def cw(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def title_row(ws, ncols, text, sub=None):
    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    c = ws["A1"]; c.value = text
    c.font = Font(bold=True, color=WHT, size=13)
    c.fill = PatternFill("solid", fgColor=DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30
    if sub:
        ws.merge_cells(f"A2:{get_column_letter(ncols)}2")
        c2 = ws["A2"]; c2.value = sub
        c2.font = Font(italic=True, size=9, color="444444")
        c2.fill = PatternFill("solid", fgColor=LIGHT)
        c2.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[2].height = 20
        return 3
    return 2

# ── Data Loading ──────────────────────────────────────────────────────────────
def load(ticker):
    fname = ticker.replace("^","").replace("-","_") + ".csv"
    path  = os.path.join(DATA_DIR, fname)
    if not os.path.exists(path): return None
    df = pd.read_csv(path)
    df["Date"] = pd.to_datetime(df["Date"])
    return df.set_index("Date").sort_index()["Close"].replace(0, np.nan).dropna()

print("Loading data...")
all_tickers = list(set(V3_TICKERS + V2_TICKERS))
price_data  = {t: load(t) for t in all_tickers if load(t) is not None}
spx         = load("spx")

# HY-IG spread (macro framework signal)
MACRO_DIR = r"C:\Users\lukas.larsson\Desktop\Privat\macro_data\portfolio_data"
def load_spread():
    hy_p = os.path.join(MACRO_DIR, "BAMLH0A0HYM2 (3).csv")
    ig_p = os.path.join(MACRO_DIR, "BAMLC0A0CM (4).csv")
    if not os.path.exists(hy_p) or not os.path.exists(ig_p): return None
    hy = pd.read_csv(hy_p, parse_dates=["observation_date"]).set_index("observation_date")["BAMLH0A0HYM2"]
    ig = pd.read_csv(ig_p, parse_dates=["observation_date"]).set_index("observation_date")["BAMLC0A0CM"]
    s  = (hy - ig) * 100  # % -> bps
    s.index.name = "Date"
    return s.dropna()

hy_ig_spread = load_spread()
SPREAD_START = str(hy_ig_spread.index[0].date()) if hy_ig_spread is not None else DCA_START
RC_DCA_MONTHLY = 25_000   # actual RC monthly DCA (vs 6k used in v2/v3 comparison)

print("\nData availability:")
for t in V3_TICKERS:
    s = price_data.get(t)
    if s is not None:
        print(f"  {t:12s}: {s.index[0].date()} -> {s.index[-1].date()} ({len(s)} rows)")
    else:
        print(f"  {t:12s}: NOT FOUND")

# ── Core Functions ────────────────────────────────────────────────────────────
def build_df(tickers, start, end=IS_END):
    avail = [t for t in tickers if t in price_data]
    df = pd.DataFrame({t: price_data[t] for t in avail})
    return df[(df.index >= start) & (df.index <= end)].dropna()

def port_metrics(df, weights):
    tickers = list(df.columns)
    w = np.array([weights.get(t, 0) for t in tickers], dtype=float)
    w /= w.sum()
    lr = np.log(df / df.shift(1)).dropna()
    mu = lr.mean().values; cov = lr.cov().values
    ann_ret = np.dot(w, mu) * 252
    ann_vol = np.sqrt(w @ cov @ w) * np.sqrt(252)
    sharpe  = ann_ret / ann_vol if ann_vol > 0 else 0
    port    = (df / df.iloc[0]).dot(w)
    max_dd  = ((port - port.cummax()) / port.cummax()).min()
    calmar  = ann_ret / abs(max_dd) if max_dd < 0 else 0
    total   = port.iloc[-1] - 1
    return {"sharpe": round(sharpe,3), "ann_ret": round(ann_ret*100,2),
            "ann_vol": round(ann_vol*100,2), "max_dd": round(max_dd*100,2),
            "calmar": round(calmar,3), "total": round(total*100,1),
            "start": str(df.index[0].date()), "end": str(df.index[-1].date())}

def asset_metrics_single(series, start, end=IS_END):
    s = series[(series.index >= start) & (series.index <= end)].dropna()
    if len(s) < 20: return {}
    lr = np.log(s / s.shift(1)).dropna()
    n = len(s) / 252
    ann_ret = (s.iloc[-1]/s.iloc[0])**(1/n) - 1 if n > 0 else 0
    ann_vol = lr.std() * np.sqrt(252)
    sharpe  = ann_ret / ann_vol if ann_vol > 0 else 0
    max_dd  = ((s - s.cummax()) / s.cummax()).min()
    return {"ann_ret": round(ann_ret*100,2), "sharpe": round(sharpe,3),
            "max_dd": round(max_dd*100,2)}

def regime_ret_single(series, rstart, rend):
    s = series.loc[rstart:rend].dropna()
    if len(s) < 5: return None
    return round((s.iloc[-1]/s.iloc[0]-1)*100, 1)

def regime_port_ret(df, weights, rstart, rend):
    s = df.loc[rstart:rend].dropna(how="all")
    if len(s) < 5: return None
    tickers = list(s.columns)
    w = np.array([weights.get(t,0) for t in tickers], dtype=float); w /= w.sum()
    port = (s / s.iloc[0]).dot(w)
    ret  = port.iloc[-1] - 1
    dd   = ((port - port.cummax()) / port.cummax()).min()
    return {"ret": round(ret*100,1), "max_dd": round(dd*100,1)}

def optimize(df, gold_cap=0.25, n=N_RESTARTS):
    tickers = list(df.columns)
    lr = np.log(df / df.shift(1)).dropna()
    mu = lr.mean().values; cov = lr.cov().values
    na = len(tickers)
    bounds = [(0.05, gold_cap if t == "xauusd" else 0.40) for t in tickers]
    cons   = [{"type":"eq","fun": lambda w: np.sum(w)-1}]
    def neg_sharpe(w):
        r = np.dot(w, mu)*252; v = np.sqrt(w @ cov @ w)*np.sqrt(252)
        return -r/v if v > 0 else 0
    best = None
    for _ in range(n):
        w0 = np.random.dirichlet(np.ones(na))
        w0 = np.clip(w0,[b[0] for b in bounds],[b[1] for b in bounds]); w0/=w0.sum()
        res = minimize(neg_sharpe, w0, method="SLSQP", bounds=bounds, constraints=cons,
                       options={"maxiter":1000,"ftol":1e-12})
        if res.success and (best is None or res.fun < best.fun): best = res
    return dict(zip(tickers, best.x)) if best else None

# ═══════════════════════════════════════════════════════════════════════════════
# STEP 1 — Core Metrics
# ═══════════════════════════════════════════════════════════════════════════════
print("\nStep 1: Core metrics...")
core_results = []

for wname, wstart in WINDOWS.items():
    # V2 fixed
    df2 = build_df(V2_TICKERS, wstart)
    if len(df2) >= 20:
        m = port_metrics(df2, V2_WEIGHTS)
        core_results.append({"portfolio":"v2","window":wname,"type":"Fixed",**m})
        opt2 = optimize(df2)
        if opt2:
            m_opt = port_metrics(df2, opt2)
            core_results.append({"portfolio":"v2","window":wname,"type":"Optimized",
                                  "opt_weights":opt2,**m_opt})
        print(f"  v2 {wname}: Fixed Sharpe {m['sharpe']}  AnnRet {m['ann_ret']}%  from {m['start']}")

    # V3 fixed (normalized)
    df3 = build_df(V3_TICKERS, wstart)
    if len(df3) >= 20:
        m = port_metrics(df3, V3_NORM)
        core_results.append({"portfolio":"v3","window":wname,"type":"Fixed",**m})
        opt3 = optimize(df3)
        if opt3:
            m_opt = port_metrics(df3, opt3)
            core_results.append({"portfolio":"v3","window":wname,"type":"Optimized",
                                  "opt_weights":opt3,**m_opt})
        print(f"  v3 {wname}: Fixed Sharpe {m['sharpe']}  AnnRet {m['ann_ret']}%  from {m['start']}")

# ═══════════════════════════════════════════════════════════════════════════════
# STEP 2 — Asset Metrics
# ═══════════════════════════════════════════════════════════════════════════════
print("\nStep 2: Asset metrics...")
asset_rows = []
for t in V3_TICKERS:
    s = price_data.get(t)
    if s is None: continue
    row = {"ticker":t,"name":V3_NAMES[t],"role":V3_ROLES[t],"weight":V3_RAW[t]}
    for wname, wstart in WINDOWS.items():
        am = asset_metrics_single(s, wstart)
        row[f"ann_ret_{wname}"] = am.get("ann_ret","N/A")
        row[f"sharpe_{wname}"]  = am.get("sharpe","N/A")
        row[f"max_dd_{wname}"]  = am.get("max_dd","N/A")
    asset_rows.append(row)

# ═══════════════════════════════════════════════════════════════════════════════
# STEP 3 — Regime Analysis
# ═══════════════════════════════════════════════════════════════════════════════
print("\nStep 3: Regime analysis...")
df_full_v3 = build_df(V3_TICKERS, IS_START)
df_full_v2 = build_df(V2_TICKERS, IS_START)

asset_regime = {}
for t in V3_TICKERS:
    s = price_data.get(t)
    if s is None: continue
    asset_regime[t] = {}
    rets_list = []
    for rname, (rs, re) in REGIMES.items():
        r = regime_ret_single(s, rs, re)
        asset_regime[t][rname] = r
        if r is not None: rets_list.append(r)
    asset_regime[t]["wins"] = sum(1 for v in rets_list if v > 0)
    asset_regime[t]["avg"]  = round(np.mean(rets_list),1) if rets_list else None

regime_v3 = []
for rname, (rs, re) in REGIMES.items():
    pm   = regime_port_ret(df_full_v3, V3_NORM, rs, re)
    spxr = regime_ret_single(spx, rs, re) if spx is not None else None
    vs   = round(pm["ret"]-spxr,1) if pm and spxr is not None else None
    regime_v3.append({"regime":rname,"start":rs,"end":re,
                       "port_ret":pm["ret"] if pm else None,
                       "port_dd": pm["max_dd"] if pm else None,
                       "spx_ret":spxr, "vs_spx":vs})

regime_v2 = []
for rname, (rs, re) in REGIMES.items():
    pm   = regime_port_ret(df_full_v2, V2_WEIGHTS, rs, re)
    spxr = regime_ret_single(spx, rs, re) if spx is not None else None
    vs   = round(pm["ret"]-spxr,1) if pm and spxr is not None else None
    regime_v2.append({"regime":rname,"start":rs,"end":re,
                       "port_ret":pm["ret"] if pm else None, "vs_spx":vs})

# ═══════════════════════════════════════════════════════════════════════════════
# STEP 4 — Gold Cap Sensitivity
# ═══════════════════════════════════════════════════════════════════════════════
print("\nStep 4: Gold cap sensitivity...")
cap_results = []
for cap in GOLD_CAPS_T:
    cap_label = "Uncapped" if cap >= 0.40 else f"{int(cap*100)}%"
    for wname, wstart in WINDOWS.items():
        df = build_df(V3_TICKERS, wstart)
        if len(df) < 20: continue
        opt = optimize(df, gold_cap=cap)
        if opt:
            m = port_metrics(df, opt)
            cap_results.append({"cap":cap,"cap_label":cap_label,"window":wname,
                                 "gold_w":round(opt.get("xauusd",0)*100,1),
                                 "silver_w":round(opt.get("xagusd",0)*100,1),
                                 "opt_weights":opt, **m})
    print(f"  Cap {cap_label} done")

# ═══════════════════════════════════════════════════════════════════════════════
# STEP 5 — Leave-One-Out Audit
# ═══════════════════════════════════════════════════════════════════════════════
print("\nStep 5: Leave-one-out...")
baseline_by_window = {}
for wname, wstart in WINDOWS.items():
    df = build_df(V3_TICKERS, wstart)
    if len(df) < 20: continue
    opt = optimize(df)
    if opt:
        m = port_metrics(df, opt)
        baseline_by_window[wname] = {"sharpe": m["sharpe"], "weights": opt}

df10 = build_df(V3_TICKERS, WINDOWS["10Y"])
lr10 = np.log(df10/df10.shift(1)).dropna()
w_arr = np.array([V3_NORM.get(t,0) for t in lr10.columns]); w_arr/=w_arr.sum()
port_rets10 = lr10.dot(w_arr)

loo_results = []
for drop_t in V3_TICKERS:
    remaining = [t for t in V3_TICKERS if t != drop_t]
    row = {"ticker":drop_t,"name":V3_NAMES[drop_t],"role":V3_ROLES[drop_t],
           "weight":V3_RAW[drop_t]}
    delta_sharpes = {}
    for wname, wstart in WINDOWS.items():
        df = build_df(remaining, wstart)
        if len(df) < 20: row[f"ds_{wname}"] = None; continue
        opt = optimize(df)
        if opt:
            m = port_metrics(df, opt)
            base_s = baseline_by_window.get(wname,{}).get("sharpe")
            delta  = round(m["sharpe"] - base_s, 3) if base_s else None
            row[f"ds_{wname}"] = delta
            delta_sharpes[wname] = delta
        else:
            row[f"ds_{wname}"] = None

    # Correlations
    if drop_t in lr10.columns and "xauusd" in lr10.columns:
        row["corr_gold"] = round(lr10[drop_t].corr(lr10["xauusd"]),3)
    else:
        row["corr_gold"] = None
    if drop_t in lr10.columns:
        row["corr_port"] = round(lr10[drop_t].corr(port_rets10),3)
    else:
        row["corr_port"] = None

    row["regime_wins"] = asset_regime.get(drop_t,{}).get("wins",0)

    # Verdict
    sharpe_pass = any(v is not None and v <= -0.03 for v in delta_sharpes.values())
    corr_pass   = row["corr_gold"] is not None and row["corr_gold"] < 0.75
    regime_pass = row["regime_wins"] >= 4
    role_count  = sum(1 for t in V3_TICKERS if t != drop_t and V3_ROLES.get(t)==V3_ROLES.get(drop_t))
    role_pass   = role_count >= 1
    passes = sum([sharpe_pass, corr_pass, regime_pass, role_pass])
    verdict = "KEEP" if passes >= 3 else ("REVIEW" if passes == 2 else "REJECT")
    row.update({"sharpe_test":"PASS" if sharpe_pass else "FAIL",
                "corr_test":"PASS" if corr_pass else "FAIL",
                "regime_test":"PASS" if regime_pass else "FAIL",
                "role_test":"PASS" if role_pass else "FAIL",
                "passes":passes, "verdict":verdict})
    loo_results.append(row)
    print(f"  {V3_NAMES[drop_t]:12s}: {verdict} ({passes}/4)  "
          f"dS 10Y={row.get('ds_10Y','?')}  dS 5Y={row.get('ds_5Y','?')}  dS 3Y={row.get('ds_3Y','?')}")

# ═══════════════════════════════════════════════════════════════════════════════
# STEP 6 — Stress Tests
# ═══════════════════════════════════════════════════════════════════════════════
print("\nStep 6: Stress tests...")
df10_full = build_df(V3_TICKERS, WINDOWS["10Y"])

# Get optimized weights per cap for 10Y
cap_weights_10y = {cr["cap"]: cr["opt_weights"]
                   for cr in cap_results if cr["window"]=="10Y"}

baseline_by_cap = {cap: port_metrics(df10_full, opt)
                   for cap, opt in cap_weights_10y.items() if opt}

stress_gold = []   # gold-only shock × cap matrix
for cap in GOLD_CAPS_T:
    opt = cap_weights_10y.get(cap)
    if not opt: continue
    base = baseline_by_cap[cap]
    for shock in SHOCK_LEVELS:
        df_s = df10_full.copy()
        df_s["xauusd"].iloc[1:] = df_s["xauusd"].iloc[1:] * (1 + shock)
        m = port_metrics(df_s, opt)
        stress_gold.append({"cap":cap,"shock":shock,
                             "sharpe":m["sharpe"],"ann_ret":m["ann_ret"],
                             "max_dd":m["max_dd"],"total":m["total"],
                             "base_sharpe":base["sharpe"],"base_total":base["total"]})

# Combined precious metals stress at 25% cap
opt_25  = cap_weights_10y.get(0.25)
base_25 = port_metrics(df10_full, opt_25) if opt_25 else {}
stress_combined = []
if opt_25:
    for shock in SHOCK_LEVELS:
        df_s = df10_full.copy()
        df_s["xauusd"].iloc[1:] = df_s["xauusd"].iloc[1:] * (1 + shock)
        df_s["xagusd"].iloc[1:] = df_s["xagusd"].iloc[1:] * (1 + shock)
        m = port_metrics(df_s, opt_25)
        stress_combined.append({"shock":shock,"sharpe":m["sharpe"],
                                 "ann_ret":m["ann_ret"],"max_dd":m["max_dd"],
                                 "total":m["total"],"base_sharpe":base_25["sharpe"],
                                 "base_total":base_25["total"]})
print("  Stress tests done")

# ═══════════════════════════════════════════════════════════════════════════════
# STEP 7 — Rolling 3Y Optimization
# ═══════════════════════════════════════════════════════════════════════════════
print("\nStep 7: Rolling 3Y optimization...")
WINDOW_DAYS = 756   # ~3Y trading days
STEP_DAYS   = 63    # ~quarterly

# Common date index for all V3 tickers
common_idx = None
for t in V3_TICKERS:
    s = price_data.get(t)
    if s is None: continue
    s_filtered = s[s.index <= IS_END]
    common_idx = s_filtered.index if common_idx is None else common_idx.intersection(s_filtered.index)
common_idx = common_idx.sort_values() if common_idx is not None else pd.DatetimeIndex([])

rolling_results = []
i = len(common_idx) - 1
end_indices = []
while i >= WINDOW_DAYS:
    end_indices.append(i); i -= STEP_DAYS
end_indices.reverse()

for ei in end_indices:
    end_date  = common_idx[ei]
    start_date= common_idx[ei - WINDOW_DAYS]
    df = pd.DataFrame({t: price_data[t] for t in V3_TICKERS if t in price_data})
    df = df[(df.index >= start_date) & (df.index <= end_date)].dropna()
    if len(df) < 200: continue
    opt = optimize(df, gold_cap=GOLD_CAP)
    if not opt: continue
    m  = port_metrics(df, opt)
    gold_ret = round((df["xauusd"].iloc[-1]/df["xauusd"].iloc[0]-1)*100,1) if "xauusd" in df else None
    row = {"end_date":end_date.strftime("%Y-%m-%d"),
           "start_date":start_date.strftime("%Y-%m-%d"),
           "n_assets":len(df.columns),
           "gold_ret_3y":gold_ret,
           "gold_price":round(df["xauusd"].iloc[-1],2) if "xauusd" in df else None,
           "sharpe":m["sharpe"],"ann_ret":m["ann_ret"],"max_dd":m["max_dd"]}
    for t in V3_TICKERS:
        row[f"w_{t}"] = round(opt.get(t,0)*100,1)
    rolling_results.append(row)

gold_ws = [r["w_xauusd"] for r in rolling_results if r.get("gold_ret_3y") is not None]
gold_rs = [r["gold_ret_3y"] for r in rolling_results if r.get("gold_ret_3y") is not None]
gold_mom_corr = round(np.corrcoef(gold_rs, gold_ws)[0,1],3) if len(gold_ws) > 2 else None
print(f"  {len(rolling_results)} windows | Gold-momentum corr: {gold_mom_corr}")

# ═══════════════════════════════════════════════════════════════════════════════
# STEP 8 — Out-of-Sample Validation
# ═══════════════════════════════════════════════════════════════════════════════
print("\nStep 8: OOS validation...")
OOS_PERIODS = [("2009-01-01","2016-03-31"),("2010-01-01","2016-03-31")]

oos_results = []
for oos_start, oos_end in OOS_PERIODS:
    avail = []
    for t in V3_TICKERS:
        s = price_data.get(t)
        if s is None: continue
        s_oos = s[(s.index >= oos_start) & (s.index <= oos_end)]
        if len(s_oos) >= 100: avail.append(t)
    excluded = [t for t in V3_TICKERS if t not in avail]
    if not avail: continue
    raw_sum  = sum(V3_RAW[t] for t in avail)
    oos_w    = {t: V3_RAW[t]/raw_sum for t in avail}
    df_oos   = build_df(avail, oos_start, oos_end)
    if len(df_oos) < 50: continue
    m = port_metrics(df_oos, oos_w)
    oos_results.append({"period":f"{oos_start[:4]}-{oos_end[:4]}",
                         "n_assets":len(avail),"excluded":excluded,
                         "weights":oos_w,"tickers":avail,**m})
    print(f"  OOS {oos_start[:4]}-{oos_end[:4]}: Sharpe {m['sharpe']}  "
          f"assets={len(avail)} excl={[V3_NAMES.get(t,t) for t in excluded]}")

# In-sample comparisons
is_full = port_metrics(build_df(V3_TICKERS, IS_START), V3_NORM)
is_full["period"] = "In-sample (all 8)"

oos_insamples = [is_full]
for oos_res in oos_results:
    df_is = build_df(oos_res["tickers"], IS_START)
    m_is  = port_metrics(df_is, oos_res["weights"])
    m_is["period"] = f"In-sample ({oos_res['n_assets']}-pos comparable)"
    oos_insamples.append(m_is)

# Year-by-year
yby_rows = []
if oos_results:
    best_oos = max(oos_results, key=lambda r: len(r["period"]))
    for year in range(int(best_oos["period"][:4]), 2027):
        ys = f"{year}-01-01"; ye = f"{year}-12-31"
        period = "OOS" if year < 2016 else "In-sample"
        ticks  = best_oos["tickers"] if year < 2016 else V3_TICKERS
        w_use  = best_oos["weights"] if year < 2016 else V3_NORM
        df_y   = build_df(ticks, ys, ye)
        if len(df_y) < 20: continue
        m_y = port_metrics(df_y, w_use)
        yby_rows.append({"year":year,"period":period,**m_y})

# ═══════════════════════════════════════════════════════════════════════════════
# STEP 9 — DCA Simulation
# ═══════════════════════════════════════════════════════════════════════════════
print("\nStep 9: DCA simulation...")

def run_dca(weights, tickers, start, end, initial, monthly):
    avail = [t for t in tickers if t in price_data]
    df = build_df(avail, start, end)
    if len(df) < 2: return pd.DataFrame()
    monthly_p = df.resample("ME").last()
    monthly_r = monthly_p.pct_change().dropna()
    ticks = list(monthly_r.columns)
    w = np.array([weights.get(t,0) for t in ticks], dtype=float)
    records = []
    value = initial; contributed = initial
    for date, row in monthly_r.iterrows():
        value    *= (1 + np.dot(w, row.values))
        value    += monthly
        contributed += monthly
        records.append({"date":date,"value":round(value),
                        "contributed":round(contributed),
                        "profit":round(value-contributed),
                        "return_pct":round((value/contributed-1)*100,1)})
    return pd.DataFrame(records).set_index("date")

sim_v2 = run_dca(V2_WEIGHTS, V2_TICKERS, DCA_START, IS_END, INITIAL_KR, MONTHLY_KR)
sim_v3 = run_dca(V3_RAW,    V3_TICKERS, DCA_START, IS_END, INITIAL_KR, MONTHLY_KR)

y_v2 = sim_v2.resample("YE").last() if len(sim_v2) > 0 else pd.DataFrame()
y_v3 = sim_v3.resample("YE").last() if len(sim_v3) > 0 else pd.DataFrame()

if len(sim_v2) > 0 and len(sim_v3) > 0:
    v2f = int(sim_v2.iloc[-1]["value"]); v3f = int(sim_v3.iloc[-1]["value"])
    cf  = int(sim_v2.iloc[-1]["contributed"])
    print(f"  v2 final: {v2f:,} kr (+{v2f-cf:,})  v3 final: {v3f:,} kr (+{v3f-cf:,})  diff: {v3f-v2f:+,}")
else:
    v2f = v3f = cf = 0

# ═══════════════════════════════════════════════════════════════════════════════
# STEP 9b — Regime-Aware DCA Comparison
# ═══════════════════════════════════════════════════════════════════════════════
print("\nStep 9b: Regime DCA comparison...")

def spread_regime(bps):
    if bps < 250:   return "SOP"
    if bps < 320:   return "CAUTION"
    if bps < 360:   return "T1"
    return "T2"

def run_dca_regime_v3(start, end, initial, monthly, strategy):
    """
    strategy:
      'A_simple'       — DCA every month regardless of spread
      'B_full_pause'   — pause all DCA during CAUTION; deploy reserve at T1/T2
      'C_hedge_cont'   — pause equity DCA during CAUTION; gold+silver always deploy
    """
    avail = [t for t in V3_TICKERS if t in price_data]
    df = build_df(avail, start, end)
    if len(df) < 2: return pd.DataFrame()

    monthly_p = df.resample("ME").last()
    monthly_r = monthly_p.pct_change().dropna()
    ticks = list(monthly_r.columns)
    w = np.array([V3_NORM.get(t, 0) for t in ticks], dtype=float)
    w /= w.sum()

    HEDGE = {"xauusd", "xagusd"}
    hedge_w_sum = sum(V3_NORM.get(t, 0) for t in ticks if t in HEDGE)
    hedge_w_sum = hedge_w_sum / sum(V3_NORM.values())  # normalised fraction

    records = []
    port_value = initial
    reserve    = 0.0
    contributed = initial

    for date, row in monthly_r.iterrows():
        port_value *= (1 + np.dot(w, row.values))

        # Current spread (last available on or before this month-end)
        if hy_ig_spread is not None:
            sl = hy_ig_spread[hy_ig_spread.index <= date]
            bps = float(sl.iloc[-1]) if len(sl) > 0 else 249.0
        else:
            bps = 249.0
        regime = spread_regime(bps)

        if strategy == "A_simple":
            deploy  = monthly
            reserve = 0.0
        elif strategy == "B_full_pause":
            if regime == "SOP":
                deploy  = monthly
            elif regime == "CAUTION":
                deploy  = 0.0
                reserve += monthly
            elif regime == "T1":
                deploy  = monthly + reserve * 0.50
                reserve = reserve * 0.50
            else:  # T2
                deploy  = monthly + reserve
                reserve = 0.0
        else:  # C_hedge_cont
            if regime == "SOP":
                deploy  = monthly
            elif regime == "CAUTION":
                deploy  = monthly * hedge_w_sum
                reserve += monthly * (1 - hedge_w_sum)
            elif regime == "T1":
                deploy  = monthly + reserve * 0.50
                reserve = reserve * 0.50
            else:  # T2
                deploy  = monthly + reserve
                reserve = 0.0

        port_value  += deploy
        contributed += monthly   # total earmarked regardless of strategy
        total_wealth = port_value + reserve

        records.append({
            "date": date, "regime": regime, "spread_bps": round(bps),
            "deployed": round(deploy), "reserve": round(reserve),
            "port_value": round(port_value), "total_wealth": round(total_wealth),
            "contributed": round(contributed),
            "profit": round(total_wealth - contributed),
            "return_pct": round((total_wealth / contributed - 1) * 100, 1),
        })

    return pd.DataFrame(records).set_index("date")

regime_start = SPREAD_START
sim_ra = run_dca_regime_v3(regime_start, IS_END, INITIAL_KR, RC_DCA_MONTHLY, "A_simple")
sim_rb = run_dca_regime_v3(regime_start, IS_END, INITIAL_KR, RC_DCA_MONTHLY, "B_full_pause")
sim_rc = run_dca_regime_v3(regime_start, IS_END, INITIAL_KR, RC_DCA_MONTHLY, "C_hedge_cont")

def sim_summary(sim, label):
    if len(sim) == 0: return
    last = sim.iloc[-1]
    print(f"  {label}: wealth {int(last['total_wealth']):,}  "
          f"port {int(last['port_value']):,}  reserve {int(last['reserve']):,}  "
          f"profit {int(last['profit']):,}  ret {last['return_pct']}%")

sim_summary(sim_ra, "A simple   ")
sim_summary(sim_rb, "B full pause")
sim_summary(sim_rc, "C hedge cont")

# Regime distribution
if len(sim_ra) > 0:
    dist = sim_ra["regime"].value_counts()
    total_m = len(sim_ra)
    print(f"  Regime months: " + "  ".join(f"{k} {v}/{total_m}" for k,v in dist.items()))

# ═══════════════════════════════════════════════════════════════════════════════
# STEP 9c — Price Action & Ratio Reserve Deployment
# ═══════════════════════════════════════════════════════════════════════════════
print("\nStep 9c: Price action reserve deployment...")

OPP_RESERVE    = 100_000   # fixed opportunistic reserve pool
RC_DCA_INITIAL = INITIAL_KR
# Fair baseline: pre-deploy the reserve immediately (no timing)
RC_DCA_INITIAL_PLUS = INITIAL_KR + OPP_RESERVE
PRIORITY_TICKERS = ["lly.us", "vrt.us", "avgo.us"]
DD_THRESH_50 = -0.20   # deploy 50% of reserve
DD_THRESH_ALL= -0.30   # deploy remaining reserve
RATIO_THRESH = 55.0    # silver entry signal

def run_dca_priceaction(start, end, initial, monthly, opp_reserve, strategy):
    """
    strategies:
      'A_base'    — no reserve; baseline (invest immediately)
      'A_plus'    — pre-deploy reserve on day 1 (fair benchmark for reserve strategies)
      'D_drawdown'— reserve deploys at -20%/-30% from 12M high on priority positions
      'E_ratio'   — reserve deploys when gold/silver ratio < 55
      'F_credit'  — reserve deploys at credit T1 (50%) / T2 (all)
      'G_combined'— D + E combined
    """
    avail = [t for t in V3_TICKERS if t in price_data]
    df = build_df(avail, start, end)
    if len(df) < 2: return pd.DataFrame()

    monthly_p = df.resample("ME").last()
    monthly_r = monthly_p.pct_change().dropna()
    ticks = list(monthly_r.columns)
    w = np.array([V3_NORM.get(t, 0) for t in ticks], dtype=float)
    w /= w.sum()

    port_value  = initial + (opp_reserve if strategy == "A_plus" else 0)
    reserve     = 0.0 if strategy in ("A_base","A_plus") else float(opp_reserve)
    total_cap   = initial + opp_reserve   # same committed capital for all
    contributed = total_cap               # committed on day 1

    # Deployment state flags (avoid re-triggering same episode)
    deployed_50 = False   # first tranche fired
    deployed_all= False   # second tranche fired
    cooldown    = 0       # months remaining before signal can re-fire

    records = []
    for i, (date, row) in enumerate(monthly_r.iterrows()):
        port_value *= (1 + np.dot(w, row.values))
        port_value += monthly

        # 12M rolling window for drawdown
        pos = monthly_p.index.get_loc(date)
        win_start = max(0, pos - 11)
        window = monthly_p.iloc[win_start:pos + 1]

        deploy = 0.0
        deploy_reason = ""
        if cooldown > 0: cooldown -= 1

        if reserve > 0:
            if strategy in ("D_drawdown", "G_combined") and cooldown == 0:
                # Check priority positions: worst drawdown from 12M high
                worst_dd = 0.0
                for pt in PRIORITY_TICKERS:
                    if pt in window.columns:
                        hi = window[pt].max()
                        curr = window[pt].iloc[-1]
                        dd = (curr - hi) / hi if hi > 0 else 0
                        worst_dd = min(worst_dd, dd)

                if not deployed_50 and worst_dd <= DD_THRESH_50:
                    deploy = reserve * 0.50
                    reserve -= deploy
                    deployed_50 = True
                    cooldown = 3
                    deploy_reason = f"DD {worst_dd*100:.1f}% (50%)"
                elif deployed_50 and not deployed_all and worst_dd <= DD_THRESH_ALL:
                    deploy = reserve
                    reserve = 0.0
                    deployed_all = True
                    cooldown = 3
                    deploy_reason = f"DD {worst_dd*100:.1f}% (all)"

                # Reset flags if position recovers above -10%
                if worst_dd > -0.10:
                    deployed_50 = False
                    deployed_all = False

            if strategy in ("E_ratio", "G_combined") and reserve > 0:
                if "xauusd" in window.columns and "xagusd" in window.columns:
                    ratio_now = window["xauusd"].iloc[-1] / window["xagusd"].iloc[-1]
                    if ratio_now < RATIO_THRESH:
                        deploy += reserve
                        reserve = 0.0
                        deploy_reason += f" Ratio {ratio_now:.1f}"

            if strategy == "F_credit" and reserve > 0:
                if hy_ig_spread is not None:
                    sl = hy_ig_spread[hy_ig_spread.index <= date]
                    bps = float(sl.iloc[-1]) if len(sl) > 0 else 249.0
                else:
                    bps = 249.0
                if bps >= 360 and not deployed_all:
                    deploy += reserve; reserve = 0.0; deployed_all = True
                    deploy_reason = f"Credit T2 {bps:.0f}bps"
                elif bps >= 320 and not deployed_50:
                    d = reserve * 0.50; deploy += d; reserve -= d; deployed_50 = True
                    deploy_reason = f"Credit T1 {bps:.0f}bps"

        port_value  += deploy
        total_wealth = port_value + reserve

        # Ratio for tracking
        ratio_track = None
        if "xauusd" in window.columns and "xagusd" in window.columns:
            ratio_track = round(window["xauusd"].iloc[-1] / window["xagusd"].iloc[-1], 1)

        records.append({
            "date": date,
            "port_value": round(port_value),
            "reserve": round(reserve),
            "total_wealth": round(total_wealth),
            "contributed": round(contributed),
            "profit": round(total_wealth - contributed),
            "return_pct": round((total_wealth / contributed - 1) * 100, 1),
            "deployed_opp": round(deploy),
            "deploy_reason": deploy_reason,
            "ratio": ratio_track,
        })

    return pd.DataFrame(records).set_index("date")

pa_start = SPREAD_START   # common start (limited by credit data for fair F comparison)

sim_pa_A    = run_dca_priceaction(pa_start, IS_END, RC_DCA_INITIAL, RC_DCA_MONTHLY, OPP_RESERVE, "A_base")
sim_pa_Aplus= run_dca_priceaction(pa_start, IS_END, RC_DCA_INITIAL, RC_DCA_MONTHLY, OPP_RESERVE, "A_plus")
sim_pa_D    = run_dca_priceaction(pa_start, IS_END, RC_DCA_INITIAL, RC_DCA_MONTHLY, OPP_RESERVE, "D_drawdown")
sim_pa_E    = run_dca_priceaction(pa_start, IS_END, RC_DCA_INITIAL, RC_DCA_MONTHLY, OPP_RESERVE, "E_ratio")
sim_pa_F    = run_dca_priceaction(pa_start, IS_END, RC_DCA_INITIAL, RC_DCA_MONTHLY, OPP_RESERVE, "F_credit")
sim_pa_G    = run_dca_priceaction(pa_start, IS_END, RC_DCA_INITIAL, RC_DCA_MONTHLY, OPP_RESERVE, "G_combined")

pa_strats = [
    ("A_base",     "No reserve — DCA only",                           sim_pa_A),
    ("A_plus",     "Pre-deploy reserve day 1 (fair benchmark)",        sim_pa_Aplus),
    ("D_drawdown", "-20%/-30% drawdown on Lilly/Vertiv/Broadcom",     sim_pa_D),
    ("E_ratio",    "Gold/silver ratio <55 (silver entry)",             sim_pa_E),
    ("F_credit",   "Credit T1/T2 (320/360 bps) — HY-IG spread",      sim_pa_F),
    ("G_combined", "Drawdown OR ratio <55",                            sim_pa_G),
]
for code, desc, sim in pa_strats:
    if len(sim) == 0: continue
    last = sim.iloc[-1]
    deploys = sim[sim["deployed_opp"] > 0]
    dep_str = f"{len(deploys)} events" if len(deploys) > 0 else "never fired"
    print(f"  {code:<12}: wealth {int(last['total_wealth']):>10,}  "
          f"profit {int(last['profit']):>9,}  ret {last['return_pct']:>6.1f}%  deploy: {dep_str}")

# ═══════════════════════════════════════════════════════════════════════════════
# EXCEL OUTPUT
# ═══════════════════════════════════════════════════════════════════════════════
print("\nBuilding Excel report...")
wb = openpyxl.Workbook()

# ──────────────────────────────────────────────────────────────────────────────
# SHEET 1: Summary
# ──────────────────────────────────────────────────────────────────────────────
ws1 = wb.active; ws1.title = "Summary"
r = title_row(ws1, 8, "Reactor Core v3 — Full Backtest",
              "Short Abundance, Long Scarcity | 8 positions + 2% cash | vs v2 baseline")

# Portfolio definition table
for ci, h in enumerate(["#","Name","Ticker","Role","Raw Weight","Norm Weight","v3 vs v2","Note"], 1):
    hdr(ws1, r, ci, h); ws1.row_dimensions[r].height = 24;
r += 1

v2_w_ref = {"xauusd":0.250,"xagusd":None,"lly.us":0.197,"wmt.us":0.227,
            "ccj.us":0.057,"vrt.us":0.090,"avgo.us":0.079,"jnj.us":0.050}
v3_notes = {"xauusd":"Unchanged","xagusd":"NEW — structural deficit thesis",
            "lly.us":"19.7%→15% (-4.7pp)","wmt.us":"22.7%→15% (-7.7pp)",
            "ccj.us":"5.7%→10% (+4.3pp)","vrt.us":"9.0%→10% (+1.0pp)",
            "avgo.us":"7.9%→8% (+0.1pp)","jnj.us":"Unchanged"}

for i, t in enumerate(V3_TICKERS, 1):
    bg = ALT if i%2==0 else WHT
    v2_ref = v2_w_ref.get(t)
    v2_str = f"{v2_ref*100:.1f}%" if v2_ref else "—"
    diff   = round((V3_RAW[t] - v2_ref)*100,1) if v2_ref else None
    diff_str = f"{diff:+.1f}pp" if diff is not None else "NEW"
    diff_bg  = GRN if diff and diff > 0 else (RED if diff and diff < 0 else YEL)
    cel(ws1, r, 1, i,                         bg=GREY, bold=True)
    cel(ws1, r, 2, V3_NAMES[t],               bg=bg, bold=True, ha="left")
    cel(ws1, r, 3, t,                          bg=bg)
    cel(ws1, r, 4, V3_ROLES[t],               bg=bg)
    cel(ws1, r, 5, f"{V3_RAW[t]*100:.1f}%",  bg=bg, bold=True)
    cel(ws1, r, 6, f"{V3_NORM[t]*100:.2f}%", bg=bg)
    cel(ws1, r, 7, diff_str,                   bg=diff_bg if diff is not None else YEL, bold=True)
    cel(ws1, r, 8, v3_notes.get(t,""),        bg=bg, ha="left")
    ws1.row_dimensions[r].height = 20; r += 1

# Cash row
cel(ws1, r, 1, 9,        bg=GREY, bold=True)
cel(ws1, r, 2, "Cash",   bg=GREY, bold=True, ha="left")
cel(ws1, r, 3, "—",      bg=GREY)
cel(ws1, r, 4, "—",      bg=GREY)
cel(ws1, r, 5, "2.0%",   bg=GREY, bold=True)
cel(ws1, r, 6, "—",      bg=GREY)
cel(ws1, r, 7, "—",      bg=GREY)
cel(ws1, r, 8, "Permanent cash floor — 55k kr, not deployed", bg=GREY, ha="left")
ws1.row_dimensions[r].height = 20; r += 2

# Headline metrics
for ci, h in enumerate(["Portfolio","Window","Type","Sharpe","Ann Ret","Ann Vol","Max DD","Calmar","Total Ret","Data From"], 1):
    hdr(ws1, r, ci, h); ws1.row_dimensions[r].height = 24
r += 1

v2_baselines = {"3Y":{"sharpe":2.684,"ann_ret":43.92,"max_dd":-28.75},
                "5Y":{"sharpe":2.033,"ann_ret":32.68,"max_dd":-22.74},
                "10Y":{"sharpe":1.851,"ann_ret":30.08,"max_dd":-24.86}}

for row in core_results:
    bg = ALT if V3_TICKERS[0] else WHT
    bg = GRN if row["portfolio"]=="v3" else LIGHT
    bold = row["type"]=="Optimized"
    v2b = v2_baselines.get(row["window"],{}) if row["portfolio"]=="v2" else {}
    s_bg = GRN if row["sharpe"]>=1.5 else (YEL if row["sharpe"]>=1.0 else ORG)
    cel(ws1, r, 1,  row["portfolio"].upper(),   bg=bg, bold=True)
    cel(ws1, r, 2,  row["window"],               bg=bg)
    cel(ws1, r, 3,  row["type"],                 bg=bg, ha="left")
    cel(ws1, r, 4,  row["sharpe"],               bg=s_bg, bold=bold)
    cel(ws1, r, 5,  f"{row['ann_ret']}%",        bg=bg, bold=bold)
    cel(ws1, r, 6,  f"{row['ann_vol']}%",        bg=bg)
    cel(ws1, r, 7,  f"{row['max_dd']}%",         bg=bg)
    cel(ws1, r, 8,  row["calmar"],               bg=bg)
    cel(ws1, r, 9,  f"+{row['total']}%",         bg=bg)
    cel(ws1, r, 10, row.get("start",""),          bg=bg)
    ws1.row_dimensions[r].height = 20; r += 1

cw(ws1, [4,18,12,12,12,12,10,10,10,12])

# ──────────────────────────────────────────────────────────────────────────────
# SHEET 2: Head-to-Head
# ──────────────────────────────────────────────────────────────────────────────
ws2 = wb.create_sheet("Head-to-Head")
r = title_row(ws2, 6, "v3 vs v2 — Head-to-Head Comparison",
              "v2 baselines from phase5_v2.py (verified). v3 computed fresh. Fixed weights used for DCA.")

def get_m(portfolio, window, mtype="Optimized"):
    return next((x for x in core_results
                 if x["portfolio"]==portfolio and x["window"]==window and x["type"]==mtype), {})

def get_oos(period_prefix):
    return next((r for r in oos_results if period_prefix in r["period"]), {})

h2h_rows = [
    # label, v2_val, v3_val, better_fn (returns 'v2','v3','~')
    ("10Y Sharpe (optimized)",
     get_m("v2","10Y"),"sharpe",   get_m("v3","10Y"),"sharpe"),
    ("10Y Sharpe (fixed weights)",
     get_m("v2","10Y","Fixed"),"sharpe", get_m("v3","10Y","Fixed"),"sharpe"),
    ("10Y Ann Return",
     get_m("v2","10Y"),"ann_ret",  get_m("v3","10Y"),"ann_ret"),
    ("10Y Ann Volatility",
     get_m("v2","10Y"),"ann_vol",  get_m("v3","10Y"),"ann_vol"),
    ("10Y Max Drawdown",
     get_m("v2","10Y"),"max_dd",   get_m("v3","10Y"),"max_dd"),
    ("10Y Calmar",
     get_m("v2","10Y"),"calmar",   get_m("v3","10Y"),"calmar"),
    ("10Y Total Return",
     get_m("v2","10Y"),"total",    get_m("v3","10Y"),"total"),
    ("5Y Sharpe (optimized)",
     get_m("v2","5Y"),"sharpe",    get_m("v3","5Y"),"sharpe"),
    ("5Y Ann Return",
     get_m("v2","5Y"),"ann_ret",   get_m("v3","5Y"),"ann_ret"),
    ("3Y Sharpe (optimized)",
     get_m("v2","3Y"),"sharpe",    get_m("v3","3Y"),"sharpe"),
    ("3Y Ann Return",
     get_m("v2","3Y"),"ann_ret",   get_m("v3","3Y"),"ann_ret"),
]

for ci, h in enumerate(["Metric","v2","v3","Delta","Better"], 1):
    hdr(ws2, r, ci, h); ws2.row_dimensions[r].height = 24
r += 1

for label, m2_dict, m2_key, m3_dict, m3_key in h2h_rows:
    v2_val = m2_dict.get(m2_key)
    v3_val = m3_dict.get(m3_key)
    if v2_val is None or v3_val is None:
        cel(ws2,r,1,label,bg=GREY,ha="left"); cel(ws2,r,2,"N/A"); cel(ws2,r,3,"N/A")
        cel(ws2,r,4,"N/A"); cel(ws2,r,5,"N/A"); ws2.row_dimensions[r].height=20; r+=1; continue
    delta = round(v3_val - v2_val, 3)
    # For max_dd and vol, lower is better for v3
    if m2_key in ("max_dd","ann_vol"):
        better = "v3" if delta < -0.5 else ("v2" if delta > 0.5 else "~")
    else:
        better = "v3" if delta > 0.02 else ("v2" if delta < -0.02 else "~")
    better_bg = GRN if better=="v3" else (RED if better=="v2" else YEL)
    suf = "%" if m2_key in ("ann_ret","ann_vol","max_dd","total") else ""
    cel(ws2,r,1,label,               bg=GREY,bold=True,ha="left")
    cel(ws2,r,2,f"{v2_val}{suf}",    bg=LIGHT)
    cel(ws2,r,3,f"{v3_val}{suf}",    bg=LIGHT)
    cel(ws2,r,4,f"{delta:+.3f}{suf}",bg=GRN if delta>0 else RED if delta<0 else YEL)
    cel(ws2,r,5,better,               bg=better_bg, bold=True)
    ws2.row_dimensions[r].height=20; r+=1

# OOS comparison
r += 1
for ci, h in enumerate(["Metric","v2","v3","Delta","Better"], 1): hdr(ws2,r,ci,h)
ws2.row_dimensions[r].height = 24; r += 1

v3_oos_10 = get_oos("2010")
oos_rows = [
    ("OOS Sharpe (2010-2016)", 0.955, v3_oos_10.get("sharpe")),
    ("OOS Ann Return (2010-2016)", None, v3_oos_10.get("ann_ret")),
    ("OOS Ann Vol (2010-2016)", None, v3_oos_10.get("ann_vol")),
    ("OOS Max DD (2010-2016)", None, v3_oos_10.get("max_dd")),
]
for label, v2_val, v3_val in oos_rows:
    v2_str = f"{v2_val}" if v2_val is not None else "—"
    v3_str = f"{v3_val}" if v3_val is not None else "N/A"
    if v2_val and v3_val:
        delta = round(v3_val-v2_val,3)
        better = "v3" if delta>0.02 else ("v2" if delta<-0.02 else "~")
    else:
        delta = None; better = "—"
    cel(ws2,r,1,label,                          bg=GREY,bold=True,ha="left")
    cel(ws2,r,2,v2_str,                         bg=LIGHT)
    cel(ws2,r,3,v3_str,                         bg=LIGHT)
    cel(ws2,r,4,f"{delta:+.3f}" if delta else "—", bg=GRN if delta and delta>0 else RED if delta and delta<0 else YEL)
    cel(ws2,r,5,better,                         bg=GRN if better=="v3" else RED if better=="v2" else YEL,bold=True)
    ws2.row_dimensions[r].height=20; r+=1

# DCA comparison
r += 1
dca_items = [
    ("DCA Final Value (SEK)", f"{int(sim_v2.iloc[-1]['value']):,}" if len(sim_v2)>0 else "—",
     f"{int(sim_v3.iloc[-1]['value']):,}" if len(sim_v3)>0 else "—"),
    ("DCA Total Profit (SEK)", f"{v2f-cf:,}" if v2f else "—", f"{v3f-cf:,}" if v3f else "—"),
    ("Precious Metals Weight", "25.0%", "35.0%"),
    ("Carry Weight", "52.4%", "35.0%"),
    ("Cyclical Weight", "5.7%", "10.0%"),
]
for label, v2_val, v3_val in dca_items:
    cel(ws2,r,1,label,bg=GREY,bold=True,ha="left")
    cel(ws2,r,2,v2_val,bg=LIGHT)
    cel(ws2,r,3,v3_val,bg=LIGHT)
    cel(ws2,r,4,"—",bg=YEL); cel(ws2,r,5,"—",bg=YEL)
    ws2.row_dimensions[r].height=20; r+=1

cw(ws2, [32,16,16,12,10])

# ──────────────────────────────────────────────────────────────────────────────
# SHEET 3: Core Metrics
# ──────────────────────────────────────────────────────────────────────────────
ws3 = wb.create_sheet("Core Metrics")
r = title_row(ws3, 10, "Core Performance Metrics — 3Y / 5Y / 10Y",
              "Row A = window-optimized weights | Row B = fixed v3 normalized weights | v2 baseline shown for reference")

for ci, h in enumerate(["Portfolio","Window","Type","Sharpe","Ann Ret","Ann Vol",
                         "Max DD","Calmar","Total Ret","Data From"], 1):
    hdr(ws3,r,ci,h); ws3.row_dimensions[r].height=24
r += 1

for row in core_results:
    is_v3 = row["portfolio"]=="v3"
    bg = ALT if is_v3 else LIGHT
    bold = row["type"]=="Optimized"
    s_bg = GRN if row["sharpe"]>=1.5 else (YEL if row["sharpe"]>=1.0 else ORG)
    cel(ws3,r,1,row["portfolio"].upper(), bg=bg,bold=True)
    cel(ws3,r,2,row["window"],            bg=bg)
    cel(ws3,r,3,row["type"],              bg=bg,ha="left")
    cel(ws3,r,4,row["sharpe"],            bg=s_bg,bold=bold)
    cel(ws3,r,5,f"{row['ann_ret']}%",     bg=bg,bold=bold)
    cel(ws3,r,6,f"{row['ann_vol']}%",     bg=bg)
    cel(ws3,r,7,f"{row['max_dd']}%",      bg=bg)
    cel(ws3,r,8,row["calmar"],            bg=bg)
    cel(ws3,r,9,f"+{row['total']}%",      bg=GRN,bold=bold)
    cel(ws3,r,10,row.get("start",""),     bg=bg)
    ws3.row_dimensions[r].height=20; r+=1

cw(ws3, [8,8,18,10,12,12,12,10,12,13])

# ──────────────────────────────────────────────────────────────────────────────
# SHEET 4: Asset Metrics
# ──────────────────────────────────────────────────────────────────────────────
ws4 = wb.create_sheet("Asset Metrics")
r = title_row(ws4, 12, "Individual Asset Metrics — 3Y / 5Y / 10Y",
              "Standalone performance (100% allocation). Price returns only.")

for ci, h in enumerate(["Name","Ticker","Role","Weight",
                         "Ret 10Y","Sharpe 10Y","DD 10Y",
                         "Ret 5Y","Sharpe 5Y","DD 5Y",
                         "Ret 3Y","Sharpe 3Y"], 1):
    hdr(ws4,r,ci,h); ws4.row_dimensions[r].height=24
r += 1

for i, row in enumerate(asset_rows, 1):
    bg = ALT if i%2==0 else WHT
    cel(ws4,r,1,row["name"],  bg=bg,ha="left",bold=True)
    cel(ws4,r,2,row["ticker"],bg=bg)
    cel(ws4,r,3,row["role"],  bg=bg)
    cel(ws4,r,4,f"{row['weight']*100:.1f}%", bg=bg)
    col = 5
    for wname in ["10Y","5Y","3Y"]:
        for key in ["ann_ret","sharpe","max_dd"]:
            if wname=="3Y" and key=="max_dd": continue  # only 2 cols for 3Y
            v = row.get(f"{key}_{wname}","N/A")
            if v == "N/A": cbg = GREY
            elif key=="max_dd": cbg = RED if v<-50 else (ORG if v<-30 else (YEL if v<-15 else GRN))
            elif key=="ann_ret": cbg = GRN if v>20 else (YEL if v>10 else (ORG if v>0 else RED))
            else: cbg = GRN if v>0.9 else (YEL if v>0.5 else (ORG if v>0 else RED))
            suf = "%" if key in ("ann_ret","max_dd") else ""
            cel(ws4,r,col,f"{v}{suf}" if v!="N/A" else "N/A", bg=cbg)
            col += 1
    ws4.row_dimensions[r].height=22; r+=1

cw(ws4,[18,12,12,9,10,10,10,10,10,10,10,10])

# ──────────────────────────────────────────────────────────────────────────────
# SHEET 5: Regime Analysis
# ──────────────────────────────────────────────────────────────────────────────
ws5 = wb.create_sheet("Regime Analysis")
rnames = list(REGIMES.keys())
r = title_row(ws5, 3+len(rnames)+2,
              "Asset Returns by Regime — v3 vs v2",
              "Regimes from instruction spec. N/A = asset had no data in that regime.")

for ci, h in enumerate(["Name","Ticker","Wins"]+rnames+["Avg Ret","Role"], 1):
    hdr(ws5,r,ci,h); ws5.row_dimensions[r].height=36
r += 1

for i, t in enumerate(V3_TICKERS, 1):
    bg  = ALT if i%2==0 else WHT
    ar  = asset_regime.get(t,{})
    wins = ar.get("wins",0)
    w_bg = GRN if wins>=5 else (YEL if wins>=4 else (ORG if wins>=3 else RED))
    cel(ws5,r,1,V3_NAMES[t],  bg=bg,ha="left",bold=True)
    cel(ws5,r,2,t,             bg=bg)
    cel(ws5,r,3,f"{wins}/6",   bg=w_bg,bold=True)
    for j, rname in enumerate(rnames,4):
        v = ar.get(rname)
        cbg = GREY if v is None else (GRN if v>20 else (YEL if v>0 else (ORG if v>-20 else RED)))
        cel(ws5,r,j,f"{v:+.1f}%" if v is not None else "N/A", bg=cbg)
    cel(ws5,r,len(rnames)+4,f"{ar.get('avg',0):+.1f}%" if ar.get('avg') else "N/A", bg=bg)
    cel(ws5,r,len(rnames)+5,V3_ROLES[t], bg=bg)
    ws5.row_dimensions[r].height=22; r+=1

# Portfolio rows (v3 and v2)
r += 1
for label, regime_list, weights in [("v3 Portfolio",regime_v3,V3_NORM),
                                      ("v2 Portfolio",regime_v2,V2_WEIGHTS)]:
    bg = ALT
    cel(ws5,r,1,label,bg=MID,bold=True,color=WHT,ha="left")
    cel(ws5,r,2,"",bg=MID); cel(ws5,r,3,"",bg=MID)
    for j, ritem in enumerate(regime_list,4):
        v = ritem.get("port_ret")
        cbg = GREY if v is None else (GRN if v>20 else (YEL if v>0 else (ORG if v>-20 else RED)))
        cel(ws5,r,j,f"{v:+.1f}%" if v is not None else "N/A", bg=cbg,bold=True)
    cel(ws5,r,len(rnames)+4,"",bg=MID); cel(ws5,r,len(rnames)+5,"",bg=MID)
    ws5.row_dimensions[r].height=22; r+=1

    # vs S&P row
    cel(ws5,r,1,f"{label} vs S&P",bg=GREY,bold=True,ha="left")
    cel(ws5,r,2,"",bg=GREY); cel(ws5,r,3,"",bg=GREY)
    for j, ritem in enumerate(regime_list,4):
        v = ritem.get("vs_spx")
        cbg = GRN if v and v>0 else (RED if v and v<-5 else YEL)
        cel(ws5,r,j,f"{v:+.1f}%" if v is not None else "N/A", bg=cbg,bold=True)
    cel(ws5,r,len(rnames)+4,"",bg=GREY); cel(ws5,r,len(rnames)+5,"",bg=GREY)
    ws5.row_dimensions[r].height=20; r+=1

cw(ws5,[18,12,7]+[13]*len(rnames)+[10,12])

# ──────────────────────────────────────────────────────────────────────────────
# SHEET 6: Gold Cap Sensitivity
# ──────────────────────────────────────────────────────────────────────────────
ws6 = wb.create_sheet("Gold Cap Sensitivity")
r = title_row(ws6, 16, "Gold Cap Sensitivity — 3Y / 5Y / 10Y",
              "80 random restarts per optimization. Silver uncapped (5%-40% bounds).")

for ci, h in enumerate(["Cap",
                         "3Y Gold%","3Y Ag%","3Y Sharpe","3Y AnnRet","3Y MaxDD",
                         "5Y Gold%","5Y Ag%","5Y Sharpe","5Y AnnRet","5Y MaxDD",
                         "10Y Gold%","10Y Ag%","10Y Sharpe","10Y AnnRet","10Y MaxDD"], 1):
    hdr(ws6,r,ci,h); ws6.row_dimensions[r].height=24
r += 1

for cap in GOLD_CAPS_T:
    cap_label = "Uncapped" if cap>=0.40 else f"{int(cap*100)}%"
    bg = YEL if cap==GOLD_CAP else WHT
    cel(ws6,r,1,cap_label,bg=GREY,bold=(cap==GOLD_CAP))
    col = 2
    for wname in ["3Y","5Y","10Y"]:
        cr = next((x for x in cap_results if x["cap"]==cap and x["window"]==wname),{})
        s_bg = GRN if cr.get("sharpe",0)>=1.5 else (YEL if cr.get("sharpe",0)>=1.0 else ORG)
        cel(ws6,r,col,  f"{cr.get('gold_w','?')}%",     bg=bg)
        cel(ws6,r,col+1,f"{cr.get('silver_w','?')}%",   bg=bg)
        cel(ws6,r,col+2,cr.get("sharpe","?"),            bg=s_bg,bold=(cap==GOLD_CAP))
        cel(ws6,r,col+3,f"{cr.get('ann_ret','?')}%",    bg=bg)
        cel(ws6,r,col+4,f"{cr.get('max_dd','?')}%",     bg=bg)
        col += 5
    ws6.row_dimensions[r].height=22; r+=1

cw(ws6,[10]+[8,7,9,10,10]*3)

# ──────────────────────────────────────────────────────────────────────────────
# SHEET 7: Floor Position Audit
# ──────────────────────────────────────────────────────────────────────────────
ws7 = wb.create_sheet("Floor Position Audit")
r = title_row(ws7, 14, "Leave-One-Out Audit — v3",
              "ΔSharpe = Sharpe(7-pos optimized) − Sharpe(8-pos baseline). KEEP if ≥3/4 criteria pass.")

# Baseline note
base_10 = baseline_by_window.get("10Y",{})
ws7.merge_cells(f"A{r}:N{r}")
c_b = ws7.cell(row=r,column=1,
    value=f"Baseline 10Y Sharpe (all 8 positions): {base_10.get('sharpe','?')} | "
          f"Constraints: 5% min / 40% max / 25% gold cap / 80 restarts")
c_b.font=Font(bold=True,size=10); c_b.fill=PatternFill("solid",fgColor=YEL)
c_b.alignment=Alignment(horizontal="left",vertical="center",wrap_text=True)
ws7.row_dimensions[r].height=24; r+=1

for ci, h in enumerate(["Name","Ticker","Role","Weight",
                         "ΔS 3Y","ΔS 5Y","ΔS 10Y",
                         "Corr Gold","Corr Port","Regime Wins",
                         "Sharpe","Corr","Regime","Role","Verdict"], 1):
    hdr(ws7,r,ci,h); ws7.row_dimensions[r].height=28
r += 1

for i, row in enumerate(loo_results, 1):
    bg = ALT if i%2==0 else WHT
    v_bg = GRN if row["verdict"]=="KEEP" else (RED if row["verdict"]=="REJECT" else YEL)
    def ds_bg(v): return GRN if v is not None and v<=-0.03 else (RED if v is not None and v>0 else YEL)
    cel(ws7,r,1, row["name"],        bg=bg,ha="left",bold=True)
    cel(ws7,r,2, row["ticker"],      bg=bg)
    cel(ws7,r,3, row["role"],        bg=bg)
    cel(ws7,r,4, f"{row['weight']*100:.1f}%", bg=bg)
    cel(ws7,r,5, row.get("ds_3Y","?"),  bg=ds_bg(row.get("ds_3Y")))
    cel(ws7,r,6, row.get("ds_5Y","?"),  bg=ds_bg(row.get("ds_5Y")))
    cel(ws7,r,7, row.get("ds_10Y","?"), bg=ds_bg(row.get("ds_10Y")))
    cel(ws7,r,8, row.get("corr_gold","?"), bg=GRN if row.get("corr_gold") and row["corr_gold"]<0.75 else RED)
    cel(ws7,r,9, row.get("corr_port","?"), bg=bg)
    wbg = GRN if row["regime_wins"]>=5 else (YEL if row["regime_wins"]>=4 else ORG)
    cel(ws7,r,10,f"{row['regime_wins']}/6", bg=wbg,bold=True)
    for col_off, key in enumerate(["sharpe_test","corr_test","regime_test","role_test"],11):
        v = row.get(key,"?")
        cel(ws7,r,col_off,v, bg=GRN if v=="PASS" else RED,bold=True)
    cel(ws7,r,15,row["verdict"], bg=v_bg,bold=True)
    ws7.row_dimensions[r].height=22; r+=1

cw(ws7,[14,10,11,8,8,8,8,10,10,11,8,8,8,8,9])

# ──────────────────────────────────────────────────────────────────────────────
# SHEET 8: Gold Stress Test
# ──────────────────────────────────────────────────────────────────────────────
ws8 = wb.create_sheet("Gold Stress Test")
r = title_row(ws8, 8, "Gold Stress Test — Permanent Level Shock",
              "Gold-only: gold price multiplied by (1+shock%). Combined: both gold AND silver shocked. Weights = cap-optimized 10Y.")

# Gold-only: Sharpe matrix
ws8.merge_cells(f"A{r}:H{r}")
c = ws8.cell(row=r,column=1,value="GOLD-ONLY STRESS — Sharpe by Cap × Shock Level")
c.font=Font(bold=True,color=WHT); c.fill=PatternFill("solid",fgColor=MID)
c.alignment=Alignment(horizontal="center"); ws8.row_dimensions[r].height=22; r+=1

shock_labels = [f"{int(s*100)}%" for s in SHOCK_LEVELS]
for ci, h in enumerate(["Gold Cap"]+shock_labels+["Unshocked"], 1): hdr(ws8,r,ci,h)
ws8.row_dimensions[r].height=24; r+=1

for cap in GOLD_CAPS_T:
    cap_label = "Uncapped" if cap>=0.40 else f"{int(cap*100)}%"
    bg = YEL if cap==GOLD_CAP else WHT
    cel(ws8,r,1,cap_label,bg=GREY,bold=(cap==GOLD_CAP))
    for col_off, shock in enumerate(SHOCK_LEVELS,2):
        sr = next((x for x in stress_gold if x["cap"]==cap and x["shock"]==shock),{})
        v = sr.get("sharpe","?")
        s_bg = GRN if isinstance(v,float) and v>=1.5 else (YEL if isinstance(v,float) and v>=1.0 else (ORG if isinstance(v,float) and v>=0.5 else RED))
        cel(ws8,r,col_off,v, bg=s_bg,bold=(cap==GOLD_CAP))
    base_v = baseline_by_cap.get(cap,{}).get("sharpe","?")
    cel(ws8,r,7,base_v, bg=GRN if isinstance(base_v,float) and base_v>=1.5 else YEL,bold=(cap==GOLD_CAP))
    ws8.row_dimensions[r].height=22; r+=1

# Gold-only: Total Return matrix
r += 1
ws8.merge_cells(f"A{r}:H{r}")
c = ws8.cell(row=r,column=1,value="GOLD-ONLY STRESS — Total Return by Cap × Shock Level")
c.font=Font(bold=True,color=WHT); c.fill=PatternFill("solid",fgColor=MID)
c.alignment=Alignment(horizontal="center"); ws8.row_dimensions[r].height=22; r+=1
for ci, h in enumerate(["Gold Cap"]+shock_labels+["Unshocked"], 1): hdr(ws8,r,ci,h)
ws8.row_dimensions[r].height=24; r+=1

for cap in GOLD_CAPS_T:
    cap_label = "Uncapped" if cap>=0.40 else f"{int(cap*100)}%"
    bg = YEL if cap==GOLD_CAP else WHT
    cel(ws8,r,1,cap_label,bg=GREY,bold=(cap==GOLD_CAP))
    for col_off, shock in enumerate(SHOCK_LEVELS,2):
        sr = next((x for x in stress_gold if x["cap"]==cap and x["shock"]==shock),{})
        v = sr.get("total","?")
        t_bg = GRN if isinstance(v,float) and v>100 else (YEL if isinstance(v,float) and v>0 else RED)
        cel(ws8,r,col_off,f"{v}%" if v!="?" else "?", bg=t_bg,bold=(cap==GOLD_CAP))
    base_v = baseline_by_cap.get(cap,{}).get("total","?")
    cel(ws8,r,7,f"{base_v}%" if base_v!="?" else "?", bg=GRN,bold=(cap==GOLD_CAP))
    ws8.row_dimensions[r].height=22; r+=1

# Combined precious metals stress
r += 2
ws8.merge_cells(f"A{r}:H{r}")
c = ws8.cell(row=r,column=1,
    value="COMBINED STRESS (Gold + Silver both shocked) — 25% Gold Cap | v3 adds 10% Silver = 35% total exposure")
c.font=Font(bold=True,color=WHT); c.fill=PatternFill("solid",fgColor=DARK)
c.alignment=Alignment(horizontal="center"); ws8.row_dimensions[r].height=22; r+=1
for ci, h in enumerate(["Shock Level","Sharpe","Ann Ret","Max DD","Total Ret",
                         "vs Gold-Only Sharpe","vs Unshocked Sharpe",""], 1):
    hdr(ws8,r,ci,h); ws8.row_dimensions[r].height=24
r += 1

for cs in stress_combined:
    shock_pct = f"{int(cs['shock']*100)}%"
    gold_only = next((x for x in stress_gold if x["cap"]==0.25 and x["shock"]==cs["shock"]),{})
    diff_vs_gold = round(cs["sharpe"]-gold_only.get("sharpe",cs["sharpe"]),3) if gold_only else "?"
    diff_vs_base = round(cs["sharpe"]-cs["base_sharpe"],3)
    bg = ORG
    cel(ws8,r,1,shock_pct,     bg=GREY,bold=True)
    cel(ws8,r,2,cs["sharpe"],  bg=GRN if cs["sharpe"]>=1.0 else (YEL if cs["sharpe"]>=0.5 else RED),bold=True)
    cel(ws8,r,3,f"{cs['ann_ret']}%", bg=bg)
    cel(ws8,r,4,f"{cs['max_dd']}%",  bg=bg)
    cel(ws8,r,5,f"{cs['total']}%",   bg=GRN if cs["total"]>0 else RED)
    cel(ws8,r,6,diff_vs_gold if diff_vs_gold!="?" else "?",
        bg=GRN if isinstance(diff_vs_gold,float) and diff_vs_gold>0 else RED)
    cel(ws8,r,7,f"{diff_vs_base:+.3f}", bg=GRN if diff_vs_base>0 else RED)
    cel(ws8,r,8,"", bg=WHT)
    ws8.row_dimensions[r].height=22; r+=1

cw(ws8,[12,10,10,10,11,18,18,6])

# ──────────────────────────────────────────────────────────────────────────────
# SHEET 9: Rolling Optimization
# ──────────────────────────────────────────────────────────────────────────────
ws9 = wb.create_sheet("Rolling Gold Weight")
r = title_row(ws9, len(V3_TICKERS)+5,
              "Rolling 3Y Optimization — Quarterly Steps",
              f"Gold cap 25% | 80 restarts | Gold-momentum correlation: {gold_mom_corr}")

cols = ["Window End","Start","Assets","Gold Ret 3Y","Gold Price","Sharpe","AnnRet","MaxDD"]
cols += [f"w_{V3_NAMES[t]}" for t in V3_TICKERS]
for ci, h in enumerate(cols, 1): hdr(ws9,r,ci,h); ws9.row_dimensions[r].height=28
r += 1

for i, row in enumerate(rolling_results, 1):
    bg = ALT if i%2==0 else WHT
    gold_ret = row.get("gold_ret_3y")
    g_bg = GRN if gold_ret and gold_ret>20 else (YEL if gold_ret and gold_ret>0 else RED)
    s_bg = GRN if row["sharpe"]>=1.5 else (YEL if row["sharpe"]>=1.0 else ORG)
    cel(ws9,r,1,row["end_date"],   bg=GREY,bold=True)
    cel(ws9,r,2,row["start_date"], bg=bg)
    cel(ws9,r,3,row["n_assets"],   bg=bg)
    cel(ws9,r,4,f"{gold_ret:+.1f}%" if gold_ret else "N/A", bg=g_bg)
    cel(ws9,r,5,row.get("gold_price","?"), bg=bg)
    cel(ws9,r,6,row["sharpe"],     bg=s_bg,bold=True)
    cel(ws9,r,7,f"{row['ann_ret']}%", bg=bg)
    cel(ws9,r,8,f"{row['max_dd']}%",  bg=bg)
    for ci_off, t in enumerate(V3_TICKERS, 9):
        w_val = row.get(f"w_{t}",0)
        w_bg = GRN if w_val>20 else (YEL if w_val>10 else (GREY if w_val>5 else WHT))
        cel(ws9,r,ci_off,f"{w_val}%", bg=w_bg)
    ws9.row_dimensions[r].height=20; r+=1

cw(ws9,[12,12,7,11,10,9,10,10]+[10]*len(V3_TICKERS))

# ──────────────────────────────────────────────────────────────────────────────
# SHEET 10: Out-of-Sample
# ──────────────────────────────────────────────────────────────────────────────
ws10 = wb.create_sheet("Out-of-Sample")
r = title_row(ws10, 9, "Out-of-Sample Validation",
              "Fixed normalized weights applied to OOS data. Vertiv excluded (IPO Feb 2020). "
              "Silver excluded if no OOS data. Weights re-normalized to available assets.")

# Summary table
for ci, h in enumerate(["Period","Assets","Sharpe","Ann Ret","Ann Vol","Max DD","Total Ret","Data From","Excluded"], 1):
    hdr(ws10,r,ci,h); ws10.row_dimensions[r].height=24
r += 1

for m in oos_insamples:
    bg = LIGHT
    s_bg = GRN if m["sharpe"]>=1.5 else (YEL if m["sharpe"]>=1.0 else ORG)
    cel(ws10,r,1,m["period"],              bg=bg,ha="left",bold=True)
    cel(ws10,r,2,len(V3_TICKERS),          bg=bg)
    cel(ws10,r,3,m["sharpe"],              bg=s_bg,bold=True)
    cel(ws10,r,4,f"{m['ann_ret']}%",       bg=bg)
    cel(ws10,r,5,f"{m['ann_vol']}%",       bg=bg)
    cel(ws10,r,6,f"{m['max_dd']}%",        bg=bg)
    cel(ws10,r,7,f"+{m['total']}%",        bg=GRN)
    cel(ws10,r,8,m.get("start",""),        bg=bg)
    cel(ws10,r,9,"—",                      bg=bg)
    ws10.row_dimensions[r].height=22; r+=1

for m in oos_results:
    s_bg = GRN if m["sharpe"]>=1.0 else (YEL if m["sharpe"]>=0.5 else RED)
    excl_str = ", ".join(V3_NAMES.get(t,t) for t in m["excluded"])
    cel(ws10,r,1,f"OOS {m['period']}",    bg=YEL,ha="left",bold=True)
    cel(ws10,r,2,m["n_assets"],            bg=YEL)
    cel(ws10,r,3,m["sharpe"],              bg=s_bg,bold=True)
    cel(ws10,r,4,f"{m['ann_ret']}%",       bg=YEL)
    cel(ws10,r,5,f"{m['ann_vol']}%",       bg=YEL)
    cel(ws10,r,6,f"{m['max_dd']}%",        bg=YEL)
    cel(ws10,r,7,f"+{m['total']}%",        bg=GRN if m["total"]>0 else RED)
    cel(ws10,r,8,m.get("start",""),        bg=YEL)
    cel(ws10,r,9,excl_str,                 bg=YEL,ha="left")
    ws10.row_dimensions[r].height=22; r+=1

# OOS interpretation thresholds
r += 1
ws10.merge_cells(f"A{r}:I{r}")
c = ws10.cell(row=r,column=1,
    value="OOS Interpretation: Sharpe >1.0 = strong structural alpha | 0.5-1.0 = partial alpha (some curve-fitting) | <0.5 = mostly curve-fitted")
c.font=Font(bold=True,size=10); c.fill=PatternFill("solid",fgColor=YEL)
c.alignment=Alignment(horizontal="left",vertical="center"); ws10.row_dimensions[r].height=22; r+=2

# Year-by-year table
for ci, h in enumerate(["Year","Period","Sharpe","Ann Ret","Ann Vol","Max DD","Total Ret","Data From","Data To"], 1):
    hdr(ws10,r,ci,h); ws10.row_dimensions[r].height=24
r += 1

for row in sorted(yby_rows, key=lambda x: x["year"]):
    is_oos = row["period"]=="OOS"
    bg = YEL if is_oos else (ALT if row["year"]%2==0 else WHT)
    s_bg = GRN if row["sharpe"]>=1.0 else (YEL if row["sharpe"]>=0.5 else RED)
    cel(ws10,r,1,row["year"],             bg=GREY,bold=True)
    cel(ws10,r,2,row["period"],           bg=bg)
    cel(ws10,r,3,row["sharpe"],           bg=s_bg,bold=is_oos)
    cel(ws10,r,4,f"{row['ann_ret']}%",    bg=bg)
    cel(ws10,r,5,f"{row['ann_vol']}%",    bg=bg)
    cel(ws10,r,6,f"{row['max_dd']}%",     bg=bg)
    cel(ws10,r,7,f"+{row['total']}%",     bg=GRN if row["total"]>0 else RED)
    cel(ws10,r,8,row.get("start",""),     bg=bg)
    cel(ws10,r,9,row.get("end",""),       bg=bg)
    ws10.row_dimensions[r].height=20; r+=1

cw(ws10,[7,22,9,10,10,10,11,12,12])

# Weight normalization table
r += 2
ws10.merge_cells(f"A{r}:I{r}")
c = ws10.cell(row=r,column=1,value="OOS Weight Normalization (longest OOS period)")
c.font=Font(bold=True,color=WHT); c.fill=PatternFill("solid",fgColor=MID)
c.alignment=Alignment(horizontal="center"); ws10.row_dimensions[r].height=22; r+=1

if oos_results:
    best_oos = max(oos_results, key=lambda x: x["n_assets"])
    for ci, h in enumerate(["Asset","Ticker","v3 Raw Weight","OOS Norm Weight","Change","Status"], 1):
        hdr(ws10,r,ci,h,bg=GREY,fg="000000"); ws10.row_dimensions[r].height=22
    r += 1
    for t in V3_TICKERS:
        raw_w = V3_RAW[t]
        oos_w = best_oos["weights"].get(t)
        excl  = t in best_oos["excluded"]
        status = "EXCLUDED" if excl else "INCLUDED"
        s_bg   = RED if excl else GRN
        change = round((oos_w - raw_w)*100,1) if oos_w else None
        cel(ws10,r,1,V3_NAMES[t],         bg=WHT,ha="left")
        cel(ws10,r,2,t,                    bg=WHT)
        cel(ws10,r,3,f"{raw_w*100:.1f}%", bg=WHT)
        cel(ws10,r,4,f"{oos_w*100:.2f}%" if oos_w else "—", bg=WHT)
        cel(ws10,r,5,f"{change:+.1f}pp" if change else "—", bg=WHT)
        cel(ws10,r,6,status,               bg=s_bg,bold=True)
        ws10.row_dimensions[r].height=20; r+=1

# ──────────────────────────────────────────────────────────────────────────────
# SHEET 11: DCA Comparison
# ──────────────────────────────────────────────────────────────────────────────
ws11 = wb.create_sheet("DCA Comparison")
r = title_row(ws11, 9, "DCA Simulation — v2 vs v3",
              f"1,000,000 kr initial + 6,000 kr/month | Start: {DCA_START} | v3: 2% cash held out (V3_RAW weights sum=0.98)")

for ci, h in enumerate(["Year","Contributed","v2 Value","v2 Profit","v2 Ret%",
                         "v3 Value","v3 Profit","v3 Ret%","v3-v2 Diff"], 1):
    hdr(ws11,r,ci,h); ws11.row_dimensions[r].height=26
r += 1

years_union = y_v2.index.union(y_v3.index)
for dt in years_union:
    if dt not in y_v2.index or dt not in y_v3.index: continue
    r2 = y_v2.loc[dt]; r3 = y_v3.loc[dt]
    diff = int(r3["value"]) - int(r2["value"])
    diff_bg = GRN if diff >= 0 else RED
    bg = ALT if r%2==0 else WHT
    cel(ws11,r,1,dt.year,           bg=GREY,bold=True)
    cel(ws11,r,2,int(r2["contributed"]), bg=bg)
    cel(ws11,r,3,int(r2["value"]),  bg=bg,bold=True)
    cel(ws11,r,4,int(r2["profit"]), bg=GRN if r2["profit"]>=0 else RED)
    cel(ws11,r,5,f"{r2['return_pct']}%", bg=bg)
    cel(ws11,r,6,int(r3["value"]),  bg=bg,bold=True)
    cel(ws11,r,7,int(r3["profit"]), bg=GRN if r3["profit"]>=0 else RED)
    cel(ws11,r,8,f"{r3['return_pct']}%", bg=bg)
    cel(ws11,r,9,f"{diff:+,}",      bg=diff_bg,bold=True)
    ws11.row_dimensions[r].height=22; r+=1

# Summary
r += 1
ws11.merge_cells(f"A{r}:I{r}")
if len(sim_v2)>0 and len(sim_v3)>0:
    c = ws11.cell(row=r,column=1,
        value=(f"Total contributed: {cf:,} kr  |  "
               f"v2 final: {v2f:,} kr (+{v2f-cf:,} kr)  |  "
               f"v3 final: {v3f:,} kr (+{v3f-cf:,} kr)  |  "
               f"v3 vs v2: {v3f-v2f:+,} kr"))
    c.font=Font(bold=True,size=10)
    c.fill=PatternFill("solid",fgColor=GRN if v3f>=v2f else RED)
    c.alignment=Alignment(horizontal="left",vertical="center",wrap_text=True)
    ws11.row_dimensions[r].height=28; r+=2

# Monthly detail
for ci, h in enumerate(["Month","Contributed","v2 Value","v2 Profit",
                         "v3 Value","v3 Profit","Diff"], 1):
    hdr(ws11,r,ci,h); ws11.row_dimensions[r].height=22
r += 1

for dt in sim_v2.index:
    if dt not in sim_v3.index: continue
    r2 = sim_v2.loc[dt]; r3 = sim_v3.loc[dt]
    diff = int(r3["value"]) - int(r2["value"])
    bg = ALT if r%2==0 else WHT
    cel(ws11,r,1,str(dt.date()),         bg=GREY)
    cel(ws11,r,2,int(r2["contributed"]), bg=bg)
    cel(ws11,r,3,int(r2["value"]),       bg=bg,bold=True)
    cel(ws11,r,4,int(r2["profit"]),      bg=GRN if r2["profit"]>=0 else RED)
    cel(ws11,r,5,int(r3["value"]),       bg=bg,bold=True)
    cel(ws11,r,6,int(r3["profit"]),      bg=GRN if r3["profit"]>=0 else RED)
    cel(ws11,r,7,diff,                   bg=GRN if diff>=0 else RED)
    ws11.row_dimensions[r].height=16; r+=1

cw(ws11,[12,14,14,14,14,14,14,12,14])

# ──────────────────────────────────────────────────────────────────────────────
# SHEET 12: Regime DCA Comparison
# ──────────────────────────────────────────────────────────────────────────────
ws12 = wb.create_sheet("Regime DCA")
r = title_row(ws12, 11, "Regime-Aware DCA — Strategy Comparison",
              (f"Base: 1,000,000 kr initial + {RC_DCA_MONTHLY:,} kr/month (actual RC DCA)  |  "
               f"Spread data from {SPREAD_START}  |  Triggers: SOP <250 | CAUTION 250-319 | T1 320-359 | T2 ≥360 bps  |  "
               f"A=Simple always | B=Full pause equity+hedges | C=Hedges continue (gold+silver), equity pauses"))

REGIME_COLORS = {"SOP": GRN, "CAUTION": YEL, "T1": ORG, "T2": RED}

# ── Summary table ─────────────────────────────────────────────────────────────
for ci, h in enumerate(["Strategy","Description",
                         "Final Total Wealth","Final Port Value","Final Reserve",
                         "Total Contributed","Total Profit","Return %",
                         "vs Strategy A","Months paused",""], 1):
    hdr(ws12, r, ci, h)
ws12.row_dimensions[r].height = 36; r += 1

strategy_rows = [
    ("A — Simple DCA",     "Deploy 25k/month regardless of spread",            sim_ra),
    ("B — Full Pause",     "Pause all DCA during CAUTION; deploy reserve at T1/T2", sim_rb),
    ("C — Hedge Continue", "Pause equity during CAUTION; gold+silver always DCA",   sim_rc),
]
ra_wealth = int(sim_ra.iloc[-1]["total_wealth"]) if len(sim_ra) > 0 else 0

for strat_name, desc, sim in strategy_rows:
    if len(sim) == 0:
        cel(ws12, r, 1, strat_name, bg=GREY, bold=True); r += 1; continue
    last = sim.iloc[-1]
    tw   = int(last["total_wealth"])
    diff = tw - ra_wealth
    diff_bg = GRN if diff >= 0 else RED
    paused = int((sim["deployed"] == 0).sum())
    bg = WHT

    cel(ws12, r, 1,  strat_name,              bg=GREY, bold=True, ha="left")
    cel(ws12, r, 2,  desc,                    bg=bg, ha="left")
    cel(ws12, r, 3,  f"{tw:,} kr",           bg=bg, bold=True)
    cel(ws12, r, 4,  f"{int(last['port_value']):,} kr", bg=bg)
    cel(ws12, r, 5,  f"{int(last['reserve']):,} kr",    bg=bg)
    cel(ws12, r, 6,  f"{int(last['contributed']):,} kr",bg=bg)
    cel(ws12, r, 7,  f"{int(last['profit']):,} kr",     bg=GRN if last["profit"]>=0 else RED, bold=True)
    cel(ws12, r, 8,  f"{last['return_pct']}%",          bg=bg, bold=True)
    cel(ws12, r, 9,  f"{diff:+,} kr",        bg=diff_bg, bold=True)
    cel(ws12, r, 10, f"{paused} months",      bg=YEL if paused > 0 else GRN)
    cel(ws12, r, 11, "",                      bg=bg)
    ws12.row_dimensions[r].height = 24; r += 1

# Regime distribution
r += 1
if len(sim_ra) > 0:
    dist = sim_ra["regime"].value_counts()
    total_m = len(sim_ra)
    ws12.merge_cells(f"A{r}:K{r}")
    dist_str = "  |  ".join(f"{k}: {v} months ({v/total_m*100:.0f}%)" for k, v in
                            sorted(dist.items(), key=lambda x: ["SOP","CAUTION","T1","T2"].index(x[0])))
    c = ws12.cell(row=r, column=1,
                  value=f"Regime distribution ({SPREAD_START} → {IS_END}):  {dist_str}")
    c.font = Font(bold=True, size=9); c.fill = PatternFill("solid", fgColor=LIGHT)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws12.row_dimensions[r].height = 22; r += 2

# ── Monthly detail ────────────────────────────────────────────────────────────
for ci, h in enumerate(["Month","Spread\nbps","Regime",
                         "A Wealth","A Port","A Ret%",
                         "B Wealth","B Port","B Reserve","B Ret%",
                         "C Wealth","C Port","C Reserve","C Ret%",
                         "B vs A","C vs A"], 1):
    hdr(ws12, r, ci, h)
ws12.row_dimensions[r].height = 30; r += 1

months_union = sim_ra.index if len(sim_ra) > 0 else []
for date in months_union:
    if date not in sim_rb.index or date not in sim_rc.index: continue
    ra = sim_ra.loc[date]; rb = sim_rb.loc[date]; rc_ = sim_rc.loc[date]
    regime  = ra["regime"]
    reg_bg  = REGIME_COLORS.get(regime, WHT)
    bg = ALT if r % 2 == 0 else WHT
    diff_b = int(rb["total_wealth"]) - int(ra["total_wealth"])
    diff_c = int(rc_["total_wealth"]) - int(ra["total_wealth"])

    cel(ws12, r, 1,  str(date.date()),          bg=GREY)
    cel(ws12, r, 2,  ra["spread_bps"],          bg=reg_bg)
    cel(ws12, r, 3,  regime,                    bg=reg_bg, bold=True)
    cel(ws12, r, 4,  int(ra["total_wealth"]),   bg=bg)
    cel(ws12, r, 5,  int(ra["port_value"]),     bg=bg)
    cel(ws12, r, 6,  f"{ra['return_pct']}%",   bg=bg)
    cel(ws12, r, 7,  int(rb["total_wealth"]),   bg=bg)
    cel(ws12, r, 8,  int(rb["port_value"]),     bg=bg)
    cel(ws12, r, 9,  int(rb["reserve"]),        bg=YEL if rb["reserve"]>0 else bg)
    cel(ws12, r, 10, f"{rb['return_pct']}%",   bg=bg)
    cel(ws12, r, 11, int(rc_["total_wealth"]),  bg=bg)
    cel(ws12, r, 12, int(rc_["port_value"]),    bg=bg)
    cel(ws12, r, 13, int(rc_["reserve"]),       bg=YEL if rc_["reserve"]>0 else bg)
    cel(ws12, r, 14, f"{rc_['return_pct']}%",  bg=bg)
    cel(ws12, r, 15, f"{diff_b:+,}",           bg=GRN if diff_b>=0 else RED, bold=True)
    cel(ws12, r, 16, f"{diff_c:+,}",           bg=GRN if diff_c>=0 else RED, bold=True)
    ws12.row_dimensions[r].height = 16; r += 1

cw(ws12, [12,9,10,14,14,9,14,14,14,9,14,14,14,9,14,14])

# ──────────────────────────────────────────────────────────────────────────────
# SHEET 13: Price Action Reserve Deployment
# ──────────────────────────────────────────────────────────────────────────────
ws13 = wb.create_sheet("Price Action Reserve")
r = title_row(ws13, 10, "Reserve Deployment — Price Action vs Credit vs Ratio",
              (f"All strategies: {RC_DCA_INITIAL:,} kr initial + {RC_DCA_MONTHLY:,} kr/month + "
               f"{OPP_RESERVE:,} kr opportunistic reserve  |  Start: {pa_start}  |  "
               f"A_plus = reserve pre-deployed day 1 (fair benchmark)  |  "
               f"D = deploy at -20%/-30% from 12M high (Lilly/Vertiv/Broadcom)  |  "
               f"E = deploy when gold/silver ratio <{RATIO_THRESH:.0f}  |  "
               f"F = deploy at credit T1/T2 (320/360 bps)  |  G = D+E combined"))

# ── Summary ───────────────────────────────────────────────────────────────────
sum_hdrs = ["Strategy","Description","Final Wealth","Final Port","Reserve Left",
            "Profit","Return %","vs A_plus","Deployments",""]
for ci, h in enumerate(sum_hdrs, 1): hdr(ws13, r, ci, h)
ws13.row_dimensions[r].height = 30; r += 1

aplus_wealth = int(sim_pa_Aplus.iloc[-1]["total_wealth"]) if len(sim_pa_Aplus) > 0 else 0

for code, desc, sim in pa_strats:
    if len(sim) == 0: continue
    last   = sim.iloc[-1]
    tw     = int(last["total_wealth"])
    diff   = tw - aplus_wealth
    diff_bg = GRN if diff > 0 else (YEL if diff == 0 else RED)
    deploys = sim[sim["deployed_opp"] > 0]
    dep_dates = ", ".join(str(d.date()) for d in deploys.index[:3])
    if len(deploys) > 3: dep_dates += f" +{len(deploys)-3} more"
    dep_str = dep_dates if len(deploys) > 0 else "never fired"
    is_bench = code in ("A_base","A_plus")
    bg = GREY if is_bench else WHT

    cel(ws13, r, 1,  code,                            bg=GREY if is_bench else LIGHT, bold=True)
    cel(ws13, r, 2,  desc,                            bg=bg, ha="left")
    cel(ws13, r, 3,  f"{tw:,} kr",                  bg=bg, bold=True)
    cel(ws13, r, 4,  f"{int(last['port_value']):,} kr", bg=bg)
    cel(ws13, r, 5,  f"{int(last['reserve']):,} kr", bg=YEL if last["reserve"]>0 else bg)
    cel(ws13, r, 6,  f"{int(last['profit']):,} kr",  bg=GRN if last["profit"]>=0 else RED, bold=True)
    cel(ws13, r, 7,  f"{last['return_pct']}%",       bg=bg, bold=True)
    cel(ws13, r, 8,  f"{diff:+,} kr" if not is_bench else "benchmark",
                                                      bg=diff_bg if not is_bench else LIGHT, bold=True)
    cel(ws13, r, 9,  dep_str,                         bg=bg, ha="left")
    cel(ws13, r, 10, "",                              bg=bg)
    ws13.row_dimensions[r].height = 24; r += 1

# Key finding note
r += 1
ws13.merge_cells(f"A{r}:J{r}")
ratio_min = min((sim_pa_A["ratio"].min() for sim in [sim_pa_A] if len(sim)>0 and "ratio" in sim.columns), default=99)
c = ws13.cell(row=r, column=1,
    value=(f"Gold/silver ratio: min {sim_pa_A['ratio'].min():.1f} in period — ratio <{RATIO_THRESH:.0f} never triggered.  "
           f"Ratio signal is forward-looking (thesis confirmation), not a historical timing tool.  "
           f"Drawdown signal fires when Lilly/Vertiv/Broadcom hit -20%/-30% from 12M high."))
c.font = Font(italic=True, size=9, color="444444")
c.fill = PatternFill("solid", fgColor=LIGHT)
c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws13.row_dimensions[r].height = 32; r += 2

# ── Monthly detail ────────────────────────────────────────────────────────────
det_hdrs = ["Month","Ratio","A_base\nWealth","A_plus\nWealth",
            "D Wealth","D Reserve","D Deploy","D Reason",
            "F Wealth","F Reserve","F Deploy"]
for ci, h in enumerate(det_hdrs, 1): hdr(ws13, r, ci, h)
ws13.row_dimensions[r].height = 30; r += 1

months_ref = sim_pa_A.index if len(sim_pa_A) > 0 else []
for date in months_ref:
    def get(sim, col):
        try: return sim.loc[date, col]
        except: return None

    ratio_v = get(sim_pa_A, "ratio")
    ratio_bg = GRN if ratio_v and ratio_v < 65 else (YEL if ratio_v and ratio_v < 75 else WHT)
    bg = ALT if r % 2 == 0 else WHT

    d_dep = get(sim_pa_D, "deployed_opp") or 0
    f_dep = get(sim_pa_F, "deployed_opp") or 0
    d_rsn = get(sim_pa_D, "deploy_reason") or ""

    cel(ws13, r, 1,  str(date.date()),                               bg=GREY)
    cel(ws13, r, 2,  ratio_v,                                        bg=ratio_bg)
    cel(ws13, r, 3,  int(get(sim_pa_A,"total_wealth") or 0),         bg=bg)
    cel(ws13, r, 4,  int(get(sim_pa_Aplus,"total_wealth") or 0),     bg=bg)
    cel(ws13, r, 5,  int(get(sim_pa_D,"total_wealth") or 0),         bg=bg)
    cel(ws13, r, 6,  int(get(sim_pa_D,"reserve") or 0),              bg=YEL if (get(sim_pa_D,"reserve") or 0)>0 else bg)
    cel(ws13, r, 7,  int(d_dep),                                     bg=GRN if d_dep > 0 else bg, bold=d_dep>0)
    cel(ws13, r, 8,  d_rsn,                                          bg=GRN if d_rsn else bg, ha="left")
    cel(ws13, r, 9,  int(get(sim_pa_F,"total_wealth") or 0),         bg=bg)
    cel(ws13, r, 10, int(get(sim_pa_F,"reserve") or 0),              bg=YEL if (get(sim_pa_F,"reserve") or 0)>0 else bg)
    cel(ws13, r, 11, int(f_dep),                                     bg=GRN if f_dep > 0 else bg, bold=f_dep>0)
    ws13.row_dimensions[r].height = 16; r += 1

cw(ws13, [12,9,14,14,14,14,12,20,14,14,12])

wb.save(REPORT)
print(f"\nReport saved: {REPORT}")
print("=" * 60)
