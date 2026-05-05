import pandas as pd
import numpy as np
import os, sys, io, warnings
from scipy.optimize import minimize
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
warnings.filterwarnings("ignore")
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

DATA_DIR = r"C:\Users\lukas.larsson\Desktop\Privat\Project Reactor Core\data"
REPORT   = r"C:\Users\lukas.larsson\Desktop\Privat\Project Reactor Core\Black_Litterman.xlsx"

BASE_TICKERS = ["xauusd","lly.us","wmt.us","avgo.us","cost.us","ccj.us","vrt.us","jnj.us"]
CANDIDATES   = ["xagusd","copx.us"]
ALL_TICKERS  = BASE_TICKERS + CANDIDATES
NAMES = {
    "xauusd":"Gold","lly.us":"Eli Lilly","wmt.us":"Walmart",
    "avgo.us":"Broadcom","cost.us":"Costco","ccj.us":"Cameco",
    "vrt.us":"Vertiv","jnj.us":"J&J","xagusd":"Silver","copx.us":"COPX",
}
ROLES = {
    "xauusd":"Hedge","lly.us":"Carry","wmt.us":"Carry","avgo.us":"Convexity",
    "cost.us":"Carry","ccj.us":"Cyclical","vrt.us":"Convexity","jnj.us":"Carry",
    "xagusd":"Hedge","copx.us":"Cyclical",
}

# Reactor Core v2 weights (8 positions)
V2_WEIGHTS = {
    "xauusd":0.250,"wmt.us":0.227,"lly.us":0.197,"vrt.us":0.090,
    "avgo.us":0.079,"ccj.us":0.057,"jnj.us":0.050,"cost.us":0.050,
}

GOLD_CAP = 0.25
MIN_W    = 0.05
MAX_W    = 0.40
TAU      = 0.05   # uncertainty in the prior
DELTA    = 2.5    # risk aversion coefficient
START    = "2016-04-01"

# ── Views: (ticker, annual_return, confidence) ────────────────────────────────
VIEWS_INPUT = [
    ("xauusd",  0.18, 0.65),   # Gold:       +18%, 65% confidence
    ("xagusd",  0.30, 0.55),   # Silver:     +30%, 55%
    ("copx.us", 0.22, 0.45),   # COPX:       +22%, 45%
    ("vrt.us",  0.40, 0.60),   # Vertiv:     +40%, 60%
    ("lly.us",  0.22, 0.60),   # Eli Lilly:  +22%, 60%
]

# ── Load data ─────────────────────────────────────────────────────────────────
def load(ticker):
    fname = ticker.replace("^","").replace("-","_") + ".csv"
    path  = os.path.join(DATA_DIR, fname)
    if not os.path.exists(path): return None
    df = pd.read_csv(path)
    df["Date"] = pd.to_datetime(df["Date"])
    df = df.set_index("Date").sort_index()
    return df["Close"].replace(0, np.nan).dropna()

df_all = pd.DataFrame({t: load(t) for t in ALL_TICKERS if load(t) is not None})
df_all = df_all[df_all.index >= START].dropna()
tickers = list(df_all.columns)
n = len(tickers)

rets     = df_all.pct_change().dropna()
mu_hist  = rets.mean() * 252          # annualized historical mean
cov_ann  = rets.cov() * 252           # annualized covariance
sigma    = cov_ann.values

# ── Equilibrium weights: V2 weights for base, floor for candidates ────────────
w_eq_raw = np.array([V2_WEIGHTS.get(t, MIN_W) for t in tickers])
w_eq     = w_eq_raw / w_eq_raw.sum()  # normalise to 1

# Implied equilibrium excess returns (CAPM reverse-optimisation)
pi = DELTA * sigma @ w_eq             # annualised

# ── Build P, Q, Omega ─────────────────────────────────────────────────────────
valid_views = [(t, r, c) for t, r, c in VIEWS_INPUT if t in tickers]
k = len(valid_views)
P           = np.zeros((k, n))
Q           = np.zeros(k)
omega_diag  = np.zeros(k)

for i, (ticker, ret, conf) in enumerate(valid_views):
    idx          = tickers.index(ticker)
    P[i, idx]    = 1.0
    Q[i]         = ret
    # Idzorek: Omega_ii = tau * (p_i Sigma p_i^T) * (1-conf)/conf
    p_i          = P[i]
    omega_diag[i] = TAU * float(p_i @ sigma @ p_i) * (1 - conf) / conf

Omega     = np.diag(omega_diag)
Omega_inv = np.diag(1.0 / omega_diag)

# ── BL posterior ──────────────────────────────────────────────────────────────
tau_sigma_inv = np.linalg.inv(TAU * sigma)
M             = np.linalg.inv(tau_sigma_inv + P.T @ Omega_inv @ P)
mu_bl         = M @ (tau_sigma_inv @ pi + P.T @ Omega_inv @ Q)
sigma_bl      = sigma + M        # posterior covariance

# ── Optimise using BL posterior returns ───────────────────────────────────────
bounds = [(MIN_W, GOLD_CAP if t == "xauusd" else MAX_W) for t in tickers]
constraints = [{"type": "eq", "fun": lambda w: np.sum(w) - 1}]

def neg_sharpe(w, mu, cov):
    r = np.dot(w, mu)
    v = np.sqrt(w @ cov @ w)
    return -r / v if v > 0 else 0

def optimise(mu, cov, n_starts=100):
    best = None
    for _ in range(n_starts):
        w0  = np.random.dirichlet(np.ones(n))
        w0  = np.clip(w0, [b[0] for b in bounds], [b[1] for b in bounds])
        w0 /= w0.sum()
        res = minimize(neg_sharpe, w0, args=(mu, cov), method="SLSQP",
                       bounds=bounds, constraints=constraints,
                       options={"maxiter":1000,"ftol":1e-12})
        if res.success and (best is None or res.fun < best.fun):
            best = res
    return best.x if best else None

w_bl = optimise(mu_bl, sigma_bl)
w_mv = optimise(mu_hist.values, sigma)   # plain MV on 10 assets for comparison

w_bl_dict = dict(zip(tickers, w_bl))
w_mv_dict = dict(zip(tickers, w_mv))

# ── Portfolio metrics ─────────────────────────────────────────────────────────
def port_metrics(df, w_dict):
    w    = np.array([w_dict.get(t, 0) for t in df.columns])
    r_   = df.pct_change().dropna()
    mu_  = r_.mean() * 252
    cov_ = r_.cov() * 252
    ar   = np.dot(w, mu_)
    av   = np.sqrt(w @ cov_.values @ w)
    sh   = ar / av if av > 0 else 0
    port = (df / df.iloc[0]).dot(w)
    mdd  = ((port - port.cummax()) / port.cummax()).min()
    tot  = port.iloc[-1] - 1
    return {"sharpe": round(sh,3), "ann_ret": round(ar*100,2),
            "ann_vol": round(av*100,2), "max_dd": round(mdd*100,2),
            "total":   round(tot*100,1)}

m_v2 = port_metrics(df_all[BASE_TICKERS], V2_WEIGHTS)
m_mv = port_metrics(df_all, w_mv_dict)
m_bl = port_metrics(df_all, w_bl_dict)

# ── Console output ────────────────────────────────────────────────────────────
print("=" * 70)
print("  Black-Litterman — Reactor Core v2 + Silver + COPX")
print(f"  τ={TAU}  δ={DELTA}  Start={START}")
print("=" * 70)

print(f"\n  {'Asset':<14} {'Equil Ret':>10}  {'BL Post Ret':>12}  {'Δ':>7}  "
      f"{'v2 w':>7}  {'MV w':>7}  {'BL w':>7}")
for i, t in enumerate(tickers):
    dr = (mu_bl[i] - pi[i]) * 100
    v2w = V2_WEIGHTS.get(t, 0) * 100
    mvw = w_mv_dict[t] * 100
    blw = w_bl_dict[t] * 100
    marker = " <-- NEW" if t in CANDIDATES else ""
    print(f"  {NAMES.get(t,t):<14} {pi[i]*100:>8.1f}%  {mu_bl[i]*100:>10.1f}%  "
          f"{dr:>+6.1f}pp  {v2w:>5.1f}%  {mvw:>5.1f}%  {blw:>5.1f}%{marker}")

print(f"\n  {'Metric':<12} {'v2 (8-pos)':>12}  {'MV (10-pos)':>13}  {'BL (10-pos)':>13}")
for k_, label in [("sharpe","Sharpe"),("ann_ret","AnnRet"),
                   ("ann_vol","AnnVol"),("max_dd","MaxDD"),("total","Total")]:
    print(f"  {label:<12} {str(m_v2[k_]):>12}  {str(m_mv[k_]):>13}  {str(m_bl[k_]):>13}")

# ── Excel output ──────────────────────────────────────────────────────────────
DARK="1F4E79"; MID="2E75B6"; LIGHT="D6E4F0"; ALT="EBF3FB"
GRN="C6EFCE"; RED="FFC7CE"; YEL="FFEB9C"; WHT="FFFFFF"; GREY="F2F2F2"; ORG="FFD966"
thin   = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr(ws, row, col, val, bg=DARK, fg=WHT, bold=True, sz=10, ha="center"):
    c = ws.cell(row=row, column=col, value=val)
    c.fill = PatternFill("solid", fgColor=bg)
    c.font = Font(bold=bold, color=fg, size=sz)
    c.alignment = Alignment(wrap_text=True, vertical="center", horizontal=ha)
    c.border = BORDER; return c

def cel(ws, row, col, val, bg=None, bold=False, ha="center", color="000000"):
    c = ws.cell(row=row, column=col, value=val)
    if bg: c.fill = PatternFill("solid", fgColor=bg)
    c.font = Font(bold=bold, color=color)
    c.alignment = Alignment(wrap_text=True, vertical="center", horizontal=ha)
    c.border = BORDER; return c

def title_row(ws, row, nc, text, bg=DARK, sz=13):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=nc)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(bold=True, color=WHT, size=sz)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 34

def cw(ws, widths):
    for i, w_ in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w_

wb = openpyxl.Workbook()

# ━━ SHEET 1: BL Summary ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ws1 = wb.active; ws1.title = "BL Summary"
title_row(ws1, 1, 9, "Black-Litterman Optimization — Reactor Core v2 + Silver + COPX")

ws1.merge_cells("A2:I2")
c2 = ws1["A2"]
c2.value = (f"Equilibrium: Reactor Core v2 weights (prior).  "
            f"Views blended via Idzorek method.  τ={TAU}  δ={DELTA}  "
            f"Gold cap {GOLD_CAP*100:.0f}%  |  Start: {START}")
c2.font = Font(italic=True, size=9, color="444444")
c2.fill = PatternFill("solid", fgColor=LIGHT)
c2.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws1.row_dimensions[2].height = 22

# Views block
title_row(ws1, 3, 9, "Investor Views", bg=MID, sz=11)
for ci, h in enumerate(["Asset","Role","View Return","Confidence",
                         "Equil Return","BL Posterior","Δ Return","View Impact",""], 1):
    hdr(ws1, 4, ci, h)
ws1.row_dimensions[4].height = 28

ri = 5
for ticker, ret, conf in valid_views:
    idx    = tickers.index(ticker)
    pi_r   = pi[idx] * 100
    bl_r   = mu_bl[idx] * 100
    delta  = bl_r - pi_r
    impact = "Pulled up" if delta > 0.5 else ("Pulled down" if delta < -0.5 else "Neutral")
    imp_bg = GRN if delta > 0.5 else (RED if delta < -0.5 else YEL)
    bg     = ALT if ri % 2 == 0 else WHT
    is_cand = ticker in CANDIDATES
    row_bg  = "FFF2CC" if is_cand else bg
    cel(ws1, ri, 1, NAMES.get(ticker, ticker), bg=row_bg, bold=True, ha="left")
    cel(ws1, ri, 2, ROLES.get(ticker,""),      bg=row_bg)
    cel(ws1, ri, 3, f"{ret*100:.0f}%",         bg=row_bg, bold=True)
    cel(ws1, ri, 4, f"{conf*100:.0f}%",        bg=row_bg)
    cel(ws1, ri, 5, f"{pi_r:.1f}%",            bg=row_bg)
    cel(ws1, ri, 6, f"{bl_r:.1f}%",            bg=row_bg)
    cel(ws1, ri, 7, f"{delta:+.1f}pp",         bg=GRN if delta>0 else RED, bold=True)
    cel(ws1, ri, 8, impact,                     bg=imp_bg, bold=True)
    cel(ws1, ri, 9, "NEW" if is_cand else "",   bg="FFD700" if is_cand else row_bg, bold=is_cand)
    ws1.row_dimensions[ri].height = 22; ri += 1

# Non-viewed assets
ri += 1
title_row(ws1, ri, 9, "Non-Viewed Assets (held at equilibrium)", bg=GREY, sz=10)
ws1.row_dimensions[ri].height = 24; ri += 1
for ci, h in enumerate(["Asset","Role","","","Equil Return","BL Posterior","Δ Return","",""], 1):
    hdr(ws1, ri, ci, h, bg=GREY, fg="000000")
ws1.row_dimensions[ri].height = 24; ri += 1

viewed_tickers = [t for t, _, _ in valid_views]
for i, t in enumerate(tickers):
    if t in viewed_tickers: continue
    pi_r  = pi[i] * 100
    bl_r  = mu_bl[i] * 100
    delta = bl_r - pi_r
    bg    = ALT if i % 2 == 0 else WHT
    cel(ws1, ri, 1, NAMES.get(t, t), bg=bg, bold=True, ha="left")
    cel(ws1, ri, 2, ROLES.get(t,""), bg=bg)
    cel(ws1, ri, 3, "",              bg=bg)
    cel(ws1, ri, 4, "",              bg=bg)
    cel(ws1, ri, 5, f"{pi_r:.1f}%", bg=bg)
    cel(ws1, ri, 6, f"{bl_r:.1f}%", bg=bg)
    cel(ws1, ri, 7, f"{delta:+.1f}pp", bg=GRN if abs(delta)>0.3 and delta>0 else (RED if abs(delta)>0.3 and delta<0 else bg))
    ws1.row_dimensions[ri].height = 20; ri += 1

cw(ws1, [14, 12, 13, 13, 14, 14, 12, 13, 8])

# ━━ SHEET 2: Weight Comparison ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ws2 = wb.create_sheet("Weight Comparison")
title_row(ws2, 1, 10, "Portfolio Weights — v2 (8-pos) vs MV (10-pos) vs BL (10-pos)")

ws2.merge_cells("A2:J2")
c2b = ws2["A2"]
c2b.value = ("v2 = current Reactor Core v2 (8 positions, historical optimizer)  |  "
             "MV = plain mean-variance on 10 assets (no views)  |  "
             "BL = Black-Litterman with investor views  |  "
             "Yellow highlight = new candidate positions")
c2b.font = Font(italic=True, size=9, color="444444")
c2b.fill = PatternFill("solid", fgColor=LIGHT)
c2b.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws2.row_dimensions[2].height = 28

for ci, h in enumerate(["#","Asset","Role","v2 Weight","MV Weight","BL Weight",
                          "BL vs v2","BL vs MV","Status",""], 1):
    hdr(ws2, 3, ci, h)
ws2.row_dimensions[3].height = 28

sorted_bl = sorted(w_bl_dict.items(), key=lambda x: x[1], reverse=True)
for i, (t, blw) in enumerate(sorted_bl, 1):
    bg      = ALT if i % 2 == 0 else WHT
    is_cand = t in CANDIDATES
    row_bg  = "FFF2CC" if is_cand else bg
    v2w     = V2_WEIGHTS.get(t, 0) * 100
    mvw     = w_mv_dict.get(t, 0) * 100
    blw_pct = blw * 100
    d_v2    = blw_pct - v2w
    d_mv    = blw_pct - mvw
    status  = "NEW" if is_cand else ("↑" if d_v2 > 0.5 else ("↓" if d_v2 < -0.5 else "≈"))
    cel(ws2, i+3, 1,  i,                         bg=row_bg)
    cel(ws2, i+3, 2,  NAMES.get(t,t),            bg=row_bg, bold=True, ha="left")
    cel(ws2, i+3, 3,  ROLES.get(t,""),           bg=row_bg)
    cel(ws2, i+3, 4,  f"{v2w:.1f}%" if v2w else "—", bg=row_bg)
    cel(ws2, i+3, 5,  f"{mvw:.1f}%",             bg=row_bg)
    cel(ws2, i+3, 6,  f"{blw_pct:.1f}%",         bg=row_bg, bold=True)
    cel(ws2, i+3, 7,  f"{d_v2:+.1f}pp",          bg=GRN if d_v2>0.5 else (RED if d_v2<-0.5 else bg), bold=True)
    cel(ws2, i+3, 8,  f"{d_mv:+.1f}pp",          bg=GRN if d_mv>0.5 else (RED if d_mv<-0.5 else bg))
    cel(ws2, i+3, 9,  status,                     bg="FFD700" if is_cand else (GRN if status=="↑" else (RED if status=="↓" else bg)), bold=True)
    ws2.row_dimensions[i+3].height = 22

# Metrics comparison
ri_m = n + 6
title_row(ws2, ri_m, 10, "Portfolio Metrics Comparison", bg=MID, sz=11)
ws2.row_dimensions[ri_m].height = 26; ri_m += 1
for ci, h in enumerate(["Metric","v2 (8-pos)","MV (10-pos)","BL (10-pos)",
                          "BL vs v2","BL vs MV","","","",""], 1):
    hdr(ws2, ri_m, ci, h, bg=GREY, fg="000000")
ws2.row_dimensions[ri_m].height = 24; ri_m += 1

metric_rows = [
    ("Sharpe",   m_v2["sharpe"],   m_mv["sharpe"],   m_bl["sharpe"],   True),
    ("Ann Ret",  m_v2["ann_ret"],  m_mv["ann_ret"],  m_bl["ann_ret"],  True),
    ("Ann Vol",  m_v2["ann_vol"],  m_mv["ann_vol"],  m_bl["ann_vol"],  False),
    ("Max DD",   m_v2["max_dd"],   m_mv["max_dd"],   m_bl["max_dd"],   False),
    ("Total Ret",m_v2["total"],    m_mv["total"],    m_bl["total"],    True),
]
for j, (label, v, mv, bl, higher_is_better) in enumerate(metric_rows):
    bg  = ALT if j % 2 == 0 else WHT
    dv  = round(bl - v, 3)
    dmv = round(bl - mv, 3)
    suffix = "pp" if label not in ("Sharpe",) else ""
    better_v   = dv > 0 if higher_is_better else dv < 0
    better_mv  = dmv > 0 if higher_is_better else dmv < 0
    cel(ws2, ri_m, 1, label,            bg=GREY, bold=True, ha="left")
    cel(ws2, ri_m, 2, str(v),           bg=bg)
    cel(ws2, ri_m, 3, str(mv),          bg=bg)
    cel(ws2, ri_m, 4, str(bl),          bg=bg, bold=True)
    cel(ws2, ri_m, 5, f"{dv:+.3f}{suffix}",  bg=GRN if better_v else RED, bold=True)
    cel(ws2, ri_m, 6, f"{dmv:+.3f}{suffix}", bg=GRN if better_mv else RED)
    ws2.row_dimensions[ri_m].height = 22; ri_m += 1

cw(ws2, [4, 14, 12, 12, 12, 12, 12, 12, 10, 8])

# ━━ SHEET 3: Assumptions ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ws3 = wb.create_sheet("Assumptions")
title_row(ws3, 1, 2, "Black-Litterman — Methodology & Assumptions")
assumptions = [
    ("Model",              "Black-Litterman with Idzorek confidence-based uncertainty"),
    ("Equilibrium prior",  f"Reactor Core v2 weights (8 positions) + candidates at {MIN_W*100:.0f}% floor, normalised"),
    ("Risk aversion δ",    f"{DELTA} (standard market assumption)"),
    ("Prior uncertainty τ",f"{TAU} (5% uncertainty in equilibrium)"),
    ("Omega method",       "Idzorek: Ω_ii = τ × (p_i Σ p_i^T) × (1−conf)/conf"),
    ("Views",              "Absolute return views (annualised) on 5 of 10 assets"),
    ("Non-viewed assets",  "Walmart, Broadcom, Cameco, J&J, Costco held at equilibrium"),
    ("Optimiser",          "Max Sharpe on BL posterior returns/covariance, SLSQP, 100 starts"),
    ("Constraints",        f"Min {MIN_W*100:.0f}% per position | Gold cap {GOLD_CAP*100:.0f}% | Sum=1"),
    ("Data window",        f"{START} to present (10Y)"),
    ("Candidates",         "Silver (xagusd) and COPX (copx.us) included as potential additions"),
    ("Gold cap",           f"Maintained at {GOLD_CAP*100:.0f}% regardless of BL view strength"),
]
hdr(ws3, 2, 1, "Parameter", bg=GREY, fg="000000")
hdr(ws3, 2, 2, "Detail",    bg=GREY, fg="000000")
ws3.row_dimensions[2].height = 24
for ri, (param, detail) in enumerate(assumptions, 3):
    bg = ALT if ri % 2 == 0 else WHT
    cel(ws3, ri, 1, param,  bg=GREY, bold=True, ha="left")
    cel(ws3, ri, 2, detail, bg=bg,   ha="left")
    ws3.row_dimensions[ri].height = 20
cw(ws3, [24, 70])

wb.save(REPORT)
print(f"\n  Report saved: {REPORT}")
print("=" * 70)
