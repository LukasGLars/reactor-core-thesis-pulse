import pandas as pd
import numpy as np
import os, sys, io, warnings, time, requests
from scipy.optimize import minimize
from io import StringIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
warnings.filterwarnings("ignore")
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

DATA_DIR   = r"C:\Users\lukas.larsson\Desktop\Privat\Project Reactor Core\data"
OUTPUT_DIR = r"C:\Users\lukas.larsson\Desktop\Privat\Project Reactor Core\output"
REPORT     = r"C:\Users\lukas.larsson\Desktop\Privat\Project Reactor Core\Portfolio_Stress_Test.xlsx"

TICKERS = ["xauusd","lly.us","wmt.us","avgo.us","lite.us","cost.us","tsla.us","ccj.us","vrt.us","jnj.us"]
NAMES   = {
    "xauusd":"Gold","lly.us":"Eli Lilly","wmt.us":"Walmart",
    "avgo.us":"Broadcom","lite.us":"Lumentum","cost.us":"Costco",
    "tsla.us":"Tesla","ccj.us":"Cameco","vrt.us":"Vertiv","jnj.us":"J&J",
}
ROLES = {
    "xauusd":"Hedge","lly.us":"Carry","wmt.us":"Carry",
    "avgo.us":"Convexity","lite.us":"Convexity","cost.us":"Carry",
    "tsla.us":"Convexity","ccj.us":"Cyclical","vrt.us":"Convexity","jnj.us":"Carry",
}
REGIMES = {
    "Pre-COVID Bull":      ("2016-04-01","2020-01-31"),
    "COVID Crash":         ("2020-02-01","2020-03-31"),
    "COVID Recovery":      ("2020-04-01","2021-12-31"),
    "Rate Hike/Inflation": ("2022-01-01","2023-07-31"),
    "Post-Hike/AI Bull":   ("2023-08-01","2024-08-31"),
    "Rate Cut":            ("2024-09-01","2026-04-01"),
}
WINDOWS = {"3Y":"2023-04-01","5Y":"2021-04-01","10Y":"2016-04-01"}
CURRENT_WEIGHTS = {
    "xauusd":0.379,"lly.us":0.156,"wmt.us":0.114,
    "avgo.us":0.050,"lite.us":0.050,"cost.us":0.050,
    "tsla.us":0.050,"ccj.us":0.050,"vrt.us":0.050,"jnj.us":0.050,
}
FLOOR_POSITIONS = [t for t in TICKERS if t != "xauusd" and
                   CURRENT_WEIGHTS[t] == 0.050]   # 7 positions at 5% min

MIN_W, MAX_W = 0.05, 0.40

# ── Helpers ───────────────────────────────────────────────────────────────────
def load(ticker):
    fname = ticker.replace("^","").replace("-","_") + ".csv"
    path  = os.path.join(DATA_DIR, fname)
    if not os.path.exists(path): return None
    df = pd.read_csv(path)
    df["Date"] = pd.to_datetime(df["Date"])
    df = df.set_index("Date").sort_index()
    return df["Close"].replace(0, np.nan).dropna()

def prices_df(tickers, start):
    d = {t: load(t) for t in tickers if load(t) is not None}
    df = pd.DataFrame({t: s[s.index >= start] for t, s in d.items()}).dropna()
    return df

def optimize_sharpe(df, min_w=MIN_W, max_w=MAX_W, gold_cap=None, n_starts=80):
    rets = df.pct_change().dropna()
    mu   = rets.mean()
    cov  = rets.cov()
    n    = len(df.columns)
    bounds = []
    for t in df.columns:
        hi = min(max_w, gold_cap) if (t == "xauusd" and gold_cap is not None) else max_w
        bounds.append((min_w, hi))
    def neg_sharpe(w):
        r = np.dot(w, mu) * 252
        v = np.sqrt(w @ cov @ w) * np.sqrt(252)
        return -r/v if v > 0 else 0
    constraints = [{"type":"eq","fun":lambda w: np.sum(w)-1}]
    best = None
    for _ in range(n_starts):
        w0 = np.random.dirichlet(np.ones(n))
        w0 = np.clip(w0, [b[0] for b in bounds], [b[1] for b in bounds])
        w0 /= w0.sum()
        res = minimize(neg_sharpe, w0, method="SLSQP", bounds=bounds,
                       constraints=constraints, options={"maxiter":1000,"ftol":1e-12})
        if res.success and (best is None or res.fun < best.fun):
            best = res
    return best.x if best else None

def port_metrics(df, weights):
    w    = np.array(weights)
    rets = df.pct_change().dropna()
    mu   = rets.mean(); cov = rets.cov()
    ann_ret = np.dot(w, mu) * 252
    ann_vol = np.sqrt(w @ cov @ w) * np.sqrt(252)
    sharpe  = ann_ret / ann_vol if ann_vol > 0 else 0
    port    = (df / df.iloc[0]).dot(w)
    max_dd  = ((port - port.cummax()) / port.cummax()).min()
    calmar  = ann_ret / abs(max_dd) if max_dd < 0 else 0
    total   = port.iloc[-1] - 1
    return {"sharpe":round(sharpe,3),"ann_ret":round(ann_ret*100,2),
            "ann_vol":round(ann_vol*100,2),"max_dd":round(max_dd*100,2),
            "calmar":round(calmar,3),"total":round(total*100,1)}

# ── 1. Gold cap sensitivity ───────────────────────────────────────────────────
print("=" * 60)
print("  1. Gold cap sensitivity")
print("=" * 60)

GOLD_CAPS = [0.15, 0.20, 0.25, 0.30, 0.35, None]  # None = unconstrained (current)
gold_results = []

for cap in GOLD_CAPS:
    row = {"gold_cap": f"{cap*100:.0f}%" if cap else "Uncapped (37.9%)"}
    for wname, wstart in WINDOWS.items():
        df = prices_df(TICKERS, wstart)
        w  = optimize_sharpe(df, gold_cap=cap)
        if w is None:
            row[f"{wname}_sharpe"] = None
            row[f"{wname}_gold_w"] = None
            continue
        w_dict = dict(zip(df.columns, w))
        m = port_metrics(df, w)
        row[f"{wname}_sharpe"]  = m["sharpe"]
        row[f"{wname}_ann_ret"] = m["ann_ret"]
        row[f"{wname}_max_dd"]  = m["max_dd"]
        row[f"{wname}_calmar"]  = m["calmar"]
        row[f"{wname}_gold_w"]  = round(w_dict.get("xauusd",0)*100, 1)
    gold_results.append(row)
    lbl = row["gold_cap"]
    print(f"  {lbl:20}  10Y Sharpe {row.get('10Y_sharpe','?')}  "
          f"Gold {row.get('10Y_gold_w','?')}%  MaxDD {row.get('10Y_max_dd','?')}%")

# ── 2. Floor position audit ───────────────────────────────────────────────────
print(f"\n{'=' * 60}")
print("  2. Floor position leave-one-out audit (10Y window)")
print("=" * 60)

df_10y = prices_df(TICKERS, "2016-04-01")
base_w = optimize_sharpe(df_10y)
base_m = port_metrics(df_10y, base_w)
print(f"  Baseline 10Y Sharpe: {base_m['sharpe']}")

# Correlation matrix (10Y)
rets_10y = df_10y.pct_change().dropna()
corr_10y = rets_10y.corr()

# Portfolio daily returns (for correlation to portfolio)
base_w_arr = np.array([dict(zip(df_10y.columns, base_w))[t] for t in df_10y.columns])
port_rets_10y = rets_10y.dot(base_w_arr)

audit_results = []

for drop_t in FLOOR_POSITIONS:
    remaining = [t for t in TICKERS if t != drop_t]
    df_sub = prices_df(remaining, "2016-04-01")
    w_sub  = optimize_sharpe(df_sub)
    if w_sub is None:
        continue
    m_sub  = port_metrics(df_sub, w_sub)

    delta_sharpe_10y = round(m_sub["sharpe"] - base_m["sharpe"], 3)

    # Multi-window delta sharpe
    ds_windows = {}
    for wname, wstart in WINDOWS.items():
        df_b = prices_df(TICKERS, wstart)
        df_s = prices_df(remaining, wstart)
        wb = optimize_sharpe(df_b, n_starts=40)
        ws = optimize_sharpe(df_s, n_starts=40)
        if wb is not None and ws is not None:
            mb = port_metrics(df_b, wb)["sharpe"]
            ms = port_metrics(df_s, ws)["sharpe"]
            ds_windows[wname] = round(ms - mb, 3)
        else:
            ds_windows[wname] = None

    # Correlation to gold
    corr_to_gold = round(corr_10y.loc[drop_t, "xauusd"], 3) if drop_t in corr_10y else None

    # Correlation to portfolio
    corr_to_port = round(rets_10y[drop_t].corr(port_rets_10y), 3)

    # Regime wins
    regime_wins = 0
    regime_details = {}
    for rname, (rs, re) in REGIMES.items():
        s = df_10y[drop_t].loc[rs:re].dropna()
        if len(s) > 5:
            ret = s.iloc[-1]/s.iloc[0] - 1
            regime_details[rname] = round(ret*100,1)
            if ret > 0: regime_wins += 1

    # Admission test
    # 1. Sharpe: does removing it hurt by >=0.03 in ALL windows?
    sharpe_passes = all(v is not None and v <= -0.03 for v in ds_windows.values())
    # 2. Correlation: corr to gold < 0.8 (general threshold)
    corr_passes = corr_to_gold is not None and abs(corr_to_gold) < 0.75
    # 3. Regime: wins >= 4/6
    regime_passes = regime_wins >= 4
    # 4. Role clarity: already assigned
    role_clear = True  # all have assigned roles; check for convexity overlap
    n_convexity = sum(1 for t in TICKERS if ROLES[t] == "Convexity")
    if ROLES[drop_t] == "Convexity" and n_convexity > 2:
        role_clear = False

    criteria_passed = sum([sharpe_passes, corr_passes, regime_passes, role_clear])
    verdict = "KEEP" if criteria_passed >= 3 else "REVIEW"

    row = {
        "ticker": drop_t, "name": NAMES[drop_t], "role": ROLES[drop_t],
        "delta_sharpe_3Y": ds_windows.get("3Y"),
        "delta_sharpe_5Y": ds_windows.get("5Y"),
        "delta_sharpe_10Y": ds_windows.get("10Y"),
        "corr_gold": corr_to_gold, "corr_port": corr_to_port,
        "regime_wins": regime_wins,
        "sharpe_test": "PASS" if sharpe_passes else "FAIL",
        "corr_test":   "PASS" if corr_passes else "FAIL",
        "regime_test": "PASS" if regime_passes else "FAIL",
        "role_test":   "PASS" if role_clear else "FAIL",
        "criteria_passed": criteria_passed,
        "verdict": verdict,
    }
    row.update({f"regime_{k.replace('/','_').replace(' ','_')}": v
                for k,v in regime_details.items()})
    audit_results.append(row)
    print(f"  {NAMES[drop_t]:<14}  dS={delta_sharpe_10y:+.3f}  "
          f"corrGold={corr_to_gold}  regimes={regime_wins}/6  "
          f"criteria={criteria_passed}/4  {verdict}")

# ── 3. FX impact (USD vs SEK) ─────────────────────────────────────────────────
print(f"\n{'=' * 60}")
print("  3. FX impact — portfolio in SEK vs USD")
print("=" * 60)

# Download USD/SEK if not cached
usdsek_path = os.path.join(DATA_DIR, "usdsek.csv")
if not os.path.exists(usdsek_path):
    print("  Downloading USD/SEK...", end=" ")
    sess = requests.Session()
    sess.headers.update({"User-Agent":"Mozilla/5.0","Referer":"https://stooq.com"})
    sess.get("https://stooq.com", timeout=15)
    r = sess.get("https://stooq.com/q/d/l/?s=usdsek&i=d", timeout=15)
    df_fx = pd.read_csv(StringIO(r.text))
    df_fx.to_csv(usdsek_path, index=False)
    print("done")

usdsek = pd.read_csv(usdsek_path)
usdsek["Date"] = pd.to_datetime(usdsek["Date"])
usdsek = usdsek.set_index("Date")["Close"].replace(0, np.nan).dropna()

# Build 5Y window (common start)
START_FX = "2021-04-01"
df_5y = prices_df(TICKERS, START_FX)
w_5y  = optimize_sharpe(df_5y, n_starts=60)
w_dict_5y = dict(zip(df_5y.columns, w_5y))
w_arr_5y  = np.array([w_dict_5y[t] for t in df_5y.columns])

# Portfolio in USD
port_usd = (df_5y / df_5y.iloc[0]).dot(w_arr_5y)

# FX series aligned to portfolio dates
fx_aligned = usdsek.reindex(df_5y.index, method="ffill").dropna()
common_idx  = port_usd.index.intersection(fx_aligned.index)
port_usd_c  = port_usd[common_idx]
fx_c        = fx_aligned[common_idx]

# Portfolio in SEK: multiply by daily FX change factor
# SEK value = USD_value * (FX_t / FX_0)
port_sek = port_usd_c * (fx_c / fx_c.iloc[0])

def series_metrics(s):
    rets = s.pct_change().dropna()
    n    = len(s)/252
    ann  = (s.iloc[-1]**(1/n)) - 1
    vol  = rets.std()*np.sqrt(252)
    dd   = ((s - s.cummax())/s.cummax()).min()
    sh   = ann/vol if vol > 0 else 0
    return {"total":round((s.iloc[-1]-1)*100,1),"ann":round(ann*100,2),
            "vol":round(vol*100,2),"sharpe":round(sh,3),"max_dd":round(dd*100,2)}

m_usd = series_metrics(port_usd_c)
m_sek = series_metrics(port_sek)
fx_drag = round(m_sek["total"] - m_usd["total"], 1)
print(f"  USD total: {m_usd['total']}%  Sharpe {m_usd['sharpe']}")
print(f"  SEK total: {m_sek['total']}%  Sharpe {m_sek['sharpe']}")
print(f"  FX drag:   {fx_drag:+.1f}pp")

# Per-position FX drag
per_pos_fx = []
for t in df_5y.columns:
    s_usd = df_5y[t] / df_5y[t].iloc[0]
    s_sek = s_usd * (fx_c.reindex(s_usd.index, method="ffill") /
                     fx_c.reindex(s_usd.index, method="ffill").iloc[0])
    ret_usd = round((s_usd.iloc[-1]-1)*100, 1)
    ret_sek = round((s_sek.iloc[-1]-1)*100, 1)
    per_pos_fx.append({
        "ticker": t, "name": NAMES[t], "role": ROLES[t],
        "weight": round(CURRENT_WEIGHTS[t]*100, 1),
        "return_usd": ret_usd, "return_sek": ret_sek,
        "fx_impact": round(ret_sek - ret_usd, 1),
    })
    print(f"  {NAMES[t]:<14}  USD {ret_usd:>+6.1f}%  SEK {ret_sek:>+6.1f}%  "
          f"FX {ret_sek-ret_usd:>+5.1f}pp")

# USD/SEK stats
fx_rets = fx_c.pct_change().dropna()
port_rets_usd = port_usd_c.pct_change().dropna()
common_r = port_rets_usd.index.intersection(fx_rets.index)
corr_port_fx = round(port_rets_usd[common_r].corr(fx_rets[common_r]), 3)
fx_ann_ret = round(((fx_c.iloc[-1]/fx_c.iloc[0])**(252/len(fx_c))-1)*100, 2)
print(f"\n  USD/SEK 5Y ann change: {fx_ann_ret:+.2f}%/yr")
print(f"  Portfolio-FX correlation: {corr_port_fx}")

# ── Excel output ──────────────────────────────────────────────────────────────
DARK="1F4E79"; MID="2E75B6"; LIGHT="D6E4F0"; ALT="EBF3FB"
GRN="C6EFCE"; RED="FFC7CE"; YEL="FFEB9C"; WHT="FFFFFF"; GREY="F2F2F2"
PASS_C=GRN; FAIL_C=RED; KEEP_C=GRN; REV_C=YEL

thin   = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr(ws, row, col, val, bg=DARK, fg=WHT, bold=True, sz=10, ha="center"):
    c = ws.cell(row=row, column=col, value=val)
    c.fill = PatternFill("solid", fgColor=bg); c.font = Font(bold=bold, color=fg, size=sz)
    c.alignment = Alignment(wrap_text=True, vertical="center", horizontal=ha)
    c.border = BORDER; return c

def cel(ws, row, col, val, bg=None, bold=False, ha="center", color="000000"):
    c = ws.cell(row=row, column=col, value=val)
    if bg: c.fill = PatternFill("solid", fgColor=bg)
    c.font = Font(bold=bold, color=color)
    c.alignment = Alignment(wrap_text=True, vertical="center", horizontal=ha)
    c.border = BORDER; return c

def title(ws, row, ncols, text, bg=DARK, sz=13):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(bold=True, color=WHT, size=sz)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 34

def cols(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

wb = openpyxl.Workbook()

# ━━ SHEET 1: Gold Cap Sensitivity ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ws1 = wb.active; ws1.title = "Gold Cap Sensitivity"
title(ws1, 1, 13, "Gold Cap Sensitivity — Sharpe & Risk Across Cap Levels")
ws1.merge_cells("A2:M2")
c = ws1["A2"]; c.value = ("Each cap tested across 3Y / 5Y / 10Y windows with 80-restart SLSQP. "
                           "Green = best in column. Current portfolio uses uncapped optimization → 37.9% gold.")
c.font = Font(italic=True, size=9, color="444444")
c.fill = PatternFill("solid", fgColor=LIGHT)
c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws1.row_dimensions[2].height = 26

hdrs1 = ["Gold Cap",
         "3Y Sharpe","3Y Gold%","3Y MaxDD%",
         "5Y Sharpe","5Y Gold%","5Y MaxDD%",
         "10Y Sharpe","10Y AnnRet%","10Y Gold%","10Y MaxDD%","10Y Calmar","10Y Total%"]
for ci, h in enumerate(hdrs1, 1): hdr(ws1, 3, ci, h)
ws1.row_dimensions[3].height = 28

# find column bests
def col_best(results, key, higher=True):
    vals = [r.get(key) for r in results if r.get(key) is not None]
    return max(vals) if higher and vals else (min(vals) if vals else None)

bests = {
    "3Y_sharpe":  col_best(gold_results, "3Y_sharpe"),
    "5Y_sharpe":  col_best(gold_results, "5Y_sharpe"),
    "10Y_sharpe": col_best(gold_results, "10Y_sharpe"),
    "10Y_max_dd": col_best(gold_results, "10Y_max_dd", higher=False),
    "10Y_calmar": col_best(gold_results, "10Y_calmar"),
}

for ri, row in enumerate(gold_results, 4):
    is_current = row["gold_cap"].startswith("Uncapped")
    bg_row = "FFF2CC" if is_current else (ALT if ri%2==0 else WHT)

    def v(key): return row.get(key)
    def bg_v(key, higher=True):
        val = v(key)
        if val is None: return bg_row
        best = bests.get(key)
        if best is None: return bg_row
        return GRN if val == best else bg_row

    cel(ws1, ri, 1,  row["gold_cap"],   bg=bg_row, bold=is_current, ha="left")
    cel(ws1, ri, 2,  v("3Y_sharpe"),    bg=bg_v("3Y_sharpe"),  bold=(v("3Y_sharpe")==bests["3Y_sharpe"]))
    cel(ws1, ri, 3,  f"{v('3Y_gold_w')}%" if v("3Y_gold_w") else "—", bg=bg_row)
    cel(ws1, ri, 4,  f"{v('3Y_max_dd')}%" if v("3Y_max_dd") else "—", bg=bg_row)
    cel(ws1, ri, 5,  v("5Y_sharpe"),    bg=bg_v("5Y_sharpe"),  bold=(v("5Y_sharpe")==bests["5Y_sharpe"]))
    cel(ws1, ri, 6,  f"{v('5Y_gold_w')}%" if v("5Y_gold_w") else "—", bg=bg_row)
    cel(ws1, ri, 7,  f"{v('5Y_max_dd')}%" if v("5Y_max_dd") else "—", bg=bg_row)
    cel(ws1, ri, 8,  v("10Y_sharpe"),   bg=bg_v("10Y_sharpe"), bold=(v("10Y_sharpe")==bests["10Y_sharpe"]))
    cel(ws1, ri, 9,  f"{v('10Y_ann_ret')}%" if v("10Y_ann_ret") else "—", bg=bg_row)
    cel(ws1, ri, 10, f"{v('10Y_gold_w')}%" if v("10Y_gold_w") else "—", bg=bg_row)
    cel(ws1, ri, 11, f"{v('10Y_max_dd')}%" if v("10Y_max_dd") else "—",
        bg=bg_v("10Y_max_dd", higher=False), bold=(v("10Y_max_dd")==bests["10Y_max_dd"]))
    cel(ws1, ri, 12, v("10Y_calmar"),   bg=bg_v("10Y_calmar"), bold=(v("10Y_calmar")==bests["10Y_calmar"]))
    cel(ws1, ri, 13, f"{v('10Y_total')}%" if v("10Y_total") else "—", bg=bg_row)
    ws1.row_dimensions[ri].height = 22

cols(ws1, [22,11,10,10, 11,10,10, 11,12,10,10,11,11])

# ━━ SHEET 2: Floor Position Audit ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ws2 = wb.create_sheet("Floor Position Audit")
title(ws2, 1, 15, "Floor Position Audit — Leave-One-Out vs Admission Criteria")
ws2.merge_cells("A2:O2")
c = ws2["A2"]; c.value = ("Leave-one-out: each position removed, remaining 9 re-optimized. "
                           "ΔSharpe = change vs baseline. PASS criteria: ΔSharpe ≤ -0.03 in all windows | "
                           "Corr to Gold < 0.75 | Regime wins ≥ 4/6 | Role not over-represented.")
c.font = Font(italic=True, size=9, color="444444")
c.fill = PatternFill("solid", fgColor=LIGHT)
c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws2.row_dimensions[2].height = 30

hdrs2 = ["Ticker","Name","Role","Weight",
         "ΔSharpe 3Y","ΔSharpe 5Y","ΔSharpe 10Y",
         "Corr Gold","Corr Port","Regime Wins",
         "Sharpe Test","Corr Test","Regime Test","Role Test",
         "Verdict"]
for ci, h in enumerate(hdrs2, 1): hdr(ws2, 3, ci, h)
ws2.row_dimensions[3].height = 30

for ri, row in enumerate(audit_results, 4):
    bg_row = ALT if ri%2==0 else WHT
    verd = row["verdict"]
    verd_bg = KEEP_C if verd == "KEEP" else REV_C

    def pf(key):
        v = row.get(key, "")
        return (PASS_C if v == "PASS" else FAIL_C) if v in ("PASS","FAIL") else bg_row

    cel(ws2, ri,  1, row["ticker"],           bg=bg_row)
    cel(ws2, ri,  2, row["name"],             bg=bg_row, ha="left")
    cel(ws2, ri,  3, row["role"],             bg=bg_row)
    cel(ws2, ri,  4, "5%", bg=bg_row)

    for ci_off, wn in enumerate(["3Y","5Y","10Y"], 5):
        ds = row.get(f"delta_sharpe_{wn}")
        ds_bg = GRN if ds is not None and ds <= -0.03 else (RED if ds is not None and ds > -0.01 else YEL)
        cel(ws2, ri, ci_off, f"{ds:+.3f}" if ds is not None else "—", bg=ds_bg)

    cel(ws2, ri,  8, row.get("corr_gold","—"), bg=bg_row)
    cel(ws2, ri,  9, row.get("corr_port","—"), bg=bg_row)
    cel(ws2, ri, 10, f"{row.get('regime_wins','?')}/6", bg=bg_row)
    cel(ws2, ri, 11, row.get("sharpe_test",""), bg=pf("sharpe_test"), bold=True)
    cel(ws2, ri, 12, row.get("corr_test",""),   bg=pf("corr_test"),   bold=True)
    cel(ws2, ri, 13, row.get("regime_test",""), bg=pf("regime_test"), bold=True)
    cel(ws2, ri, 14, row.get("role_test",""),   bg=pf("role_test"),   bold=True)
    cel(ws2, ri, 15, verd, bg=verd_bg, bold=True)
    ws2.row_dimensions[ri].height = 22

# Baseline note
ri_note = len(audit_results) + 6
ws2.merge_cells(f"A{ri_note}:O{ri_note}")
c = ws2.cell(row=ri_note, column=1,
             value=f"Baseline 10Y Sharpe (all 10 positions): {base_m['sharpe']}  |  "
                   f"Ann Return: {base_m['ann_ret']}%  |  MaxDD: {base_m['max_dd']}%  |  Calmar: {base_m['calmar']}")
c.font = Font(bold=True, size=10)
c.fill = PatternFill("solid", fgColor=LIGHT)
c.alignment = Alignment(horizontal="center")
ws2.row_dimensions[ri_note].height = 22

cols(ws2, [10,14,12,8, 11,11,11, 10,10,12, 11,10,12,10, 10])

# ━━ SHEET 3: FX Impact ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ws3 = wb.create_sheet("FX Impact (SEK)")
title(ws3, 1, 7, "FX Impact — Portfolio & Holdings USD vs SEK (5Y, 2021-2026)")
ws3.merge_cells("A2:G2")
c = ws3["A2"]; c.value = (f"USD/SEK 5Y change: {fx_ann_ret:+.2f}%/yr annualised  |  "
                           f"Portfolio-FX correlation: {corr_port_fx}  |  "
                           f"FX drag = SEK return minus USD return (negative = SEK strengthened)")
c.font = Font(italic=True, size=9, color="444444")
c.fill = PatternFill("solid", fgColor=LIGHT)
c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws3.row_dimensions[2].height = 26

# Portfolio summary
for ci, h in enumerate(["","Metric","Value (USD)","Value (SEK)","FX Impact (pp)","",""], 1):
    hdr(ws3, 3, ci, h)
ws3.row_dimensions[3].height = 26

summ = [
    ("Total Return",  f"{m_usd['total']}%",  f"{m_sek['total']}%",  f"{m_sek['total']-m_usd['total']:+.1f}pp"),
    ("Ann. Return",   f"{m_usd['ann']}%",    f"{m_sek['ann']}%",    f"{m_sek['ann']-m_usd['ann']:+.2f}pp"),
    ("Sharpe Ratio",  str(m_usd['sharpe']),  str(m_sek['sharpe']),  f"{m_sek['sharpe']-m_usd['sharpe']:+.3f}"),
    ("Max Drawdown",  f"{m_usd['max_dd']}%", f"{m_sek['max_dd']}%", f"{m_sek['max_dd']-m_usd['max_dd']:+.2f}pp"),
    ("Ann. Vol",      f"{m_usd['vol']}%",    f"{m_sek['vol']}%",    f"{m_sek['vol']-m_usd['vol']:+.2f}pp"),
]
for ri, (metric, usd, sek, fx) in enumerate(summ, 4):
    bg = ALT if ri%2==0 else WHT
    fx_val = float(fx.replace("pp","").replace("%","").replace("+",""))
    fx_bg = GRN if fx_val > 0 else (RED if fx_val < -1 else bg)
    cel(ws3, ri, 1, "", bg=bg)
    cel(ws3, ri, 2, metric, bg=GREY, bold=True, ha="left")
    cel(ws3, ri, 3, usd,    bg=bg)
    cel(ws3, ri, 4, sek,    bg=bg)
    cel(ws3, ri, 5, fx,     bg=fx_bg, bold=True)
    cel(ws3, ri, 6, "", bg=bg); cel(ws3, ri, 7, "", bg=bg)
    ws3.row_dimensions[ri].height = 22

# Per-position
ri = len(summ) + 7
ws3.merge_cells(f"A{ri}:G{ri}")
c = ws3.cell(row=ri, column=1, value="Per-Position FX Impact (5Y, current weights)")
c.font = Font(bold=True, color=WHT, size=11)
c.fill = PatternFill("solid", fgColor=MID)
c.alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[ri].height = 26
ri += 1

for ci, h in enumerate(["Ticker","Name","Role","Weight","Return USD","Return SEK","FX Impact"], 1):
    hdr(ws3, ri, ci, h)
ws3.row_dimensions[ri].height = 26
ri += 1

for i, row in enumerate(sorted(per_pos_fx, key=lambda x: x["fx_impact"])):
    bg = ALT if i%2==0 else WHT
    fx_v = row["fx_impact"]
    fx_bg = GRN if fx_v > 1 else (RED if fx_v < -3 else bg)
    cel(ws3, ri, 1, row["ticker"],              bg=bg)
    cel(ws3, ri, 2, row["name"],                bg=bg, ha="left")
    cel(ws3, ri, 3, row["role"],                bg=bg)
    cel(ws3, ri, 4, f"{row['weight']}%",        bg=bg)
    cel(ws3, ri, 5, f"{row['return_usd']:+.1f}%", bg=bg)
    cel(ws3, ri, 6, f"{row['return_sek']:+.1f}%", bg=bg)
    cel(ws3, ri, 7, f"{fx_v:+.1f}pp",          bg=fx_bg, bold=(abs(fx_v)>3))
    ws3.row_dimensions[ri].height = 20
    ri += 1

cols(ws3, [10, 16, 12, 9, 14, 14, 14])

wb.save(REPORT)
print(f"\n  Report saved: {REPORT}")
print("=" * 60)
