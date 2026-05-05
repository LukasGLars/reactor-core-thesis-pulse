import pandas as pd
import numpy as np
import os, sys, io, warnings
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
warnings.filterwarnings("ignore")
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

DATA_DIR = r"C:\Users\lukas.larsson\Desktop\Privat\Project Reactor Core\data"
REPORT   = r"C:\Users\lukas.larsson\Desktop\Privat\Project Reactor Core\Out_of_Sample.xlsx"

# Original v2 weights (8 positions)
W_ORIG = {
    "xauusd": 0.250, "wmt.us": 0.227, "lly.us": 0.197,
    "vrt.us": 0.090, "avgo.us": 0.079, "ccj.us": 0.057,
    "jnj.us": 0.050, "cost.us": 0.050,
}

# 7-position weights — Vertiv removed, remainder normalized
W_7 = {t: w for t, w in W_ORIG.items() if t != "vrt.us"}
total_7 = sum(W_7.values())
W_7 = {t: round(w / total_7, 6) for t, w in W_7.items()}

NAMES = {
    "xauusd":"Gold","lly.us":"Eli Lilly","wmt.us":"Walmart",
    "avgo.us":"Broadcom","cost.us":"Costco","ccj.us":"Cameco",
    "jnj.us":"J&J",
}

# Periods
IN_SAMPLE    = ("2016-04-01", "2026-03-31")   # optimization window
OOS_FULL     = ("2010-01-01", "2016-03-31")   # out-of-sample (pre-optimization)
OOS_09       = ("2009-08-05", "2016-03-31")   # max available (Broadcom IPO limit)

def load(ticker):
    df = pd.read_csv(os.path.join(DATA_DIR, ticker.replace("-","_") + ".csv"))
    df["Date"] = pd.to_datetime(df["Date"])
    return df.set_index("Date").sort_index()["Close"].replace(0, np.nan).dropna()

def build(tickers, start, end):
    df = pd.DataFrame({t: load(t) for t in tickers})
    return df[(df.index >= start) & (df.index <= end)].dropna()

def metrics(df, w_dict):
    w    = np.array([w_dict[t] for t in df.columns])
    r_   = df.pct_change().dropna()
    mu_  = r_.mean() * 252
    cov_ = r_.cov() * 252
    ar   = np.dot(w, mu_)
    av   = np.sqrt(w @ cov_.values @ w)
    sh   = ar / av if av > 0 else 0
    port = (df / df.iloc[0]).dot(w)
    mdd  = ((port - port.cummax()) / port.cummax()).min()
    tot  = port.iloc[-1] - 1
    cal  = ar / abs(mdd) if mdd < 0 else 0
    return {
        "sharpe":  round(sh, 3),
        "ann_ret": round(ar * 100, 2),
        "ann_vol": round(av * 100, 2),
        "max_dd":  round(mdd * 100, 2),
        "calmar":  round(cal, 3),
        "total":   round(tot * 100, 1),
        "port":    port,
    }

tickers_7 = list(W_7.keys())

# ── In-sample (8-pos, original weights) ──────────────────────────────────────
df_is8 = build(list(W_ORIG.keys()), *IN_SAMPLE)
m_is8  = metrics(df_is8, W_ORIG)

# ── In-sample (7-pos, normalized weights) ────────────────────────────────────
df_is7 = build(tickers_7, *IN_SAMPLE)
m_is7  = metrics(df_is7, W_7)

# ── Out-of-sample 2010-2016 (7-pos) ──────────────────────────────────────────
df_oos = build(tickers_7, *OOS_FULL)
m_oos  = metrics(df_oos, W_7)

# ── Out-of-sample 2009-2016 max window ───────────────────────────────────────
df_oos_max = build(tickers_7, *OOS_09)
m_oos_max  = metrics(df_oos_max, W_7)

# ── Year-by-year OOS breakdown ────────────────────────────────────────────────
yearly_oos = []
for yr in range(2010, 2017):
    s = f"{yr}-01-01"; e = f"{yr}-12-31"
    try:
        df_y = build(tickers_7, s, e)
        if len(df_y) < 50: continue
        m_y  = metrics(df_y, W_7)
        yearly_oos.append({"year": yr, **m_y})
    except: continue

# ── Year-by-year IS breakdown ─────────────────────────────────────────────────
yearly_is = []
for yr in range(2016, 2027):
    s = f"{yr}-01-01"; e = f"{yr}-12-31"
    try:
        df_y = build(list(W_ORIG.keys()), s, e)
        if len(df_y) < 50: continue
        m_y  = metrics(df_y, W_ORIG)
        yearly_is.append({"year": yr, **m_y})
    except: continue

# ── Console output ────────────────────────────────────────────────────────────
print("=" * 68)
print("  Out-of-Sample Validation — Reactor Core v2")
print("  7-position test (Vertiv excluded, weights normalized)")
print("=" * 68)

print(f"\n  7-pos normalized weights:")
for t, w in sorted(W_7.items(), key=lambda x: -x[1]):
    orig = W_ORIG[t]
    print(f"    {NAMES.get(t,t):<14} {w*100:.2f}%  (was {orig*100:.1f}%)")

print(f"\n  {'Period':<30} {'Sharpe':>8}  {'AnnRet':>8}  {'AnnVol':>8}  {'MaxDD':>8}  {'Total':>8}")
print(f"  {'-'*70}")
print(f"  {'IN-SAMPLE 8-pos (2016-2026)':<30} {m_is8['sharpe']:>8}  {m_is8['ann_ret']:>7}%  {m_is8['ann_vol']:>7}%  {m_is8['max_dd']:>7}%  {m_is8['total']:>7}%")
print(f"  {'IN-SAMPLE 7-pos (2016-2026)':<30} {m_is7['sharpe']:>8}  {m_is7['ann_ret']:>7}%  {m_is7['ann_vol']:>7}%  {m_is7['max_dd']:>7}%  {m_is7['total']:>7}%")
print(f"  {'OUT-OF-SAMPLE (2010-2016)':<30} {m_oos['sharpe']:>8}  {m_oos['ann_ret']:>7}%  {m_oos['ann_vol']:>7}%  {m_oos['max_dd']:>7}%  {m_oos['total']:>7}%")
print(f"  {'OUT-OF-SAMPLE (2009-2016)':<30} {m_oos_max['sharpe']:>8}  {m_oos_max['ann_ret']:>7}%  {m_oos_max['ann_vol']:>7}%  {m_oos_max['max_dd']:>7}%  {m_oos_max['total']:>7}%")

print(f"\n  Year-by-year OUT-OF-SAMPLE:")
for y in yearly_oos:
    print(f"    {y['year']}:  Sharpe {y['sharpe']:>6}  AnnRet {y['ann_ret']:>7}%  MaxDD {y['max_dd']:>7}%  Total {y['total']:>7}%")

print(f"\n  Year-by-year IN-SAMPLE:")
for y in yearly_is:
    print(f"    {y['year']}:  Sharpe {y['sharpe']:>6}  AnnRet {y['ann_ret']:>7}%  MaxDD {y['max_dd']:>7}%  Total {y['total']:>7}%")

# ── Excel ─────────────────────────────────────────────────────────────────────
DARK="1F4E79"; MID="2E75B6"; LIGHT="D6E4F0"; ALT="EBF3FB"
GRN="C6EFCE"; RED="FFC7CE"; YEL="FFEB9C"; WHT="FFFFFF"; GREY="F2F2F2"
thin   = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr(ws, row, col, val, bg=DARK, fg=WHT):
    c = ws.cell(row=row, column=col, value=val)
    c.fill = PatternFill("solid", fgColor=bg)
    c.font = Font(bold=True, color=fg, size=10)
    c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    c.border = BORDER; return c

def cel(ws, row, col, val, bg=None, bold=False, ha="center"):
    c = ws.cell(row=row, column=col, value=val)
    if bg: c.fill = PatternFill("solid", fgColor=bg)
    c.font = Font(bold=bold)
    c.alignment = Alignment(vertical="center", horizontal=ha)
    c.border = BORDER; return c

def title_row(ws, row, nc, text, bg=DARK, sz=13):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=nc)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(bold=True, color=WHT, size=sz)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 34

def cw(ws, widths):
    for i, w_ in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w_

wb = openpyxl.Workbook()

# ━━ SHEET 1: Summary ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ws1 = wb.active; ws1.title = "OOS Summary"
title_row(ws1, 1, 7, "Out-of-Sample Validation — Reactor Core v2 (7-pos, Vertiv excluded)")

ws1.merge_cells("A2:G2")
c2 = ws1["A2"]
c2.value = ("Weights optimized on 2016-2026 data. Tested on 2010-2016 (unseen).  "
            "Vertiv excluded (IPO 2018) — remaining 7 weights normalized.  "
            "If OOS Sharpe >> 0, portfolio has structural alpha beyond data-fitting.")
c2.font = Font(italic=True, size=9, color="444444")
c2.fill = PatternFill("solid", fgColor=LIGHT)
c2.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws1.row_dimensions[2].height = 28

for ci, h in enumerate(["Period","Positions","Sharpe","Ann Ret","Ann Vol","Max DD","Total Ret"], 1):
    hdr(ws1, 3, ci, h)
ws1.row_dimensions[3].height = 28

summary_rows = [
    ("In-Sample 2016-2026",   "8-pos (full)",      m_is8, False),
    ("In-Sample 2016-2026",   "7-pos (no Vertiv)", m_is7, False),
    ("OUT-OF-SAMPLE 2010-2016","7-pos (no Vertiv)", m_oos, True),
    ("OUT-OF-SAMPLE 2009-2016","7-pos (no Vertiv)", m_oos_max, True),
]
for ri, (period, pos, m, is_oos) in enumerate(summary_rows, 4):
    bg = "E8F5E9" if is_oos else (ALT if ri%2==0 else WHT)
    sh_bg = GRN if m["sharpe"] >= 1.0 else (YEL if m["sharpe"] >= 0.5 else RED)
    cel(ws1, ri, 1, period,              bg=bg, bold=is_oos, ha="left")
    cel(ws1, ri, 2, pos,                 bg=bg)
    cel(ws1, ri, 3, m["sharpe"],         bg=sh_bg, bold=True)
    cel(ws1, ri, 4, f"{m['ann_ret']}%",  bg=bg)
    cel(ws1, ri, 5, f"{m['ann_vol']}%",  bg=bg)
    cel(ws1, ri, 6, f"{m['max_dd']}%",   bg=bg)
    cel(ws1, ri, 7, f"{m['total']}%",    bg=bg)
    ws1.row_dimensions[ri].height = 24

# Verdict
ri_v = 9
ws1.merge_cells(f"A{ri_v}:G{ri_v}")
oos_sharpe = m_oos["sharpe"]
if oos_sharpe >= 1.0:
    verdict = f"STRUCTURAL — OOS Sharpe {oos_sharpe} >= 1.0. Portfolio holds up on unseen data. Alpha is not purely data-fitted."
    vbg = GRN
elif oos_sharpe >= 0.5:
    verdict = f"PARTIAL — OOS Sharpe {oos_sharpe}. Some structural alpha but weaker than in-sample. Partial data-fitting likely."
    vbg = YEL
else:
    verdict = f"FITTED — OOS Sharpe {oos_sharpe} < 0.5. Portfolio does not generalize. In-sample performance is largely a backtest artifact."
    vbg = RED

c_v = ws1.cell(row=ri_v, column=1, value=verdict)
c_v.font = Font(bold=True, size=11)
c_v.fill = PatternFill("solid", fgColor=vbg)
c_v.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws1.row_dimensions[ri_v].height = 36

cw(ws1, [28, 20, 10, 10, 10, 10, 10])

# ━━ SHEET 2: Year-by-Year ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ws2 = wb.create_sheet("Year by Year")
title_row(ws2, 1, 7, "Annual Performance — Out-of-Sample (2010-2016) vs In-Sample (2016-2026)")

for ci, h in enumerate(["Year","Period","Sharpe","Ann Ret","Ann Vol","Max DD","Total Ret"], 1):
    hdr(ws2, 2, ci, h)
ws2.row_dimensions[2].height = 28

ri = 3
for y in yearly_oos:
    bg = "E8F5E9"
    sh_bg = GRN if y["sharpe"] >= 1.0 else (YEL if y["sharpe"] >= 0.5 else RED)
    ret_bg = GRN if y["ann_ret"] > 0 else RED
    cel(ws2, ri, 1, y["year"],           bg=GREY, bold=True)
    cel(ws2, ri, 2, "Out-of-Sample",     bg=bg)
    cel(ws2, ri, 3, y["sharpe"],         bg=sh_bg, bold=True)
    cel(ws2, ri, 4, f"{y['ann_ret']}%",  bg=ret_bg)
    cel(ws2, ri, 5, f"{y['ann_vol']}%",  bg=bg)
    cel(ws2, ri, 6, f"{y['max_dd']}%",   bg=bg)
    cel(ws2, ri, 7, f"{y['total']}%",    bg=ret_bg)
    ws2.row_dimensions[ri].height = 22; ri += 1

ri += 1  # spacer
for y in yearly_is:
    bg = ALT if ri%2==0 else WHT
    sh_bg = GRN if y["sharpe"] >= 1.0 else (YEL if y["sharpe"] >= 0.5 else RED)
    ret_bg = GRN if y["ann_ret"] > 0 else RED
    cel(ws2, ri, 1, y["year"],           bg=GREY, bold=True)
    cel(ws2, ri, 2, "In-Sample",         bg=bg)
    cel(ws2, ri, 3, y["sharpe"],         bg=sh_bg, bold=True)
    cel(ws2, ri, 4, f"{y['ann_ret']}%",  bg=ret_bg)
    cel(ws2, ri, 5, f"{y['ann_vol']}%",  bg=bg)
    cel(ws2, ri, 6, f"{y['max_dd']}%",   bg=bg)
    cel(ws2, ri, 7, f"{y['total']}%",    bg=ret_bg)
    ws2.row_dimensions[ri].height = 22; ri += 1

cw(ws2, [8, 16, 10, 10, 10, 10, 10])

# ━━ SHEET 3: Weight Comparison ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ws3 = wb.create_sheet("Weight Adjustment")
title_row(ws3, 1, 4, "Weight Adjustment — Original v2 vs 7-pos OOS Test")
for ci, h in enumerate(["Asset","Original Weight","OOS Weight","Change"], 1):
    hdr(ws3, 2, ci, h)
ws3.row_dimensions[2].height = 26

for ri, (t, w_orig) in enumerate(sorted(W_ORIG.items(), key=lambda x: -x[1]), 3):
    bg = ALT if ri%2==0 else WHT
    w_oos = W_7.get(t, 0)
    delta = (w_oos - w_orig) * 100
    excluded = t == "vrt.us"
    row_bg = RED if excluded else bg
    cel(ws3, ri, 1, NAMES.get(t,t),          bg=row_bg, bold=excluded, ha="left")
    cel(ws3, ri, 2, f"{w_orig*100:.1f}%",    bg=row_bg)
    cel(ws3, ri, 3, f"{w_oos*100:.2f}%" if not excluded else "EXCLUDED", bg=row_bg, bold=excluded)
    cel(ws3, ri, 4, f"{delta:+.2f}pp" if not excluded else "—",
        bg=GRN if delta>0 else (RED if excluded else bg))
    ws3.row_dimensions[ri].height = 22

cw(ws3, [16, 18, 18, 12])

wb.save(REPORT)
print(f"\n  Report saved: {REPORT}")
print("=" * 68)
