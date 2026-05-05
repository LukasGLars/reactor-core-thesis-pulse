import pandas as pd
import numpy as np
import os, sys, io, warnings
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
warnings.filterwarnings("ignore")
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

DATA_DIR = r"C:\Users\lukas.larsson\Desktop\Privat\Project Reactor Core\data"
REPORT   = r"C:\Users\lukas.larsson\Desktop\Privat\Project Reactor Core\Thesis_Test.xlsx"

# v2 original
W_V2 = {
    "xauusd": 0.250, "wmt.us": 0.227, "lly.us": 0.197,
    "vrt.us": 0.090, "avgo.us": 0.079, "ccj.us": 0.057,
    "jnj.us": 0.050, "cost.us": 0.050,
}

# Thesis portfolio — 98% invested, 2% cash
W_THESIS = {
    "xauusd": 0.250, "xagusd": 0.100, "lly.us": 0.150,
    "wmt.us": 0.150, "ccj.us": 0.100, "vrt.us": 0.100,
    "avgo.us": 0.080, "jnj.us": 0.050,
}
CASH_THESIS = 0.02

NAMES = {
    "xauusd":"Gold","lly.us":"Eli Lilly","wmt.us":"Walmart",
    "avgo.us":"Broadcom","cost.us":"Costco","ccj.us":"Cameco",
    "vrt.us":"Vertiv","jnj.us":"J&J","xagusd":"Silver",
}

PERIODS = {
    "10Y In-Sample":      ("2016-04-01", "2026-03-31"),
    "5Y In-Sample":       ("2021-04-01", "2026-03-31"),
    "3Y In-Sample":       ("2023-04-01", "2026-03-31"),
    "OOS 2010-2016":      ("2010-01-01", "2016-03-31"),
    "OOS 2009-2016":      ("2009-08-05", "2016-03-31"),
}

def load(ticker):
    df = pd.read_csv(os.path.join(DATA_DIR, ticker.replace("-","_") + ".csv"))
    df["Date"] = pd.to_datetime(df["Date"])
    return df.set_index("Date").sort_index()["Close"].replace(0, np.nan).dropna()

def build(tickers, start, end):
    df = pd.DataFrame({t: load(t) for t in tickers})
    return df[(df.index >= start) & (df.index <= end)].dropna()

def metrics(df, w_dict, cash=0.0):
    # cash portion earns 0 — simply scales down returns
    w = np.array([w_dict[t] for t in df.columns])
    r_   = df.pct_change().dropna()
    mu_  = r_.mean() * 252 * (1 - cash)
    cov_ = r_.cov() * 252 * (1 - cash) ** 2
    ar   = np.dot(w, mu_)
    av   = np.sqrt(w @ cov_.values @ w)
    sh   = ar / av if av > 0 else 0
    port = ((df / df.iloc[0]).dot(w)) * (1 - cash) + cash
    mdd  = ((port - port.cummax()) / port.cummax()).min()
    tot  = port.iloc[-1] - 1
    cal  = ar / abs(mdd) if mdd < 0 else 0
    return {
        "sharpe":  round(sh, 3),
        "ann_ret": round(ar * 100, 2),
        "ann_vol": round(av * 100, 2),
        "max_dd":  round(mdd * 100, 2),
        "calmar":  round(cal, 3),
        "total":   round(tot * 100, 1),
    }

tickers_v2     = list(W_V2.keys())
tickers_thesis = list(W_THESIS.keys())

# ── Run all periods ───────────────────────────────────────────────────────────
print("=" * 72)
print("  Thesis Portfolio Test — Short Abundance, Long Scarcity")
print("  vs Reactor Core v2")
print("=" * 72)

print(f"\n  Thesis weights (98% invested, 2% cash):")
for t, w in sorted(W_THESIS.items(), key=lambda x: -x[1]):
    v2w = W_V2.get(t, 0)
    delta = (w - v2w) * 100
    marker = " NEW" if t not in W_V2 else ""
    print(f"    {NAMES.get(t,t):<14} {w*100:.1f}%  (v2: {v2w*100:.1f}%  d={delta:+.1f}pp){marker}")
print(f"    {'Cash':<14}  2.0%")

results = {}
print(f"\n  {'Period':<22} {'v2 Sharpe':>10} {'v2 Ret':>8} {'v2 DD':>8} | {'TH Sharpe':>10} {'TH Ret':>8} {'TH DD':>8} {'Winner':>8}")
print(f"  {'-'*85}")

for pname, (start, end) in PERIODS.items():
    is_oos = "OOS" in pname
    try:
        df_v2 = build(tickers_v2, start, end)
        m_v2  = metrics(df_v2, W_V2)
    except:
        m_v2 = None

    try:
        df_th = build(tickers_thesis, start, end)
        m_th  = metrics(df_th, W_THESIS, cash=CASH_THESIS)
    except:
        m_th = None

    if m_v2 and m_th:
        winner = "THESIS" if m_th["sharpe"] > m_v2["sharpe"] else "v2"
        results[pname] = {"v2": m_v2, "thesis": m_th, "winner": winner, "oos": is_oos}
        tag = " *OOS*" if is_oos else ""
        print(f"  {pname+tag:<22} {m_v2['sharpe']:>10} {m_v2['ann_ret']:>7}% {m_v2['max_dd']:>7}% | "
              f"{m_th['sharpe']:>10} {m_th['ann_ret']:>7}% {m_th['max_dd']:>7}% {winner:>8}")
    else:
        print(f"  {pname:<22} {'insufficient data':>40}")

# Year-by-year
print(f"\n  Year-by-year (thesis vs v2):")
print(f"  {'Year':<6} {'v2 Sharpe':>10} {'v2 Ret':>8} | {'TH Sharpe':>10} {'TH Ret':>8} {'Winner':>8}")
yearly = []
for yr in range(2010, 2027):
    s = f"{yr}-01-01"; e = f"{yr}-12-31"
    try:
        df_v2 = build(tickers_v2, s, e)
        m_v2  = metrics(df_v2, W_V2) if len(df_v2) >= 50 else None
    except: m_v2 = None
    try:
        df_th = build(tickers_thesis, s, e)
        m_th  = metrics(df_th, W_THESIS, cash=CASH_THESIS) if len(df_th) >= 50 else None
    except: m_th = None

    if m_v2 and m_th:
        winner = "THESIS" if m_th["sharpe"] > m_v2["sharpe"] else "v2"
        is_oos = yr < 2016
        yearly.append({"year":yr,"v2":m_v2,"thesis":m_th,"winner":winner,"oos":is_oos})
        tag = "*" if is_oos else " "
        print(f"  {yr}{tag:<5} {m_v2['sharpe']:>10} {m_v2['ann_ret']:>7}% | "
              f"{m_th['sharpe']:>10} {m_th['ann_ret']:>7}% {winner:>8}")

# ── Excel ─────────────────────────────────────────────────────────────────────
DARK="1F4E79"; MID="2E75B6"; LIGHT="D6E4F0"; ALT="EBF3FB"
GRN="C6EFCE"; RED="FFC7CE"; YEL="FFEB9C"; WHT="FFFFFF"; GREY="F2F2F2"
thin   = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr(ws, row, col, val, bg=DARK, fg=WHT):
    c = ws.cell(row=row, column=col, value=val)
    c.fill = PatternFill("solid", fgColor=bg)
    c.font = Font(bold=True, color=fg, size=10)
    c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    c.border = BORDER; return c

def cel(ws, row, col, val, bg=None, bold=False, ha="center"):
    c = ws.cell(row=row, column=col, value=val)
    if bg: c.fill = PatternFill("solid", fgColor=bg)
    c.font = Font(bold=bold)
    c.alignment = Alignment(vertical="center", horizontal=ha)
    c.border = BORDER; return c

def title_row(ws, row, nc, text, bg=DARK, sz=13):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=nc)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(bold=True, color=WHT, size=sz)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 34

def cw(ws, widths):
    for i, w_ in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w_

wb = openpyxl.Workbook()

# ━━ SHEET 1: Head to Head ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ws1 = wb.active; ws1.title = "Head to Head"
title_row(ws1, 1, 11, "Thesis Portfolio vs Reactor Core v2 — Head to Head")

ws1.merge_cells("A2:K2")
c2 = ws1["A2"]
c2.value = ("Thesis: Short Abundance Long Scarcity — Gold 25% Silver 10% Cameco 10% Vertiv 10% "
            "Broadcom 8% Lilly 15% Walmart 15% J&J 5% Cash 2%  |  "
            "OOS rows marked — weights optimized on 2016-2026, tested on 2010-2016")
c2.font = Font(italic=True, size=9, color="444444")
c2.fill = PatternFill("solid", fgColor=LIGHT)
c2.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws1.row_dimensions[2].height = 28

for ci, h in enumerate(["Period","v2 Sharpe","v2 AnnRet","v2 MaxDD","v2 Total",
                          "TH Sharpe","TH AnnRet","TH MaxDD","TH Total","Winner",""], 1):
    hdr(ws1, 3, ci, h)
ws1.row_dimensions[3].height = 28

for ri, (pname, r) in enumerate(results.items(), 4):
    bg     = "E8F5E9" if r["oos"] else (ALT if ri%2==0 else WHT)
    win_v2 = r["winner"] == "v2"
    win_th = r["winner"] == "THESIS"
    cel(ws1, ri, 1,  pname,                        bg=bg, bold=r["oos"], ha="left")
    cel(ws1, ri, 2,  r["v2"]["sharpe"],             bg=GRN if win_v2 else bg, bold=win_v2)
    cel(ws1, ri, 3,  f"{r['v2']['ann_ret']}%",      bg=bg)
    cel(ws1, ri, 4,  f"{r['v2']['max_dd']}%",       bg=bg)
    cel(ws1, ri, 5,  f"{r['v2']['total']}%",        bg=bg)
    cel(ws1, ri, 6,  r["thesis"]["sharpe"],          bg=GRN if win_th else bg, bold=win_th)
    cel(ws1, ri, 7,  f"{r['thesis']['ann_ret']}%",   bg=bg)
    cel(ws1, ri, 8,  f"{r['thesis']['max_dd']}%",    bg=bg)
    cel(ws1, ri, 9,  f"{r['thesis']['total']}%",     bg=bg)
    cel(ws1, ri, 10, r["winner"],
        bg=GRN if win_th else YEL, bold=True)
    ws1.row_dimensions[ri].height = 24

cw(ws1, [22,11,11,10,10,11,11,10,10,10,4])

# ━━ SHEET 2: Year by Year ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ws2 = wb.create_sheet("Year by Year")
title_row(ws2, 1, 8, "Year-by-Year — Thesis vs v2  (* = Out-of-Sample)")

for ci, h in enumerate(["Year","OOS","v2 Sharpe","v2 AnnRet","v2 MaxDD",
                          "TH Sharpe","TH AnnRet","TH MaxDD"], 1):
    hdr(ws2, 2, ci, h)
ws2.row_dimensions[2].height = 28

for ri, y in enumerate(yearly, 3):
    bg     = "E8F5E9" if y["oos"] else (ALT if ri%2==0 else WHT)
    win_th = y["winner"] == "THESIS"
    win_v2 = y["winner"] == "v2"
    cel(ws2, ri, 1, y["year"],                    bg=GREY, bold=True)
    cel(ws2, ri, 2, "OOS" if y["oos"] else "",    bg="E8F5E9" if y["oos"] else bg)
    cel(ws2, ri, 3, y["v2"]["sharpe"],             bg=GRN if win_v2 else bg, bold=win_v2)
    cel(ws2, ri, 4, f"{y['v2']['ann_ret']}%",      bg=GRN if y["v2"]["ann_ret"]>0 else RED)
    cel(ws2, ri, 5, f"{y['v2']['max_dd']}%",       bg=bg)
    cel(ws2, ri, 6, y["thesis"]["sharpe"],          bg=GRN if win_th else bg, bold=win_th)
    cel(ws2, ri, 7, f"{y['thesis']['ann_ret']}%",   bg=GRN if y["thesis"]["ann_ret"]>0 else RED)
    cel(ws2, ri, 8, f"{y['thesis']['max_dd']}%",    bg=bg)
    ws2.row_dimensions[ri].height = 22

cw(ws2, [8,8,11,11,10,11,11,10])

# ━━ SHEET 3: Weight Comparison ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ws3 = wb.create_sheet("Weight Comparison")
title_row(ws3, 1, 5, "Weight Comparison — v2 vs Thesis Portfolio")

for ci, h in enumerate(["Asset","Role","v2 Weight","Thesis Weight","Change"], 1):
    hdr(ws3, 2, ci, h)
ws3.row_dimensions[2].height = 26

ROLES = {
    "xauusd":"Hedge","lly.us":"Carry","wmt.us":"Carry","avgo.us":"Convexity",
    "cost.us":"Carry","ccj.us":"Cyclical","vrt.us":"Convexity","jnj.us":"Carry",
    "xagusd":"Hedge",
}
all_assets = sorted(set(list(W_V2.keys()) + list(W_THESIS.keys())),
                    key=lambda t: -W_THESIS.get(t, 0))

for ri, t in enumerate(all_assets, 3):
    bg    = ALT if ri%2==0 else WHT
    v2w   = W_V2.get(t, 0)
    thw   = W_THESIS.get(t, 0)
    delta = (thw - v2w) * 100
    is_new = t not in W_V2
    is_removed = t not in W_THESIS
    row_bg = "FFF2CC" if is_new else ("FFC7CE" if is_removed else bg)
    cel(ws3, ri, 1, NAMES.get(t,t),               bg=row_bg, bold=is_new or is_removed, ha="left")
    cel(ws3, ri, 2, ROLES.get(t,""),              bg=row_bg)
    cel(ws3, ri, 3, f"{v2w*100:.1f}%" if v2w else "—", bg=row_bg)
    cel(ws3, ri, 4, f"{thw*100:.1f}%" if thw else "REMOVED", bg=row_bg, bold=True)
    cel(ws3, ri, 5, f"{delta:+.1f}pp" if not is_removed else "REMOVED",
        bg=GRN if delta>0 else (RED if is_removed else bg), bold=True)
    ws3.row_dimensions[ri].height = 22

# Cash row
ri = len(all_assets) + 3
cel(ws3, ri, 1, "Cash",  bg=YEL, bold=True, ha="left")
cel(ws3, ri, 2, "—",     bg=YEL)
cel(ws3, ri, 3, "0.0%",  bg=YEL)
cel(ws3, ri, 4, "2.0%",  bg=YEL, bold=True)
cel(ws3, ri, 5, "+2.0pp",bg=YEL, bold=True)
ws3.row_dimensions[ri].height = 22

cw(ws3, [16, 12, 14, 14, 12])

wb.save(REPORT)
print(f"\n  Report saved: {REPORT}")
print("=" * 72)
