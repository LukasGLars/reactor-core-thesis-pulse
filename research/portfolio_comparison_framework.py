import pandas as pd
import numpy as np
import os, sys, io, warnings
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
warnings.filterwarnings("ignore")
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

MACRO_DIR = r"C:\Users\lukas.larsson\Desktop\Privat\macro_data\portfolio_data"
DATA_DIR  = r"C:\Users\lukas.larsson\Desktop\Privat\Project Reactor Core\data"
REPORT    = r"C:\Users\lukas.larsson\Desktop\Privat\Project Reactor Core\Portfolio_Comparison_Framework_v2.xlsx"

# ── Framework thresholds (v4.12) ─────────────────────────────────────────────
SOP_MAX    = 249   # bps  <=249 = SOP
T1_MIN     = 320   # bps  320-359 = T1
T2_MIN     = 360   # bps  >=360 = T2
# CAUTION = 250-319

# ── Simulation parameters ─────────────────────────────────────────────────────
INITIAL_VALUE   = 100_000   # starting invested capital ($)
MONTHLY_CONTRIB =   6_000   # monthly contribution, midpoint of 5-7k
START_DATE      = "2021-04-01"

# ── Portfolio definitions ─────────────────────────────────────────────────────
CURRENT_PORTFOLIO = {
    "acwi.us": 0.70 * 0.82,
    "xmld.uk": 0.15 * 0.82,
    "gld.us":  0.07 * 0.82,
    "ura.us":  0.30 * 0.18,
    "smh.us":  0.20 * 0.18,
    "0rq9.uk": 0.18 * 0.18,
    "paas.us": 0.07 * 0.18,
}
_total = sum(CURRENT_PORTFOLIO.values())
CURRENT_PORTFOLIO = {t: w/_total for t, w in CURRENT_PORTFOLIO.items()}

REACTOR_PORTFOLIO = {
    # Reactor Core v2 — gold capped 25%, Tesla + Lumentum removed, 10Y optimized
    "xauusd": 0.250, "wmt.us": 0.227, "lly.us": 0.197,
    "vrt.us": 0.090, "avgo.us": 0.079, "ccj.us": 0.057,
    "jnj.us": 0.050, "cost.us": 0.050,
}

# ── Load helpers ──────────────────────────────────────────────────────────────
def load_price(ticker):
    for fname in [
        ticker.replace("^","").replace("-","_") + ".csv",
        ticker.replace("^","").replace("-","_").replace(".","_") + ".csv",
    ]:
        path = os.path.join(DATA_DIR, fname)
        if os.path.exists(path):
            df = pd.read_csv(path)
            df["Date"] = pd.to_datetime(df["Date"])
            df = df.set_index("Date").sort_index()
            return df["Close"].replace(0, np.nan).dropna()
    return None

def build_port_returns(weights, start=START_DATE):
    price_dict = {}
    for t in weights:
        s = load_price(t)
        if s is not None:
            s = s[s.index >= start]
            if len(s) > 10:
                price_dict[t] = s
    df = pd.DataFrame(price_dict).ffill().dropna()
    used = list(df.columns)
    w_arr = np.array([weights[t] for t in used])
    w_arr /= w_arr.sum()
    normed = df / df.iloc[0]
    port = normed.dot(w_arr)
    return port.pct_change().dropna()

# ── Load HY-IG spread ─────────────────────────────────────────────────────────
hy = pd.read_csv(os.path.join(MACRO_DIR, "BAMLH0A0HYM2 (3).csv"))
ig = pd.read_csv(os.path.join(MACRO_DIR, "BAMLC0A0CM (4).csv"))
hy["observation_date"] = pd.to_datetime(hy["observation_date"])
ig["observation_date"] = pd.to_datetime(ig["observation_date"])
hy = hy.set_index("observation_date")["BAMLH0A0HYM2"]
ig = ig.set_index("observation_date")["BAMLC0A0CM"]

spread_bps = (hy - ig).dropna() * 100   # % points -> bps

# Forward-fill gap to end of backtest (spread data ends ~Feb 26 2026)
END = pd.Timestamp("2026-04-01")
if spread_bps.index[-1] < END:
    extra = pd.date_range(spread_bps.index[-1] + pd.Timedelta(days=1), END, freq="B")
    spread_bps = pd.concat([spread_bps, pd.Series(spread_bps.iloc[-1], index=extra)])

def get_regime(bps):
    if bps <= SOP_MAX:      return "SOP"
    elif bps < T1_MIN:      return "CAUTION"
    elif bps < T2_MIN:      return "T1"
    else:                   return "T2"

# ── Simulation ────────────────────────────────────────────────────────────────
def simulate(port_rets, label="Portfolio"):
    """
    Runs two simulations side by side:
      - simple_dca  : always invests monthly contribution (no framework)
      - framework   : follows SOP/CAUTION/T1/T2 rules
    Returns (simple_values, framework_values, event_log)
    """
    common = port_rets.index.intersection(spread_bps.index)
    rets   = port_rets[common]
    sp     = spread_bps.reindex(common, method="ffill")

    # State — simple DCA
    sd_invested = float(INITIAL_VALUE)

    # State — framework
    fw_invested = float(INITIAL_VALUE)
    fw_cash     = 0.0

    prev_regime = get_regime(sp.iloc[0])
    prev_month  = rets.index[0].to_period("M")

    # RE-ENTRY state
    re_cash_to_deploy = 0.0   # snapshot of cash at RE-ENTRY trigger
    re_days_done      = 0     # trading days elapsed since RE-ENTRY

    sd_vals = []
    fw_vals = []
    log     = []

    for date, ret in rets.items():
        sp_val  = sp.get(date, sp.iloc[-1])
        regime  = get_regime(sp_val)
        this_m  = date.to_period("M")

        # ── Monthly contribution ──────────────────────────────────────────
        if this_m != prev_month:
            prev_month = this_m

            # Simple DCA: always buy
            sd_invested += MONTHLY_CONTRIB

            # Framework: invest only in SOP
            if regime == "SOP" and re_cash_to_deploy <= 0:
                fw_invested += MONTHLY_CONTRIB
            else:
                fw_cash += MONTHLY_CONTRIB
                if regime != "SOP":
                    log.append({"Date": date.date(), "Event": f"Monthly cash held ({regime})",
                                "Cash_deployed": 0, "Cash_accumulated": round(fw_cash),
                                "Spread_bps": round(sp_val,1), "Regime": regime})

        # ── Regime transition: trigger events ─────────────────────────────
        if prev_regime in ("CAUTION","T1","T2") and regime == "SOP":
            # RE-ENTRY: start deploying 25% of cash per week
            if fw_cash > 0:
                re_cash_to_deploy = fw_cash   # snapshot
                re_days_done      = 0
                log.append({"Date": date.date(), "Event": "RE-ENTRY triggered (25%/week for 4 weeks)",
                            "Cash_deployed": 0, "Cash_accumulated": round(fw_cash),
                            "Spread_bps": round(sp_val,1), "Regime": regime})

        elif prev_regime == "CAUTION" and regime == "T1":
            # T1: deploy 50% of accumulated cash NOW
            deploy = fw_cash * 0.50
            if deploy > 0:
                fw_invested += deploy
                fw_cash     -= deploy
                re_cash_to_deploy = 0  # cancel any pending RE-ENTRY
                log.append({"Date": date.date(), "Event": "T1 — deploy 50% of cash",
                            "Cash_deployed": round(deploy), "Cash_accumulated": round(fw_cash),
                            "Spread_bps": round(sp_val,1), "Regime": regime})

        elif prev_regime in ("CAUTION","T1") and regime == "T2":
            # T2: deploy ALL remaining cash
            deploy = fw_cash
            if deploy > 0:
                fw_invested += deploy
                fw_cash      = 0
                re_cash_to_deploy = 0
                log.append({"Date": date.date(), "Event": "T2 — deploy ALL cash",
                            "Cash_deployed": round(deploy), "Cash_accumulated": 0,
                            "Spread_bps": round(sp_val,1), "Regime": regime})

        # ── RE-ENTRY gradual deployment: 25%/week ─────────────────────────
        if re_cash_to_deploy > 0 and fw_cash > 0:
            re_days_done += 1
            # 4 weeks = 20 trading days, deploy 25% per week = 1/20 per day pro-rated
            deploy_today = min(fw_cash, re_cash_to_deploy / 20)
            fw_invested += deploy_today
            fw_cash     -= deploy_today
            if fw_cash <= 1 or re_days_done >= 20:
                re_cash_to_deploy = 0
                re_days_done      = 0

        # ── Apply daily return to invested portion ────────────────────────
        sd_invested *= (1 + ret)
        fw_invested *= (1 + ret)
        # Cash earns 0%

        sd_vals.append(sd_invested)
        fw_vals.append(fw_invested + fw_cash)

        prev_regime = regime

    idx = rets.index
    return (pd.Series(sd_vals, index=idx),
            pd.Series(fw_vals, index=idx),
            pd.DataFrame(log))

def calc_metrics(val_series):
    """Compute metrics from a total-value series (initial + contributions)."""
    if val_series is None or len(val_series) < 10:
        return {}
    n_years = len(val_series) / 252
    # Daily returns on total value (includes contribution bumps — minor distortion)
    d_rets = val_series.pct_change().dropna()
    # Remove contribution-day spikes > 10% for vol/Sharpe calc
    d_rets_clean = d_rets[d_rets.abs() < 0.10]
    ann_vol = d_rets_clean.std() * np.sqrt(252)
    final   = val_series.iloc[-1]
    total_contrib = INITIAL_VALUE + MONTHLY_CONTRIB * n_years * 12
    profit = final - total_contrib
    roi    = profit / total_contrib
    cagr   = (final / INITIAL_VALUE) ** (1/n_years) - 1  # vs initial only
    # MaxDD on value series
    roll_max = val_series.cummax()
    max_dd   = ((val_series - roll_max) / roll_max).min()
    sharpe   = (cagr - 0) / ann_vol if ann_vol > 0 else 0
    calmar   = cagr / abs(max_dd) if max_dd < 0 else 0
    return {
        "final_value":   round(final),
        "total_contrib": round(total_contrib),
        "profit":        round(profit),
        "roi_pct":       round(roi*100, 1),
        "cagr_pct":      round(cagr*100, 2),
        "ann_vol_pct":   round(ann_vol*100, 2),
        "sharpe":        round(sharpe, 3),
        "max_dd_pct":    round(max_dd*100, 2),
        "calmar":        round(calmar, 3),
    }

# ── Run simulations ───────────────────────────────────────────────────────────
print("Loading portfolio returns...")
rets_current = build_port_returns(CURRENT_PORTFOLIO)
rets_reactor  = build_port_returns(REACTOR_PORTFOLIO)

print("Running simulations...")
cur_simple, cur_fw, cur_log = simulate(rets_current, "Current")
rxr_simple, rxr_fw, rxr_log = simulate(rets_reactor,  "Reactor")

m_cur_simple = calc_metrics(cur_simple)
m_cur_fw     = calc_metrics(cur_fw)
m_rxr_simple = calc_metrics(rxr_simple)
m_rxr_fw     = calc_metrics(rxr_fw)

print(f"\n  {'':30} {'Simple DCA':>14} {'+ Framework':>14}")
print(f"  {'Current Portfolio':30} ${m_cur_simple['final_value']:>13,} ${m_cur_fw['final_value']:>13,}")
print(f"  {'Reactor Core':30} ${m_rxr_simple['final_value']:>13,} ${m_rxr_fw['final_value']:>13,}")
print(f"\n  Total contributed: ${round(INITIAL_VALUE + MONTHLY_CONTRIB * (len(cur_simple)/252) * 12):,}")

# ── Regime calendar ───────────────────────────────────────────────────────────
regime_cal = []
common_idx = rets_current.index.intersection(spread_bps.index)
sp_aligned = spread_bps.reindex(common_idx, method="ffill")
prev = None
seg_start = None
for date, bps in sp_aligned.items():
    r = get_regime(bps)
    if r != prev:
        if prev is not None:
            regime_cal.append({"start": seg_start, "end": date, "regime": prev,
                                "avg_spread": round(sp_aligned[seg_start:date].mean(),1),
                                "days": (date - seg_start).days})
        prev = r; seg_start = date
if prev:
    regime_cal.append({"start": seg_start, "end": sp_aligned.index[-1], "regime": prev,
                        "avg_spread": round(sp_aligned[seg_start:].mean(),1),
                        "days": (sp_aligned.index[-1] - seg_start).days})

# ── Excel output ──────────────────────────────────────────────────────────────
DARK  = "1F4E79"; MID   = "2E75B6"; LIGHT = "D6E4F0"
ALT   = "EBF3FB"; GRN   = "C6EFCE"; RED   = "FFC7CE"
YEL   = "FFEB9C"; ORG   = "FFCC99"; WHT   = "FFFFFF"
GREY  = "F2F2F2"; C_CUR = "FFF2CC"; C_RXR = "E2EFDA"
CAUTION_C = "FFE0CC"; T1_C = "FFD0D0"; T2_C = "FF9999"; SOP_C = "CCFFCC"

thin   = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr(ws, row, col, val, bg=DARK, fg=WHT, bold=True, sz=11, ha="center"):
    c = ws.cell(row=row, column=col, value=val)
    c.fill = PatternFill("solid", fgColor=bg)
    c.font = Font(bold=bold, color=fg, size=sz)
    c.alignment = Alignment(wrap_text=True, vertical="center", horizontal=ha)
    c.border = BORDER; return c

def cel(ws, row, col, val, bg=None, bold=False, ha="center", color="000000", fmt=None):
    c = ws.cell(row=row, column=col, value=val)
    if bg: c.fill = PatternFill("solid", fgColor=bg)
    c.font = Font(bold=bold, color=color)
    c.alignment = Alignment(wrap_text=True, vertical="center", horizontal=ha)
    c.border = BORDER
    if fmt: c.number_format = fmt
    return c

def cols(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def title_row(ws, row, ncols, text, bg=DARK, sz=13):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(bold=True, color=WHT, size=sz)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 34

wb = openpyxl.Workbook()

# ━━ SHEET 1: Head-to-Head ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ws1 = wb.active; ws1.title = "Framework Head-to-Head"
title_row(ws1, 1, 5, "Portfolio Comparison — Framework-Managed vs Simple DCA")

ws1.merge_cells("A2:E2")
note = (f"Start: {START_DATE}  |  Initial invested: ${INITIAL_VALUE:,}  |  "
        f"Monthly contribution: ${MONTHLY_CONTRIB:,}  |  Cash earns 0%  |  "
        f"Framework v4.12: SOP ≤{SOP_MAX} / CAUTION 250-{T1_MIN-1} / T1 {T1_MIN}-{T2_MIN-1} / T2 ≥{T2_MIN} bps")
c = ws1["A2"]; c.value = note
c.font = Font(italic=True, size=9, color="444444")
c.fill = PatternFill("solid", fgColor=LIGHT)
c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws1.row_dimensions[2].height = 28

for ci, (h, bg) in enumerate([
    ("Metric",              GREY),
    ("Current — Simple DCA", C_CUR),
    ("Current + Framework",  C_CUR),
    ("Reactor — Simple DCA", C_RXR),
    ("Reactor + Framework",  C_RXR),
], 1):
    hdr(ws1, 3, ci, h, bg=bg, fg="000000", sz=10)
ws1.row_dimensions[3].height = 30

metrics_rows = [
    ("Final Portfolio Value ($)", "final_value",   True,  lambda v: f"${v:,.0f}"),
    ("Total Contributed ($)",     "total_contrib",  None,  lambda v: f"${v:,.0f}"),
    ("Total Profit ($)",          "profit",         True,  lambda v: f"${v:,.0f}"),
    ("Return on Investment",      "roi_pct",        True,  lambda v: f"{v:.1f}%"),
    ("CAGR (vs initial capital)", "cagr_pct",       True,  lambda v: f"{v:.2f}%"),
    ("Ann. Volatility",           "ann_vol_pct",    False, lambda v: f"{v:.2f}%"),
    ("Sharpe Ratio",              "sharpe",         True,  lambda v: f"{v:.3f}"),
    ("Max Drawdown",              "max_dd_pct",     False, lambda v: f"{v:.2f}%"),
    ("Calmar Ratio",              "calmar",         True,  lambda v: f"{v:.3f}"),
]

def best_bg(vals, higher_better):
    if higher_better is None: return [None]*len(vals)
    if higher_better:
        best = max(vals)
        return [GRN if v == best else RED for v in vals]
    else:
        best = min(vals)
        return [GRN if v == best else RED for v in vals]

for ri, (label, key, hb, fmt) in enumerate(metrics_rows, 4):
    bg_row = ALT if ri % 2 == 0 else WHT
    cel(ws1, ri, 1, label, bg=GREY, bold=True, ha="left")
    vals = [m_cur_simple.get(key), m_cur_fw.get(key),
            m_rxr_simple.get(key), m_rxr_fw.get(key)]
    bgs  = best_bg(vals, hb)
    for ci, (v, b) in enumerate(zip(vals, bgs), 2):
        cel(ws1, ri, ci, fmt(v) if v is not None else "N/A",
            bg=b or bg_row, bold=(b==GRN))
    ws1.row_dimensions[ri].height = 22

# Framework advantage rows
ri = len(metrics_rows) + 5
ws1.merge_cells(f"A{ri}:E{ri}")
c = ws1.cell(row=ri, column=1, value="Framework Advantage (Framework vs Simple DCA)")
c.font = Font(bold=True, color=WHT, size=11)
c.fill = PatternFill("solid", fgColor=MID)
c.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[ri].height = 26
ri += 1

for ci, h in enumerate(["Metric","—","Current gain","—","Reactor gain"], 1):
    hdr(ws1, ri, ci, h, bg=GREY if ci in (2,4) else MID, fg="000000" if ci in (2,4) else WHT)
ws1.row_dimensions[ri].height = 26
ri += 1

adv_rows = [
    ("Extra final value", "final_value",   lambda v: f"+${v:,.0f}"),
    ("Extra profit",      "profit",        lambda v: f"+${v:,.0f}"),
    ("ROI improvement",   "roi_pct",       lambda v: f"+{v:.1f}pp"),
    ("Sharpe improvement","sharpe",        lambda v: f"+{v:.3f}"),
    ("MaxDD improvement", "max_dd_pct",    lambda v: f"{v:.2f}pp"),
]
for label, key, fmt_fn in adv_rows:
    bg_row = ALT if ri % 2 == 0 else WHT
    cur_adv = (m_cur_fw.get(key,0) or 0) - (m_cur_simple.get(key,0) or 0)
    rxr_adv = (m_rxr_fw.get(key,0) or 0) - (m_rxr_simple.get(key,0) or 0)
    cel(ws1, ri, 1, label, bg=GREY, bold=True, ha="left")
    cel(ws1, ri, 2, "—", bg=bg_row)
    cel(ws1, ri, 3, fmt_fn(cur_adv), bg=GRN if cur_adv > 0 else RED, bold=True)
    cel(ws1, ri, 4, "—", bg=bg_row)
    cel(ws1, ri, 5, fmt_fn(rxr_adv), bg=GRN if rxr_adv > 0 else RED, bold=True)
    ws1.row_dimensions[ri].height = 22
    ri += 1

cols(ws1, [28, 20, 20, 20, 20])

# ━━ SHEET 2: Regime Calendar ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ws2 = wb.create_sheet("Regime Calendar")
title_row(ws2, 1, 6, "HY-IG Spread Regime History — Framework Trigger Log")

for ci, h in enumerate(["Start","End","Regime","Avg Spread (bps)","Duration (days)","Framework Action"], 1):
    hdr(ws2, 2, ci, h)
ws2.row_dimensions[2].height = 28

REGIME_COLORS = {"SOP": SOP_C, "CAUTION": CAUTION_C, "T1": T1_C, "T2": T2_C}
REGIME_ACTIONS = {
    "SOP":     "Deploy monthly contribution into portfolio",
    "CAUTION": "Hold monthly contributions as cash — no new investment",
    "T1":      "Deploy 50% of accumulated cash into portfolio",
    "T2":      "Deploy ALL remaining cash into portfolio",
}

for ri, row in enumerate(regime_cal, 3):
    rg = row["regime"]
    bg = REGIME_COLORS.get(rg, WHT)
    cel(ws2, ri, 1, str(row["start"].date()), bg=bg)
    cel(ws2, ri, 2, str(row["end"].date()),   bg=bg)
    cel(ws2, ri, 3, rg,                        bg=bg, bold=True)
    cel(ws2, ri, 4, row["avg_spread"],          bg=bg)
    cel(ws2, ri, 5, row["days"],                bg=bg)
    cel(ws2, ri, 6, REGIME_ACTIONS.get(rg,""), bg=bg, ha="left")
    ws2.row_dimensions[ri].height = 20

cols(ws2, [14, 14, 12, 18, 16, 45])

# ━━ SHEET 3: Event Log ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ws3 = wb.create_sheet("Event Log")
title_row(ws3, 1, 6, "Framework Deployment Events — When Cash Was Held or Deployed")

# Combine logs
all_log = []
for row in cur_log.to_dict("records"):
    row["portfolio"] = "Current"
    all_log.append(row)
for row in rxr_log.to_dict("records"):
    row["portfolio"] = "Reactor"
    all_log.append(row)
all_log.sort(key=lambda x: str(x["Date"]))

for ci, h in enumerate(["Date","Portfolio","Event","Cash Deployed ($)","Cash Accumulated ($)","Spread (bps)"], 1):
    hdr(ws3, 2, ci, h)
ws3.row_dimensions[2].height = 28

for ri, row in enumerate(all_log, 3):
    ev  = row.get("Event","")
    bg  = WHT
    if "T2"        in ev: bg = T2_C
    elif "T1"      in ev: bg = T1_C
    elif "RE-ENTRY" in ev: bg = SOP_C
    elif "held"    in ev: bg = CAUTION_C
    alt = ALT if ri % 2 == 0 else bg
    cel(ws3, ri, 1, str(row.get("Date","")),           bg=alt)
    cel(ws3, ri, 2, row.get("portfolio",""),            bg=alt, bold=True)
    cel(ws3, ri, 3, ev,                                 bg=bg,  ha="left")
    cd = row.get("Cash_deployed",0) or 0
    ca = row.get("Cash_accumulated",0) or 0
    cel(ws3, ri, 4, f"${cd:,.0f}" if cd else "—",      bg=alt)
    cel(ws3, ri, 5, f"${ca:,.0f}",                      bg=alt)
    cel(ws3, ri, 6, row.get("Spread_bps",""),           bg=alt)
    ws3.row_dimensions[ri].height = 18

cols(ws3, [13, 12, 48, 16, 18, 14])

# ━━ SHEET 4: Methodology ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ws4 = wb.create_sheet("Methodology")
title_row(ws4, 1, 2, "Simulation Methodology & Assumptions")

notes = [
    ("Parameter", "Value / Assumption"),
    ("Start date",               START_DATE),
    ("Initial invested capital", f"${INITIAL_VALUE:,}"),
    ("Monthly contribution",     f"${MONTHLY_CONTRIB:,} (midpoint of 5-7k framework guideline)"),
    ("Cash return",              "0% — conservative; ignores MMF/HYSA yield (~4-5% in 2022-2024)"),
    ("Spread source",            "FRED BAMLH0A0HYM2 - BAMLC0A0CM (OAS, % pts * 100 = bps)"),
    ("Spread gap handling",      "Forward-filled from last known value (2026-02-26) to 2026-04-01"),
    ("Framework version",        "v4.12: SOP ≤249 / CAUTION 250-319 / T1 320-359 / T2 ≥360 bps"),
    ("SOP rule",                 "Monthly contribution invested immediately in portfolio"),
    ("CAUTION rule",             "Monthly contribution held as cash — existing holdings untouched"),
    ("T1 rule",                  "Deploy 50% of accumulated cash on first day spread ≥320 bps (from CAUTION)"),
    ("T2 rule",                  "Deploy ALL remaining cash on first day spread ≥360 bps"),
    ("RE-ENTRY rule",            "On first day spread returns ≤249 from higher regime: deploy 25% of cash/week over 4 weeks"),
    ("Portfolio proxies (Current)", "UK ETFs (XMLD, 0RQ9) loaded from Reactor data dir; US proxies used for EU ETFs"),
    ("CAGR calculation",         "Annualised return of final_value vs INITIAL_VALUE only (not total contributions)"),
    ("Sharpe calculation",       "CAGR / ann_vol; contribution-day return spikes >10% excluded from vol calc"),
    ("No transaction costs",     "Bid/ask spreads, FX costs, and broker fees not modelled"),
    ("No tax",                   "Capital gains tax not modelled"),
]

hdr(ws4, 2, 1, "Parameter", sz=10)
hdr(ws4, 2, 2, "Value / Assumption", sz=10)
ws4.row_dimensions[2].height = 24

for ri, (k, v) in enumerate(notes[1:], 3):
    bg = ALT if ri % 2 == 0 else WHT
    cel(ws4, ri, 1, k, bg=GREY, bold=True, ha="left")
    cel(ws4, ri, 2, v, bg=bg, ha="left")
    ws4.row_dimensions[ri].height = 20

cols(ws4, [32, 70])

wb.save(REPORT)
print(f"\nReport saved: {REPORT}")
