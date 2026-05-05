import pandas as pd
import numpy as np
import os, sys, io, warnings
from scipy.optimize import minimize
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
warnings.filterwarnings("ignore")
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

DATA_DIR = r"C:\Users\lukas.larsson\Desktop\Privat\Project Reactor Core\data"
REPORT   = r"C:\Users\lukas.larsson\Desktop\Privat\Project Reactor Core\Admission_Test.xlsx"

# ── Portfolio after cuts (Tesla + Lumentum removed, gold capped at 35%) ───────
BASE_TICKERS = ["xauusd","lly.us","wmt.us","avgo.us","cost.us","ccj.us","vrt.us","jnj.us"]
CANDIDATES   = {
    "xagusd":  {"name": "Silver",  "role": "Hedge",      "corr_gold_threshold": 0.80},
    "asml.us": {"name": "ASML",    "role": "Convexity",  "corr_gold_threshold": 0.99},
    "copx.us": {"name": "COPX",    "role": "Cyclical",   "corr_gold_threshold": 0.60},
}
NAMES = {
    "xauusd":"Gold","lly.us":"Eli Lilly","wmt.us":"Walmart",
    "avgo.us":"Broadcom","cost.us":"Costco","ccj.us":"Cameco",
    "vrt.us":"Vertiv","jnj.us":"J&J",
    "xagusd":"Silver","asml.us":"ASML","copx.us":"COPX",
}
REGIMES = {
    "Pre-COVID Bull":      ("2016-04-01","2020-01-31"),
    "COVID Crash":         ("2020-02-01","2020-03-31"),
    "COVID Recovery":      ("2020-04-01","2021-12-31"),
    "Rate Hike/Inflation": ("2022-01-01","2023-07-31"),
    "Post-Hike/AI Bull":   ("2023-08-01","2024-08-31"),
    "Rate Cut":            ("2024-09-01","2026-04-01"),
}
WINDOWS  = {"3Y":"2023-04-01","5Y":"2021-04-01","10Y":"2016-04-01"}
GOLD_CAP = 0.35
MIN_W    = 0.05
MAX_W    = 0.40

# ── Helpers ───────────────────────────────────────────────────────────────────
def load(ticker):
    fname = ticker.replace("^","").replace("-","_") + ".csv"
    path  = os.path.join(DATA_DIR, fname)
    if not os.path.exists(path): return None
    df = pd.read_csv(path)
    df["Date"] = pd.to_datetime(df["Date"])
    df = df.set_index("Date").sort_index()
    return df["Close"].replace(0, np.nan).dropna()

def build_df(tickers, start):
    d = {t: load(t) for t in tickers if load(t) is not None}
    df = pd.DataFrame({t: s[s.index >= start] for t, s in d.items()}).dropna()
    return df

def optimize(df, gold_cap=GOLD_CAP, n=80):
    rets = df.pct_change().dropna()
    mu   = rets.mean(); cov = rets.cov()
    n_a  = len(df.columns)
    bounds = [(MIN_W, min(MAX_W, gold_cap) if t == "xauusd" else MAX_W)
              for t in df.columns]
    def neg_sharpe(w):
        r = np.dot(w, mu)*252
        v = np.sqrt(w @ cov @ w)*np.sqrt(252)
        return -r/v if v > 0 else 0
    constraints = [{"type":"eq","fun":lambda w: np.sum(w)-1}]
    best = None
    for _ in range(n):
        w0 = np.random.dirichlet(np.ones(n_a))
        w0 = np.clip(w0, [b[0] for b in bounds], [b[1] for b in bounds])
        w0 /= w0.sum()
        res = minimize(neg_sharpe, w0, method="SLSQP", bounds=bounds,
                       constraints=constraints, options={"maxiter":1000,"ftol":1e-12})
        if res.success and (best is None or res.fun < best.fun):
            best = res
    return best.x if best else None

def metrics(df, w):
    wa   = np.array(w)
    rets = df.pct_change().dropna()
    mu   = rets.mean(); cov = rets.cov()
    ann_ret = np.dot(wa, mu)*252
    ann_vol = np.sqrt(wa @ cov @ wa)*np.sqrt(252)
    sharpe  = ann_ret/ann_vol if ann_vol > 0 else 0
    port    = (df/df.iloc[0]).dot(wa)
    max_dd  = ((port - port.cummax())/port.cummax()).min()
    calmar  = ann_ret/abs(max_dd) if max_dd < 0 else 0
    total   = port.iloc[-1]-1
    return {"sharpe":round(sharpe,3),"ann_ret":round(ann_ret*100,2),
            "ann_vol":round(ann_vol*100,2),"max_dd":round(max_dd*100,2),
            "calmar":round(calmar,3),"total":round(total*100,1),
            "weights": dict(zip(df.columns, wa.round(4)))}

def regime_ret(ticker, rstart, rend):
    s = load(ticker)
    if s is None: return None
    s = s.loc[rstart:rend].dropna()
    if len(s) < 5: return None
    return round((s.iloc[-1]/s.iloc[0]-1)*100, 1)

# ── Baselines ─────────────────────────────────────────────────────────────────
cand_names = ", ".join(v["name"] for v in CANDIDATES.values())
print("=" * 65)
print(f"  Admission Test — {cand_names}")
print(f"  Base portfolio: 8 positions (post-cut), gold cap {GOLD_CAP*100:.0f}%")
print("=" * 65)

baselines = {}
for wname, wstart in WINDOWS.items():
    df  = build_df(BASE_TICKERS, wstart)
    w   = optimize(df)
    m   = metrics(df, w)
    baselines[wname] = m
    print(f"  Baseline {wname}:  Sharpe {m['sharpe']}  AnnRet {m['ann_ret']}%  "
          f"MaxDD {m['max_dd']}%  Gold {m['weights'].get('xauusd',0)*100:.1f}%")

# ── Admission test per candidate ──────────────────────────────────────────────
results = {}

for cand_t, cand_info in CANDIDATES.items():
    print(f"\n  --- Testing: {cand_info['name']} ({cand_t}) ---")
    test_tickers = BASE_TICKERS + [cand_t]
    row = {"ticker": cand_t, "name": cand_info["name"], "role": cand_info["role"]}

    # ΔSharpe across all windows
    delta_sharpes = {}
    window_metrics = {}
    for wname, wstart in WINDOWS.items():
        df_b = build_df(BASE_TICKERS, wstart)
        df_t = build_df(test_tickers, wstart)
        if cand_t not in df_t.columns:
            print(f"    {wname}: candidate data missing"); continue
        wb  = optimize(df_b)
        wt  = optimize(df_t)
        mb  = metrics(df_b, wb)
        mt  = metrics(df_t, wt)
        ds  = round(mt["sharpe"] - mb["sharpe"], 3)
        delta_sharpes[wname] = ds
        window_metrics[wname] = {"base": mb, "test": mt, "delta": ds,
                                  "cand_weight": round(mt["weights"].get(cand_t,0)*100,1)}
        print(f"    {wname}: base Sharpe {mb['sharpe']}  +candidate {mt['sharpe']}  "
              f"dS={ds:+.3f}  cand_w={mt['weights'].get(cand_t,0)*100:.1f}%  "
              f"MaxDD {mt['max_dd']}%  Gold {mt['weights'].get('xauusd',0)*100:.1f}%")

    # Correlation analysis (10Y)
    df_10y = build_df(BASE_TICKERS + [cand_t], "2016-04-01")
    if cand_t in df_10y.columns:
        rets_10y = df_10y.pct_change().dropna()
        corr_to_gold = round(rets_10y[cand_t].corr(rets_10y["xauusd"]), 3)
        # Portfolio returns (base weights)
        w_base_10y = optimize(build_df(BASE_TICKERS, "2016-04-01"))
        base_df_10y = build_df(BASE_TICKERS, "2016-04-01")
        port_rets = base_df_10y.pct_change().dropna().dot(w_base_10y)
        corr_to_port = round(rets_10y[cand_t].corr(
            port_rets.reindex(rets_10y.index).dropna()), 3)

        # Vol per unit return (does it improve efficiency?)
        base_w_10 = optimize(build_df(BASE_TICKERS,"2016-04-01"))
        test_w_10 = optimize(df_10y)
        mb10 = metrics(build_df(BASE_TICKERS,"2016-04-01"), base_w_10)
        mt10 = metrics(df_10y, test_w_10)
        vol_per_ret_base = round(mb10["ann_vol"]/mb10["ann_ret"], 3) if mb10["ann_ret"] > 0 else None
        vol_per_ret_test = round(mt10["ann_vol"]/mt10["ann_ret"], 3) if mt10["ann_ret"] > 0 else None
    else:
        corr_to_gold = corr_to_port = vol_per_ret_base = vol_per_ret_test = None

    print(f"    Corr to Gold: {corr_to_gold}  (threshold: <{cand_info['corr_gold_threshold']})")
    print(f"    Corr to Port: {corr_to_port}")
    print(f"    Vol/AnnRet (base→test): {vol_per_ret_base} → {vol_per_ret_test}")

    # Regime wins (candidate individually)
    regime_rets = {}
    regime_wins = 0
    for rname, (rs, re) in REGIMES.items():
        r = regime_ret(cand_t, rs, re)
        regime_rets[rname] = r
        if r is not None and r > 0: regime_wins += 1

    print(f"    Regime wins: {regime_wins}/6")
    for rn, rv in regime_rets.items():
        flag = "+" if rv and rv > 0 else ""
        print(f"      {rn:<24} {flag}{rv}%" if rv is not None else f"      {rn:<24} N/A")

    # ── Admission criteria evaluation ─────────────────────────────────────────
    # 1. ΔSharpe ≥ +0.03 in ALL windows
    sharpe_pass = all(v >= 0.03 for v in delta_sharpes.values() if v is not None)

    # 2. Correlation: reduces vol per unit return AND corr < threshold
    corr_pass = (corr_to_gold is not None and
                 corr_to_gold < cand_info["corr_gold_threshold"] and
                 vol_per_ret_test is not None and
                 vol_per_ret_base is not None and
                 vol_per_ret_test <= vol_per_ret_base)

    # 3. Regime wins ≥ 4/6
    regime_pass = regime_wins >= 4

    # 4. System improvement: ≥ 3 of 4 sub-criteria
    #    a. Sharpe ↑ (10Y)
    #    b. MaxDD not worse by > 2pp
    #    c. Improves ≥4/6 regimes (same as regime_pass)
    #    d. Reduces gold concentration
    ds10 = delta_sharpes.get("10Y", 0)
    mt10 = window_metrics.get("10Y", {}).get("test", {})
    mb10 = window_metrics.get("10Y", {}).get("base", {})
    sys_a = ds10 > 0
    sys_b = (mt10.get("max_dd",0) - mb10.get("max_dd",0)) <= 2.0
    sys_c = regime_wins >= 4
    gold_w_base = baselines.get("10Y",{}).get("weights",{}).get("xauusd",1)
    gold_w_test = window_metrics.get("10Y",{}).get("test",{}).get("weights",{}).get("xauusd",1) if window_metrics.get("10Y") else 1
    sys_d = gold_w_test < gold_w_base
    sys_score = sum([sys_a, sys_b, sys_c, sys_d])
    system_pass = sys_score >= 3

    criteria = sum([sharpe_pass, corr_pass, regime_pass, system_pass])
    verdict  = "ADMIT" if criteria >= 3 else ("CONDITIONAL" if criteria == 2 else "REJECT")

    print(f"\n    CRITERIA SUMMARY:")
    print(f"      1. ΔSharpe ≥+0.03 all windows: {'PASS' if sharpe_pass else 'FAIL'}  {delta_sharpes}")
    print(f"      2. Correlation impact:          {'PASS' if corr_pass else 'FAIL'}  "
          f"corrGold={corr_to_gold} vol/ret {vol_per_ret_base}->{vol_per_ret_test}")
    print(f"      3. Regime wins ≥4/6:            {'PASS' if regime_pass else 'FAIL'}  {regime_wins}/6")
    print(f"      4. System improvement ≥3/4:     {'PASS' if system_pass else 'FAIL'}  "
          f"[Sharpe↑:{sys_a} | DD ok:{sys_b} | Regimes:{sys_c} | GoldRedux:{sys_d}]")
    print(f"    => {criteria}/4 criteria  VERDICT: {verdict}")

    results[cand_t] = {
        "name": cand_info["name"], "role": cand_info["role"],
        "delta_sharpes": delta_sharpes,
        "window_metrics": window_metrics,
        "corr_gold": corr_to_gold, "corr_port": corr_to_port,
        "vol_per_ret_base": vol_per_ret_base, "vol_per_ret_test": vol_per_ret_test,
        "regime_wins": regime_wins, "regime_rets": regime_rets,
        "sharpe_pass": sharpe_pass, "corr_pass": corr_pass,
        "regime_pass": regime_pass, "system_pass": system_pass,
        "sys_detail": {"sharpe_up": sys_a, "dd_ok": sys_b, "regimes": sys_c, "gold_redux": sys_d},
        "criteria": criteria, "verdict": verdict,
    }

# ── Final portfolio (8 base + all candidates) ─────────────────────────────────
ALL_TICKERS = BASE_TICKERS + list(CANDIDATES.keys())
print(f"\n{'=' * 65}")
print(f"  Final portfolio: 8 base + {cand_names} ({len(ALL_TICKERS)} positions)")
print("=" * 65)

final_metrics = {}
for wname, wstart in WINDOWS.items():
    df  = build_df(ALL_TICKERS, wstart)
    w   = optimize(df)
    m   = metrics(df, w)
    final_metrics[wname] = m
    ws  = {t: round(v*100,1) for t,v in m["weights"].items()}
    cand_str = "  ".join(f"{NAMES.get(ct,ct)} {ws.get(ct,0)}%" for ct in CANDIDATES)
    print(f"  {wname}: Sharpe {m['sharpe']}  AnnRet {m['ann_ret']}%  "
          f"MaxDD {m['max_dd']}%  Gold {ws.get('xauusd',0)}%  {cand_str}")
    print(f"       Weights: " + "  ".join(f"{NAMES.get(t,t)} {v}%" for t,v in
          sorted(ws.items(), key=lambda x: x[1], reverse=True)))

# ── Excel output ──────────────────────────────────────────────────────────────
DARK="1F4E79"; MID="2E75B6"; LIGHT="D6E4F0"; ALT="EBF3FB"
GRN="C6EFCE"; RED="FFC7CE"; YEL="FFEB9C"; WHT="FFFFFF"; GREY="F2F2F2"
ADMIT_C="C6EFCE"; COND_C="FFEB9C"; REJECT_C="FFC7CE"
PASS_C=GRN; FAIL_C=RED

thin   = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr(ws, row, col, val, bg=DARK, fg=WHT, bold=True, sz=10, ha="center"):
    c = ws.cell(row=row, column=col, value=val)
    c.fill = PatternFill("solid", fgColor=bg); c.font = Font(bold=bold, color=fg, size=sz)
    c.alignment = Alignment(wrap_text=True, vertical="center", horizontal=ha)
    c.border = BORDER; return c

def cel(ws, row, col, val, bg=None, bold=False, ha="center", color="000000"):
    c = ws.cell(row=row, column=col, value=val)
    if bg: c.fill = PatternFill("solid", fgColor=bg)
    c.font = Font(bold=bold, color=color)
    c.alignment = Alignment(wrap_text=True, vertical="center", horizontal=ha)
    c.border = BORDER; return c

def title(ws, row, nc, text, bg=DARK, sz=13):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=nc)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(bold=True, color=WHT, size=sz)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 34

def cols(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

wb = openpyxl.Workbook()

# ━━ SHEET 1: Admission Scorecard ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ws1 = wb.active; ws1.title = "Admission Scorecard"
title(ws1, 1, 11, f"Admission Test — {cand_names} vs Reactor Core Criteria")

ws1.merge_cells("A2:K2")
c = ws1["A2"]
c.value = (f"Base: 8-position portfolio (Tesla + Lumentum removed, gold capped {GOLD_CAP*100:.0f}%)  |  "
           f"Criteria: ΔSharpe ≥+0.03 all windows | Corr to gold <threshold | Regime wins ≥4/6 | System improvement ≥3/4")
c.font = Font(italic=True, size=9, color="444444")
c.fill = PatternFill("solid", fgColor=LIGHT)
c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws1.row_dimensions[2].height = 28

hdrs1 = ["Candidate","Role",
         "ΔSharpe 3Y","ΔSharpe 5Y","ΔSharpe 10Y",
         "Corr Gold","Criterion 1\nΔSharpe","Criterion 2\nCorrelation",
         "Criterion 3\nRegimes","Criterion 4\nSystem","Verdict"]
for ci, h in enumerate(hdrs1, 1): hdr(ws1, 3, ci, h)
ws1.row_dimensions[3].height = 36

for ri, (ct, r) in enumerate(results.items(), 4):
    bg = ALT if ri%2==0 else WHT
    vbg = ADMIT_C if r["verdict"]=="ADMIT" else (COND_C if r["verdict"]=="CONDITIONAL" else REJECT_C)

    def pf(passed): return (PASS_C, "PASS") if passed else (FAIL_C, "FAIL")

    cel(ws1, ri, 1,  r["name"], bg=bg, bold=True, ha="left")
    cel(ws1, ri, 2,  r["role"], bg=bg)
    for ci_off, wn in enumerate(["3Y","5Y","10Y"], 3):
        ds = r["delta_sharpes"].get(wn)
        ds_bg = GRN if ds is not None and ds >= 0.03 else (RED if ds is not None and ds < 0 else YEL)
        cel(ws1, ri, ci_off, f"{ds:+.3f}" if ds is not None else "—", bg=ds_bg)
    cel(ws1, ri, 6, r["corr_gold"], bg=GRN if r["corr_gold"] and r["corr_gold"]<0.80 else RED)
    b, t = pf(r["sharpe_pass"]);  cel(ws1, ri, 7,  t, bg=b, bold=True)
    b, t = pf(r["corr_pass"]);    cel(ws1, ri, 8,  t, bg=b, bold=True)
    b, t = pf(r["regime_pass"]);  cel(ws1, ri, 9,  t, bg=b, bold=True)
    b, t = pf(r["system_pass"]); cel(ws1, ri, 10, t, bg=b, bold=True)
    cel(ws1, ri, 11, r["verdict"], bg=vbg, bold=True)
    ws1.row_dimensions[ri].height = 22

cols(ws1, [12,12,11,11,11,11,13,13,13,13,13])

# ━━ SHEET 2: Window Metrics Detail ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ws2 = wb.create_sheet("Window Metrics")
title(ws2, 1, 11, "Detailed Metrics — Base vs +Candidate Across Windows")

hdrs2 = ["Candidate","Window",
          "Base Sharpe","Test Sharpe","ΔSharpe",
          "Base AnnRet","Test AnnRet",
          "Base MaxDD","Test MaxDD","ΔMaxDD",
          "Cand Weight"]
for ci, h in enumerate(hdrs2, 1): hdr(ws2, 2, ci, h)
ws2.row_dimensions[2].height = 30

ri = 3
for ct, r in results.items():
    for wname in ["3Y","5Y","10Y"]:
        wm = r["window_metrics"].get(wname)
        if not wm: continue
        bg = ALT if ri%2==0 else WHT
        ds = wm["delta"]
        dd_delta = round(wm["test"]["max_dd"] - wm["base"]["max_dd"], 2)
        cel(ws2, ri, 1,  r["name"], bg=bg, bold=True)
        cel(ws2, ri, 2,  wname, bg=bg)
        cel(ws2, ri, 3,  wm["base"]["sharpe"], bg=bg)
        cel(ws2, ri, 4,  wm["test"]["sharpe"], bg=GRN if ds>=0.03 else (YEL if ds>=0 else RED))
        cel(ws2, ri, 5,  f"{ds:+.3f}", bg=GRN if ds>=0.03 else (YEL if ds>=0 else RED), bold=True)
        cel(ws2, ri, 6,  f"{wm['base']['ann_ret']}%", bg=bg)
        cel(ws2, ri, 7,  f"{wm['test']['ann_ret']}%", bg=bg)
        cel(ws2, ri, 8,  f"{wm['base']['max_dd']}%", bg=bg)
        cel(ws2, ri, 9,  f"{wm['test']['max_dd']}%", bg=GRN if dd_delta<=0 else RED)
        cel(ws2, ri, 10, f"{dd_delta:+.2f}pp", bg=GRN if dd_delta<=0 else RED)
        cel(ws2, ri, 11, f"{wm['cand_weight']}%", bg=bg, bold=True)
        ws2.row_dimensions[ri].height = 20
        ri += 1
    ri += 1  # spacer

cols(ws2, [12,8,12,12,10,12,12,12,12,11,12])

# ━━ SHEET 3: Regime Analysis ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ws3 = wb.create_sheet("Regime Analysis")
cand_list  = list(CANDIDATES.keys())
ncols_ws3  = 1 + len(cand_list) + 2 + len(cand_list)  # regime + cands + gold+port + wins
title(ws3, 1, ncols_ws3, "Candidate Regime Performance vs Gold & Portfolio Benchmark")

hdrs3 = (["Regime"] + [NAMES.get(ct, ct) for ct in cand_list] +
         ["Gold","Portfolio (10Y)"] +
         [f"{NAMES.get(ct,ct)} Wins" for ct in cand_list])
for ci, h in enumerate(hdrs3, 1): hdr(ws3, 2, ci, h)
ws3.row_dimensions[2].height = 28

df_base_10y   = build_df(BASE_TICKERS, "2016-04-01")
w_base_10y    = optimize(df_base_10y)
port_base_10y = (df_base_10y/df_base_10y.iloc[0]).dot(w_base_10y)

for ri, (rname, (rs, re)) in enumerate(REGIMES.items(), 3):
    bg = ALT if ri%2==0 else WHT
    cel(ws3, ri, 1, rname, bg=GREY, bold=True, ha="left")
    cand_rets = [results[ct]["regime_rets"].get(rname) for ct in cand_list]
    gold_r    = regime_ret("xauusd", rs, re)
    port_s    = port_base_10y.loc[rs:re]
    port_r    = round((port_s.iloc[-1]/port_s.iloc[0]-1)*100,1) if len(port_s)>5 else None
    for ci, val in enumerate(cand_rets + [gold_r, port_r], 2):
        vbg = GRN if val is not None and val > 0 else (RED if val is not None and val < 0 else bg)
        cel(ws3, ri, ci, f"{val:+.1f}%" if val is not None else "N/A", bg=vbg)
    win_col_start = 2 + len(cand_list) + 2
    for ci_off, (ct, cr) in enumerate(zip(cand_list, cand_rets)):
        cel(ws3, ri, win_col_start + ci_off,
            "+" if cr and cr > 0 else "-",
            bg=GRN if cr and cr > 0 else RED, bold=True)
    ws3.row_dimensions[ri].height = 20

ri_s = len(REGIMES) + 4
cel(ws3, ri_s, 1, "Total wins (>0)", bg=GREY, bold=True, ha="left")
for ci_off, ct in enumerate(cand_list, 2):
    wins = results[ct]["regime_wins"]
    cel(ws3, ri_s, ci_off, f"{wins}/6", bg=GRN if wins>=4 else RED, bold=True)
ws3.row_dimensions[ri_s].height = 22

ws3_widths = [24] + [14]*len(cand_list) + [14, 18] + [14]*len(cand_list)
cols(ws3, ws3_widths)

# ━━ SHEET 4: Final 10-Position Portfolio ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ws4 = wb.create_sheet("Final 10-Position Portfolio")
title(ws4, 1, 6, f"Final Portfolio — 8 Base + {cand_names} (Gold cap {GOLD_CAP*100:.0f}%)")

for ci, h in enumerate(["Window","Sharpe","Ann Return","Max DD","Calmar","Total Return"], 1):
    hdr(ws4, 2, ci, h)
ws4.row_dimensions[2].height = 26

for ri, (wname, m) in enumerate(final_metrics.items(), 3):
    bg = ALT if ri%2==0 else WHT
    cel(ws4, ri, 1, wname, bg=GREY, bold=True)
    cel(ws4, ri, 2, m["sharpe"],         bg=bg, bold=True)
    cel(ws4, ri, 3, f"{m['ann_ret']}%",  bg=bg)
    cel(ws4, ri, 4, f"{m['max_dd']}%",   bg=bg)
    cel(ws4, ri, 5, m["calmar"],         bg=bg)
    cel(ws4, ri, 6, f"{m['total']}%",    bg=bg)
    ws4.row_dimensions[ri].height = 22

# Final weights (10Y)
ri = len(final_metrics) + 5
ws4.merge_cells(f"A{ri}:F{ri}")
c = ws4.cell(row=ri, column=1, value="Optimized Weights — 10Y Window")
c.font = Font(bold=True, color=WHT, size=11); c.fill = PatternFill("solid", fgColor=MID)
c.alignment = Alignment(horizontal="center", vertical="center")
ws4.row_dimensions[ri].height = 26; ri += 1

for ci, h in enumerate(["Ticker","Name","Role","Weight","vs Current",""], 1):
    hdr(ws4, ri, ci, h)
ws4.row_dimensions[ri].height = 24; ri += 1

CURRENT_W = {"xauusd":0.379,"lly.us":0.156,"wmt.us":0.114,
             "avgo.us":0.050,"cost.us":0.050,"ccj.us":0.050,
             "vrt.us":0.050,"jnj.us":0.050}
ROLES_ALL  = {"xauusd":"Hedge","lly.us":"Carry","wmt.us":"Carry","avgo.us":"Convexity",
              "cost.us":"Carry","ccj.us":"Cyclical","vrt.us":"Convexity","jnj.us":"Carry",
              "xagusd":"Hedge","asml.us":"Convexity","copx.us":"Cyclical"}

final_w = final_metrics["10Y"]["weights"]
for i, (t, w) in enumerate(sorted(final_w.items(), key=lambda x: x[1], reverse=True), 1):
    bg = ALT if i%2==0 else WHT
    curr = CURRENT_W.get(t)
    delta = round((w - curr)*100, 1) if curr else None
    delta_str = f"{delta:+.1f}pp" if delta is not None else "NEW"
    delta_bg  = GRN if delta and delta > 0 else (RED if delta and delta < 0 else "FFE0CC")
    cel(ws4, ri, 1, t,                    bg=bg)
    cel(ws4, ri, 2, NAMES.get(t,t),      bg=bg, ha="left")
    cel(ws4, ri, 3, ROLES_ALL.get(t,""), bg=bg)
    cel(ws4, ri, 4, f"{w*100:.1f}%",     bg=bg, bold=True)
    cel(ws4, ri, 5, delta_str,           bg=delta_bg, bold=True)
    cel(ws4, ri, 6, "",                  bg=bg)
    ws4.row_dimensions[ri].height = 22; ri += 1

# Comparison vs baseline
ri += 2
ws4.merge_cells(f"A{ri}:F{ri}")
c = ws4.cell(row=ri, column=1, value="Metrics: Final 10-pos vs Original 10-pos (10Y)")
c.font = Font(bold=True, color=WHT, size=11); c.fill = PatternFill("solid", fgColor=MID)
c.alignment = Alignment(horizontal="center", vertical="center")
ws4.row_dimensions[ri].height = 26; ri += 1

for ci, h in enumerate(["Metric","Original (Gold uncapped)",f"Final (35% cap + {cand_names})","Change","",""], 1):
    hdr(ws4, ri, ci, h)
ws4.row_dimensions[ri].height = 24; ri += 1

orig = baselines["10Y"]
fin  = final_metrics["10Y"]
comp_rows = [
    ("Sharpe",     orig["sharpe"],   fin["sharpe"],   lambda a,b: f"{b-a:+.3f}"),
    ("Ann Return", orig["ann_ret"],  fin["ann_ret"],  lambda a,b: f"{b-a:+.2f}pp"),
    ("Max DD",     orig["max_dd"],   fin["max_dd"],   lambda a,b: f"{b-a:+.2f}pp"),
    ("Calmar",     orig["calmar"],   fin["calmar"],   lambda a,b: f"{b-a:+.3f}"),
    ("Total Ret",  orig["total"],    fin["total"],    lambda a,b: f"{b-a:+.1f}pp"),
]
for i, (label, ov, fv, dfn) in enumerate(comp_rows):
    bg = ALT if i%2==0 else WHT
    diff = dfn(ov, fv)
    diff_v = float(diff.replace("pp","").replace("+",""))
    better = diff_v > 0 if label != "Max DD" else diff_v < 0
    cel(ws4, ri, 1, label,   bg=GREY, bold=True, ha="left")
    cel(ws4, ri, 2, str(ov), bg=bg)
    cel(ws4, ri, 3, str(fv), bg=bg)
    cel(ws4, ri, 4, diff,    bg=GRN if better else RED, bold=True)
    cel(ws4, ri, 5, "", bg=bg); cel(ws4, ri, 6, "", bg=bg)
    ws4.row_dimensions[ri].height = 22; ri += 1

cols(ws4, [10,18,12,14,16,8])

wb.save(REPORT)
print(f"\n  Report saved: {REPORT}")
print("=" * 65)
